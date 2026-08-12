#!/usr/bin/env python3
"""
TCM-PA – Downloader de Licitações
Adaptável a qualquer município/órgão e qualquer ano.

Uso:
    python tcmpa_licitacoes.py              ← menu interativo
    python tcmpa_licitacoes.py --ajuda      ← mostra IDs de municípios/órgãos
"""

import os, re, io, sys, time, json, requests
from bs4 import BeautifulSoup
from urllib.parse import urljoin, urlparse, parse_qs, urlencode
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from datetime import datetime

# ═══════════════════════════════════════════════════════════════════════════════
#  CONSTANTES FIXAS DO SISTEMA TCM-PA
# ═══════════════════════════════════════════════════════════════════════════════

BASE_URL  = "https://www.tcm.pa.gov.br"
MURAL_BASE = "https://www.tcm.pa.gov.br/mural-de-licitacoes/licitacoes/listagem"
DELAY     = 1.5   # segundos entre requisições

HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
        "AppleWebKit/537.36 (KHTML, like Gecko) "
        "Chrome/124.0.0.0 Safari/537.36"
    ),
    "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
    "Accept-Language": "pt-BR,pt;q=0.9",
}

_EXT_MAP = {
    "application/pdf": ".pdf",
    "application/msword": ".doc",
    "application/vnd.openxmlformats-officedocument.wordprocessingml.document": ".docx",
    "application/vnd.ms-excel": ".xls",
    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet": ".xlsx",
    "application/zip": ".zip",
    "text/plain": ".txt",
}

_S3_PATTERN = re.compile(r"/arquivo-s3/", re.I)
_FILE_EXT   = re.compile(r"\.(pdf|doc|docx|xls|xlsx|zip|rar|odt|rtf)(\?|#|$)", re.I)
_NAV_SKIP   = re.compile(
    r"(javascript:|mailto:|^#|/mural-de-licitacoes/licitacoes/listagem|[?&](page|tab|filtro|search)=)",
    re.I
)


# ═══════════════════════════════════════════════════════════════════════════════
#  OCR — EXTRAÇÃO DO NÚMERO DA LICITAÇÃO
#  Captura o número da LICITAÇÃO (Pregão, Dispensa, Tomada de Preços...),
#  NUNCA o número do PROCESSO ADMINISTRATIVO.
# ═══════════════════════════════════════════════════════════════════════════════


# ═══════════════════════════════════════════════════════════════════════════════
#  PADRÕES
# ═══════════════════════════════════════════════════════════════════════════════

# Modalidades que identificam um número de LICITAÇÃO
_MOD_KW = (
    r"(?:"
    r"PREG[ÃA]O\s+ELETR[ÔO]NICO(?:\s+SRP)?"
    r"|PREG[ÃA]O\s+PRESENCIAL(?:\s+SRP)?"
    r"|PREG[ÃA]O(?:\s+SRP)?"
    r"|TOMADA\s+DE\s+PRE[ÇC]OS"
    r"|CARTA\s+CONVITE|CONVITE"
    r"|CONCORR[ÊE]NCIA(?:\s+P[ÚU]BLICA)?"
    r"|CONCURSO"
    r"|LEIL[ÃA]O"
    r"|DISPENSA(?:\s+DE\s+LICITA[ÇC][ÃA]O)?(?:\s+ELETR[ÔO]NICA)?"
    r"|INEXIGIBILIDADE(?:\s+DE\s+LICITA[ÇC][ÃA]O)?"
    r"|CHAMADA\s+P[ÚU]BLICA"
    r"|CHAMAMENTO\s+P[ÚU]BLICO"
    r"|CREDENCIAMENTO"
    r"|DI[ÁA]LOGO\s+COMPETITIVO"
    r"|ADES[ÃA]O\s+(?:A|À)\s+ATA"
    r"|LICITA[ÇC][ÃA]O"
    r")"
)

# Formatos de número aceitos: 024/2023, 24-2023, 9/2023-00037, 001.2023
_NUM = r"(\d{1,6}\s*[./\-]\s*\d{2,4}(?:\s*[\-/]\s*\d{1,6})?)"

# Separador entre a modalidade e o número (Nº, N°, No, nr, #, :, espaço)
_SEP = r"(?:\s*(?:N[º°ᵒoO\.]*|NUM(?:ERO)?|NR|#|:)?\s*)"

# Padrão principal: MODALIDADE [Nº] NUMERO
_RE_MOD_NUM = re.compile(_MOD_KW + _SEP + _NUM, re.I)

# Termos que indicam que o número NÃO é da licitação
_RE_PROCESSO = re.compile(
    r"(?:PROCESSO(?:\s+(?:ADMINISTRATIVO|LICITAT[ÓO]RIO|DE\s+COMPRA))?"
    r"|PROTOCOLO|EMPENHO|CONTRATO|ATA\s+DE\s+REGISTRO|ARP"
    r"|CNPJ|CPF|LEI|DECRETO|PORTARIA|OF[ÍI]CIO|MEMORANDO|CONV[ÊE]NIO)"
    r"[^\n]{0,25}$",
    re.I
)

_RE_ANO = re.compile(r"(20\d{2})")


def _normalizar_numero(bruto: str) -> str:
    """'024 / 2023' → '024/2023'"""
    n = re.sub(r"\s+", "", bruto)
    n = n.replace(".", "/")
    return n


def limpar_ruido_ocr(texto: str) -> str:
    """
    Corrige ruídos típicos de OCR que atrapalham a captura do número:
      - dígitos espaçados:  '0 2 4 / 2 0 2 3'  →  '024/2023'
      - letras confundidas com dígitos dentro de números: O→0, l/I→1, S→5
      - 'N o' / 'N °' / 'N.º' normalizados para 'Nº'
    """
    if not texto:
        return texto

    t = texto

    # 'N o 024' / 'N.º 024' / 'N ° 024'  →  'Nº 024'
    t = re.sub(r"\bN\s*[.ºo°]{1,2}\s*", "Nº ", t, flags=re.I)

    # Junta sequências de dígitos separados por espaço simples:
    # '0 2 4 / 2 0 2 3' → '024/2023'  (aplica repetidamente)
    for _ in range(6):
        novo = re.sub(r"(?<=\d) (?=\d)", "", t)
        novo = re.sub(r"(?<=\d) ?([/\-.]) ?(?=\d)", r"\1", novo)
        if novo == t:
            break
        t = novo

    # Corrige letras dentro de blocos claramente numéricos (ex.: 'O24/2O23')
    def _fix(m):
        bloco = m.group(0)
        tab   = str.maketrans({"O": "0", "o": "0", "l": "1", "I": "1",
                               "S": "5", "B": "8"})
        return bloco.translate(tab)

    t = re.sub(r"\b(?=[\dOolISB]{1,6}[/\-][\dOolISB]{2,4}\b)[\dOolISB/\-]+\b", _fix, t)

    return t


def _tem_processo_antes(texto: str, pos: int, janela: int = 60) -> bool:
    """
    Verifica se o número é, na verdade, de PROCESSO/CONTRATO/ARP.

    Só considera o texto da MESMA LINHA imediatamente antes do match —
    um 'PROCESSO Nº ...' numa linha anterior não invalida o número da
    licitação da linha seguinte.
    """
    inicio_linha = texto.rfind("\n", 0, pos) + 1        # começo da linha atual
    inicio       = max(inicio_linha, pos - janela)
    trecho       = texto[inicio:pos]
    return bool(_RE_PROCESSO.search(trecho))


def extrair_numero_licitacao(texto: str, ano_ref: int | str = None,
                             numero_mural: str = "") -> dict:
    """
    Encontra o número da LICITAÇÃO no texto.

    Retorna:
      {
        'numero':      '024/2023'   ← melhor candidato
        'modalidade':  'PREGÃO ELETRÔNICO'
        'confianca':   0-100
        'candidatos':  [(numero, score, modalidade), ...]
      }
    """
    if not texto:
        return {"numero": "", "modalidade": "", "confianca": 0, "candidatos": []}

    # Normaliza espaços e corrige ruídos de OCR antes de procurar
    txt = re.sub(r"[ \t]+", " ", texto)
    txt = limpar_ruido_ocr(txt)

    candidatos = {}   # numero_normalizado -> dict(score, modalidade, ocorrencias)

    for m in _RE_MOD_NUM.finditer(txt):
        bruto      = m.group(1)
        modalidade = m.group(0)[: m.start(1) - m.start()].strip(" :nºN°.#")
        numero     = _normalizar_numero(bruto)

        # Rejeita se vier precedido de PROCESSO/CONTRATO/etc.
        if _tem_processo_antes(txt, m.start()):
            continue

        # Rejeita números implausíveis (ano fora de faixa)
        ano_m = _RE_ANO.search(numero)
        if ano_m:
            ano_num = int(ano_m.group(1))
            if not (2000 <= ano_num <= 2100):
                continue

        info = candidatos.setdefault(numero, {
            "score": 0, "modalidade": modalidade, "ocorrencias": 0
        })
        info["ocorrencias"] += 1

        # ── Pontuação ────────────────────────────────────────────────────────
        info["score"] += 10                       # casou modalidade + número
        if m.start() < 1500:
            info["score"] += 8                    # aparece no topo do documento
        if ano_ref and ano_m and str(ano_ref) == ano_m.group(1):
            info["score"] += 12                   # ano bate com o da licitação
        if numero_mural:
            so_dig_a = re.sub(r"\D", "", numero)
            so_dig_b = re.sub(r"\D", "", numero_mural)
            if so_dig_a and so_dig_a == so_dig_b:
                info["score"] += 25               # idêntico ao número do mural
            elif so_dig_a and so_dig_b and (so_dig_a in so_dig_b or so_dig_b in so_dig_a):
                info["score"] += 10               # contido/contém

    if not candidatos:
        return {"numero": "", "modalidade": "", "confianca": 0, "candidatos": []}

    # Frequência conta, mas com peso menor que os sinais acima
    for num, info in candidatos.items():
        info["score"] += min(info["ocorrencias"], 5) * 2

    ordenados = sorted(candidatos.items(), key=lambda kv: kv[1]["score"], reverse=True)
    melhor_num, melhor = ordenados[0]

    confianca = min(100, melhor["score"] * 2)

    return {
        "numero":     melhor_num,
        "modalidade": melhor["modalidade"],
        "confianca":  confianca,
        "candidatos": [(n, i["score"], i["modalidade"]) for n, i in ordenados[:5]],
    }


# ═══════════════════════════════════════════════════════════════════════════════
#  EXTRAÇÃO DE TEXTO (nativo + OCR)
# ═══════════════════════════════════════════════════════════════════════════════

def extrair_texto_pdf(caminho: str, max_paginas: int = 3,
                      usar_ocr: bool = True, dpi: int = 300,
                      min_chars: int = 250) -> tuple:
    """
    Extrai texto de um PDF.
      1. Tenta a camada de texto nativa (rápido, exato)
      2. Se vier pouco texto (PDF escaneado), aplica OCR

    Retorna (texto, metodo) — metodo em {'nativo','ocr','nativo+ocr','erro'}
    """
    try:
        import fitz  # PyMuPDF
    except ImportError:
        return "", "erro:pymupdf-ausente"

    texto  = ""
    metodo = "nativo"

    try:
        doc = fitz.open(caminho)
        n   = min(len(doc), max_paginas)

        for i in range(n):
            texto += doc[i].get_text() + "\n"

        # Pouco texto → provavelmente escaneado → OCR
        if usar_ocr and len(texto.strip()) < min_chars:
            texto_ocr = _ocr_paginas(doc, n, dpi)
            if texto_ocr.strip():
                metodo = "ocr" if len(texto.strip()) < 40 else "nativo+ocr"
                texto += "\n" + texto_ocr

        doc.close()
    except Exception as e:
        return texto, f"erro:{e}"

    return texto, metodo


def _ocr_paginas(doc, n_paginas: int, dpi: int = 300) -> str:
    """Rasteriza e aplica OCR (PyMuPDF + pytesseract, sem depender de Poppler)."""
    try:
        import pytesseract
        from PIL import Image, ImageOps
    except ImportError:
        return ""

    texto = ""
    for i in range(n_paginas):
        try:
            pix = doc[i].get_pixmap(dpi=dpi)
            img = Image.open(io.BytesIO(pix.tobytes("png")))

            # Pré-processamento: cinza + autocontraste (ajuda muito em digitalizações)
            img = img.convert("L")
            img = ImageOps.autocontrast(img)

            # psm 6 = bloco uniforme de texto; funciona bem em editais/capas
            texto += pytesseract.image_to_string(img, lang="por", config="--psm 6") + "\n"
        except Exception:
            continue
    return texto


def _mesmo_numero(a: str, b: str) -> bool:
    """Compara dois números de licitação ignorando zeros à esquerda e separadores."""
    if not a or not b:
        return False
    pa = [p.lstrip("0") or "0" for p in re.split(r"[^\d]+", a) if p]
    pb = [p.lstrip("0") or "0" for p in re.split(r"[^\d]+", b) if p]
    if not pa or not pb:
        return False
    # Compara os dois primeiros grupos (número e ano), que é o que identifica
    return pa[:2] == pb[:2]


def descobrir_numero_nos_arquivos(pasta: str, ano_ref=None, numero_mural: str = "",
                                  max_arquivos: int = 6, usar_ocr: bool = True) -> dict:
    """
    Varre os PDFs de uma pasta e devolve o melhor número de licitação encontrado.
    Prioriza arquivos cujo nome sugere edital/aviso/capa.
    """
    if not os.path.isdir(pasta):
        return {"numero": "", "confianca": 0, "arquivo": "", "metodo": ""}

    pdfs = [f for f in os.listdir(pasta) if f.lower().endswith(".pdf")]
    if not pdfs:
        return {"numero": "", "confianca": 0, "arquivo": "", "metodo": ""}

    # Ordem de prioridade: edital > aviso > termo > ata > resto
    def prioridade(nome: str) -> int:
        n = nome.lower()
        if "edital"        in n: return 0
        if "aviso"         in n: return 1
        if "ato de abertura" in n or "abertura" in n: return 2
        if "termo de refer" in n: return 3
        if "ata"           in n: return 4
        if "contrato"      in n: return 5
        return 6

    pdfs.sort(key=prioridade)

    melhor = {"numero": "", "confianca": 0, "arquivo": "", "metodo": ""}

    for nome in pdfs[:max_arquivos]:
        caminho       = os.path.join(pasta, nome)
        texto, metodo = extrair_texto_pdf(caminho, usar_ocr=usar_ocr)
        if not texto.strip():
            continue

        r = extrair_numero_licitacao(texto, ano_ref, numero_mural)
        if r["numero"] and r["confianca"] > melhor["confianca"]:
            melhor = {
                "numero":     r["numero"],
                "confianca":  r["confianca"],
                "arquivo":    nome,
                "metodo":     metodo,
                "modalidade": r.get("modalidade", ""),
            }
            # Confiança alta o bastante: não precisa varrer o resto
            if melhor["confianca"] >= 90:
                break

    return melhor


# ═══════════════════════════════════════════════════════════════════════════════

# ═══════════════════════════════════════════════════════════════════════════════
#  NORMALIZAÇÃO DE MODALIDADES
#  Cada regra é (padrão_regex, nome_oficial).
#  A ordem importa: regras mais específicas primeiro.
# ═══════════════════════════════════════════════════════════════════════════════

_MODALIDADES = [
    # ── Adesão / Carona ───────────────────────────────────────────────────────
    (r"ades.o.+(ata|srp|registro)",         "Adesão a Ata de Registro de Preço"),
    (r"carona",                              "Carona"),

    # ── Registro de Preços originários ───────────────────────────────────────
    (r"(inten.+registro|irp)",               "Intenção de Registro de Preços"),
    (r"srp.+chamamento|chamamento.+srp|registro.+pre.+chamamento",
                                             "Registro de Preços Originário de Chamamento Público"),
    (r"(srp|registro.+pre.+).+eletr|pregão.+eletr.+(srp|registro)",
                                             "Registro de Preços Originário de Pregão Eletrônico"),
    (r"(srp|registro.+pre.+).+presencial|pregão.+presencial.+(srp|registro)",
                                             "Registro de Preços Originário de Pregão Presencial"),

    # ── Pregão ────────────────────────────────────────────────────────────────
    (r"preg.+eletr",                         "Pregão Eletrônico"),
    (r"preg.+presencial",                    "Pregão Presencial"),
    (r"preg",                                "Pregão Eletrônico"),   # genérico → eletrônico

    # ── Dispensa / Inexigibilidade / Contratação ──────────────────────────────
    (r"dispensa",                            "Dispensa de Licitação"),
    (r"inexigib",                            "Inexigibilidade de Licitação"),
    (r"contrata.+diret",                     "Contratação Direta"),

    # ── Demais modalidades ────────────────────────────────────────────────────
    (r"concorr",                             "Concorrência"),
    (r"concurso",                            "Concurso"),
    (r"convite",                             "Convite"),
    (r"chamada.+p.blica|chamamento",         "Chamada Pública"),
    (r"di.logo.+compet",                     "Diálogo Competitivo"),
    (r"leil.o",                              "Leilão"),
    (r"tomada.+pre",                         "Tomada de Preços"),
    (r"credenciamento",                      "Credenciamento"),
]

def normalizar_modalidade(texto: str) -> str:
    """
    Recebe o texto bruto da modalidade vindo do mural (ex: 'PREGÃO ELETRÔNICO',
    'DISPENSA, ART. 75, INCISOS I', 'ADESÃO A ATA DE SRP') e retorna o
    nome oficial padronizado da lista acima.
    Se não bater com nenhuma regra, devolve o texto original limpo.
    """
    t = texto.strip().lower()
    for pattern, nome in _MODALIDADES:
        if re.search(pattern, t, re.I):
            return nome
    # fallback: devolve capitalizado mas sem código/artigo
    limpo = re.sub(r"[\.,]\s*(art|inc|lei|§).*$", "", texto, flags=re.I).strip()
    return limpo.title() if limpo else texto

# ═══════════════════════════════════════════════════════════════════════════════
#  ✏  CONFIGURAÇÃO  –  edite apenas estas 2 linhas
# ═══════════════════════════════════════════════════════════════════════════════

LINK_MURAL = (
    "https://www.tcm.pa.gov.br/mural-de-licitacoes/licitacoes/listagem"
    "?LINCEMVWLICITACOESSearch%5BLEGISLACAO_ID%5D=&"
    "LINCEMVWLICITACOESSearch%5BNUMERO_DOCUMENTO%5D=&"
    "LINCEMVWLICITACOESSearch%5BMODALIDADE_ID%5D=90&"
    "LINCEMVWLICITACOESSearch%5BTIPO_ID%5D=&"
    "LINCEMVWLICITACOESSearch%5BOBJETO%5D=&"
    "LINCEMVWLICITACOESSearch%5BDATA_ABERTURA%5D=&"
    "LINCEMVWLICITACOESSearch%5BDATA_PUBLICACAO%5D=&"
    "LINCEMVWLICITACOESSearch%5BID_MUNICIPIO%5D=21&"
    "LINCEMVWLICITACOESSearch%5BORGAO_ID%5D=21001&"
    "LINCEMVWLICITACOESSearch%5BSTATUS_ID%5D=&"
    "LINCEMVWLICITACOESSearch%5BVL_REFERENCIADO%5D=&"
    "LINCEMVWLICITACOESSearch%5BVL_ADJUDICADO%5D="
    "&per-page=30"
)

PASTA_SAIDA = r"C:\Downloads"   # arquivos salvos diretamente aqui

# ── Nome da pasta da entidade ────────────────────────────────────────────────
# Deixe "" para o script derivar automaticamente do órgão
# (PREFEITURA MUNICIPAL DE X → "PM X" | CÂMARA MUNICIPAL DE X → "CM X")
NOME_ENTIDADE = "PM Cametá"

# ── Faixa de anos ────────────────────────────────────────────────────────────
# ANO_MAXIMO = 2023 → baixa 2023 e anos anteriores ("de 2023 para trás")
# Use None em qualquer um para não limitar aquele extremo.
ANO_MAXIMO = 2026
ANO_MINIMO = 2023

# ── OCR ──────────────────────────────────────────────────────────────────────
# Lê os PDFs baixados e extrai o NÚMERO DA LICITAÇÃO (não o do processo).
OCR_ATIVO        = True
OCR_MAX_ARQUIVOS = 6      # quantos PDFs analisar por licitação
OCR_MAX_PAGINAS  = 3      # páginas por PDF (o número fica na capa/cabeçalho)
OCR_DPI          = 300    # 300 é bom equilíbrio; 400 melhora digitalizações ruins

# ═══════════════════════════════════════════════════════════════════════════════
#  LEITURA AUTOMÁTICA DO LINK
# ═══════════════════════════════════════════════════════════════════════════════

def _parse_link(url: str) -> dict:
    """
    Lê o link do mural e extrai automaticamente:
      - ID do município
      - ID do órgão
      - Ano (EXERCICIO)
      - Nome legível para nomear pastas/excel
    Funciona com qualquer link copiado do mural TCM-PA,
    independente da ordem ou codificação dos parâmetros.
    """
    from urllib.parse import unquote
    # Normaliza %5B → [ e %5D → ]
    url_dec = unquote(url)
    qs      = parse_qs(urlparse(url_dec).query)

    # Chaves podem aparecer como  LINCEMVWLICITACOESSearch[CAMPO]
    def get(campo):
        for k, v in qs.items():
            if campo.upper() in k.upper():
                val = v[0].strip()
                return val if val else ""
        return ""

    id_municipio = get("ID_MUNICIPIO")
    id_orgao     = get("ORGAO_ID")
    ano_str      = get("EXERCICIO")

    # Ano: tenta extrair do parâmetro ou da URL literal.
    # Se não houver EXERCICIO na URL, ano = None → baixa TODOS os anos.
    if not ano_str:
        m = re.search(r"EXERCICIO[%5D=]+(\d{4})", url, re.I)
        ano_str = m.group(1) if m else ""

    ano = int(ano_str) if ano_str.isdigit() else None

    # Nome do município: busca o texto da página de municípios ou usa o ID
    nome = f"municipio_{id_municipio}" if id_municipio else "tcmpa"

    return {
        "id_municipio": id_municipio,
        "id_orgao":     id_orgao,
        "ano":          ano,           # None = todos os anos
        "nome":         nome,
        "url_base":     url,           # URL original — usada como base para paginação
    }

def _enriquecer_nome(cfg: dict, session) -> dict:
    """
    Descobre o nome real do município lendo a coluna 'Município' da tabela
    do mural JÁ FILTRADA pelo ID correto — garante que o nome bate com o filtro.
    Usa a URL completa (com todos os filtros) para não pegar linha errada.
    """
    resp = safe_get(session, build_mural_url(cfg, 1))
    if not resp:
        return cfg

    soup  = BeautifulSoup(resp.text, "html.parser")
    table = soup.find("table")
    if not table:
        return cfg

    id_mun = cfg.get("id_municipio", "")
    id_org = cfg.get("id_orgao", "")

    for row in table.find_all("tr")[1:]:
        cols = row.find_all("td")
        if len(cols) < 9:
            continue

        municipio = cols[7].get_text(strip=True)
        orgao     = cols[8].get_text(strip=True)

        if not municipio:
            continue

        # Só aceita se o link desta linha aponta para o município/órgão correto
        link = row.find("a", href=True)
        href = link["href"] if link else ""

        # Valida pelo ID do município na URL da licitação
        if id_mun and f"ID_MUNICIPIO%5D={id_mun}" not in href and f"municipio/{id_mun}" not in href:
            # Sem ID na URL da linha — valida pelo ID do órgão
            if id_org and id_org not in orgao and id_org not in href:
                continue  # linha de outro município, pula

        cfg["nome"] = municipio
        print(f"  [i] Município identificado: {municipio}")
        if orgao:
            cfg["nome_orgao"] = orgao
            print(f"  [i] Órgão identificado: {orgao}")
        break

    # ── Nome da entidade para a pasta ────────────────────────────────────────
    cfg["entidade"] = NOME_ENTIDADE.strip() or derivar_nome_entidade(
        cfg.get("nome_orgao", ""), cfg.get("nome", "")
    )
    print(f"  [i] Entidade (pasta): {cfg['entidade']}")

    return cfg


def derivar_nome_entidade(nome_orgao: str, nome_municipio: str) -> str:
    """
    'PREFEITURA MUNICIPAL DE SANTA IZABEL DO PARÁ'  → 'PM Santa Izabel do Pará'
    'CÂMARA MUNICIPAL DE SANTA IZABEL DO PARÁ'      → 'CM Santa Izabel do Pará'
    Sem órgão reconhecível, usa o nome do município.
    """
    org = (nome_orgao or "").strip()
    # Remove código numérico do início: '068002 - CÂMARA MUNICIPAL DE ...'
    org = re.sub(r"^\s*\d+\s*[-–]\s*", "", org)

    if re.search(r"c[âa]mara", org, re.I):
        sigla = "CM"
    elif re.search(r"prefeitura", org, re.I):
        sigla = "PM"
    elif re.search(r"fundo", org, re.I):
        sigla = "FM"
    else:
        sigla = ""

    # Pega o que vem depois de "DE/DO/DA"
    m = re.search(r"\b(?:DE|DO|DA)\s+(.+)$", org, re.I)
    local = m.group(1).strip() if m else (nome_municipio or org)
    local = " ".join(p.capitalize() if p.lower() not in ("de","do","da","dos","das")
                     else p.lower() for p in local.split())

    return f"{sigla} {local}".strip() if sigla else (local or "entidade")

def build_mural_url(cfg: dict, page: int = 1) -> str:
    """
    Constrói a URL de cada página do mural.
    Remove page= e per-page= da URL base e reconstrói limpo.
    """
    # Remove parâmetros de paginação existentes
    base = re.sub(r"[&?]page=\d+", "", cfg["url_base"])
    base = re.sub(r"[&?]per-page=\d+", "", base)
    base = base.rstrip("&").rstrip("?")

    sep = "&" if "?" in base else "?"
    url = f"{base}{sep}per-page=30"
    if page > 1:
        url += f"&page={page}"
    return url

# ═══════════════════════════════════════════════════════════════════════════════
#  UTILITÁRIOS
# ═══════════════════════════════════════════════════════════════════════════════

def sanitize_pasta(name: str) -> str:
    """Para nomes de pasta: remove caracteres inválidos, mantém espaços."""
    name = re.sub(r'[<>:"/\\|?*\x00-\x1f]', '', name)
    name = re.sub(r'\s+', ' ', name.strip())
    return name[:80]

def sanitize(name: str) -> str:
    """Para nomes de arquivo: remove caracteres inválidos, mantém espaços."""
    name = re.sub(r'[<>:"/\\|?*\x00-\x1f]', '', name)
    name = re.sub(r'\s+', ' ', name.strip())
    name = name.strip(". ")
    return name[:100]

def make_session() -> requests.Session:
    s = requests.Session()
    s.headers.update(HEADERS)
    return s

def safe_get(session, url, retries=3):
    for attempt in range(retries):
        try:
            r = session.get(url, timeout=30)
            r.raise_for_status()
            return r
        except Exception as e:
            print(f"  [!] Tentativa {attempt+1}/{retries}: {e}")
            time.sleep(2 ** attempt)
    return None

# ═══════════════════════════════════════════════════════════════════════════════
#  FILTRO DE ANO
# ═══════════════════════════════════════════════════════════════════════════════

def _ano_da_licitacao(lic: dict):
    """Descobre o ano da licitação pela data de abertura, publicação ou número."""
    for campo in ("data_abertura", "data_publicacao"):
        m = re.search(r"\d{2}/\d{2}/(\d{4})", lic.get(campo, "").strip())
        if m:
            return int(m.group(1))
    m = re.search(r"[/\-](20\d{2})[\-/]", lic.get("numero", ""))
    return int(m.group(1)) if m else None


def is_ano_filtro(lic: dict, ano=None) -> bool:
    """
    Aceita a licitação conforme a faixa ANO_MINIMO..ANO_MAXIMO.

    Com ANO_MAXIMO = 2023 e ANO_MINIMO = None, aceita 2023 e anos anteriores
    ("de 2023 para trás"). Se o ano não for identificável, aceita por segurança.
    """
    ano_lic = _ano_da_licitacao(lic)
    if ano_lic is None:
        return True                      # indeterminado: não descarta

    if ANO_MAXIMO is not None and ano_lic > ANO_MAXIMO:
        return False
    if ANO_MINIMO is not None and ano_lic < ANO_MINIMO:
        return False
    return True

# ═══════════════════════════════════════════════════════════════════════════════
#  COLETA DO MURAL
# ═══════════════════════════════════════════════════════════════════════════════

def get_all_licitacoes(session, cfg: dict) -> list:
    licitacoes      = []
    page            = 1
    numeros_vistos  = set()   # números de licitação já coletados (detecta repetição)
    paginas_vazias  = 0       # páginas consecutivas sem NADA novo
    MAX_PAGINAS     = 200     # trava de segurança absoluta

    while page <= MAX_PAGINAS:
        url = build_mural_url(cfg, page)
        print(f"\n[*] Página {page}...")
        resp = safe_get(session, url)
        if not resp:
            print("  [!] Sem resposta. Encerrando.")
            break

        soup  = BeautifulSoup(resp.text, "html.parser")
        table = soup.find("table")
        if not table:
            print("  [!] Tabela não encontrada. Fim.")
            break

        # ── Total de páginas pelo paginador «1 2 3...10» ──────────────────────
        total_pages = None
        for tag in soup.find_all(["ul", "div", "nav"], class_=re.compile(r"pag", re.I)):
            nums = [int(a.get_text(strip=True)) for a in tag.find_all("a")
                    if a.get_text(strip=True).isdigit()]
            if nums:
                total_pages = max(nums)
                break
        if total_pages and page > total_pages:
            print(f"  [.] Página {page} > total de {total_pages}. Fim.")
            break

        rows = table.find_all("tr")
        novos_nesta_pagina = 0

        for row in rows[1:]:
            cols = row.find_all("td")
            if len(cols) < 5:
                continue
            link_tag = next(
                (col.find("a", href=True)
                 for col in cols
                 if col.find("a", href=True)
                 and "/mural-de-licitacoes/" in col.find("a")["href"]),
                None
            )
            if not link_tag:
                continue

            numero     = link_tag.get_text(strip=True)
            detail_url = urljoin(BASE_URL, link_tag["href"])

            # ── Chave única da licitação (número + ID da ficha) ───────────────
            chave = f"{numero}|{detail_url}"
            if chave in numeros_vistos:
                continue  # já coletada → ignora (evita repetir página)
            numeros_vistos.add(chave)

            ct = [c.get_text(" ", strip=True) for c in cols]
            lic = {
                "numero":              numero,
                "url":                 detail_url,
                "legislacao":          ct[0]  if len(ct) >  0 else "",
                "modalidade_original": ct[2]  if len(ct) >  2 else "",
                "modalidade":          normalizar_modalidade(ct[2] if len(ct) > 2 else ""),
                "tipo":                ct[3]  if len(ct) >  3 else "",
                "objeto":              ct[4]  if len(ct) >  4 else "",
                "data_abertura":       ct[5]  if len(ct) >  5 else "",
                "data_publicacao":     ct[6]  if len(ct) >  6 else "",
                "municipio":           ct[7]  if len(ct) >  7 else "",
                "orgao":               ct[8]  if len(ct) >  8 else "",
                "status":              ct[9]  if len(ct) >  9 else "",
                "vl_referenciado":     ct[10] if len(ct) > 10 else "",
                "vl_adjudicado":       ct[11] if len(ct) > 11 else "",
            }

            if not is_ano_filtro(lic, cfg["ano"]):
                print(f"  [~] Ano diferente, ignorado: {numero}")
                continue

            licitacoes.append(lic)
            novos_nesta_pagina += 1
            print(f"  [+] {numero} | {lic['modalidade']} | {lic['status']}")

        # ── Critério de parada infalível ──────────────────────────────────────
        # Se a página não trouxe NENHUMA licitação nova (todas repetidas ou vazia),
        # significa que o site está repetindo a última página → fim.
        if novos_nesta_pagina == 0:
            paginas_vazias += 1
            print(f"  [.] Nenhuma licitação nova nesta página ({paginas_vazias}).")
            if paginas_vazias >= 1:
                print("  [.] Fim da paginação (sem itens novos).")
                break
        else:
            paginas_vazias = 0

        page += 1
        time.sleep(DELAY)

    if page > MAX_PAGINAS:
        print(f"  [!] Limite de segurança de {MAX_PAGINAS} páginas atingido.")

    rotulo_ano = cfg['ano'] if cfg['ano'] is not None else "todos os anos"
    print(f"\n[✓] Total de licitações ({rotulo_ano}): {len(licitacoes)}")
    return licitacoes

# ═══════════════════════════════════════════════════════════════════════════════
#  ABAS DA LICITAÇÃO
# ═══════════════════════════════════════════════════════════════════════════════

def get_licitacao_id(url: str) -> str | None:
    m = re.search(r"/ficha/(\d+)", url)
    if m:
        return m.group(1)
    qs = parse_qs(urlparse(url).query)
    if "id" in qs:
        return qs["id"][0]
    m = re.search(r"/(\d+)(?:[#?]|$)", url)
    if m:
        return m.group(1)
    return None

def build_tab_urls(licitacao_url: str, soup: BeautifulSoup, lic_id: str) -> dict:
    abas = {}

    for a in soup.find_all("a", href=True):
        href     = a["href"]
        texto    = a.get_text(strip=True).lower()
        fragment = href.split("#")[1].lower() if "#" in href else ""
        clean    = href.split("#")[0]
        full_url = urljoin(licitacao_url, clean) if clean else licitacao_url

        if fragment:
            if re.search(r"doc", fragment):
                abas.setdefault("doc_base", full_url)
            elif re.search(r"contrato", fragment) and "aditivo" not in fragment:
                abas.setdefault("cont_base", full_url)

        if re.match(r"documentos?\s*\d*$", texto):
            abas.setdefault("doc_link", full_url)
        elif re.match(r"contratos?\s*\d*$", texto) and "aditivo" not in texto:
            abas.setdefault("cont_link", full_url)

        if re.search(r"[?&]tab=doc", href, re.I):
            abas["doc_tab"] = full_url
        if re.search(r"[?&]tab=contrato", href, re.I) and "aditivo" not in href.lower():
            abas["cont_tab"] = full_url

    for tag in soup.find_all(attrs={"data-url": True}):
        du = tag["data-url"]
        if re.search(r"doc", du, re.I):
            abas.setdefault("doc_data", urljoin(licitacao_url, du))
        elif re.search(r"contrato", du, re.I) and "aditivo" not in du.lower():
            abas.setdefault("cont_data", urljoin(licitacao_url, du))

    result = {}

    # Documentos: prioridade tab > link > data > candidatos derivados do ID
    result["documentos"] = (abas.get("doc_tab") or abas.get("doc_link")
                            or abas.get("doc_data") or abas.get("doc_base"))
    if not result["documentos"] and lic_id:
        result["documentos"] = f"{BASE_URL}/mural-de-licitacoes/licitacoes/ficha/{lic_id}"
        result["doc_candidatos"] = [
            f"{BASE_URL}/mural-de-licitacoes/licitacoes/documentos/{lic_id}",
            f"{BASE_URL}/mural-de-licitacoes/licitacoes/ficha/{lic_id}?tab=documentos",
        ]

    # Contratos
    result["contratos"] = (abas.get("cont_tab") or abas.get("cont_link")
                           or abas.get("cont_data") or abas.get("cont_base"))
    if not result["contratos"] and lic_id:
        result["contratos"] = f"{BASE_URL}/mural-de-licitacoes/licitacoes/ficha/{lic_id}"
        result["cont_candidatos"] = [
            f"{BASE_URL}/mural-de-licitacoes/licitacoes/contratos/{lic_id}",
            f"{BASE_URL}/mural-de-licitacoes/licitacoes/ficha/{lic_id}?tab=contratos",
        ]

    return result

def fetch_tab(session, primary: str, candidates: list = None):
    urls = [primary] + (candidates or [])
    seen = set()
    for url in urls:
        if url in seen or not url:
            continue
        seen.add(url)
        resp = safe_get(session, url)
        if resp and len(resp.content) > 50:
            return resp, url
    return None, None

# ═══════════════════════════════════════════════════════════════════════════════
#  EXTRAÇÃO DE LINKS
# ═══════════════════════════════════════════════════════════════════════════════

def is_file_url(href: str) -> bool:
    return bool(_S3_PATTERN.search(href) or _FILE_EXT.search(href))

def _num_contrato(texto: str) -> str:
    """Extrai o número do contrato de um texto ('Contrato nº 123/2023' → '123/2023')."""
    m = re.search(r"Contrato\s+n[ºo°]\s*([\d/.\-]+)", texto or "", re.I)
    return m.group(1).strip() if m else ""


def extract_contratos(soup: BeautifulSoup, base_url: str) -> list:
    """
    Extrai TODOS os contratos da seção 'Contratos', tratando múltiplos contratos.

    Estrutura real do TCM-PA (cada contrato é um bloco/accordion):
        Contratos2                              ← heading da seção (2 contratos)
        TRANSFORMAT COMÉRCIO E SERVIÇOS LTDA  CNPJ  ← contrato 1
          Contrato nº 20230001
          R$ ...
          VIGÊNCIA INÍCIO ... FIM ...
          CONTRATO → CONTRATO.PDF
          OUTROS DOCUMENTOS → ... (ignorar)
        AUTOCAR COMERCIO DE VEICULOS LTDA  CNPJ      ← contrato 2
          Contrato nº 20230002
          ...
          CONTRATO → CONTRATO.PDF
        Aditivos de Contrato                    ← fim da seção

    Retorna lista de dicts, um por contrato:
        {
          'numero_contrato', 'contratado', 'cnpj', 'valor',
          'vigencia_inicio', 'vigencia_fim',
          'arquivos': [{'url','nome'}, ...]   ← só o(s) CONTRATO.PDF
        }
    """
    # 1. Localiza o heading 'Contratos' do conteúdo (não o do menu de abas)
    headings   = soup.find_all(re.compile(r"h[1-6]"))
    candidatos = [t for t in headings
                  if re.match(r"contratos\s*\d*$", t.get_text(strip=True), re.I)]

    tag_contratos = None
    for cand in reversed(candidatos):
        for elem in cand.find_all_next():
            if elem.name and re.match(r"h[1-6]$", elem.name):
                th = elem.get_text(strip=True)
                if re.search(r"aditivo", th, re.I): break
                if re.match(r"contratos\s*\d*$", th, re.I): break
                continue
            if elem.name == "a" and elem.get("href") and is_file_url(elem["href"]):
                tag_contratos = cand
                break
        if tag_contratos:
            break
    if tag_contratos is None and candidatos:
        tag_contratos = candidatos[-1]

    if tag_contratos is None:
        return []

    # 2. Coleta os elementos da seção (do heading 'Contratos' até 'Aditivos').
    #    IMPORTANTE: usamos apenas nós FOLHA (sem filhos-tag) e os <a>.
    #    find_all_next() devolve pai E filhos — sem esse filtro o mesmo
    #    conteúdo é contado várias vezes, gerando blocos fantasma e
    #    arquivos duplicados.
    elementos = []
    for elem in tag_contratos.find_all_next():
        if elem.name and re.match(r"h[1-6]$", elem.name):
            if re.search(r"aditivo", elem.get_text(strip=True), re.I):
                break            # fim da seção de contratos
            elementos.append(elem)
            continue
        if elem.name == "a":
            elementos.append(elem)
            continue
        # Só folhas: elementos sem nenhuma tag filha
        if elem.name and not elem.find(True):
            elementos.append(elem)

    # 3. Segmenta em blocos, um por contrato.
    #    Marcador: 'Contrato nº XXXX' — sempre presente. NÃO usamos CNPJ,
    #    pois falha com pessoa física (CPF) ou cabeçalho sem documento.
    #    O contratado aparece ANTES do 'Contrato nº', então cada bloco puxa
    #    alguns elementos anteriores (lookback), sem invadir o bloco anterior.
    LOOKBACK = 8

    marcos = []
    for idx, elem in enumerate(elementos):
        texto = elem.get_text(" ", strip=True) if hasattr(elem, "get_text") else ""
        if len(texto) < 150 and re.search(r"Contrato\s+n[ºo°]\s*[\d/.\-]+", texto, re.I):
            marcos.append(idx)

    blocos = []
    for i, ini_idx in enumerate(marcos):
        fim_idx  = marcos[i + 1] if i + 1 < len(marcos) else len(elementos)

        # Lookback: recua para capturar o cabeçalho (contratado + documento),
        # mas PARA ao encontrar um link de arquivo — ele pertence ao bloco
        # anterior e não pode ser recontado aqui.
        limite   = marcos[i - 1] + 1 if i > 0 else 0
        ini_look = ini_idx
        passos   = 0
        while ini_look > limite and passos < LOOKBACK:
            anterior = elementos[ini_look - 1]
            if getattr(anterior, "name", "") == "a" and anterior.get("href"):
                break            # arquivo do contrato anterior → não recua mais
            ini_look -= 1
            passos   += 1

        trecho = elementos[ini_look:fim_idx]
        texto  = "\n".join(e.get_text(" ", strip=True) for e in trecho
                           if hasattr(e, "get_text"))
        blocos.append({"elementos": trecho, "texto": texto})

    # 4. Para cada bloco, extrai dados + arquivo CONTRATO.PDF
    contratos = []
    for bloco in blocos:
        info = _parse_bloco_contrato(bloco, base_url)
        if info:
            contratos.append(info)

    # ── 5. REDE DE SEGURANÇA ─────────────────────────────────────────────────
    # O heading informa quantos contratos existem ("Contratos3"). Se achamos
    # menos que isso, algum bloco escapou da segmentação — recuperamos por
    # varredura direta da seção, sem depender da estrutura de blocos.
    qtd_esperada = _qtd_contratos_esperada(soup)
    if qtd_esperada and len(contratos) < qtd_esperada:
        # 5a. Um bloco pode ter absorvido contratos vizinhos — nesse caso ele
        #     fica com vários arquivos de contrato. Separamos cada arquivo
        #     excedente em um contrato próprio.
        contratos = _desdobrar_blocos_multiplos(contratos, qtd_esperada)

        # 5b. Ainda faltando? Varre a seção atrás de arquivos que não entraram
        #     em nenhum bloco (estrutura fora do padrão).
        if len(contratos) < qtd_esperada:
            recuperados = _recuperar_contratos_faltantes(
                elementos, contratos, base_url, qtd_esperada
            )
            if recuperados:
                contratos.extend(recuperados)

    return contratos


def _desdobrar_blocos_multiplos(contratos: list, qtd_esperada: int) -> list:
    """
    Quando um bloco ficou com mais de um arquivo de CONTRATO, é porque dois
    contratos foram grudados (o segundo não tinha 'Contrato nº'). Cada arquivo
    extra vira um contrato próprio, preservando os dados do bloco de origem.
    """
    resultado = []
    for c in contratos:
        arquivos = c.get("arquivos", [])
        if len(arquivos) <= 1 or len(resultado) >= qtd_esperada:
            resultado.append(c)
            continue

        # Primeiro arquivo mantém os dados extraídos do bloco
        primeiro = dict(c)
        primeiro["arquivos"] = [arquivos[0]]
        resultado.append(primeiro)

        # Demais viram contratos separados, sem herdar dados do primeiro
        for a in arquivos[1:]:
            resultado.append({
                "numero_contrato": "",
                "contratado":      "(não identificado)",
                "cnpj":            "",
                "valor":           "",
                "vigencia_inicio": "",
                "vigencia_fim":    "",
                "arquivos":        [a],
            })

    if len(resultado) != len(contratos):
        print(f"      [i] Bloco com contratos grudados desdobrado: "
              f"{len(contratos)} → {len(resultado)}")
    return resultado


def _qtd_contratos_esperada(soup: BeautifulSoup):
    """Lê a quantidade indicada no heading 'ContratosN' (aba ou seção)."""
    melhor = None
    for h in soup.find_all(re.compile(r"h[1-6]")):
        m = re.match(r"contratos\s*(\d+)$", h.get_text(strip=True), re.I)
        if m:
            n = int(m.group(1))
            melhor = n if melhor is None else max(melhor, n)
    return melhor


def _recuperar_contratos_faltantes(elementos, ja_achados, base_url, qtd_esperada):
    """
    Fallback: varre TODOS os arquivos de contrato da seção e recupera os que
    não entraram em nenhum bloco. Garante que nenhum contrato fique de fora
    mesmo se a estrutura da página mudar.
    """
    urls_ja = {a["url"] for c in ja_achados for a in c.get("arquivos", [])}

    orfaos      = []
    secao_outros = False

    for elem in elementos:
        texto = elem.get_text(" ", strip=True) if hasattr(elem, "get_text") else ""

        if re.search(r"outros\s+documentos", texto, re.I):
            secao_outros = True
        elif re.match(r"^\s*contrato\s*$", texto, re.I):
            secao_outros = False

        if elem.name == "a" and elem.get("href"):
            href = elem["href"].strip()
            nome = elem.get_text(strip=True)
            if (is_file_url(href)
                    and not _NAV_SKIP.search(href)
                    and "aditivo" not in nome.lower()
                    and "aditivo" not in href.lower()
                    and not secao_outros
                    and not re.search(r"(parecer|ato de designa|controle interno|apostilamento)", nome, re.I)):
                url = urljoin(base_url, href)
                if url not in urls_ja:
                    orfaos.append({"url": url, "nome": nome or "CONTRATO"})
                    urls_ja.add(url)

    if not orfaos:
        return []

    print(f"      [!] {len(orfaos)} arquivo(s) de contrato fora dos blocos — recuperando.")
    return [{
        "numero_contrato": "",
        "contratado":      "(não identificado)",
        "cnpj":            "",
        "valor":           "",
        "vigencia_inicio": "",
        "vigencia_fim":    "",
        "arquivos":        [o],
    } for o in orfaos]


def _parse_bloco_contrato(bloco: dict, base_url: str) -> dict:
    """Extrai dados e arquivos de um único bloco de contrato."""
    texto = bloco["texto"]
    obj   = {}

    m = re.search(r"Contrato\s+n[ºo°]\s*([\d/.\-]+)", texto, re.I)
    if m:
        obj["numero_contrato"] = m.group(1).strip()

    # Documento: CNPJ ou CPF (contratado pode ser pessoa física)
    m = re.search(r"(\d{2}\.\d{3}\.\d{3}/\d{4}-\d{2})", texto)
    if m:
        obj["cnpj"] = m.group(1)
    else:
        m = re.search(r"(\d{3}\.\d{3}\.\d{3}-\d{2})", texto)   # CPF
        if m:
            obj["cnpj"] = m.group(1)

    # Contratado: nome imediatamente antes do CNPJ/CPF
    m = re.search(
        r"([A-ZÁÉÍÓÚÀÂÃÊÔÇ][A-ZÁÉÍÓÚÀÂÃÊÔÇa-záéíóúàâãêôç0-9\s\.\-&]{3,80}?)"
        r"\s*(?:\d{2}\.\d{3}\.\d{3}/\d{4}-\d{2}|\d{3}\.\d{3}\.\d{3}-\d{2})",
        texto)
    if m:
        obj["contratado"] = re.sub(r"\s+", " ", m.group(1).strip())
    else:
        # Sem documento: pega o nome após o rótulo CONTRATADO
        m = re.search(r"CONTRATADO\s*[:\-]?\s*([^\n]{3,80})", texto, re.I)
        if m:
            obj["contratado"] = re.sub(r"\s+", " ", m.group(1).strip())

    m = re.search(r"R\$\s*([\d\.]+,\d{2})", texto)
    if m:
        obj["valor"] = m.group(1)

    m = re.search(r"IN[ÍI]CIO\s*(\d{2}/\d{2}/\d{4}).*?FIM\s*(\d{2}/\d{2}/\d{4})", texto, re.I | re.S)
    if m:
        obj["vigencia_inicio"] = m.group(1)
        obj["vigencia_fim"]    = m.group(2)
    else:
        datas = re.findall(r"\d{2}/\d{2}/\d{4}", texto)
        if len(datas) >= 2:
            obj["vigencia_inicio"] = datas[0]
            obj["vigencia_fim"]    = datas[1]

    # ── Arquivos: só o(s) sob o rótulo CONTRATO, ignorando OUTROS DOCUMENTOS ──
    arquivos     = []
    urls_vistas  = set()
    secao_outros = False

    for elem in bloco["elementos"]:
        texto_elem = elem.get_text(" ", strip=True) if hasattr(elem, "get_text") else ""

        # Detecta seções de "outros documentos" e "aditivos" dentro do bloco
        if re.search(r"outros\s+documentos", texto_elem, re.I) and elem.name in ("div","span","p","strong","b","td","th","li","h4","h5","h6"):
            secao_outros = True
        elif re.search(r"^\s*aditivos?\s*$", texto_elem, re.I) and elem.name in ("div","span","p","strong","b","td","th","li","h4","h5","h6"):
            secao_outros = True
        elif re.match(r"^\s*contrato\s*$", texto_elem, re.I) and elem.name in ("div","span","p","strong","b","td","th","li","a","h4","h5","h6"):
            secao_outros = False

        if elem.name == "a" and elem.get("href"):
            href = elem["href"].strip()
            nome = elem.get_text(strip=True)
            if (is_file_url(href)
                    and not _NAV_SKIP.search(href)
                    and "aditivo" not in nome.lower()
                    and "aditivo" not in href.lower()):
                if secao_outros:
                    continue
                if re.search(r"(parecer|ato de designa|controle interno|apostilamento|outros)", nome, re.I):
                    continue
                url = urljoin(base_url, href)
                if url not in urls_vistas:
                    arquivos.append({"url": url, "nome": nome or "CONTRATO"})
                    urls_vistas.add(url)

    obj["arquivos"] = arquivos
    # Aceita o bloco se tiver qualquer identificação OU arquivos
    return obj if (obj.get("numero_contrato") or obj.get("cnpj")
                   or obj.get("contratado") or arquivos) else None

def extract_section_links(soup: BeautifulSoup, base_url: str, keyword: str) -> list:
    """
    Lê somente a tabela da seção identificada pelo heading (h1-h4) que contém keyword.
    Usa o texto do link como nome do arquivo.
    """
    section = None
    for tag in soup.find_all(re.compile(r"h[1-4]")):
        if re.search(keyword, tag.get_text(), re.I):
            section = tag
            break
    if not section:
        return []

    table = section.find_next("table")
    if not table:
        return []

    result = []
    for tr in table.find_all("tr"):
        for a in tr.find_all("a", href=True):
            href = a["href"].strip()
            if not href or _NAV_SKIP.search(href) or not is_file_url(href):
                continue
            nome = a.get_text(strip=True) or "arquivo"
            result.append({"url": urljoin(base_url, href), "nome": nome})
    return result

def extract_file_links(soup: BeautifulSoup, base_url: str) -> list:
    """Fallback genérico: todos os links de arquivo da página."""
    found = {}
    for a in soup.find_all("a", href=True):
        href = a["href"].strip()
        if not href or _NAV_SKIP.search(href) or not is_file_url(href):
            continue
        url  = urljoin(base_url, href)
        nome = a.get_text(strip=True) or os.path.basename(href.split("?")[0]) or "arquivo"
        found[url] = nome
    return [{"url": u, "nome": n} for u, n in found.items()]

# ═══════════════════════════════════════════════════════════════════════════════
#  DOWNLOAD
# ═══════════════════════════════════════════════════════════════════════════════

def _ext_from_ct(ct: str) -> str:
    for mime, ext in _EXT_MAP.items():
        if mime in ct:
            return ext
    return ""

def download_file(session, url: str, dest_dir: str, nome_sugerido: str = "") -> bool:
    """
    Prioridade do nome:
      1. nome_sugerido (texto do link na tabela = nome bonito do sistema)
      2. Content-Disposition
      3. Segmento curto da URL com extensão
      4. "arquivo" + extensão do Content-Type
    """
    try:
        r = session.get(url, timeout=60, stream=True)
        r.raise_for_status()

        ct = r.headers.get("Content-Type", "").lower()
        if "text/html" in ct and "application" not in ct:
            print(f"    [~] Ignorado (HTML): {url[:80]}")
            return False

        ext = _ext_from_ct(ct)

        if nome_sugerido:
            base  = sanitize(nome_sugerido)
            fname = base if re.search(r"\.[a-z]{2,5}$", base, re.I) else base + ext
        else:
            cd = r.headers.get("Content-Disposition", "")
            m  = re.search(r"filename\*?=(?:UTF-8'')?[\"'\s]*([^;\n\"']+)", cd, re.I)
            cd_name = sanitize(m.group(1).strip()) if m else ""
            if cd_name and len(cd_name) < 120:
                fname = cd_name if re.search(r"\.[a-z]{2,5}$", cd_name, re.I) else cd_name + ext
            else:
                seg = os.path.basename(urlparse(url).path).split("?")[0].strip()
                if seg and len(seg) < 80 and re.search(r"\.[a-z]{2,5}$", seg, re.I):
                    fname = seg
                else:
                    fname = "arquivo" + ext

        fname    = sanitize(fname) or ("arquivo" + ext) or "arquivo"
        filepath = os.path.join(dest_dir, fname)

        base_fp, ext_fp = os.path.splitext(filepath)
        n = 1
        while os.path.exists(filepath):
            filepath = f"{base_fp}_{n}{ext_fp}"
            n += 1

        os.makedirs(dest_dir, exist_ok=True)
        size = 0
        with open(filepath, "wb") as f:
            for chunk in r.iter_content(8192):
                f.write(chunk)
                size += len(chunk)

        if size < 100:
            os.remove(filepath)
            print(f"    [~] Ignorado (muito pequeno, {size}B)")
            return False

        print(f"    [↓] {os.path.basename(filepath)}  ({size/1024:.1f} KB)")
        return True

    except Exception as e:
        print(f"    [!] Erro: {e}  →  {url[:80]}")
        return False

# ═══════════════════════════════════════════════════════════════════════════════
#  DADOS DO CONTRATO
# ═══════════════════════════════════════════════════════════════════════════════

def extract_contract_info(soup: BeautifulSoup) -> list:
    """
    Extrai os dados de cada contrato conforme a estrutura do Print 3:
      - Contratado:  'F SERVICOS E LOCACOES LTDA'
      - CNPJ:        '22.807.008/0001-05'
      - Nº contrato: 'Contrato nº 20269009'
      - Valor:       'R$165.543,60'
      - Vigência:    INÍCIO 17/04/2026  /  FIM 17/04/2027
    Cada contrato é um bloco que começa em 'Contrato nº ...'.
    """
    contracts = []
    full_text = soup.get_text("\n", strip=True)

    # Localiza cada bloco "Contrato nº XXXX"
    blocos = re.split(r"(?=Contrato\s+n[ºo°]\s*\d+)", full_text, flags=re.I)

    for bloco in blocos:
        if not re.search(r"Contrato\s+n[ºo°]\s*\d+", bloco, re.I):
            continue

        obj = {}

        # Nº do contrato
        m = re.search(r"Contrato\s+n[ºo°]\s*(\d+)", bloco, re.I)
        if m:
            obj["numero_contrato"] = m.group(1).strip()

        # CNPJ (formato XX.XXX.XXX/XXXX-XX)
        m = re.search(r"(\d{2}\.\d{3}\.\d{3}/\d{4}-\d{2})", bloco)
        if m:
            obj["cnpj"] = m.group(1)

        # Contratado: aparece após o rótulo CONTRATADO, ou antes do CNPJ
        m = re.search(r"CONTRATADO\s*\n?\s*([A-ZÁÉÍÓÚÀÂÃÊÔÇ][^\n]{2,80}?)\s*(?:\d{2}\.\d{3}\.\d{3}/|\n)", bloco, re.I)
        if not m:
            # Fallback: nome em maiúsculas seguido do CNPJ
            m = re.search(r"([A-ZÁÉÍÓÚÀÂÃÊÔÇ][A-ZÁÉÍÓÚÀÂÃÊÔÇa-z0-9\s\.\-&]{4,80}?)\s*\d{2}\.\d{3}\.\d{3}/\d{4}-\d{2}", bloco)
        if m:
            obj["contratado"] = re.sub(r"\s+", " ", m.group(1).strip())

        # Valor (R$ ...)
        m = re.search(r"R\$\s*([\d\.]+,\d{2})", bloco)
        if m:
            obj["valor"] = m.group(1)

        # Vigência início / fim — busca duas datas DD/MM/YYYY próximas a INÍCIO/FIM
        m = re.search(r"IN[ÍI]CIO\s*\n?\s*(\d{2}/\d{2}/\d{4}).*?FIM\s*\n?\s*(\d{2}/\d{2}/\d{4})", bloco, re.I | re.S)
        if m:
            obj["vigencia_inicio"] = m.group(1)
            obj["vigencia_fim"]    = m.group(2)
        else:
            # Fallback: duas primeiras datas do bloco
            datas = re.findall(r"\d{2}/\d{2}/\d{4}", bloco)
            if len(datas) >= 2:
                obj["vigencia_inicio"] = datas[0]
                obj["vigencia_fim"]    = datas[1]

        if obj.get("cnpj") or obj.get("contratado"):
            contracts.append(obj)

    # Fallback: tabelas (caso a estrutura mude)
    if not contracts:
        for table in soup.find_all("table"):
            rows = table.find_all("tr")
            hdrs = []
            for i, row in enumerate(rows):
                cells = row.find_all(["th", "td"])
                txts  = [c.get_text(" ", strip=True) for c in cells]
                if not txts:
                    continue
                if i == 0 or any(re.search(r"(contratad|fornecedor|empresa|cnpj|vig)", t, re.I) for t in txts):
                    hdrs = [t.lower() for t in txts]
                    continue
                if len(txts) > 1:
                    obj = {hdrs[j] if j < len(hdrs) else f"col_{j}": v for j, v in enumerate(txts)}
                    if any(obj.values()):
                        contracts.append(obj)

    return contracts

# ═══════════════════════════════════════════════════════════════════════════════
#  PROCESSAMENTO DE UMA LICITAÇÃO
# ═══════════════════════════════════════════════════════════════════════════════

def process_licitacao(session, lic: dict, output_dir: str) -> dict:
    modalidade  = sanitize_pasta(lic.get("modalidade", "SEM_MODALIDADE"))
    numero      = sanitize_pasta(lic.get("numero",     "SEM_NUMERO"))
    folder_name = f"{modalidade} – {numero}"
    folder_path = os.path.join(output_dir, folder_name)
    docs_dir    = os.path.join(folder_path, "Documentos")
    cont_dir    = os.path.join(folder_path, "Contratos")

    print(f"\n{'='*60}")
    print(f"[>] {folder_name}")
    os.makedirs(folder_path, exist_ok=True)

    time.sleep(DELAY)
    resp = safe_get(session, lic["url"])
    if not resp:
        print("  [!] Não foi possível acessar.")
        return {}

    soup   = BeautifulSoup(resp.text, "html.parser")
    lic_id = get_licitacao_id(lic["url"])
    abas   = build_tab_urls(lic["url"], soup, lic_id)

    with open(os.path.join(folder_path, "_dados.json"), "w", encoding="utf-8") as f:
        json.dump(lic, f, ensure_ascii=False, indent=2)

    doc_urls_baixadas = set()

    # ── DOCUMENTOS ────────────────────────────────────────────────────────────
    doc_links = []

    doc_resp, doc_url_usado = fetch_tab(
        session, abas.get("documentos", ""),
        abas.get("doc_candidatos", [])
    )

    if doc_resp and "text/html" in doc_resp.headers.get("Content-Type","").lower():
        dsoup     = BeautifulSoup(doc_resp.text, "html.parser")
        doc_links = extract_section_links(dsoup, doc_url_usado, r"Documentos")
        if not doc_links:
            doc_links = extract_file_links(dsoup, doc_url_usado)
        print(f"  [i] {len(doc_links)} documento(s) encontrado(s).")
    elif doc_resp:
        os.makedirs(docs_dir, exist_ok=True)
        download_file(session, doc_url_usado, docs_dir)
        doc_urls_baixadas.add(doc_url_usado)
    else:
        doc_links = extract_section_links(soup, lic["url"], r"Documentos")
        if not doc_links:
            doc_links = extract_file_links(soup, lic["url"])
        print(f"  [i] {len(doc_links)} documento(s) (via página principal).")

    if doc_links:
        os.makedirs(docs_dir, exist_ok=True)
        print(f"  [*] Baixando documentos...")
        for i, d in enumerate(doc_links, 1):
            nome = f"{i:02d} - {sanitize(d['nome'])}"
            download_file(session, d["url"], docs_dir, nome)
            doc_urls_baixadas.add(d["url"])
            time.sleep(0.4)

    # ── OCR: número REAL da licitação dentro dos arquivos ────────────────────
    ocr_info = {"numero": "", "confianca": 0, "arquivo": "", "metodo": ""}
    if OCR_ATIVO and os.path.isdir(docs_dir):
        print("  [*] Lendo documentos (OCR) para achar o nº da licitação...")
        try:
            ocr_info = descobrir_numero_nos_arquivos(
                docs_dir,
                ano_ref      = _ano_da_licitacao(lic),
                numero_mural = lic.get("numero", ""),
                max_arquivos = OCR_MAX_ARQUIVOS,
                usar_ocr     = True,
            )
            if ocr_info["numero"]:
                confere = "confere" if _mesmo_numero(ocr_info["numero"], lic.get("numero","")) else "DIVERGE"
                print(f"      [OCR] nº {ocr_info['numero']} "
                      f"(conf. {ocr_info['confianca']}%, {ocr_info['metodo']}, "
                      f"{ocr_info['arquivo']}) → {confere}")
            else:
                print("      [OCR] número não localizado nos documentos.")
        except Exception as e:
            print(f"      [!] Falha no OCR: {e}")

    # ── CONTRATOS ─────────────────────────────────────────────────────────────
    contract_info = []

    cont_resp, cont_url_usado = fetch_tab(
        session, abas.get("contratos", ""),
        abas.get("cont_candidatos", [])
    )

    if not cont_resp:
        print("  [.] Aba Contratos não acessível.")
    elif "text/html" not in cont_resp.headers.get("Content-Type","").lower():
        if cont_url_usado not in doc_urls_baixadas:
            print("  [i] Contrato é arquivo direto. Baixando...")
            os.makedirs(cont_dir, exist_ok=True)
            download_file(session, cont_url_usado, cont_dir, "Contrato")
        else:
            print("  [~] Contrato já baixado como documento. Pulando.")
    else:
        csoup = BeautifulSoup(cont_resp.text, "html.parser")

        # Extrai TODOS os contratos da seção (pode haver vários)
        contratos = extract_contratos(csoup, cont_url_usado)

        # Quantidade esperada conforme o número exibido no heading "ContratosN"
        qtd_esperada = None
        for h in csoup.find_all(re.compile(r"h[1-6]")):
            m = re.match(r"contratos\s*(\d+)$", h.get_text(strip=True), re.I)
            if m:
                qtd_esperada = int(m.group(1))
                break

        print(f"  [i] Contratos encontrados na seção: {len(contratos)}"
              + (f" (interface indica: {qtd_esperada})" if qtd_esperada is not None else ""))

        if qtd_esperada is not None and len(contratos) != qtd_esperada:
            print(f"  [!] ATENÇÃO: divergência entre encontrados ({len(contratos)}) "
                  f"e esperados ({qtd_esperada}).")

        # ── Percorre TODOS os contratos ───────────────────────────────────────
        processados = 0
        for idx, ct in enumerate(contratos, 1):
            contratado = ct.get("contratado", "SEM NOME")
            cnpj       = ct.get("cnpj", "sem CNPJ")
            num_ct     = ct.get("numero_contrato", str(idx))
            print(f"  ┌─ Contrato {idx}/{len(contratos)}: {contratado} | CNPJ {cnpj} | nº {num_ct}")

            # Pasta dedicada por contrato (evita arquivos sobrescritos quando há vários)
            sub_nome = sanitize_pasta(f"{idx:02d} - {contratado}")[:60]
            ct_dir   = os.path.join(cont_dir, sub_nome) if len(contratos) > 1 else cont_dir

            # Filtra arquivos já baixados como documento
            arquivos = [a for a in ct.get("arquivos", [])
                        if a["url"] not in doc_urls_baixadas]

            if not arquivos:
                print(f"  │  [.] Nenhum arquivo de contrato para baixar.")
            else:
                os.makedirs(ct_dir, exist_ok=True)
                baixados = 0
                for j, a in enumerate(arquivos, 1):
                    nome = f"{j:02d} - {sanitize(a['nome'])}"
                    try:
                        ok = download_file(session, a["url"], ct_dir, nome)
                        if ok:
                            baixados += 1
                            print(f"  │  [✓] {a['nome']}")
                        else:
                            print(f"  │  [~] Falhou (vazio/HTML): {a['nome']}")
                    except Exception as e:
                        # Erro individual não interrompe os demais contratos
                        print(f"  │  [!] Erro ao baixar {a['nome']}: {e}")
                    time.sleep(0.4)
                print(f"  │  Status: {baixados}/{len(arquivos)} arquivo(s) baixado(s).")

            # Guarda os dados do contrato (sem a lista interna de arquivos)
            contract_info.append({k: v for k, v in ct.items() if k != "arquivos"})
            processados += 1
            print(f"  └─ Contrato {idx} concluído.")

        # ── Validação final ───────────────────────────────────────────────────
        print(f"  [i] Contratos processados: {processados}/{len(contratos)}")
        if qtd_esperada is not None and processados < qtd_esperada:
            print(f"  [!] AVISO: processados {processados} de {qtd_esperada} indicados na interface.")

    if contract_info:
        print(f"  [✓] {len(contract_info)} contrato(s) com dados extraídos.")

    return {
        "numero":              lic.get("numero",""),
        "modalidade":          lic.get("modalidade",""),
        "modalidade_original": lic.get("modalidade_original",""),
        "tipo":                lic.get("tipo",""),
        "objeto":              lic.get("objeto",""),
        "data_abertura":       lic.get("data_abertura",""),
        "data_publicacao":     lic.get("data_publicacao",""),
        "status":              lic.get("status",""),
        "vl_referenciado":     lic.get("vl_referenciado",""),
        "vl_adjudicado":       lic.get("vl_adjudicado",""),
        "municipio":           lic.get("municipio",""),
        "orgao":               lic.get("orgao",""),
        "num_ocr":             ocr_info.get("numero", ""),
        "ocr_confianca":        ocr_info.get("confianca", 0),
        "ocr_arquivo":          ocr_info.get("arquivo", ""),
        "ocr_metodo":           ocr_info.get("metodo", ""),
        "ocr_confere":          ("Sim" if _mesmo_numero(ocr_info.get("numero",""), lic.get("numero",""))
                                 else ("Não" if ocr_info.get("numero") else "")),
        "pasta":               folder_name,
        "url":                 lic["url"],
        "contratos":           contract_info,
    }

# ═══════════════════════════════════════════════════════════════════════════════
#  GERAÇÃO DO EXCEL
# ═══════════════════════════════════════════════════════════════════════════════

def gerar_excel(resultados: list, output_path: str, cfg: dict):
    wb    = openpyxl.Workbook()
    hfill = PatternFill("solid", fgColor="1F4E79")
    hfont = Font(color="FFFFFF", bold=True, size=11)
    ctr   = Alignment(horizontal="center", vertical="center", wrap_text=True)
    zebra = PatternFill("solid", fgColor="EBF3FB")

    def estiliza(ws, headers, widths):
        ws.append(headers)
        for cell in ws[1]:
            cell.fill = hfill; cell.font = hfont; cell.alignment = ctr
        ws.row_dimensions[1].height = 30
        for i, w in enumerate(widths, 1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

    def zebrar(ws):
        for idx, row in enumerate(ws.iter_rows(min_row=2), 2):
            fill = zebra if idx % 2 == 0 else None
            for cell in row:
                if fill: cell.fill = fill
                cell.alignment = Alignment(vertical="center", wrap_text=True)

    # ── Aba Licitações ────────────────────────────────────────────────────────
    ws1 = wb.active; ws1.title = "Licitações"
    def to_number(valor: str):
        """Converte string de valor brasileiro para float. Ex: '2.917.991,23' → 2917991.23"""
        v = str(valor).strip()
        if not v:
            return ""
        # Remove pontos de milhar e troca vírgula decimal por ponto
        v = re.sub(r"[^\d,]", "", v)   # mantém só dígitos e vírgula
        v = v.replace(",", ".")
        try:
            return float(v)
        except ValueError:
            return valor   # devolve original se não conseguir converter

    NUM_FMT = '#,##0.00'   # formato Excel: separador de milhar + 2 decimais

    estiliza(ws1,
        ["Nº Licitação","Nº Licitação (OCR)","Confere?","Confiança OCR","Arquivo (OCR)",
         "Modalidade","Modalidade (TCM)","Tipo","Objeto","Abertura","Publicação",
         "Status","Município","Órgão","Vlr Referenciado","Vlr Adjudicado","Pasta","URL"],
        [18,20,10,13,35,30,30,20,60,16,16,15,20,35,18,18,45,60])
    for r in resultados:
        row = [r.get(k,"") for k in
            ("numero","num_ocr","ocr_confere","ocr_confianca","ocr_arquivo",
             "modalidade","modalidade_original","tipo","objeto","data_abertura","data_publicacao",
             "status","municipio","orgao","vl_referenciado","vl_adjudicado","pasta","url")]
        ws1.append(row)
        # Destaca em vermelho quando o número do OCR diverge do mural
        if r.get("ocr_confere") == "Não":
            ws1.cell(ws1.max_row, 3).font = Font(color="C00000", bold=True)
        # Aplica formato numérico nas colunas de valor (agora 15 e 16)
        for col in (15, 16):
            cell = ws1.cell(ws1.max_row, col)
            val  = to_number(cell.value)
            cell.value       = val
            if isinstance(val, float):
                cell.number_format = NUM_FMT
    zebrar(ws1)

    # ── Aba Contratos ─────────────────────────────────────────────────────────
    ws2 = wb.create_sheet("Contratos")
    estiliza(ws2,
        ["Nº Licitação","Modalidade","Município","Órgão","Nº Contrato","Contratado","CNPJ",
         "Vigência Início","Vigência Fim","Valor do Contrato","URL"],
        [18,25,20,35,18,50,22,16,16,20,60])

    def pick(d, *pats):
        for pat in pats:
            for k, v in d.items():
                if re.search(pat, str(k), re.I):
                    return str(v).strip()
        return ""

    for r in resultados:
        for c in r.get("contratos", []):
            num_cont   = pick(c, r"numero_contrato", r"n.*contrato")
            contratado = pick(c, r"contratad", r"fornecedor", r"empresa", r"raz.o", r"nome")
            cnpj       = pick(c, r"cnpj", r"cpf")
            valor      = pick(c, r"valor", r"montante")

            # Vigência: campos já separados ou extraídos de texto único
            vi = pick(c, r"vigencia_inicio", r"in.cio")
            vf = pick(c, r"vigencia_fim", r"fim")
            if not vi and not vf:
                vig = pick(c, r"vig.ncia", r"prazo", r"per.odo")
                m = re.search(r"(\d{2}/\d{2}/\d{4})\s*(?:a|até|[–\-])\s*(\d{2}/\d{2}/\d{4})", vig, re.I)
                if m:
                    vi, vf = m.group(1), m.group(2)

            ws2.append([r["numero"], r["modalidade"], r.get("municipio",""), r.get("orgao",""),
                        num_cont, contratado, cnpj, vi, vf, valor, r["url"]])
            # Formato numérico na coluna Valor do Contrato (col 10)
            cell = ws2.cell(ws2.max_row, 10)
            val  = to_number(cell.value)
            cell.value = val
            if isinstance(val, float):
                cell.number_format = NUM_FMT
    zebrar(ws2)

    wb.save(output_path)
    print(f"\n[✓] Excel salvo: {output_path}")

# ═══════════════════════════════════════════════════════════════════════════════
#  MAIN
# ═══════════════════════════════════════════════════════════════════════════════

def main():
    # Lê o link e monta a configuração automaticamente
    cfg     = _parse_link(LINK_MURAL)
    session = make_session()
    cfg     = _enriquecer_nome(cfg, session)   # lê nome real do município da tabela filtrada

    # Rótulo da faixa de anos
    if ANO_MAXIMO is not None and ANO_MINIMO is not None:
        rotulo_ano, ano_pasta = f"{ANO_MINIMO} a {ANO_MAXIMO}", f"{ANO_MINIMO}-{ANO_MAXIMO}"
    elif ANO_MAXIMO is not None:
        rotulo_ano, ano_pasta = f"até {ANO_MAXIMO}", f"ate_{ANO_MAXIMO}"
    elif ANO_MINIMO is not None:
        rotulo_ano, ano_pasta = f"de {ANO_MINIMO} em diante", f"desde_{ANO_MINIMO}"
    else:
        rotulo_ano, ano_pasta = "todos os anos", "todos_anos"

    # Pasta com o NOME DA ENTIDADE (ex.: "CM Santa Izabel")
    nome_pasta = sanitize_pasta(cfg.get("entidade") or cfg["nome"])
    output_dir = os.path.join(PASTA_SAIDA, f"{nome_pasta} {ano_pasta}")
    cfg["output_dir"] = output_dir

    # Confirmação visual antes de começar
    print(f"\n  Entidade  : {cfg.get('entidade','?')}")
    print(f"  Município : {cfg['nome']}")
    print(f"  Órgão     : {cfg.get('nome_orgao', cfg['id_orgao'])}")
    print(f"  Anos      : {rotulo_ano}")
    print(f"  OCR       : {'ativo' if OCR_ATIVO else 'desligado'}")
    print(f"  Saída     : {output_dir}\n")

    os.makedirs(output_dir, exist_ok=True)

    licitacoes = get_all_licitacoes(session, cfg)
    if not licitacoes:
        print(f"\n[!] Nenhuma licitação encontrada ({rotulo_ano}).")
        return

    print(f"\n[✓] {len(licitacoes)} licitação(ões) para processar.")

    resultados = []
    for i, lic in enumerate(licitacoes, 1):
        print(f"\n[{i}/{len(licitacoes)}]")
        try:
            r = process_licitacao(session, lic, cfg["output_dir"])
            if r:
                resultados.append(r)
        except Exception as e:
            print(f"  [!] Erro em {lic.get('numero','?')}: {e}")
        time.sleep(DELAY)

    excel = os.path.join(
        cfg["output_dir"],
        f"licitacoes_{nome_pasta}_{ano_pasta}_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
    )
    gerar_excel(resultados, excel, cfg)

    total_cont = sum(len(r.get("contratos",[])) for r in resultados)
    print(f"\n  Concluído: {len(resultados)} licitações | {total_cont} contratos")
    print(f"  Excel: {os.path.basename(excel)}")

if __name__ == "__main__":
    main()
