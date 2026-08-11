# =====================================================================
#  Publicação de Sessão — portal CR2 (Playwright / Bubble)
# =====================================================================
#
# Ritmo (igual RGF):
#   botão "Criar Publicação" → modal "Cadastrar Sessão"
#   → Tipo / Data / Número → uploads → Publicar → próximo
#
# Campos do modal:
#   Tipo *, Data *, Número *, Pauta, Ata, Lista de Presença,
#   Votações Nominais (arquivo), Votações Nominais (Link),
#   checkbox "Não houve publicações", Transmissão (opcional)
#
# Entrada (prioridade):
#   1) REGISTRO_UNICO (painel — 1 sessao)
#   2) Planilha Sessoes.xlsx / Sessoes.csv (gerada no download)
#   3) PASTA_SESSOES — pastas "18ª Sessão Ordinária - 16-11-2023"
#   4) CSV_FILA
#
# Flags: --test  --yes  --headless  --pasta CAMINHO  --csv CAMINHO

from __future__ import annotations

import argparse
import calendar
import csv
import os
import re
import shutil
import subprocess
import sys
import time
import unicodedata
from pathlib import Path
from typing import Any
from urllib.parse import urlparse

try:
    from playwright.sync_api import TimeoutError as PWTimeout
    from playwright.sync_api import sync_playwright
except ImportError:
    PWTimeout = None
    sync_playwright = None


class Cancelado(Exception):
    """Fila interrompida pelo usuario (centro-automacoes)."""


def pedido_cancelado():
    return False


def _abortar_se_cancelado():
    if pedido_cancelado():
        print("[AVISO] Fila cancelada pelo usuario.")
        raise Cancelado()


# ---------------------------------------------------------------------
#  CONFIG
# ---------------------------------------------------------------------

URL_LOGIN = "https://www.portalcr2.com.br/?view=login"
URL_PORTAL_SESSAO = ""  # ex.: https://www.portalcr2.com.br/sessoes/sessoes-entidade

PORTAL_USUARIO = ""
PORTAL_SENHA = ""

HEADLESS = False
MODO_TESTE = False
ABRIR_LOGIN_ANTES_DO_PORTAL = True
PORTAL_LOGIN_BOTAO = "Entrar"

CSV_FILA = Path(__file__).resolve().parent / "fila_sessoes.csv"
# Pasta com subpastas de sessao (ex.: ...\sessoes_2021\33ª Ordinária - 14-10-2021)
PASTA_SESSOES = Path(r"C:\Users\tobia\Documents\mds\missao_baixar_sessao\sessoes_2021")
# Preenchido pelo painel para 1 sessao avulsa (dict). None = usa pasta/CSV.
REGISTRO_UNICO = None

# IA local (Ollama) para Declaracao: ler o PDF e achar o mes de referencia
REFINAR_IA_DECLARACAO = True
MODELO_IA = "llama3.2:3b"
OLLAMA_URL = "http://127.0.0.1:11434"

OPERA_EXE = None
PASTA_SCREENSHOTS = Path(__file__).resolve().parent / "screenshots_pub"

PAUSA_APOS_ANEXAR = 0.48
PAUSA_POLL_UPLOAD_UI = 0.24
MAX_TENTATIVAS_POLL_UPLOAD = 18
PAUSA_APOS_CONFIRMAR_UPLOAD = 0.7
TIMEOUT_PUBLICAR_HABILITADO_S = 75
# Max espera por feedback; se o modal nao fechar, segue sem marcar erro
TIMEOUT_RESULTADO_PUBLICACAO_S = 18
TIMEOUT_LOADER_TOPO_S = 120
# Pausa curta so para a UI assentar apos clicar Publicar
PAUSA_APOS_CLICAR_PUBLICAR = 0.35
# Entre uma sessao e a proxima (apos concluir a atual)
PAUSA_ENTRE_SESSOES = 2.0

# Portal: botão "Criar Publicação" → modal "Cadastrar Sessão"
MODAL_TITULO_REGEX = r"(Criar|Cadastrar).*Sess[aã]o"

LABELS_TIPO = ("Tipo",)
LABELS_DATA = ("Data", "Data da sessão", "Data da Sessão", "Data da sessao")
LABELS_NUMERO = ("Número", "Numero", "Nº", "N°", "Sessão", "Sessao")
LABELS_NAO_HOUVE = (
    "Não houve publicações",
    "Nao houve publicacoes",
    "Não houve publicação",
)
LABELS_LINK_VOTACOES = (
    "Votações Nominais (Link)",
    "Votacoes Nominais (Link)",
    "Votações Nominais Link",
    "Link Votações Nominais",
)

UPLOADS = (
    ("pauta", ("Pauta",)),
    ("ata", ("Ata",)),
    ("presenca", ("Lista de Presença", "Lista de Presenca", "Presença", "Presenca")),
    (
        "votacoes_arquivo",
        (
            "Votações Nominais (arquivo)",
            "Votacoes Nominais (arquivo)",
            "Votações Nominais",
            "Votacoes Nominais",
        ),
    ),
)

# Modal de Declaracao: um unico campo "Arquivo" (sem Pauta/Ata/Numero)
UPLOADS_DECLARACAO = (("arquivo", ("Arquivo", "Documento", "Anexar")),)

_TEXTO_ERRO_APOS_PUBLICAR_RX = re.compile(
    r"(?:\berro\b|\bfalha\b|inv[aá]lid|obrigat[oó]rio|"
    r"n[aã]o\s+foi\s+poss[ií]vel|tente\s+novamente|"
    r"j[aá]\s+existe|duplicad|\bduplicat)",
    re.I,
)
# Evita falso positivo em labels do form ("Cadastrar Sessão", "Não houve publicações")
_TEXTO_SUCESSO_MODAL_RX = re.compile(
    r"(publicado\s+com\s+sucesso|publicada\s+com\s+sucesso|"
    r"salvo\s+com\s+sucesso|salva\s+com\s+sucesso|"
    r"cadastrado\s+com\s+sucesso|cadastrada\s+com\s+sucesso|"
    r"registrado\s+com\s+sucesso|enviado\s+com\s+sucesso|"
    r"sess[aã]o\s+(?:publicada|cadastrada|salva)|"
    r"(?:opera[cç][aã]o|registro)\s+realizad[ao]\s+com\s+sucesso|"
    r"\bsucesso\b)",
    re.I,
)
# Labels fixos do formulário — não contam como mensagem de erro
_TEXTO_LABEL_FORM_RX = re.compile(
    r"n[aã]o\s+houve\s+publica|"
    r"declara[cç][aã]o:\s*n[aã]o\s+houve\s+sess|"
    r"cadastrar\s+sess|"
    r"criar\s+publica",
    re.I,
)


# ---------------------------------------------------------------------
#  Playwright / Opera
# ---------------------------------------------------------------------

def _recarregar_playwright():
    global PWTimeout, sync_playwright
    try:
        from playwright.sync_api import TimeoutError as PWTimeout
        from playwright.sync_api import sync_playwright
        return True
    except ImportError:
        PWTimeout = None
        sync_playwright = None
        return False


def _python_venv_do_projeto():
    # automacoes/venv ou centro-automacoes/venv
    aqui = Path(__file__).resolve().parent
    for cand in (
        aqui.parent / "venv" / "Scripts" / "python.exe",
        aqui.parent.parent / "centro-automacoes" / "venv" / "Scripts" / "python.exe",
    ):
        if cand.is_file():
            return cand
    return aqui.parent / "venv" / "Scripts" / "python.exe"


def garantir_playwright_pronto():
    if sync_playwright is not None:
        return
    subprocess.run(
        [sys.executable, "-m", "pip", "install", "playwright"],
        check=False,
    )
    if _recarregar_playwright():
        return
    venv_py = _python_venv_do_projeto()
    if venv_py.is_file():
        subprocess.run([str(venv_py), "-m", "pip", "install", "playwright"], check=False)
        if Path(sys.executable).resolve() != venv_py.resolve():
            os.execv(str(venv_py), [str(venv_py)] + sys.argv)
        _recarregar_playwright()


def verificar_playwright_instalado():
    if sync_playwright is not None:
        return
    garantir_playwright_pronto()
    if sync_playwright is None:
        print("[ERRO] Playwright indisponivel. pip install playwright")
        sys.exit(1)


def _opera_via_program_files():
    local = os.environ.get("LOCALAPPDATA", "")
    prog = os.environ.get("PROGRAMFILES", "")
    prog_x86 = os.environ.get("PROGRAMFILES(X86)", "")
    for p in (
        Path(local) / "Programs" / "Opera" / "opera.exe",
        Path(local) / "Programs" / "Opera GX" / "opera.exe",
        Path(prog) / "Opera" / "opera.exe",
        Path(prog_x86) / "Opera" / "opera.exe",
    ):
        if p.is_file():
            return p.resolve()
    base = Path(local) / "Programs"
    if base.is_dir():
        for folder in sorted(base.glob("Opera*")):
            exe = folder / "opera.exe"
            if exe.is_file():
                return exe.resolve()
    return None


def resolver_caminho_opera():
    if OPERA_EXE:
        p = Path(OPERA_EXE)
        if p.is_file():
            return p.resolve()
    found = shutil.which("opera") or shutil.which("opera.exe")
    if found:
        return Path(found).resolve()
    return _opera_via_program_files()


# ---------------------------------------------------------------------
#  Utilitarios
# ---------------------------------------------------------------------

def normalizar(texto):
    nfd = unicodedata.normalize("NFD", str(texto or ""))
    return "".join(c for c in nfd if unicodedata.category(c) != "Mn").lower().strip()


def url_portal_ativa(url):
    return bool((url or "").strip())


def salvar_screenshot(page, nome):
    try:
        PASTA_SCREENSHOTS.mkdir(parents=True, exist_ok=True)
        page.screenshot(path=str(PASTA_SCREENSHOTS / "{}.png".format(nome)), full_page=True)
    except Exception:
        pass


def preencher_campo(page, locator, valor):
    locator.click()
    time.sleep(0.08)
    page.keyboard.press("Control+a")
    time.sleep(0.05)
    page.keyboard.press("Delete")
    time.sleep(0.05)
    locator.fill(valor)
    time.sleep(0.08)


def preencher_campo_rapido(page, locator, valor):
    try:
        locator.focus(timeout=4000)
    except Exception:
        locator.click(timeout=6000)
    locator.fill(valor, timeout=8000)


def _fill_by_label_candidates(scope, labels, valor, page):
    for lb in labels:
        loc = scope.get_by_label(lb, exact=False).first
        try:
            loc.wait_for(state="visible", timeout=2500)
            preencher_campo(page, loc, valor)
            return True
        except Exception:
            continue
    return False


def _caminho_arquivo(valor):
    if not valor:
        return None
    p = Path(str(valor).strip().strip('"'))
    if not p.is_file():
        return None
    return p


# ---------------------------------------------------------------------
#  Leitura de fila
# ---------------------------------------------------------------------

CSV_COLS = (
    "tipo",
    "data",
    "numero",
    "pauta",
    "ata",
    "presenca",
    "votacoes_arquivo",
    "votacoes_link",
    "nao_houve_publicacoes",
)

PLANILHA_COLS_UI = (
    "Tipo",
    "Data",
    "Número",
    "Pauta",
    "Ata",
    "Lista de Presença",
    "Votações Nominais (arquivo)",
    "Votações Nominais (Link)",
    "Não houve publicações",
)


def _registro_vazio():
    return {k: "" for k in CSV_COLS}


_RE_DATA_FIM = re.compile(
    r"^(.*?)\s*[-–—]\s*(\d{1,2})[-/.](\d{1,2})[-/.](\d{4})\s*$",
    re.I,
)

# Tipos exatos do select "Tipo" no modal Cadastrar Sessão (portal CR2)
_TIPOS_PORTAL_SELECT = (
    "Audiência Pública",
    "Especial",
    "Extraordinária",
    "Itinerante",
    "Ordinária",
    "Preparatória",
    "Solene",
    "Tribuna Popular",
    "Declaração: não houve sessão",
)

# Tipos do portal / download-normas (ordem: compostos primeiro)
_TIPOS_PUB = (
    ("Audiência Pública", re.compile(r"^audi[eê]ncia\s+p[uú]blica$", re.I)),
    ("Tribuna Popular", re.compile(r"^tribuna\s+popular$", re.I)),
    ("Extraordinária", re.compile(r"^extra\s*ordin[aá]ria$", re.I)),
    ("Preparatória", re.compile(r"^preparat[oó]ria$", re.I)),
    ("Itinerante", re.compile(r"^itinerante$", re.I)),
    ("Especial", re.compile(r"^especial$", re.I)),
    ("Solene", re.compile(r"^solene$", re.I)),
    ("Ordinária", re.compile(r"^ordin[aá]ria$", re.I)),
)

_TIPOS_SEM_SESSAO = frozenset({"Audiência Pública", "Tribuna Popular"})
_TIPO_DECLARACAO = "Declaração: não houve sessão"

_MESES_NOME = {
    "janeiro": 1,
    "fevereiro": 2,
    "marco": 3,
    "março": 3,
    "abril": 4,
    "maio": 5,
    "junho": 6,
    "julho": 7,
    "agosto": 8,
    "setembro": 9,
    "outubro": 10,
    "novembro": 11,
    "dezembro": 12,
    "jan": 1,
    "fev": 2,
    "mar": 3,
    "abr": 4,
    "mai": 5,
    "jun": 6,
    "jul": 7,
    "ago": 8,
    "set": 9,
    "out": 10,
    "nov": 11,
    "dez": 12,
}


def _eh_declaracao(item: dict | None = None, tipo: str = "") -> bool:
    t = tipo or ((item or {}).get("tipo") or "")
    return normalizar(t).startswith("declaracao") or t == _TIPO_DECLARACAO


def ultimo_dia_mes(mes: int, ano: int) -> str:
    """Retorna DD/MM/AAAA no ultimo dia do mes (ex.: 31/01/2026)."""
    mes = int(mes)
    ano = int(ano)
    if mes < 1 or mes > 12 or ano < 1900:
        raise ValueError("Mes/ano invalidos: {}/{}".format(mes, ano))
    dia = calendar.monthrange(ano, mes)[1]
    return "{:02d}/{:02d}/{}".format(dia, mes, ano)


def _ler_texto_pdf_declaracao(caminho: Path, max_paginas: int = 4) -> str:
    try:
        import pdfplumber

        with pdfplumber.open(str(caminho)) as pdf:
            pages = pdf.pages[: max(1, max_paginas)]
            return "\n".join((p.extract_text() or "") for p in pages)
    except Exception as e:
        print("    [AVISO] Nao li o PDF ({}): {}".format(caminho.name, str(e)[:80]))
        return ""


def _mes_ano_por_heuristica(texto: str, nome_arquivo: str = "") -> tuple[int, int] | None:
    """Tenta achar mes/ano de competencia no nome ou no texto (nao so data de assinatura)."""
    fontes = [nome_arquivo or "", texto or ""]

    # Prioridade: "referente a janeiro de 2026", "mes de janeiro/2026", "competencia 01/2026"
    padroes = [
        re.compile(
            r"(?:referente\s+a[o]?\s+|compet[eê]ncia\s+(?:de\s+)?|m[eê]s\s+de\s+|"
            r"relativa\s+a[o]?\s+|no\s+m[eê]s\s+de\s+|do\s+m[eê]s\s+de\s+)"
            r"(janeiro|fevereiro|mar[cç]o|abril|maio|junho|julho|agosto|"
            r"setembro|outubro|novembro|dezembro)"
            r"\s*(?:de\s+|/\s*|-\s*)?((?:20|19)\d{2})",
            re.I,
        ),
        re.compile(
            r"(?:compet[eê]ncia|refer[eê]ncia|per[ií]odo)\s*[:\-]?\s*"
            r"(0?[1-9]|1[0-2])\s*[/\-.]\s*((?:20|19)\d{2})",
            re.I,
        ),
        re.compile(
            r"\b(janeiro|fevereiro|mar[cç]o|abril|maio|junho|julho|agosto|"
            r"setembro|outubro|novembro|dezembro)"
            r"[\s/\-_]+((?:20|19)\d{2})\b",
            re.I,
        ),
        re.compile(
            r"\b(0?[1-9]|1[0-2])[\s/\-_]((?:20|19)\d{2})\b"
        ),
    ]
    for fonte in fontes:
        if not fonte:
            continue
        for rx in padroes:
            m = rx.search(fonte)
            if not m:
                continue
            g1, g2 = m.group(1), m.group(2)
            if g1.isdigit():
                mes, ano = int(g1), int(g2)
            else:
                mes = _MESES_NOME.get(normalizar(g1), 0)
                ano = int(g2)
            if 1 <= mes <= 12 and 1990 <= ano <= 2100:
                return mes, ano
    return None


def _mes_ano_por_ia(texto: str, nome_arquivo: str = "") -> tuple[int, int] | None:
    """Pede ao Ollama o mes/ano a que a declaracao SE REFERE (nao a data de assinatura)."""
    if not (texto or "").strip():
        return None
    try:
        auto = Path(__file__).resolve().parent.parent
        if str(auto) not in sys.path:
            sys.path.insert(0, str(auto))
        from _comum.ia_ollama import chamar_json
    except Exception as e:
        print("    [AVISO] IA indisponivel: {}".format(str(e)[:100]))
        return None

    trecho = re.sub(r"\s+", " ", (texto or "").strip())[:3500]
    prompt = (
        "Voce analisa uma DECLARACAO de que NAO HOUVE SESSAO legislativa.\n"
        "Objetivo: descobrir o MES e ANO de COMPETENCIA (o periodo a que o documento "
        "se refere), NAO a data de assinatura (que pode ser de outro mes).\n"
        "Arquivo: {nome}\n"
        "Texto do PDF:\n{txt}\n\n"
        "Responda APENAS JSON: {{\"mes\": <1-12>, \"ano\": <AAAA>, "
        "\"confianca\": <0a1>, \"motivo\": \"...\"}}\n"
        "Se nao souber, use mes=0 e ano=0."
    ).format(nome=nome_arquivo or "declaracao.pdf", txt=trecho)

    try:
        dados = chamar_json(
            prompt,
            modelo=MODELO_IA,
            base_url=OLLAMA_URL,
            temperatura=0.05,
            timeout=120,
        )
        mes = int(dados.get("mes") or 0)
        ano = int(dados.get("ano") or 0)
        conf = float(dados.get("confianca") or 0)
        motivo = str(dados.get("motivo") or "")[:120]
        print(
            "    IA declaracao: mes={}/{} conf={:.2f} ({})".format(
                mes, ano, conf, motivo
            )
        )
        if 1 <= mes <= 12 and 1990 <= ano <= 2100 and conf >= 0.35:
            return mes, ano
    except Exception as e:
        print("    [AVISO] IA nao resolveu o mes: {}".format(str(e)[:120]))
    return None


def resolver_data_declaracao(item: dict) -> str:
    """
    Data do portal para Declaracao = ultimo dia do mes de competencia.
    Usa texto do PDF + IA (se ligada); fallback: nome do arquivo / data ja existente.
    """
    pdf = _caminho_arquivo(item.get("arquivo") or item.get("pauta") or "")
    texto = ""
    nome = ""
    if pdf is not None:
        nome = pdf.name
        texto = _ler_texto_pdf_declaracao(pdf)

    mes_ano = None
    if REFINAR_IA_DECLARACAO:
        print("    Lendo PDF com IA para achar o mes de referencia...")
        mes_ano = _mes_ano_por_ia(texto, nome)
    if mes_ano is None:
        mes_ano = _mes_ano_por_heuristica(texto, nome)
    if mes_ano is None:
        # data ja preenchida (ex.: planilha) — usa so mes/ano, forca ultimo dia
        m = re.match(r"(\d{1,2})/(\d{1,2})/(\d{4})", (item.get("data") or "").strip())
        if m:
            mes_ano = (int(m.group(2)), int(m.group(3)))
    if mes_ano is None:
        raise ValueError(
            "Nao consegui descobrir o mes/ano da declaracao "
            "({}). Ajuste o nome do PDF (ex.: Declaracao-Janeiro-2026.pdf) "
            "ou ligue o Ollama.".format(nome or "sem arquivo")
        )

    data = ultimo_dia_mes(mes_ano[0], mes_ano[1])
    print(
        "    Declaracao → competencia {:02d}/{} → data portal {}".format(
            mes_ano[0], mes_ano[1], data
        )
    )
    return data


def preparar_item_declaracao(item: dict) -> dict:
    """Ajusta campos para o modal de Declaracao (Arquivo + Data ultimo dia)."""
    item = dict(item)
    item["tipo"] = _TIPO_DECLARACAO
    item["nao_houve_publicacoes"] = "sim"
    # PDF fica em arquivo (campo do modal); mantem pauta como alias
    pdf = item.get("arquivo") or item.get("pauta") or ""
    if pdf:
        item["arquivo"] = pdf
        item["pauta"] = pdf
    item["numero"] = ""  # campo Numero nao existe neste modal
    item["ata"] = ""
    item["presenca"] = ""
    item["votacoes_arquivo"] = ""
    item["votacoes_link"] = ""
    item["data"] = resolver_data_declaracao(item)
    return item


def _normalizar_tipo_select(tipo: str) -> str:
    """Converte 'Sessão Ordinária' / 'Ordinária' → rótulo exato do select."""
    t = re.sub(r"\s+", " ", (tipo or "").strip())
    if not t:
        return ""
    if normalizar(t).startswith("declaracao"):
        return _TIPO_DECLARACAO
    # Remove prefixo Sessão
    t2 = re.sub(r"^sess[aã]o\s+", "", t, flags=re.I).strip()
    for nome, rx in _TIPOS_PUB:
        if nome.lower() == t.lower() or nome.lower() == t2.lower() or rx.match(t2) or rx.match(t):
            return nome
    for nome in _TIPOS_PORTAL_SELECT:
        if normalizar(nome) == normalizar(t) or normalizar(nome) == normalizar(t2):
            return nome
    return t2 or t


def _classificar_tipo_bruto(bruto: str) -> tuple[str, str]:
    """
    Retorna (tipo_select, evento).
    tipo_select: rótulo EXATO do dropdown (ex. 'Ordinária', 'Especial').
    """
    bruto = re.sub(r"\s+", " ", (bruto or "").strip(" -–—,"))
    if not bruto:
        return "", ""

    resto = re.sub(r"^sess[aã]o\s+", "", bruto, flags=re.I).strip()
    evento = ""

    m = re.match(r"^(.+?)\s*[-–—,:]\s*(.+)$", resto)
    if m:
        cand_tipo, cand_evt = m.group(1).strip(), m.group(2).strip()
    else:
        cand_tipo, cand_evt = resto, ""

    tipo_curto = ""
    for nome, rx in _TIPOS_PUB:
        if rx.match(cand_tipo):
            tipo_curto = nome
            evento = cand_evt
            break
        if rx.match(resto):
            tipo_curto = nome
            break

    if not tipo_curto:
        tipo_curto = _normalizar_tipo_select(cand_tipo or resto)
        if cand_evt and tipo_curto in dict(_TIPOS_PUB):
            evento = cand_evt

    return tipo_curto, evento


def _formatar_numero(num_raw: str, tipo_select: str, evento: str = "") -> str:
    """
    Campo Número do portal, ex.:
      '1ª Sessão Ordinária'
      'Sessão Especial - Dia dos Pais'
      '4ª Audiência Pública'
    """
    num_raw = (num_raw or "").strip()
    if num_raw:
        m = re.match(r"^(\d+)\s*[ªºa°]?\s*$", num_raw, re.I)
        if m:
            num_raw = "{}ª".format(int(m.group(1)))

    tipo_select = _normalizar_tipo_select(tipo_select)
    if tipo_select == _TIPO_DECLARACAO:
        base = "Declaração: não houve sessão"
    elif tipo_select in _TIPOS_SEM_SESSAO:
        base = "{} {}".format(num_raw, tipo_select).strip() if num_raw else tipo_select
    else:
        rotulo = "Sessão {}".format(tipo_select) if tipo_select else "Sessão"
        base = "{} {}".format(num_raw, rotulo).strip() if num_raw else rotulo

    evento = re.sub(r"\s+", " ", (evento or "").strip(" -–—,"))
    if evento and evento.lower() not in base.lower():
        return "{} - {}".format(base, evento)
    return base


def parse_nome_pasta_sessao(nome_pasta: str) -> dict:
    """
    Pastas do download-normas / legado → campos do modal CR2:

      '18ª Sessão Ordinária - 16-11-2023'
        → Tipo=Ordinária  Data=16/11/2023  Número=18ª Sessão Ordinária

      'Sessão Especial - Dia dos Pais - 12-08-2023'
        → Tipo=Especial   Data=12/08/2023  Número=Sessão Especial - Dia dos Pais
    """
    nome = (nome_pasta or "").strip()
    item = _registro_vazio()

    resto = nome
    m_data = _RE_DATA_FIM.match(nome)
    if m_data:
        resto = (m_data.group(1) or "").strip()
        d, mo, y = m_data.group(2), m_data.group(3), m_data.group(4)
        item["data"] = "{:02d}/{:02d}/{}".format(int(d), int(mo), y)

    num_raw = ""
    m_num = re.match(r"^(\d+\s*[ªºa°]?)\s+(.+)$", resto, re.I)
    if m_num:
        num_raw = m_num.group(1).strip()
        resto = m_num.group(2).strip()

    tipo_select, evento = _classificar_tipo_bruto(resto)
    if not tipo_select:
        item["tipo"] = nome
        item["numero"] = nome
        return item

    item["tipo"] = tipo_select
    item["numero"] = _formatar_numero(num_raw, tipo_select, evento)
    return item


def _achar_pdf_por_palavras(pasta: Path, palavras: tuple[str, ...], excluir: tuple[str, ...] = ()):
    if not pasta.is_dir():
        return ""
    candidatos = []
    for f in pasta.iterdir():
        if not f.is_file():
            continue
        if f.suffix.lower() not in (".pdf", ".PDF"):
            continue
        stem = normalizar(f.stem)
        if any(ex in stem for ex in excluir if ex):
            continue
        if any(p in stem for p in palavras):
            candidatos.append(f)
    if not candidatos:
        return ""
    # Prefere nome exato (Pauta.pdf) depois o menor nome
    candidatos.sort(key=lambda p: (0 if normalizar(p.stem) in palavras else 1, len(p.name), p.name.lower()))
    return str(candidatos[0])


def arquivos_da_pasta_sessao(pasta: Path) -> dict:
    return {
        "pauta": _achar_pdf_por_palavras(pasta, ("pauta",)),
        "ata": _achar_pdf_por_palavras(pasta, ("ata",), excluir=("pauta",)),
        "presenca": _achar_pdf_por_palavras(
            pasta, ("presenca", "lista de presenca", "lista_presenca", "frequencia")
        ),
        "votacoes_arquivo": _achar_pdf_por_palavras(
            pasta, ("votac", "votacoes", "votacao")
        ),
    }


def _eh_pasta_ano(nome: str) -> bool:
    return bool(re.fullmatch(r"\d{4}", (nome or "").strip()))


def _tipo_portal_valido(tipo: str) -> bool:
    t = _normalizar_tipo_select(tipo)
    if not t:
        return False
    return t in _TIPOS_PORTAL_SELECT or any(
        t == nome or rx.match(t) for nome, rx in _TIPOS_PUB
    )

def _eh_pasta_declaracoes(nome: str) -> bool:
    return normalizar(nome) in ("declaracoes", "declaracao")


def _eh_pasta_comissoes(nome: str) -> bool:
    return normalizar(nome) in ("comissoes", "comissao")


def _pastas_sessao_recursivas(root: Path) -> list[Path]:
    """
    Aceita:
      pasta/18ª Sessão Ordinária - 16-11-2023/
      pasta/2023/18ª Sessão Ordinária - 16-11-2023/
    Ignora pastas só de ano (2019) mesmo com PDF solto.
    Ignora Declarações e Comissões (não vão no modal de sessão do portal).
    """
    root = Path(root)
    if not root.is_dir():
        return []
    achadas = []
    for dirpath, dirnames, filenames in os.walk(root):
        p = Path(dirpath)
        if _eh_pasta_declaracoes(p.name) or _eh_pasta_comissoes(p.name):
            dirnames.clear()
            continue
        if _eh_pasta_ano(p.name):
            # ano: continua descendo nas subpastas de sessão
            continue
        pdfs = [f for f in filenames if f.lower().endswith(".pdf")]
        if not pdfs:
            continue
        meta = parse_nome_pasta_sessao(p.name)
        if not _tipo_portal_valido(meta.get("tipo") or ""):
            continue
        if meta.get("data") or meta.get("numero"):
            achadas.append(p)
            dirnames.clear()
    achadas.sort(key=lambda x: x.as_posix().lower())
    return achadas


def _chave_sessao(item: dict) -> str:
    return "|".join(
        [
            normalizar(item.get("tipo") or ""),
            normalizar(item.get("numero") or ""),
            (item.get("data") or "").strip(),
        ]
    )


def _itens_pdfs_soltos(root: Path, ja_cobertos: set[str], itens_existentes: list | None = None) -> list:
    """
    PDFs soltos em pastas de ano → preenche Pauta/Ata faltantes nas sessões
    já detectadas, ou cria linha nova.
    """
    root = Path(root)
    if not root.is_dir():
        return []
    por_chave: dict[str, dict] = {}
    if itens_existentes:
        for it in itens_existentes:
            por_chave[_chave_sessao(it)] = it

    novos: list = []
    for dirpath, _dirnames, filenames in os.walk(root):
        p = Path(dirpath)
        if _eh_pasta_declaracoes(p.name):
            continue
        if not (_eh_pasta_ano(p.name) or p.resolve() == root.resolve()):
            continue
        for f in filenames:
            if not f.lower().endswith(".pdf"):
                continue
            pdf = p / f
            meta = parse_nome_pasta_sessao(pdf.stem)
            if not _tipo_portal_valido(meta.get("tipo") or ""):
                continue
            if not (meta.get("data") or meta.get("numero")):
                continue
            item = _registro_vazio()
            item.update(meta)
            item["nao_houve_publicacoes"] = ""
            item["pasta"] = str(p)
            n = normalizar(pdf.stem)
            if re.search(r"(^|[^a-z])ata([^a-z]|$)", n) and "pauta" not in n:
                slot = "ata"
            elif "presen" in n:
                slot = "presenca"
            elif "votac" in n:
                slot = "votacoes_arquivo"
            else:
                slot = "pauta"
            item[slot] = str(pdf)
            chave = _chave_sessao(item)
            if chave in por_chave:
                dest = por_chave[chave]
                if not dest.get(slot):
                    dest[slot] = str(pdf)
                continue
            por_chave[chave] = item
            novos.append(item)
    return novos


def _itens_declaracoes(root: Path) -> list:
    """PDFs em .../Declarações/ → tipo 'Declaração: não houve sessão'."""
    itens = []
    root = Path(root)
    if not root.is_dir():
        return itens
    for dirpath, _dirnames, filenames in os.walk(root):
        p = Path(dirpath)
        if not _eh_pasta_declaracoes(p.name):
            continue
        for f in sorted(filenames):
            if not f.lower().endswith(".pdf"):
                continue
            pdf = p / f
            item = _registro_vazio()
            # Heuristica leve pelo nome (IA refine na hora de publicar)
            mes_ano = _mes_ano_por_heuristica("", pdf.stem)
            if mes_ano:
                try:
                    item["data"] = ultimo_dia_mes(mes_ano[0], mes_ano[1])
                except ValueError:
                    item["data"] = ""
            item["tipo"] = _TIPO_DECLARACAO
            item["numero"] = ""
            item["nao_houve_publicacoes"] = "sim"
            item["pasta"] = str(p)
            item["arquivo"] = str(pdf)
            item["pauta"] = str(pdf)  # alias legado
            itens.append(item)
    return itens


def ler_pasta_sessoes(pasta_base) -> list:
    """Varre pastas de sessão (com ou sem ano) + PDFs soltos + Declarações."""
    root = Path(pasta_base)
    if not root.is_dir():
        print("[AVISO] Pasta de sessoes nao encontrada: {}".format(root))
        return []
    itens = []
    cobertos: set[str] = set()
    for i, sub in enumerate(_pastas_sessao_recursivas(root), start=1):
        item = parse_nome_pasta_sessao(sub.name)
        if not _tipo_portal_valido(item.get("tipo") or ""):
            print("[AVISO] Ignorando pasta com tipo invalido: {}".format(sub.name))
            continue
        arquivos = arquivos_da_pasta_sessao(sub)
        item.update(arquivos)
        item["nao_houve_publicacoes"] = ""
        item["linha"] = i
        item["pasta"] = str(sub)
        if not item.get("data") and not item.get("numero"):
            print("[AVISO] Ignorando pasta sem meta: {}".format(sub.name))
            continue
        if not any(arquivos.values()):
            print("[AVISO] Pasta sem PDF (Pauta/Ata/...): {}".format(sub.name))
        itens.append(item)
        cobertos.add(_chave_sessao(item))

    for item in _itens_pdfs_soltos(root, cobertos, itens_existentes=itens):
        item["linha"] = len(itens) + 1
        itens.append(item)
        cobertos.add(_chave_sessao(item))

    for item in _itens_declaracoes(root):
        item["linha"] = len(itens) + 1
        itens.append(item)

    # Preferir sessões com data; depois as que têm Pauta+Ata (melhor p/ modo teste)
    def _sort_key(it):
        d = it.get("data") or ""
        m = re.match(r"(\d{2})/(\d{2})/(\d{4})", d)
        if m:
            data_k = "0{2}{1}{0}".format(m.group(1), m.group(2), m.group(3))
        else:
            data_k = "9" + "99999999"
        tem_ambos = 0 if (it.get("pauta") and it.get("ata")) else 1
        return (tem_ambos, data_k, normalizar(it.get("numero") or ""))

    itens.sort(key=_sort_key)
    for i, it in enumerate(itens, start=1):
        it["linha"] = i

    print("[INFO] {} sessao(oes) na pasta {}".format(len(itens), root))
    return itens


def salvar_planilha_sessoes(itens: list, pasta: str | Path) -> Path | None:
    """Grava Sessoes.csv + Sessoes.xlsx com os campos do modal CR2."""
    if not itens:
        return None
    pasta = Path(pasta)
    pasta.mkdir(parents=True, exist_ok=True)
    csv_path = pasta / "Sessoes.csv"
    xlsx_path = pasta / "Sessoes.xlsx"

    rows = []
    for it in itens:
        rows.append(
            [
                it.get("tipo", ""),
                it.get("data", ""),
                it.get("numero", ""),
                it.get("pauta", ""),
                it.get("ata", ""),
                it.get("presenca", ""),
                it.get("votacoes_arquivo", ""),
                it.get("votacoes_link", ""),
                it.get("nao_houve_publicacoes", ""),
            ]
        )

    with csv_path.open("w", encoding="utf-8-sig", newline="") as fh:
        w = csv.writer(fh, delimiter=";")
        w.writerow(PLANILHA_COLS_UI)
        w.writerows(rows)

    try:
        from openpyxl import Workbook

        wb = Workbook()
        ws = wb.active
        ws.title = "Sessões"
        ws.append(list(PLANILHA_COLS_UI))
        for row in rows:
            ws.append(row)
        for col in ws.columns:
            maxlen = 14
            letter = col[0].column_letter
            for cell in col:
                maxlen = max(maxlen, min(56, len(str(cell.value or ""))))
            ws.column_dimensions[letter].width = maxlen + 2
        wb.save(xlsx_path)
        print("[INFO] Planilha: {} ({} linhas)".format(xlsx_path, len(rows)))
        return xlsx_path
    except Exception:
        print("[INFO] Planilha CSV: {} ({} linhas)".format(csv_path, len(rows)))
        return csv_path


def gerar_planilha_da_pasta(pasta_base) -> Path | None:
    """
    Varre pasta e grava Sessoes.xlsx dentro da pasta de Pautas/Atas
    (não só na raiz), para a publicação ler no mesmo caminho.
    """
    base = Path(pasta_base)
    itens = ler_pasta_sessoes(base)
    if not itens:
        return None

    por_destino: dict[Path, list] = {}
    for it in itens:
        pasta = Path(it.get("pasta") or "")
        dest = base
        try:
            rel = pasta.relative_to(base)
            if rel.parts:
                topo = rel.parts[0]
                # Ano ou Declarações direto sob a base → planilha na base
                if re.fullmatch(r"\d{4}", topo) or _eh_pasta_declaracoes(topo):
                    dest = base
                else:
                    # Ex.: .../Pautas e Atas.../2023/18ª... → planilha em Pautas e Atas...
                    dest = base / topo
        except ValueError:
            dest = base
        por_destino.setdefault(dest, []).append(it)

    saida = None
    for dest, grupo in sorted(por_destino.items(), key=lambda kv: str(kv[0]).lower()):
        out = salvar_planilha_sessoes(grupo, dest)
        if out and saida is None:
            saida = out
    return saida


def _item_de_mapa(mapa: dict, linha: int) -> dict | None:
    item = _registro_vazio()
    aliases = {
        "tipo": ("tipo",),
        "data": ("data", "data da sessao"),
        "numero": ("numero", "nº", "n", "sessao"),
        "pauta": ("pauta",),
        "ata": ("ata",),
        "presenca": ("presenca", "lista de presenca", "lista_presenca"),
        "votacoes_arquivo": (
            "votacoes_arquivo",
            "votacoes nominais arquivo",
            "votacoes",
        ),
        "votacoes_link": (
            "votacoes_link",
            "votacoes nominais link",
            "link",
        ),
        "nao_houve_publicacoes": (
            "nao houve publicacoes",
            "nao_houve_publicacoes",
            "sem publicacoes",
        ),
    }
    for dest, keys in aliases.items():
        for k in keys:
            if k in mapa and mapa[k] is not None and str(mapa[k]).strip():
                item[dest] = str(mapa[k]).strip()
                break
    if item.get("tipo"):
        item["tipo"] = _normalizar_tipo_select(item["tipo"])
    if not (item["tipo"] or item["data"] or item["numero"]):
        return None
    if item.get("tipo") and not _tipo_portal_valido(item["tipo"]):
        return None
    item["linha"] = linha
    return item


def ler_planilha_fila(caminho) -> list:
    """Lê Sessoes.xlsx ou .csv gerada na extração."""
    path = Path(caminho)
    if not path.is_file():
        return []
    if path.suffix.lower() in (".xlsx", ".xlsm"):
        try:
            from openpyxl import load_workbook

            wb = load_workbook(path, read_only=True, data_only=True)
            ws = wb.active
            rows = list(ws.iter_rows(values_only=True))
            wb.close()
            if not rows:
                return []
            headers = [normalizar(str(h or "")) for h in rows[0]]
            itens = []
            for i, row in enumerate(rows[1:], start=2):
                if not row or all(c is None or str(c).strip() == "" for c in row):
                    continue
                mapa = {
                    headers[j]: ("" if row[j] is None else str(row[j]).strip())
                    for j in range(min(len(headers), len(row)))
                }
                item = _item_de_mapa(mapa, i)
                if item:
                    itens.append(item)
            print("[INFO] {} sessao(oes) na planilha {}".format(len(itens), path))
            return itens
        except Exception as e:
            print("[AVISO] Falha ao ler xlsx ({}): {}".format(path, str(e)[:80]))
            return []
    return ler_csv_fila(path)


def ler_csv_fila(caminho):
    path = Path(caminho)
    if not path.is_file():
        print("[AVISO] CSV nao encontrado: {}".format(path))
        return []
    itens = []
    with open(path, "r", encoding="utf-8-sig", newline="") as fh:
        sample = fh.read(2048)
        fh.seek(0)
        delim = ";" if sample.count(";") >= sample.count(",") else ","
        reader = csv.DictReader(fh, delimiter=delim)
        for i, row in enumerate(reader, start=2):
            if not row:
                continue
            mapa = {normalizar(k): v for k, v in row.items() if k}
            item = _item_de_mapa(mapa, i)
            if item:
                itens.append(item)
    print("[INFO] {} sessao(oes) no CSV {}".format(len(itens), path))
    return itens


def montar_fila():
    """
    Prioridade (ritmo RGF: lote → publicar um a um via Criar Publicação):
      1) REGISTRO_UNICO
      2) Planilha Sessoes.xlsx/csv
      3) Varredura de pastas (+ regenera planilha)
    """
    if isinstance(REGISTRO_UNICO, dict) and any(
        str(REGISTRO_UNICO.get(k) or "").strip() for k in ("tipo", "data", "numero")
    ):
        item = _registro_vazio()
        for k in CSV_COLS:
            item[k] = str(REGISTRO_UNICO.get(k) or "").strip()
        item["linha"] = 1
        print("[INFO] Fila: 1 sessao (registro unico).")
        return [item]

    pasta = Path(PASTA_SESSOES) if PASTA_SESSOES else None

    # Sempre regenera a planilha a partir das pastas (fonte da verdade)
    if pasta and pasta.is_dir():
        itens = ler_pasta_sessoes(pasta)
        if itens:
            salvar_planilha_sessoes(itens, pasta)
            return itens

    candidatos = []
    if CSV_FILA:
        candidatos.append(Path(CSV_FILA))
    if pasta and pasta.is_dir():
        candidatos.extend(
            [
                pasta / "Sessoes.xlsx",
                pasta / "Sessoes.csv",
                pasta / "fila_sessoes.csv",
            ]
        )

    for cand in candidatos:
        if cand and cand.is_file():
            itens = ler_planilha_fila(cand)
            if itens:
                return itens
    return []


# ---------------------------------------------------------------------
#  Login / navegacao
# ---------------------------------------------------------------------

def credenciais_portal_configuradas():
    return bool((PORTAL_USUARIO or "").strip() and (PORTAL_SENHA or "").strip())


def navegar_para_url(page, url, etiqueta, pausa_apos_carregar=0.5):
    print("[INFO] Carregando {}...".format(etiqueta))
    ultimo = None
    for wt in ("domcontentloaded", "load"):
        try:
            page.goto(url, wait_until=wt, timeout=120000)
            ultimo = None
            break
        except Exception as e:
            ultimo = e
    if ultimo:
        raise ultimo
    time.sleep(pausa_apos_carregar)


def aguardar_barra_carregamento_topo(page, timeout_s=None, etiqueta=""):
    if timeout_s is None:
        timeout_s = TIMEOUT_LOADER_TOPO_S
    tag = " [{}]".format(etiqueta) if etiqueta else ""
    limite = time.monotonic() + float(timeout_s)
    viu = False
    js = """
        () => {
            function ativa(el) {
                if (!el) return false;
                var s = window.getComputedStyle(el);
                if (s.display === 'none' || parseFloat(s.opacity) < 0.08) return false;
                var r = el.getBoundingClientRect();
                if (r.top > 22 || r.height > 28) return false;
                return r.width > 12;
            }
            if (ativa(document.querySelector('#nprogress .bar'))) return true;
            if (ativa(document.querySelector('.turbo-progress-bar'))) return true;
            return false;
        }
    """
    while time.monotonic() < limite:
        try:
            ativo = page.evaluate(js)
        except Exception:
            ativo = False
        if ativo:
            viu = True
            time.sleep(0.14)
            continue
        return
    if viu:
        raise TimeoutError("Barra de progresso ativa apos {}s{}.".format(timeout_s, tag))


def _resolver_escopo_login(page):
    ultimo = None
    for _ in range(50):
        ordem = [page]
        try:
            for fr in page.frames:
                if fr != page.main_frame:
                    ordem.append(fr)
        except Exception:
            pass
        for scope in ordem:
            try:
                scope.locator("input[type='password']").first.wait_for(
                    state="visible", timeout=300
                )
                return scope
            except Exception as e:
                ultimo = e
        time.sleep(0.08)
    raise TimeoutError("Formulario de login nao pronto: {}".format(ultimo))


def login_automatico_portal(page):
    usuario = PORTAL_USUARIO.strip()
    senha = PORTAL_SENHA.strip()
    scope = _resolver_escopo_login(page)

    preenchido = False
    for rx in (
        re.compile(r"informe seu e-?\s*mail\s*:?", re.I),
        re.compile(r"seu e-?\s*mail", re.I),
    ):
        try:
            loc = scope.get_by_label(rx).first
            loc.wait_for(state="visible", timeout=3500)
            preencher_campo_rapido(page, loc, usuario)
            preenchido = True
            break
        except Exception:
            continue
    if not preenchido:
        for sel in (
            "input[type='email']",
            "input[autocomplete='username']",
            "input[autocomplete='email']",
            "input[type='text']",
        ):
            try:
                loc = scope.locator(sel).first
                loc.wait_for(state="visible", timeout=2500)
                preencher_campo_rapido(page, loc, usuario)
                preenchido = True
                break
            except Exception:
                continue
    if not preenchido:
        raise RuntimeError("Campo de usuario/email nao encontrado")

    try:
        loc = scope.get_by_label(re.compile(r"informe sua senha", re.I)).first
        loc.wait_for(state="visible", timeout=3500)
        preencher_campo_rapido(page, loc, senha)
    except Exception:
        campo = scope.locator("input[type='password']").first
        campo.wait_for(state="visible", timeout=10000)
        preencher_campo_rapido(page, campo, senha)

    time.sleep(0.07)
    clicou = False
    for rotulo in (
        (PORTAL_LOGIN_BOTAO or "").strip(),
        "Acessar",
        "Entrar",
        "Login",
    ):
        if not rotulo:
            continue
        try:
            b = scope.get_by_role("button", name=re.compile(re.escape(rotulo), re.I)).first
            b.wait_for(state="visible", timeout=3500)
            b.click(force=True)
            clicou = True
            break
        except Exception:
            continue
    if not clicou:
        scope.locator("button[type='submit'], input[type='submit']").first.click(force=True)


def aguardar_login_usuario(page, pular_enter=False):
    navegar_para_url(page, URL_LOGIN, "login", 0.15)
    if credenciais_portal_configuradas():
        try:
            login_automatico_portal(page)
            print("[INFO] Formulario de login enviado.")
            time.sleep(0.55)
        except Exception as e:
            print("[AVISO] Login automatico falhou ({})".format(str(e)[:160]))
    if pular_enter:
        time.sleep(1.25)
    else:
        input("[INFO] Quando estiver logado, pressione Enter...\n>>> ")


def garantir_pagina_portal(page, url_alvo, etiqueta_log):
    navegar_para_url(page, url_alvo, etiqueta_log, 0.55)
    path = (urlparse(url_alvo).path or "").strip("/")
    slug = path.split("/")[-1] if path else ""
    if slug and slug not in page.url:
        try:
            page.goto(url_alvo, wait_until="domcontentloaded", timeout=120000)
        except Exception:
            pass
        time.sleep(0.25)
    print("[INFO] URL atual: {}".format(page.url))
    try:
        page.locator("button:has-text('Criar Publicação')").wait_for(
            state="visible", timeout=45000
        )
        print("[INFO] Botao Criar Publicacao visivel.")
        aguardar_barra_carregamento_topo(page, etiqueta=etiqueta_log)
    except Exception as e:
        print("[AVISO] Criar Publicacao nao apareceu: {}".format(str(e)[:120]))


def criar_navegador_e_login(pular_enter_pos_login=False):
    verificar_playwright_instalado()
    opera = resolver_caminho_opera()
    pw = sync_playwright().start()
    launch_kwargs = {"headless": HEADLESS}
    if opera:
        print("[INFO] Abrindo Opera: {}".format(opera))
        launch_kwargs["executable_path"] = str(opera)
    else:
        print("[INFO] Opera nao encontrado — usando Chromium do Playwright.")
    try:
        browser = pw.chromium.launch(**launch_kwargs)
    except Exception as e:
        print("[ERRO] Falha ao iniciar navegador: {}".format(e))
        pw.stop()
        sys.exit(1)
    page = browser.new_context().new_page()
    if ABRIR_LOGIN_ANTES_DO_PORTAL:
        aguardar_login_usuario(page, pular_enter=pular_enter_pos_login)
    return pw, browser, page


# ---------------------------------------------------------------------
#  Modal Sessao
# ---------------------------------------------------------------------

def _loc_modal_titulo(page):
    return page.locator("text=/{}/i".format(MODAL_TITULO_REGEX)).first


def abrir_modal(page):
    """Mesmo ritmo do RGF: botão Criar Publicação → modal Cadastrar Sessão."""
    criar_btn = page.locator("button:has-text('Criar Publicação')").first
    criar_btn.wait_for(state="visible", timeout=15000)
    criar_btn.scroll_into_view_if_needed()
    time.sleep(0.06)
    criar_btn.click()
    time.sleep(0.26)
    try:
        _loc_modal_titulo(page).wait_for(state="visible", timeout=10000)
    except Exception:
        page.locator("button:has-text('Publicar')").first.wait_for(
            state="visible", timeout=10000
        )
    time.sleep(0.11)


def _modal_bubble_sessao(page):
    try:
        page.evaluate(
            """
            () => {
                document.querySelectorAll('[data-cr2-pub-modal-marker]').forEach(function (el) {
                    el.removeAttribute('data-cr2-pub-modal-marker');
                });
                var pubs = Array.from(
                    document.querySelectorAll('button, div[role="button"], .bubble-element.Button')
                ).filter(function (b) {
                    var t = ((b.innerText || b.textContent || '') + '').trim();
                    return (t === 'Publicar' || t.indexOf('Publicar') >= 0) && t.length < 48;
                });
                for (var i = pubs.length - 1; i >= 0; i--) {
                    var node = pubs[i];
                    var depth = 0;
                    while (node && depth < 28) {
                        depth++;
                        node = node.parentElement;
                        if (!node || !node.querySelectorAll) continue;
                        var txt = (node.innerText || '').slice(0, 6000);
                        var files = node.querySelectorAll('input[type=file]');
                        var temSessao =
                            /Cadastrar\s+Sess/i.test(txt) ||
                            /Sess[aã]o/i.test(txt) ||
                            /Pauta/i.test(txt);
                        if (!temSessao) continue;
                        if (!files.length) continue;
                        if (txt.indexOf('Publicar') < 0) continue;
                        node.setAttribute('data-cr2-pub-modal-marker', '1');
                        return true;
                    }
                }
                return false;
            }
            """
        )
        root = page.locator('[data-cr2-pub-modal-marker="1"]').first
        root.wait_for(state="visible", timeout=8000)
        return root
    except Exception:
        pass
    try:
        cand = (
            page.locator("div.bubble-element.Group")
            .filter(has=page.locator("button:has-text('Publicar')"))
            .filter(has=page.locator("input[type=file]"))
        )
        if cand.count() > 0:
            root = cand.first
            root.wait_for(state="visible", timeout=8000)
            return root
    except Exception:
        pass
    return None


def fechar_modal(page):
    try:
        root = _modal_bubble_sessao(page)
        if root is not None:
            root.locator("button:has-text('Fechar')").first.click(timeout=4000)
            time.sleep(0.25)
    except Exception:
        try:
            page.keyboard.press("Escape")
            time.sleep(0.22)
        except Exception:
            pass
    try:
        _loc_modal_titulo(page).wait_for(state="hidden", timeout=8000)
    except Exception:
        pass


def _truthy_flag(valor) -> bool:
    v = normalizar(str(valor or ""))
    return v in ("1", "sim", "s", "true", "yes", "x", "nao houve", "nao_houve")


def _marcar_nao_houve_publicacoes(page, modal_root, marcar: bool):
    if not marcar:
        return
    scope = modal_root if modal_root is not None else page
    for lb in LABELS_NAO_HOUVE:
        try:
            loc = scope.get_by_label(lb, exact=False).first
            loc.wait_for(state="visible", timeout=2500)
            if not loc.is_checked():
                loc.check(force=True)
            print("    Checkbox: Não houve publicações")
            return
        except Exception:
            pass
        try:
            loc = scope.get_by_text(lb, exact=False).first
            loc.click(timeout=2500)
            print("    Checkbox (texto): Não houve publicações")
            return
        except Exception:
            continue
    print("    [AVISO] Checkbox 'Não houve publicações' nao encontrado.")


def _aguardar_confirmacao_upload(page, modal_root, path: Path):
    nome_pdf_lower = path.name.lower()
    print("    Aguardando confirmacao do upload ({})...".format(path.name))
    for _ in range(MAX_TENTATIVAS_POLL_UPLOAD):
        try:
            if modal_root is not None:
                areas = modal_root.locator(".file-input-text")
            else:
                areas = page.locator(".file-input-text")
            n = areas.count()
            for i in range(n):
                txt = areas.nth(i).inner_text().strip()
                tl = txt.lower()
                if nome_pdf_lower and nome_pdf_lower in tl:
                    print("    Upload confirmado (arquivo na UI).")
                    time.sleep(PAUSA_APOS_CONFIRMAR_UPLOAD)
                    return
                if len(txt) > 5 and "clique aqui" not in tl and path.stem.lower()[:20] in tl:
                    print("    Upload confirmado: '{}'".format(txt[:50]))
                    time.sleep(PAUSA_APOS_CONFIRMAR_UPLOAD)
                    return
        except Exception:
            pass
        time.sleep(PAUSA_POLL_UPLOAD_UI)
    time.sleep(PAUSA_APOS_CONFIRMAR_UPLOAD)


def _preencher_tipo(page, modal_root, tipo_ui):
    if not (tipo_ui or "").strip():
        raise ValueError("Tipo e obrigatorio.")
    # Preferir rótulo EXATO do select do portal
    tipo_ui = _normalizar_tipo_select(tipo_ui.strip())
    candidatos = [tipo_ui]
    # Legado / fallback
    if tipo_ui not in _TIPOS_SEM_SESSAO and tipo_ui != _TIPO_DECLARACAO:
        if not tipo_ui.lower().startswith("sess"):
            candidatos.append("Sessão {}".format(tipo_ui))
    m = re.match(r"^sess[aã]o\s+(.+)$", tipo_ui, re.I)
    if m:
        candidatos.insert(0, m.group(1).strip())

    scope = modal_root if modal_root is not None else page
    for cand in candidatos:
        if modal_root is not None:
            try:
                selects = modal_root.locator("select")
                for i in range(selects.count()):
                    sel = selects.nth(i)
                    try:
                        sel.select_option(label=cand, timeout=3500)
                        print("    Tipo (select): {}".format(cand))
                        return
                    except Exception:
                        try:
                            sel.select_option(
                                label=re.compile(
                                    r"^\s*{}\s*$".format(re.escape(cand)), re.I
                                ),
                                timeout=3000,
                            )
                            print("    Tipo (select regex): {}".format(cand))
                            return
                        except Exception:
                            continue
            except Exception:
                pass
        for lb in LABELS_TIPO:
            try:
                loc = scope.get_by_label(lb, exact=False).first
                loc.wait_for(state="visible", timeout=4000)
                tag = loc.evaluate("el => el.tagName")
                if tag == "SELECT":
                    loc.select_option(label=cand, timeout=4000)
                else:
                    loc.click()
                    time.sleep(0.28)
                    page.get_by_text(cand, exact=True).last.click(timeout=5000)
                print("    Tipo: {}".format(cand))
                return
            except Exception:
                continue
    raise RuntimeError(
        "Campo Tipo nao encontrado (valor={!r}). Opcoes do portal: {}".format(
            tipo_ui, ", ".join(_TIPOS_PORTAL_SELECT)
        )
    )


def _preencher_data(page, modal_root, data_ui):
    if not (data_ui or "").strip():
        raise ValueError("Data e obrigatoria.")
    data_ui = data_ui.strip()
    # Aceita DD-MM-YYYY da pasta → DD/MM/YYYY do portal
    m = re.match(r"^(\d{1,2})[-/.](\d{1,2})[-/.](\d{4})$", data_ui)
    if m:
        data_ui = "{:02d}/{:02d}/{}".format(int(m.group(1)), int(m.group(2)), m.group(3))

    scopes = []
    if modal_root is not None:
        scopes.append(modal_root)
    scopes.append(page)

    css_data = (
        "input.bubble-element.Input[placeholder*='01/01'], "
        "input.bubble-element.Input[placeholder*='01/01/2024'], "
        "input[type='input'][placeholder*='01/01'], "
        "input[placeholder*='01/01/2024'], "
        "input[placeholder*='Ex.: 01'], "
        "input[placeholder*='dd/mm'], "
        "input[placeholder*='DD/MM']"
    )

    for scope in scopes:
        # 1) CSS Bubble (type=input / bubble-element) — mesmo padrão do RGF
        try:
            loc = scope.locator(css_data).first
            loc.wait_for(state="visible", timeout=4000)
            preencher_campo(page, loc, data_ui)
            print("    Data (CSS placeholder): {}".format(data_ui))
            return
        except Exception:
            pass

        # 2) get_by_placeholder
        for ph in (
            re.compile(r"01/01/2024", re.I),
            re.compile(r"ex\.?\s*:\s*01", re.I),
            re.compile(r"\d{2}/\d{2}/\d{4}"),
            re.compile(r"dd\s*/\s*mm", re.I),
        ):
            try:
                loc = scope.get_by_placeholder(ph).first
                loc.wait_for(state="visible", timeout=2500)
                preencher_campo(page, loc, data_ui)
                print("    Data (placeholder): {}".format(data_ui))
                return
            except Exception:
                continue

        # 3) label
        if _fill_by_label_candidates(scope, LABELS_DATA, data_ui, page):
            print("    Data (label): {}".format(data_ui))
            return

        # 4) texto "Data" → próximo input
        try:
            loc = scope.locator(
                "xpath=(//*[contains(translate(normalize-space(.),"
                "'ÁÀÃÂáàãâ','AAAAaaaa'),'Data') "
                "or contains(translate(normalize-space(.),"
                "'ÁÀÃÂáàãâ','AAAAaaaa'),'DATA')]"
                "[not(self::script)][not(self::style)])[1]"
                "/following::input[not(@type='file') and not(@type='hidden')"
                " and not(@type='checkbox') and not(@type='radio')][1]"
            ).first
            loc.wait_for(state="visible", timeout=3000)
            preencher_campo(page, loc, data_ui)
            print("    Data (xpath apos rotulo): {}".format(data_ui))
            return
        except Exception:
            pass

        # 5) 1º input de texto do modal cujo placeholder parece data
        try:
            inputs = scope.locator(
                "input.bubble-element.Input, input[type='input'], "
                "input[type='text'], input:not([type])"
            )
            n = min(inputs.count(), 12)
            for i in range(n):
                inp = inputs.nth(i)
                try:
                    ph = (inp.get_attribute("placeholder") or "").lower()
                except Exception:
                    continue
                if "01/01" in ph or "/202" in ph or "dd/mm" in ph or (
                    "ex" in ph and "/" in ph
                ):
                    preencher_campo(page, inp, data_ui)
                    print("    Data (scan input): {}".format(data_ui))
                    return
        except Exception:
            pass

    salvar_screenshot(page, "ERRO_CAMPO_DATA")
    raise RuntimeError("Campo Data nao encontrado")


def _preencher_numero(page, modal_root, numero_ui):
    if not (numero_ui or "").strip():
        raise ValueError("Numero e obrigatorio.")
    numero_ui = numero_ui.strip()
    scopes = []
    if modal_root is not None:
        scopes.append(modal_root)
    scopes.append(page)

    css_num = (
        "input.bubble-element.Input[placeholder*='Sess'], "
        "input.bubble-element.Input[placeholder*='1ª'], "
        "input[type='input'][placeholder*='Sess'], "
        "input[placeholder*='1ª Sessão'], "
        "input[placeholder*='1a Sess'], "
        "input[placeholder*='Ordinária']"
    )

    for scope in scopes:
        try:
            loc = scope.locator(css_num).first
            loc.wait_for(state="visible", timeout=4000)
            preencher_campo(page, loc, numero_ui)
            print("    Numero (CSS placeholder): {}".format(numero_ui))
            return
        except Exception:
            pass

        if _fill_by_label_candidates(scope, LABELS_NUMERO, numero_ui, page):
            print("    Numero (label): {}".format(numero_ui))
            return

        for ph in (
            re.compile(r"1[ªa]\s*Sess", re.I),
            re.compile(r"Sess[aã]o\s+Ordin", re.I),
            re.compile(r"Ex\.:\s*1", re.I),
        ):
            try:
                loc = scope.get_by_placeholder(ph).first
                loc.wait_for(state="visible", timeout=2500)
                preencher_campo(page, loc, numero_ui)
                print("    Numero (placeholder): {}".format(numero_ui))
                return
            except Exception:
                continue

        try:
            loc = scope.locator(
                "xpath=(//*[contains(translate(normalize-space(.),"
                "'Úú','Uu'),'Numero') or contains(normalize-space(.),'Número')"
                " or contains(normalize-space(.),'Nº')])[1]"
                "/following::input[not(@type='file') and not(@type='hidden')"
                " and not(@type='checkbox')][1]"
            ).first
            loc.wait_for(state="visible", timeout=3000)
            preencher_campo(page, loc, numero_ui)
            print("    Numero (xpath): {}".format(numero_ui))
            return
        except Exception:
            pass

    salvar_screenshot(page, "ERRO_CAMPO_NUMERO")
    raise RuntimeError("Campo Numero nao encontrado")


def _revelar_input_file(page, alvo) -> None:
    """Revela só o input que vamos usar (não empilha todos em cima do Publicar)."""
    try:
        if hasattr(alvo, "evaluate"):
            alvo.evaluate(
                """
                (el) => {
                    if (!el || el.tagName !== 'INPUT') return;
                    if (!el.getAttribute('data-cr2-file-style')) {
                        el.setAttribute('data-cr2-file-style', JSON.stringify({
                            display: el.style.display,
                            opacity: el.style.opacity,
                            visibility: el.style.visibility,
                            position: el.style.position,
                            top: el.style.top,
                            left: el.style.left,
                            zIndex: el.style.zIndex
                        }));
                    }
                    el.style.display = 'block';
                    el.style.opacity = '0.01';
                    el.style.visibility = 'visible';
                    el.style.position = 'fixed';
                    el.style.top = '4px';
                    el.style.left = '4px';
                    el.style.width = '2px';
                    el.style.height = '2px';
                    el.style.zIndex = '1';
                }
                """
            )
    except Exception:
        pass


def _restaurar_inputs_file(page) -> None:
    """Remove overlays de file input para o clique em Publicar não cair neles."""
    try:
        page.evaluate(
            """
            () => {
                document.querySelectorAll('input[type=file]').forEach(function (el) {
                    var raw = el.getAttribute('data-cr2-file-style');
                    if (raw) {
                        try {
                            var s = JSON.parse(raw);
                            el.style.display = s.display || '';
                            el.style.opacity = s.opacity || '';
                            el.style.visibility = s.visibility || '';
                            el.style.position = s.position || '';
                            el.style.top = s.top || '';
                            el.style.left = s.left || '';
                            el.style.zIndex = s.zIndex || '';
                            el.style.width = '';
                            el.style.height = '';
                        } catch (e) {}
                        el.removeAttribute('data-cr2-file-style');
                    } else {
                        el.style.position = '';
                        el.style.top = '';
                        el.style.left = '';
                        el.style.zIndex = '';
                        el.style.opacity = '';
                        el.style.width = '';
                        el.style.height = '';
                    }
                });
            }
            """
        )
    except Exception:
        pass


def _revelar_inputs_file(page):
    """Legado — preferir _revelar_input_file no alvo específico."""
    _restaurar_inputs_file(page)


def _mapear_inputs_file_por_rotulo(page, modal_root) -> dict[str, Any]:
    """
    Associa cada input[type=file] do modal ao rótulo (Pauta, Ata, …).
    Evita o bug de 'Ata' casar com 'Data' via contains().
    """
    try:
        marcados = page.evaluate(
            """
            () => {
                function norm(s) {
                    return (s || '')
                        .normalize('NFD')
                        .replace(/[\\u0300-\\u036f]/g, '')
                        .toLowerCase()
                        .replace(/\\s+/g, ' ')
                        .trim();
                }
                function rotuloDe(inp) {
                    var n = inp;
                    for (var d = 0; d < 10 && n; d++) {
                        var txt = norm(n.innerText || n.textContent || '');
                        if (txt.indexOf('pauta') >= 0 && txt.indexOf('ata') < 0)
                            return 'pauta';
                        // 'ata' sozinho — nao confundir com 'data' / 'cadastrar'
                        if (/(^|[^a-z])ata([^a-z]|$)/.test(txt) &&
                            txt.indexOf('data') < 0 &&
                            txt.indexOf('pauta') < 0 &&
                            txt.indexOf('cadastr') < 0)
                            return 'ata';
                        if (txt.indexOf('lista de presenca') >= 0 ||
                            (txt.indexOf('presenca') >= 0 && txt.indexOf('votac') < 0))
                            return 'presenca';
                        if (txt.indexOf('votacoes nominais (arquivo)') >= 0 ||
                            (txt.indexOf('votac') >= 0 && txt.indexOf('arquivo') >= 0) ||
                            (txt.indexOf('votacoes nominais') >= 0 &&
                             txt.indexOf('link') < 0 && txt.length < 80))
                            return 'votacoes_arquivo';
                        // Declaracao: rotulo "Arquivo" / area "Clique aqui para subir"
                        if ((txt.indexOf('arquivo') >= 0 && txt.indexOf('votac') < 0) ||
                            txt.indexOf('clique aqui para subir') >= 0 ||
                            txt.indexOf('tamanho maximo') >= 0)
                            return 'arquivo';
                        n = n.parentElement;
                    }
                    // texto imediatamente anterior no DOM
                    var prev = inp.previousElementSibling;
                    for (var k = 0; k < 6 && prev; k++) {
                        var t2 = norm(prev.innerText || prev.textContent || '');
                        if (t2 === 'pauta' || t2.indexOf('pauta') === 0) return 'pauta';
                        if (t2 === 'ata') return 'ata';
                        if (t2 === 'arquivo' || t2.indexOf('arquivo') === 0) return 'arquivo';
                        if (t2.indexOf('lista de presenca') >= 0) return 'presenca';
                        if (t2.indexOf('votacoes nominais') >= 0 && t2.indexOf('link') < 0)
                            return 'votacoes_arquivo';
                        prev = prev.previousElementSibling;
                    }
                    return '';
                }
                var root = document.querySelector('[data-cr2-pub-modal-marker="1"]') || document.body;
                var files = Array.from(root.querySelectorAll('input[type=file]'));
                var out = [];
                files.forEach(function (inp, idx) {
                    out.push({ idx: idx, key: rotuloDe(inp) });
                });
                // Se algum ficou sem key, preenche na ordem Pauta/Ata/Presenca/Votacoes
                // (Declaracao com 1 input sem rotulo → 'arquivo')
                var ordem = ['pauta', 'ata', 'presenca', 'votacoes_arquivo', 'arquivo'];
                if (files.length === 1 && !out[0].key) {
                    out[0].key = 'arquivo';
                    return out;
                }
                var usados = {};
                out.forEach(function (o) { if (o.key) usados[o.key] = true; });
                out.forEach(function (o) {
                    if (o.key) return;
                    for (var i = 0; i < ordem.length; i++) {
                        if (!usados[ordem[i]]) {
                            o.key = ordem[i];
                            usados[ordem[i]] = true;
                            break;
                        }
                    }
                });
                return out;
            }
            """
        )
    except Exception:
        marcados = []

    mapa: dict[str, Any] = {}
    scope = modal_root if modal_root is not None else page
    files = scope.locator("input[type=file]")
    for item in marcados or []:
        key = (item or {}).get("key") or ""
        idx = (item or {}).get("idx")
        if not key or idx is None:
            continue
        try:
            loc = files.nth(int(idx))
            loc.wait_for(state="attached", timeout=2000)
            mapa[key] = loc
        except Exception:
            continue
    return mapa


def _input_file_por_rotulo(scope, page, labels, mapa_cache: dict | None = None):
    """Localiza input[type=file] pelo rotulo — preferindo mapa JS (preciso)."""
    key_alias = {
        "pauta": "pauta",
        "ata": "ata",
        "lista de presenca": "presenca",
        "lista de presença": "presenca",
        "presenca": "presenca",
        "presença": "presenca",
        "votacoes nominais (arquivo)": "votacoes_arquivo",
        "votações nominais (arquivo)": "votacoes_arquivo",
        "votacoes nominais": "votacoes_arquivo",
        "votações nominais": "votacoes_arquivo",
        "arquivo": "arquivo",
        "documento": "arquivo",
        "anexar": "arquivo",
    }
    mapa = mapa_cache if mapa_cache is not None else {}
    for lb in labels:
        k = key_alias.get(normalizar(lb))
        if k and k in mapa:
            return mapa[k]

    # Fallback: texto EXATO do rotulo (nunca contains 'ata' em 'data')
    for lb in labels:
        try:
            loc = scope.get_by_text(re.compile(r"^\s*{}\s*$".format(re.escape(lb)), re.I)).first
            loc.wait_for(state="visible", timeout=2000)
            handle = loc.element_handle(timeout=1500)
            if not handle:
                continue
            file_handle = handle.evaluate_handle(
                """
                (el) => {
                    var p = el.parentElement;
                    for (var i = 0; i < 10 && p; i++) {
                        var near = p.querySelector('input[type=file]');
                        if (near) return near;
                        p = p.parentElement;
                    }
                    var n = el.nextElementSibling;
                    for (var j = 0; j < 12 && n; j++) {
                        if (n.matches && n.matches('input[type=file]')) return n;
                        var q = n.querySelector && n.querySelector('input[type=file]');
                        if (q) return q;
                        n = n.nextElementSibling;
                    }
                    return null;
                }
                """
            )
            el = file_handle.as_element() if file_handle else None
            if el:
                return el
        except Exception:
            continue
    return None


def fazer_upload_por_rotulo(page, modal_root, labels, caminho, mapa_cache=None):
    path = _caminho_arquivo(caminho)
    if path is None:
        return False
    scope = modal_root if modal_root is not None else page

    if mapa_cache is None:
        mapa_cache = _mapear_inputs_file_por_rotulo(page, modal_root)

    alvo = _input_file_por_rotulo(scope, page, labels, mapa_cache=mapa_cache)
    if alvo is None and "arquivo" in mapa_cache:
        alvo = mapa_cache["arquivo"]
    if alvo is None:
        ordem = [u[0] for u in UPLOADS] + ["arquivo"]
        key = None
        for k, lbs in list(UPLOADS) + list(UPLOADS_DECLARACAO):
            if labels[0] in lbs or any(normalizar(labels[0]) == normalizar(x) for x in lbs):
                key = k
                break
        idx = ordem.index(key) if key in ordem else -1
        if idx >= 0 and key != "arquivo":
            try:
                alvo = scope.locator("input[type=file]").nth(idx)
                alvo.wait_for(state="attached", timeout=3000)
                print("    Upload '{}' via indice {}".format(labels[0], idx))
            except Exception:
                alvo = None
        elif key == "arquivo" or normalizar(labels[0]) in ("arquivo", "documento"):
            try:
                alvo = scope.locator("input[type=file]").first
                alvo.wait_for(state="attached", timeout=3000)
                print("    Upload '{}' via primeiro input".format(labels[0]))
            except Exception:
                alvo = None

    if alvo is None:
        print("    [AVISO] Input file nao achado para '{}'".format(labels[0]))
        return False

    try:
        _revelar_input_file(page, alvo)
        time.sleep(0.05)
        alvo.set_input_files(str(path))
        time.sleep(PAUSA_APOS_ANEXAR)
        print("    Upload {}: {}".format(labels[0], path.name))
        _aguardar_confirmacao_upload(page, modal_root, path)
        _restaurar_inputs_file(page)
        return True
    except Exception as e:
        _restaurar_inputs_file(page)
        print("    [Upload] Falhou {}: {}".format(labels[0], str(e)[:80]))
        return False


def _preencher_link_votacoes(page, modal_root, link):
    if not (link or "").strip():
        return
    link = link.strip()
    scope = modal_root if modal_root is not None else page
    if _fill_by_label_candidates(scope, LABELS_LINK_VOTACOES, link, page):
        print("    Link votacoes preenchido.")
        return
    for ph in (
        re.compile(r"votacaonominal", re.I),
        re.compile(r"www\.votacao", re.I),
        re.compile(r"https?://", re.I),
    ):
        try:
            loc = scope.get_by_placeholder(ph).first
            loc.wait_for(state="visible", timeout=2000)
            preencher_campo(page, loc, link)
            print("    Link votacoes (placeholder).")
            return
        except Exception:
            continue
    print("    [AVISO] Campo link de votacoes nao encontrado.")


def preencher_modal_sessao(page, item):
    """
    Preenche o modal conforme o tipo:
      - Sessao normal: Tipo + Data + Numero + Pauta/Ata/...
      - Declaracao: Tipo + checkbox + Data (ultimo dia do mes) + Arquivo
        (sem campo Numero / sem Pauta)
    """
    modal_root = _modal_bubble_sessao(page)
    eh_decl = _eh_declaracao(item)

    if eh_decl:
        item = preparar_item_declaracao(item)

    _preencher_tipo(page, modal_root, item.get("tipo", ""))
    time.sleep(0.35)
    # Apos trocar o Tipo, o Bubble remonta o formulario
    modal_root = _modal_bubble_sessao(page)

    if eh_decl or _truthy_flag(item.get("nao_houve_publicacoes")):
        _marcar_nao_houve_publicacoes(page, modal_root, True)
        time.sleep(0.15)
        modal_root = _modal_bubble_sessao(page)

    _preencher_data(page, modal_root, item.get("data", ""))
    time.sleep(0.06)

    if not eh_decl:
        _preencher_numero(page, modal_root, item.get("numero", ""))
        time.sleep(0.06)

    uploads_pendentes = []
    if eh_decl:
        caminho_arq = item.get("arquivo") or item.get("pauta")
        if _caminho_arquivo(caminho_arq):
            uploads_pendentes.append(("arquivo", ("Arquivo", "Documento"), caminho_arq))
    else:
        for key, labels in UPLOADS:
            if _caminho_arquivo(item.get(key)):
                uploads_pendentes.append((key, labels, item.get(key)))

    if uploads_pendentes:
        mapa_files = _mapear_inputs_file_por_rotulo(page, modal_root)
        print(
            "    Anexos a subir: {}".format(
                ", ".join(u[1][0] for u in uploads_pendentes)
            )
        )
        for _key, labels, caminho in uploads_pendentes:
            ok_up = fazer_upload_por_rotulo(
                page, modal_root, labels, caminho, mapa_cache=mapa_files
            )
            # Declaracao: se nao achou rotulo, tenta o unico input file do modal
            if not ok_up and eh_decl:
                ok_up = _upload_unico_arquivo_modal(page, modal_root, caminho)
            time.sleep(0.06)
    else:
        print("    Nenhum anexo nesta linha — seguindo para Publicar.")

    if not eh_decl and (item.get("votacoes_link") or "").strip():
        _preencher_link_votacoes(page, modal_root, item.get("votacoes_link", ""))

    _restaurar_inputs_file(page)
    time.sleep(0.1)
    return modal_root


def _upload_unico_arquivo_modal(page, modal_root, caminho) -> bool:
    """Fallback da Declaracao: sobe no unico input[type=file] visivel."""
    path = _caminho_arquivo(caminho)
    if path is None:
        return False
    scope = modal_root if modal_root is not None else page
    try:
        files = scope.locator("input[type=file]")
        n = files.count()
        if n < 1:
            print("    [AVISO] Nenhum input file no modal de Declaracao.")
            return False
        alvo = files.first
        alvo.wait_for(state="attached", timeout=3000)
        _revelar_input_file(page, alvo)
        time.sleep(0.05)
        alvo.set_input_files(str(path))
        time.sleep(PAUSA_APOS_ANEXAR)
        print("    Upload Arquivo (unico input): {}".format(path.name))
        _aguardar_confirmacao_upload(page, modal_root, path)
        _restaurar_inputs_file(page)
        return True
    except Exception as e:
        _restaurar_inputs_file(page)
        print("    [Upload] Falhou Arquivo: {}".format(str(e)[:80]))
        return False


def _ler_valores_form_sessao(page, modal_root):
    """Snapshot leve dos campos Data/Número para detectar reset apos Publicar."""
    escopos = []
    if modal_root is not None:
        escopos.append(modal_root)
    escopos.append(page)
    data_v = num_v = ""
    for esc in escopos:
        if data_v and num_v:
            break
        for labels, destino in (
            (LABELS_DATA, "data"),
            (LABELS_NUMERO, "numero"),
        ):
            if destino == "data" and data_v:
                continue
            if destino == "numero" and num_v:
                continue
            for lab in labels:
                try:
                    loc = esc.get_by_label(re.compile(re.escape(lab), re.I)).first
                    if not loc.is_visible(timeout=200):
                        continue
                    val = (loc.input_value(timeout=400) or "").strip()
                    if destino == "data":
                        data_v = val
                    else:
                        num_v = val
                    break
                except Exception:
                    continue
    return data_v, num_v


def _parece_erro_real(texto: str) -> bool:
    """True só se houver padrao de erro fora dos labels do formulario."""
    if not texto:
        return False
    # Remove pedacos que sao so labels do form Bubble
    limpo = _TEXTO_LABEL_FORM_RX.sub(" ", texto)
    return bool(_TEXTO_ERRO_APOS_PUBLICAR_RX.search(limpo))


def _toast_sucesso_na_pagina(page) -> bool:
    """Alguns temas Bubble mostram alerta/toast fora do modal."""
    try:
        body = page.locator("body").inner_text(timeout=1200) or ""
    except Exception:
        return False
    # Preferir trechos curtos no topo (toast), nao o form inteiro
    trecho = body[:2500]
    if _TEXTO_SUCESSO_MODAL_RX.search(trecho):
        # Evitar casar so com titulo "Cadastrar..."
        if re.search(r"cadastrar\s+sess", trecho, re.I) and not re.search(
            r"sucesso|publicad[ao]\s+com|salvo\s+com", trecho, re.I
        ):
            return False
        return True
    return False


def aguardar_resultado_apos_publicar(page, modal_root, snapshot_antes=None):
    """
    Confirma Publicar. No portal de Sessao o Bubble as vezes:
      - fecha o modal (ok)
      - mostra toast de sucesso (ok)
      - deixa o modal aberto e limpa os campos (ok — publicou)
      - deixa o modal igual, sem mensagem (ainda publicou; nao travar a fila)
    """
    titulo_loc = _loc_modal_titulo(page)
    fim = time.monotonic() + TIMEOUT_RESULTADO_PUBLICACAO_S
    ultimo = ""
    data_antes, num_antes = snapshot_antes or ("", "")
    viu_loader = False
    estabilizou_sem_loader_desde = None

    while time.monotonic() < fim:
        # Barra de progresso do Bubble = ainda processando
        try:
            loader_ativo = page.evaluate(
                """
                () => {
                    function ativa(el) {
                        if (!el) return false;
                        var s = window.getComputedStyle(el);
                        if (s.display === 'none' || parseFloat(s.opacity) < 0.08) return false;
                        var r = el.getBoundingClientRect();
                        return r.width > 12 && r.height > 0;
                    }
                    return !!(ativa(document.querySelector('#nprogress .bar'))
                        || ativa(document.querySelector('.turbo-progress-bar')));
                }
                """
            )
        except Exception:
            loader_ativo = False
        if loader_ativo:
            viu_loader = True
            estabilizou_sem_loader_desde = None
            time.sleep(0.2)
            continue
        if viu_loader and estabilizou_sem_loader_desde is None:
            estabilizou_sem_loader_desde = time.monotonic()

        try:
            visivel = titulo_loc.is_visible(timeout=400)
        except Exception:
            visivel = False
        if not visivel:
            time.sleep(0.28)
            try:
                if titulo_loc.is_visible(timeout=400):
                    continue
            except Exception:
                pass
            print("    Modal fechou — publicacao aceita.")
            return

        try:
            if modal_root is not None:
                ultimo = modal_root.inner_text(timeout=1500)
            else:
                ultimo = page.locator("body").inner_text(timeout=1500)
        except Exception:
            ultimo = ""

        if _parece_erro_real(ultimo or ""):
            raise RuntimeError(
                "Resposta no modal apos Publicar: {}".format(
                    (ultimo or "").replace("\n", " ").strip()[:260]
                )
            )
        if _TEXTO_SUCESSO_MODAL_RX.search(ultimo or "") or _toast_sucesso_na_pagina(page):
            print("    Mensagem de sucesso detectada.")
            return

        # Form resetado (Data/Numero limpos) com modal ainda aberto = publicou
        if data_antes or num_antes:
            data_agora, num_agora = _ler_valores_form_sessao(page, modal_root)
            limpou_data = bool(data_antes) and not data_agora
            limpou_num = bool(num_antes) and not num_agora
            if limpou_data or limpou_num:
                print("    Formulario limpo apos Publicar — publicacao aceita.")
                return

        # Loader ja passou e passaram ~4s sem erro → Bubble costuma ter gravado
        if (
            viu_loader
            and estabilizou_sem_loader_desde is not None
            and (time.monotonic() - estabilizou_sem_loader_desde) >= 4.0
        ):
            print(
                "    Sem mensagem no modal, mas o envio terminou — "
                "assumindo publicacao OK (portal de Sessao)."
            )
            return

        time.sleep(0.42)

    # Timeout: se nao houve erro claro, nao derruba a fila (publicacao costuma ter ido)
    if _parece_erro_real(ultimo or ""):
        raise TimeoutError(
            "Sem confirmacao apos Publicar ({}s). Ultimo texto: {}".format(
                TIMEOUT_RESULTADO_PUBLICACAO_S,
                (ultimo or "").replace("\n", " ").strip()[:200],
            )
        )
    print(
        "    [AVISO] Modal ainda aberto apos {}s sem erro visivel — "
        "seguindo (confira no portal se a sessao entrou).".format(
            TIMEOUT_RESULTADO_PUBLICACAO_S
        )
    )
    return


def clicar_publicar(page):
    _restaurar_inputs_file(page)
    time.sleep(0.12)
    modal_root = _modal_bubble_sessao(page)
    snapshot = _ler_valores_form_sessao(page, modal_root)
    if modal_root is not None:
        btn = modal_root.locator("button:has-text('Publicar')").first
    else:
        btn = page.locator("button:has-text('Publicar')").first
    btn.wait_for(state="visible", timeout=15000)
    btn.scroll_into_view_if_needed()
    limite = time.monotonic() + TIMEOUT_PUBLICAR_HABILITADO_S
    while time.monotonic() < limite:
        try:
            if btn.is_enabled():
                break
        except Exception:
            pass
        time.sleep(0.35)
    else:
        salvar_screenshot(page, "TIMEOUT_PUBLICAR_SESSAO")
        raise TimeoutError("Botao Publicar desabilitado por demais tempo.")
    aguardar_barra_carregamento_topo(page, etiqueta="antes de Publicar")
    time.sleep(0.15)
    # Garante que nenhum input file transparente cobre o botao
    _restaurar_inputs_file(page)
    try:
        box = btn.bounding_box()
        if box:
            page.mouse.click(
                box["x"] + box["width"] / 2,
                box["y"] + box["height"] / 2,
            )
        else:
            btn.click(timeout=15000)
    except Exception:
        btn.click(force=True, timeout=15000)
    print("    Clicou em Publicar.")
    time.sleep(0.25)
    aguardar_resultado_apos_publicar(page, modal_root, snapshot_antes=snapshot)
    time.sleep(PAUSA_APOS_CLICAR_PUBLICAR)


def publicar_um(page, item, idx, total):
    _abortar_se_cancelado()
    rotulo = item.get("numero") or item.get("data") or "sessao"
    print(
        "[-> SESSAO] [{}/{}] {} | {} | {}".format(
            idx, total, item.get("tipo"), item.get("data"), rotulo
        )
    )
    abrir_modal(page)
    preencher_modal_sessao(page, item)
    clicar_publicar(page)
    try:
        if _loc_modal_titulo(page).is_visible():
            fechar_modal(page)
    except Exception:
        try:
            fechar_modal(page)
        except Exception:
            pass
    print("    [OK] Concluido.")


# ---------------------------------------------------------------------
#  Main
# ---------------------------------------------------------------------

def parse_args(argv=None):
    p = argparse.ArgumentParser(description="Publicacao de Sessao — portal CR2")
    p.add_argument("--test", action="store_true", help="Publica so a 1a sessao da fila")
    p.add_argument("--yes", action="store_true", help="Pula Enter pos-login")
    p.add_argument("--headless", action="store_true")
    p.add_argument(
        "--pasta",
        type=str,
        default="",
        help="Pasta com subpastas de sessao (ex.: ...\\sessoes_2021)",
    )
    p.add_argument("--csv", type=str, default="", help="Caminho do CSV da fila (fallback)")
    return p.parse_args(argv)


def main(argv=None):
    args = parse_args(argv)
    global HEADLESS, MODO_TESTE, CSV_FILA, PASTA_SESSOES
    if args.headless:
        HEADLESS = True
    if args.test:
        MODO_TESTE = True
    if args.pasta:
        PASTA_SESSOES = Path(args.pasta)
    if args.csv:
        CSV_FILA = Path(args.csv)

    if not url_portal_ativa(URL_PORTAL_SESSAO):
        raise ValueError(
            "URL_PORTAL_SESSAO vazia. Informe a URL admin do modulo Sessao "
            "(ex.: https://www.portalcr2.com.br/sessoes/...)."
        )

    fila = montar_fila()
    if not fila:
        raise ValueError(
            "Nenhuma sessao na fila. Informe PASTA_SESSOES "
            "(ex.: ...\\sessoes_2021), REGISTRO_UNICO ou CSV ({}).".format(CSV_FILA)
        )
    if MODO_TESTE:
        fila = fila[:1]
        print("[INFO] Modo teste: 1 sessao.")

    print("=" * 60)
    print("  PUBLICACAO DE SESSAO — portal CR2")
    print("  Total: {}".format(len(fila)))
    print("  Pasta: {}".format(PASTA_SESSOES))
    print("  URL: {}".format(URL_PORTAL_SESSAO))
    print("  Ritmo: apos cada sessao, {}s e abre a proxima".format(PAUSA_ENTRE_SESSOES))
    print("=" * 60)

    pw = browser = page = None
    ok = erros = 0
    try:
        pw, browser, page = criar_navegador_e_login(pular_enter_pos_login=args.yes)
        garantir_pagina_portal(page, URL_PORTAL_SESSAO, "Sessao")
        for i, item in enumerate(fila, 1):
            try:
                publicar_um(page, item, i, len(fila))
                ok += 1
                if i < len(fila):
                    print(
                        "    Proxima sessao em {:.0f}s...".format(PAUSA_ENTRE_SESSOES)
                    )
                    time.sleep(PAUSA_ENTRE_SESSOES)
            except Cancelado:
                raise
            except Exception as e:
                erros += 1
                print("    [ERRO] {}".format(e))
                salvar_screenshot(page, "sessao_erro_{}".format(i))
                try:
                    fechar_modal(page)
                except Exception:
                    pass
                if i < len(fila):
                    time.sleep(PAUSA_ENTRE_SESSOES)
    finally:
        if browser:
            try:
                browser.close()
            except Exception:
                pass
        if pw:
            try:
                pw.stop()
            except Exception:
                pass

    print("")
    print("=" * 60)
    print("  CONCLUIDO! OK: {} | Erros: {} | Total: {}".format(ok, erros, len(fila)))
    print("=" * 60)
    if erros:
        raise RuntimeError("Publicacao de sessao com {} erro(s).".format(erros))


if __name__ == "__main__":
    main()
