# -*- coding: utf-8 -*-
"""Planilhas da Extração Pro: relatório por categoria + auditoria global."""

from __future__ import annotations

import re
from pathlib import Path
from typing import Any

# Colunas do relatório por categoria / consolidado
CAMPOS_RELATORIO = [
    ("categoria", "Categoria"),
    ("tipo", "Tipo"),
    ("numero", "Número"),
    ("ano", "Ano"),
    ("data", "Data"),
    ("titulo", "Título / ementa"),
    ("arquivo", "Arquivo"),
    ("caminho", "Caminho"),
    ("url_pdf", "URL do PDF"),
    ("url_fonte", "URL da fonte"),
    ("status", "Status"),
    ("observacao", "Observação"),
]

# Planilha pronta para subir em Matérias Legislativas (portal CR2 / Automações)
CAMPOS_MATERIAS = [
    ("tipo", "Tipo"),
    ("numero_pub", "Número"),
    ("descricao", "Descrição"),
    ("data_publicacao", "Data de Publicação"),
    ("autoria", "Autoria"),
    ("situacao", "Situação"),
    ("arquivo", "Arquivo"),
    ("caminho", "Caminho"),
]

CAMPOS_AUDITORIA = [
    ("categoria", "Categoria"),
    ("arquivo", "Arquivo"),
    ("campo", "Campo"),
    ("valor", "Valor"),
    ("origem", "Origem do dado"),
    ("metodo", "Método"),
    ("trecho", "Trecho / evidência"),
    ("status_doc", "Status do documento"),
]

# Tipos do módulo Matérias Legislativas no portal CR2
TIPOS_MATERIAS_LEGISLATIVAS = {
    "Anteprojeto de Lei",
    "Emendas Impositivas",
    "Indicação",
    "Iniciativa Popular",
    "Moção",
    "Moção de Aplauso",
    "Moção de Pesar",
    "Moção de Reconhecimento",
    "Pedido de Providência",
    "Projeto de Decreto Legislativo",
    "Projeto de Emenda à Lei Orgânica",
    "Projeto de Emenda ao Regimento Interno",
    "Projeto de Emenda ao RI",
    "Projeto de Indicação",
    "Projeto de Lei",
    "Projeto de Lei Complementar",
    "Projeto de Resolução",
    "Proposições",
    "Requerimento",
    "Veto",
    "Devolvido ao Executivo",
}

_SITUACOES_CONHECIDAS = [
    "Matéria Lida",
    "Em Tramitação",
    "Em tramitação",
    "Aprovado",
    "Aprovada",
    "Arquivado",
    "Arquivada",
    "Sancionado",
    "Sancionada",
    "Vetado",
    "Vetada",
    "Retirado",
    "Retirada",
    "Prejudicado",
    "Prejudicada",
    "Transformado em Norma Jurídica",
    "Aguardando",
    "Publicado",
    "Publicada",
]

_MESES = {
    "janeiro": 1,
    "fevereiro": 2,
    "marco": 3,
    "março": 3,
    "abril": 4,
    "maio": 5,
    "junho": 6,
    "julho": 7,
    "agosto": 8,
    "setembro": 9,
    "outubro": 10,
    "novembro": 11,
    "dezembro": 12,
}

_RE_NUM_ANO = re.compile(
    r"(?:n[º°o\.]*\s*)?(\d{1,4})\s*/\s*((?:20|19)\d{2})",
    re.I,
)
_RE_DATA_EXTENSO = re.compile(
    r"(?:de\s+)?(\d{1,2})\s+de\s+"
    r"(janeiro|fevereiro|mar[cç]o|abril|maio|junho|julho|agosto|"
    r"setembro|outubro|novembro|dezembro)\s+de\s+((?:20|19)\d{2})",
    re.I,
)
_RE_DATA_NUM = re.compile(
    r"\b(\d{1,2})[./\-](\d{1,2})[./\-]((?:20|19)\d{2})\b"
)


def _norm(texto: str) -> str:
    return re.sub(r"\s+", " ", (texto or "").strip())


def _limpar_nome_pasta(nome: str) -> str:
    t = _norm(nome)
    t = re.sub(r'[<>:"/\\|?*]', " ", t)
    t = re.sub(r"\s+", " ", t).strip(" .-_")
    return t[:120] if t else "Categoria"


def extrair_data(*textos: str) -> tuple[str, str, str]:
    """Retorna (data dd/mm/aaaa, origem_label, trecho) ou ('', '', '')."""
    for rotulo, texto in (
        ("listagem", textos[0] if textos else ""),
        ("titulo", textos[1] if len(textos) > 1 else ""),
        ("pdf", textos[2] if len(textos) > 2 else ""),
        ("url", textos[3] if len(textos) > 3 else ""),
    ):
        t = _norm(texto)
        if not t:
            continue
        m = _RE_DATA_EXTENSO.search(t)
        if m:
            dia = int(m.group(1))
            mes = _MESES.get(m.group(2).lower().replace("ç", "c"), 0)
            if not mes:
                mes = _MESES.get(m.group(2).lower(), 0)
            ano = int(m.group(3))
            if mes and 1 <= dia <= 31 and 1900 <= ano <= 2100:
                data = f"{dia:02d}/{mes:02d}/{ano}"
                return data, rotulo, m.group(0)[:120]
        m = _RE_DATA_NUM.search(t)
        if m:
            dia, mes, ano = int(m.group(1)), int(m.group(2)), int(m.group(3))
            if 1 <= dia <= 31 and 1 <= mes <= 12 and 1900 <= ano <= 2100:
                data = f"{dia:02d}/{mes:02d}/{ano}"
                return data, rotulo, m.group(0)
    return "", "", ""


def data_para_epoch(data: str) -> float | None:
    """Converte dd/mm/aaaa ou dd-mm-aaaa em timestamp Unix (meio-dia local)."""
    from datetime import datetime

    t = _norm(data).replace("-", "/")
    m = re.match(r"^(\d{1,2})/(\d{1,2})/((?:20|19)\d{2})$", t)
    if not m:
        return None
    try:
        dt = datetime(int(m.group(3)), int(m.group(2)), int(m.group(1)), 12, 0, 0)
        return dt.timestamp()
    except ValueError:
        return None


def ano_para_epoch(ano: int | str | None) -> float | None:
    """Fallback: 01/01/AAAA ao meio-dia — pelo menos o ano fica certo no Explorer."""
    from datetime import datetime

    try:
        a = int(str(ano).strip())
    except (TypeError, ValueError):
        return None
    if a < 1900 or a > 2100:
        return None
    try:
        return datetime(a, 1, 1, 12, 0, 0).timestamp()
    except ValueError:
        return None


def extrair_numero_ano_com_origem(
    *pares: tuple[str, str],
) -> tuple[str, str, str, str, str]:
    """
    pares: (rotulo_origem, texto)
    Retorna (numero, ano, origem, metodo, trecho)
    """
    for rotulo, texto in pares:
        t = _norm(texto)
        if not t:
            continue
        m = _RE_NUM_ANO.search(t)
        if m:
            return (
                str(int(m.group(1))).zfill(3),
                m.group(2),
                rotulo,
                "padrão Nºxxx/aaaa",
                m.group(0)[:120],
            )
        m = re.search(
            r"n[º°o\.]*\s*(\d{1,4}).{0,40}?(20\d{2}|19\d{2})",
            t,
            re.I,
        )
        if m:
            return (
                str(int(m.group(1))).zfill(3),
                m.group(2),
                rotulo,
                "Nº … de aaaa",
                m.group(0)[:120],
            )
        m = re.search(r"(?<!\d)(\d{1,4})[.\-](20\d{2}|19\d{2})(?!\d)", t)
        if m:
            return (
                str(int(m.group(1))).zfill(3),
                m.group(2),
                rotulo,
                "basename nnn.aaaa",
                m.group(0)[:120],
            )
    return "", "", "", "", ""


def eh_materia_legislativa(tipo: str, categoria: str = "") -> bool:
    """True se o documento entra no módulo Matérias Legislativas do CR2."""
    t = _norm(tipo)
    if t in TIPOS_MATERIAS_LEGISLATIVAS:
        return True
    # aliases / variações
    low = t.lower()
    for nome in TIPOS_MATERIAS_LEGISLATIVAS:
        if low == nome.lower():
            return True
    cat = _norm(categoria).lower()
    if any(
        x in cat
        for x in (
            "matéria",
            "materia",
            "proposi",
            "projeto de lei",
            "requerimento",
            "moção",
            "mocao",
            "indicação",
            "indicacao",
            "veto",
        )
    ):
        # categoria sugere matérias, mesmo se tipo genérico
        if t and t not in ("Portaria", "Decreto", "Lei", "Lei Complementar", "Ofício"):
            return True
    return False


def formatar_numero_publicacao(numero: str, ano: str) -> str:
    """Número no formato do portal: 001/2024."""
    n = re.sub(r"\D", "", str(numero or ""))
    a = re.sub(r"\D", "", str(ano or ""))
    if n and a and len(a) == 4:
        return f"{int(n):03d}/{a}"
    if n and a:
        return f"{n}/{a}"
    if n:
        return n
    return ""


def extrair_descricao(titulo: str, tipo: str = "") -> tuple[str, str, str]:
    """
    Ementa/descrição a partir do título da listagem.
    Ex.: 'PROJETO DE LEI Nº 001/2024, DE 08 DE JANEIRO DE 2024 (Autoriza...)'
         → 'Autoriza...'
    Retorna (descricao, origem, trecho).
    """
    t = _norm(titulo)
    if not t:
        return "", "", ""

    # Remove cauda de Autoria/Situação se vieram coladas no mesmo texto
    t_limpo = re.split(
        r"\b(?:Autoria|Tipo\s+de\s+[Aa]utoria|Situa[cç][aã]o)\s*:",
        t,
        maxsplit=1,
        flags=re.I,
    )[0].strip()

    # Conteúdo entre o primeiro '(' e o último ')' (ementa com parênteses internos)
    m = re.search(r"\((.+)\)\s*$", t_limpo)
    if m and len(_norm(m.group(1))) >= 15:
        desc = _norm(m.group(1))
        return desc, "listagem", m.group(0)[:120]

    desc = t_limpo
    if tipo:
        desc = re.sub(re.escape(tipo), "", desc, count=1, flags=re.I)
    desc = re.sub(
        r"n[º°o\.]*\s*\d{1,4}\s*/\s*(?:20|19)\d{2}",
        "",
        desc,
        count=1,
        flags=re.I,
    )
    desc = re.sub(
        r",?\s*de\s+\d{1,2}\s+de\s+"
        r"(?:janeiro|fevereiro|mar[cç]o|abril|maio|junho|julho|agosto|"
        r"setembro|outubro|novembro|dezembro)\s+de\s+(?:20|19)\d{2}",
        "",
        desc,
        count=1,
        flags=re.I,
    )
    desc = re.sub(r"^\s*[,.\-–—:\s\(]+", "", desc)
    desc = desc.rstrip(" )")
    desc = _norm(desc)
    if len(desc) < 8:
        return t_limpo[:500], "listagem", t_limpo[:120]
    return desc[:500], "listagem", desc[:120]


def extrair_autoria(*textos: str) -> tuple[str, str, str]:
    """Retorna (autoria, origem, trecho)."""
    for rotulo, texto in (
        ("listagem", textos[0] if textos else ""),
        ("titulo", textos[1] if len(textos) > 1 else ""),
        ("pdf", textos[2] if len(textos) > 2 else ""),
    ):
        t = _norm(texto)
        if not t:
            continue
        for rx in (
            r"autoria\s*:\s*([^\n|;]{3,120})",
            r"autor(?:a|es)?\s*:\s*([^\n|;]{3,120})",
            r"tipo\s+de\s+autoria\s*:\s*([^\n|;]{3,80})",
            r"(poder\s+executivo)",
            r"(mesa\s+diretora)",
            r"(presidente\s+da\s+c[aâ]mara(?:\s+municipal)?)",
            r"(sem\s+autoria)",
            r"vereador(?:a)?\s+([A-ZÁÉÍÓÚÂÊÔÃÕÇ][\wÀ-ú\.\'\-\s]{2,80})",
            r"deputad[oa]\s+([A-ZÁÉÍÓÚÂÊÔÃÕÇ][\wÀ-ú\.\'\-\s]{2,80})",
        ):
            m = re.search(rx, t, re.I)
            if m:
                valor = _norm(m.group(1))
                valor = re.split(r"\s{2,}|Situa[cç][aã]o\s*:|Ementa\s*:", valor)[0]
                valor = valor.strip(" .,;|-")
                if len(valor) >= 3:
                    return valor[:120], rotulo, m.group(0)[:120]
    return "", "", ""


def extrair_situacao(*textos: str) -> tuple[str, str, str]:
    """Retorna (situacao, origem, trecho)."""
    blob = " | ".join(_norm(x) for x in textos if x)
    if not blob:
        return "", "", ""
    for sit in sorted(_SITUACOES_CONHECIDAS, key=len, reverse=True):
        if re.search(re.escape(sit), blob, re.I):
            m = re.search(re.escape(sit), blob, re.I)
            return sit, "listagem/pdf", (m.group(0) if m else sit)
    m = re.search(r"situa[cç][aã]o\s*:\s*([^\n|;]{3,80})", blob, re.I)
    if m:
        return _norm(m.group(1))[:80], "listagem/pdf", m.group(0)[:120]
    return "", "", ""


def montar_registro(
    *,
    categoria: str,
    tipo: str,
    tipo_origem: str,
    tipo_trecho: str,
    numero: str,
    ano: str,
    num_origem: str,
    num_metodo: str,
    num_trecho: str,
    data: str,
    data_origem: str,
    data_trecho: str,
    titulo: str,
    arquivo: str,
    caminho: str,
    url_pdf: str,
    url_fonte: str,
    status: str,
    observacao: str = "",
    descricao: str = "",
    descricao_origem: str = "",
    descricao_trecho: str = "",
    autoria: str = "",
    autoria_origem: str = "",
    autoria_trecho: str = "",
    situacao: str = "",
    situacao_origem: str = "",
    situacao_trecho: str = "",
) -> dict[str, Any]:
    """Monta linha do relatório + itens de auditoria (_auditoria)."""
    auditoria: list[dict[str, str]] = []

    def _aud(campo: str, valor: str, origem: str, metodo: str, trecho: str) -> None:
        if not (valor or "").strip():
            return
        auditoria.append(
            {
                "categoria": categoria or "",
                "arquivo": arquivo or "",
                "campo": campo,
                "valor": valor,
                "origem": origem or "—",
                "metodo": metodo or "",
                "trecho": (trecho or "")[:200],
                "status_doc": status or "",
            }
        )

    numero_pub = formatar_numero_publicacao(numero, ano)
    desc = (descricao or "").strip()
    if not desc:
        desc, descricao_origem, descricao_trecho = extrair_descricao(titulo, tipo)

    _aud("Tipo", tipo, tipo_origem, "detecção por regras/catálogo", tipo_trecho)
    _aud("Número", numero_pub or numero, num_origem, num_metodo, num_trecho)
    _aud("Ano", ano, num_origem or "pasta/filtro", num_metodo or "ano_fallback", num_trecho or str(ano))
    _aud(
        "Data de Publicação",
        data,
        data_origem,
        "data por extenso ou numérica",
        data_trecho,
    )
    _aud(
        "Descrição",
        (desc or "")[:180],
        descricao_origem or "listagem/página",
        "ementa / texto após tipo e data",
        (descricao_trecho or desc or "")[:120],
    )
    _aud("Autoria", autoria, autoria_origem, "padrão Autoria:/Autor:/Vereador", autoria_trecho)
    _aud("Situação", situacao, situacao_origem, "rótulo conhecido ou Situação:", situacao_trecho)
    if titulo and titulo != desc:
        _aud(
            "Título bruto",
            (titulo or "")[:180],
            "listagem/página",
            "texto do card ou título do post",
            (titulo or "")[:120],
        )

    return {
        "categoria": categoria or "",
        "tipo": tipo or "",
        "numero": numero or "",
        "numero_pub": numero_pub or (numero or ""),
        "ano": str(ano or ""),
        "data": data or "",
        "data_publicacao": data or "",
        "titulo": (titulo or "")[:500],
        "descricao": (desc or "")[:500],
        "autoria": (autoria or "")[:120],
        "situacao": (situacao or "")[:80],
        "arquivo": arquivo or "",
        "caminho": caminho or "",
        "url_pdf": url_pdf or "",
        "url_fonte": url_fonte or "",
        "status": status or "",
        "observacao": observacao or "",
        "eh_materia": eh_materia_legislativa(tipo, categoria),
        "_auditoria": auditoria,
    }


def _escrever_aba(ws, registros: list[dict[str, Any]], campos: list[tuple[str, str]]) -> None:
    from openpyxl.styles import Font

    for i, (_, rotulo) in enumerate(campos, 1):
        ws.cell(row=1, column=i, value=rotulo).font = Font(name="Arial", size=10, bold=True)
    for r, reg in enumerate(registros, 2):
        for c, (chave, _) in enumerate(campos, 1):
            ws.cell(row=r, column=c, value=reg.get(chave) or "")


def salvar_relatorio_categoria(
    registros: list[dict[str, Any]],
    pasta_categoria: str | Path,
    nome_categoria: str = "",
) -> Path | None:
    """Grava Relatorio.xlsx dentro da pasta da categoria.

    Se forem Matérias Legislativas, a aba principal usa os campos do portal CR2.
    """
    from openpyxl import Workbook

    pasta = Path(pasta_categoria)
    pasta.mkdir(parents=True, exist_ok=True)
    da_cat = [
        r
        for r in registros
        if _limpar_nome_pasta(r.get("categoria") or "")
        == _limpar_nome_pasta(nome_categoria or pasta.name)
        or (r.get("categoria") or "") == (nome_categoria or pasta.name)
    ]
    if not da_cat:
        da_cat = list(registros)
    if not da_cat:
        return None

    materias = [r for r in da_cat if r.get("eh_materia")]
    path = pasta / "Relatorio.xlsx"
    wb = Workbook()

    if materias and len(materias) >= max(1, len(da_cat) // 2):
        # Categoria predominantemente de matérias → planilha no formato Automações
        ws = wb.active
        ws.title = "Matérias Legislativas"
        _escrever_aba(ws, materias, CAMPOS_MATERIAS)
        if len(da_cat) > len(materias):
            wo = wb.create_sheet("Outros")
            _escrever_aba(wo, [r for r in da_cat if not r.get("eh_materia")], CAMPOS_RELATORIO)
    else:
        ws = wb.active
        ws.title = "Documentos"
        _escrever_aba(ws, da_cat, CAMPOS_RELATORIO)
        if materias:
            wm = wb.create_sheet("Matérias Legislativas")
            _escrever_aba(wm, materias, CAMPOS_MATERIAS)

    aud: list[dict[str, Any]] = []
    for r in da_cat:
        aud.extend(r.get("_auditoria") or [])
    wa = wb.create_sheet("Auditoria")
    _escrever_aba(wa, aud, CAMPOS_AUDITORIA)
    wb.save(path)
    return path


def salvar_planilha_materias(
    registros: list[dict[str, Any]],
    pasta_base: str | Path,
) -> Path | None:
    """Grava Materias_Legislativas.xlsx (pronta para Automações / portal CR2)."""
    from openpyxl import Workbook

    materias = [
        r
        for r in registros
        if r.get("eh_materia") and (r.get("status") or "") in ("ok", "pulado", "")
    ]
    if not materias:
        return None

    pasta = Path(pasta_base)
    pasta.mkdir(parents=True, exist_ok=True)
    path = pasta / "Materias_Legislativas.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "Matérias Legislativas"
    _escrever_aba(ws, materias, CAMPOS_MATERIAS)
    aud: list[dict[str, Any]] = []
    for r in materias:
        aud.extend(r.get("_auditoria") or [])
    wa = wb.create_sheet("Auditoria")
    _escrever_aba(wa, aud, CAMPOS_AUDITORIA)
    wb.save(path)
    return path


def salvar_planilhas_normas(
    registros: list[dict[str, Any]],
    pasta_base: str | Path,
) -> dict[str, str]:
    """
    Grava:
      - {PASTA_BASE}/{Categoria}/Relatorio.xlsx  (por categoria)
      - {PASTA_BASE}/Materias_Legislativas.xlsx  (campos do portal CR2)
      - {PASTA_BASE}/Normas.xlsx                 (consolidado)
      - {PASTA_BASE}/Auditoria_Normas.xlsx       (auditoria global)
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font

    pasta = Path(pasta_base)
    pasta.mkdir(parents=True, exist_ok=True)
    saidas: dict[str, str] = {}
    if not registros:
        return saidas

    # Por categoria
    por_cat: dict[str, list[dict[str, Any]]] = {}
    for r in registros:
        cat = _limpar_nome_pasta(r.get("categoria") or "Geral") or "Geral"
        por_cat.setdefault(cat, []).append(r)

    for cat, regs in por_cat.items():
        dest = pasta / cat
        p = salvar_relatorio_categoria(regs, dest, cat)
        if p:
            saidas[f"categoria:{cat}"] = str(p)
            print(f"  [PLANILHA] {cat}: {p} ({len(regs)} docs)")

    # Matérias Legislativas (upload Automações)
    p_mat = salvar_planilha_materias(registros, pasta)
    if p_mat:
        n_mat = sum(1 for r in registros if r.get("eh_materia"))
        saidas["materias"] = str(p_mat)
        print(f"  [PLANILHA] Matérias Legislativas: {p_mat} ({n_mat} docs)")

    # Consolidado
    path_normas = pasta / "Normas.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "Documentos"
    _escrever_aba(ws, registros, CAMPOS_RELATORIO)
    materias = [r for r in registros if r.get("eh_materia")]
    if materias:
        wm = wb.create_sheet("Matérias Legislativas")
        _escrever_aba(wm, materias, CAMPOS_MATERIAS)
    aud_all: list[dict[str, Any]] = []
    for r in registros:
        aud_all.extend(r.get("_auditoria") or [])
    wa = wb.create_sheet("Auditoria")
    _escrever_aba(wa, aud_all, CAMPOS_AUDITORIA)
    wb.save(path_normas)
    saidas["normas"] = str(path_normas)
    print(f"  [PLANILHA] Consolidado: {path_normas} ({len(registros)} docs)")

    # Auditoria dedicada
    path_aud = pasta / "Auditoria_Normas.xlsx"
    wb2 = Workbook()
    w2 = wb2.active
    w2.title = "Auditoria"
    _escrever_aba(w2, aud_all, CAMPOS_AUDITORIA)
    wr = wb2.create_sheet("Resumo")
    wr["A1"] = "Campo"
    wr["B1"] = "Preenchidos"
    wr["A1"].font = Font(bold=True)
    wr["B1"].font = Font(bold=True)
    cont: dict[str, int] = {}
    for item in aud_all:
        c = item.get("campo") or "?"
        cont[c] = cont.get(c, 0) + 1
    for i, (campo, n) in enumerate(sorted(cont.items()), 2):
        wr.cell(row=i, column=1, value=campo)
        wr.cell(row=i, column=2, value=n)
    wr.cell(row=len(cont) + 3, column=1, value="Documentos no relatório")
    wr.cell(row=len(cont) + 3, column=2, value=len(registros))
    wr.cell(row=len(cont) + 4, column=1, value="Matérias Legislativas")
    wr.cell(row=len(cont) + 4, column=2, value=sum(1 for r in registros if r.get("eh_materia")))
    wb2.save(path_aud)
    saidas["auditoria"] = str(path_aud)
    print(f"  [PLANILHA] Auditoria: {path_aud} ({len(aud_all)} campos)")

    return saidas
