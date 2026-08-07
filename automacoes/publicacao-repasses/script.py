# -*- coding: utf-8 -*-
"""
Publicação de Repasses — portal CR2 (Playwright / Bubble)

Fluxo (diferente das outras publicações):
  1) "Criar Publicação" → modal "Cadastrar Repasse"
  2) Preenche Mês/Ano, Data, Valores e Descrição (= "Repasse") — SEM link e SEM arquivo
  3) Publicar → diálogo "Deseja anexar documentos…?" → Anexar
  4) "Criar Documento" → sobe o PDF → lápis Editar
  5) "Editar Documento" → Data de Publicação (+ descrição do arquivo) → Editar
  6) Finalizar

Entrada (prioridade):
  1) REGISTRO_UNICO (painel)
  2) PASTA_BASE\\Repasses.xlsx (+ PDFs em PASTA_BASE\\Repasses\\<ano>\\)
  3) PASTA_BASE\\Repasses.csv

Flags: --test  --yes  --headless  --pasta CAMINHO
"""

from __future__ import annotations

import argparse
import csv
import os
import re
import shutil
import subprocess
import sys
import time
import unicodedata
from pathlib import Path
from typing import Any

try:
    from playwright.sync_api import TimeoutError as PWTimeout
    from playwright.sync_api import sync_playwright
except ImportError:
    PWTimeout = None
    sync_playwright = None


class Cancelado(Exception):
    """Fila interrompida pelo usuario (centro-automacoes)."""


def pedido_cancelado():
    return False


def _abortar_se_cancelado():
    if pedido_cancelado():
        print("[AVISO] Fila cancelada pelo usuario.")
        raise Cancelado()


# ---------------------------------------------------------------------
# CONFIG
# ---------------------------------------------------------------------

URL_LOGIN = "https://www.portalcr2.com.br/?view=login"
URL_PORTAL_REPASSE = ""  # ex.: https://www.portalcr2.com.br/.../repasses

PORTAL_USUARIO = ""
PORTAL_SENHA = ""

HEADLESS = False
MODO_TESTE = False
ABRIR_LOGIN_ANTES_DO_PORTAL = True
PORTAL_LOGIN_BOTAO = "Entrar"

# Pasta com Repasses.xlsx e subpasta Repasses\ANO\*.pdf (saida do download-repasses)
PASTA_BASE = Path(r"C:\Downloads\repasses")
REGISTRO_UNICO = None

OPERA_EXE = None
PASTA_SCREENSHOTS = Path(__file__).resolve().parent / "screenshots_pub"

PAUSA_APOS_ANEXAR = 0.28
PAUSA_POLL_UPLOAD_UI = 0.18
MAX_TENTATIVAS_POLL_UPLOAD = 14
PAUSA_APOS_CONFIRMAR_UPLOAD = 0.35
TIMEOUT_PUBLICAR_HABILITADO_S = 60
TIMEOUT_RESULTADO_PUBLICACAO_S = 22
TIMEOUT_LOADER_TOPO_S = 120
PAUSA_APOS_CLICAR_PUBLICAR = 0.18
PAUSA_ENTRE_ITENS = 0.9

DESCRICAO_PADRAO_REPASSE = "Repasse"
TIMEOUT_DIALOGO_ANEXAR_S = 18
TIMEOUT_CRIAR_DOCUMENTO_S = 22
TIMEOUT_EDITAR_DOCUMENTO_S = 15
PAUSA_APOS_ANEXAR_DIALOGO = 0.3
PAUSA_APOS_UPLOAD_DOC = 0.55
PAUSA_APOS_EDITAR_DOC = 0.35
PAUSA_APOS_FINALIZAR = 0.4
PAUSA_ENTRE_CAMPOS = 0.05
PAUSA_SAIR_CAMPO = 0.1
PAUSA_APOS_DATA_DOC_S = 1.0

MODAL_TITULO_REGEX = r"(Criar|Cadastrar).*Repasse"

LABELS_LINK = ("Link",)
LABELS_MES_ANO = ("Mês e Ano", "Mes e Ano", "Mês/Ano", "Mes/Ano", "Competência", "Competencia")
LABELS_DATA = ("Data", "Data do repasse", "Data do Repasse")
LABELS_VALOR_PREVISTO = (
    "Valor Previsto (R$)",
    "Valor Previsto",
    "Valor previsto",
    "Previsto",
)
LABELS_VALOR_REALIZADO = (
    "Valor Realizado (R$)",
    "Valor Realizado",
    "Valor realizado",
    "Realizado",
)
LABELS_DESCRICAO = ("Descrição", "Descricao")
LABELS_ARQUIVO = ("Arquivo", "Documento", "Anexar")

_TEXTO_ERRO_APOS_PUBLICAR_RX = re.compile(
    r"(?:\berro\b|\bfalha\b|inv[aá]lid|obrigat[oó]rio|"
    r"n[aã]o\s+foi\s+poss[ií]vel|tente\s+novamente|"
    r"j[aá]\s+existe|duplicad|\bduplicat)",
    re.I,
)
_TEXTO_SUCESSO_MODAL_RX = re.compile(
    r"(publicado\s+com\s+sucesso|publicada\s+com\s+sucesso|"
    r"salvo\s+com\s+sucesso|salva\s+com\s+sucesso|"
    r"cadastrado\s+com\s+sucesso|cadastrada\s+com\s+sucesso|"
    r"\bsucesso\b)",
    re.I,
)
_TEXTO_LABEL_FORM_RX = re.compile(
    r"cadastrar\s+repasse|criar\s+publica|valor\s+previsto|valor\s+realizado",
    re.I,
)

CAMPOS = (
    "link",
    "mes_ano",
    "data",
    "valor_previsto",
    "valor_realizado",
    "descricao",
    "arquivo",
)


# ---------------------------------------------------------------------
# Playwright bootstrap
# ---------------------------------------------------------------------

def _recarregar_playwright():
    global PWTimeout, sync_playwright
    try:
        from playwright.sync_api import TimeoutError as PWTimeout
        from playwright.sync_api import sync_playwright
        return True
    except ImportError:
        PWTimeout = None
        sync_playwright = None
        return False


def garantir_playwright_pronto():
    if sync_playwright is not None:
        return
    subprocess.run([sys.executable, "-m", "pip", "install", "playwright"], check=False)
    if _recarregar_playwright():
        return
    subprocess.run([sys.executable, "-m", "playwright", "install", "chromium"], check=False)
    _recarregar_playwright()


def verificar_playwright_instalado():
    if sync_playwright is not None:
        return
    garantir_playwright_pronto()
    if sync_playwright is None:
        print("[ERRO] Playwright indisponivel. pip install playwright")
        sys.exit(1)


def _opera_via_program_files():
    local = os.environ.get("LOCALAPPDATA", "")
    prog = os.environ.get("PROGRAMFILES", "")
    prog_x86 = os.environ.get("PROGRAMFILES(X86)", "")
    for p in (
        Path(local) / "Programs" / "Opera" / "opera.exe",
        Path(local) / "Programs" / "Opera GX" / "opera.exe",
        Path(prog) / "Opera" / "opera.exe",
        Path(prog_x86) / "Opera" / "opera.exe",
    ):
        if p.is_file():
            return p.resolve()
    return None


def resolver_caminho_opera():
    if OPERA_EXE:
        p = Path(OPERA_EXE)
        if p.is_file():
            return p.resolve()
    found = shutil.which("opera") or shutil.which("opera.exe")
    if found:
        return Path(found).resolve()
    return _opera_via_program_files()


# ---------------------------------------------------------------------
# Utilitarios
# ---------------------------------------------------------------------

def normalizar(texto):
    nfd = unicodedata.normalize("NFD", str(texto or ""))
    return "".join(c for c in nfd if unicodedata.category(c) != "Mn").lower().strip()


def url_portal_ativa(url):
    return bool((url or "").strip())


def salvar_screenshot(page, nome):
    try:
        PASTA_SCREENSHOTS.mkdir(parents=True, exist_ok=True)
        page.screenshot(
            path=str(PASTA_SCREENSHOTS / "{}.png".format(nome)), full_page=True
        )
    except Exception:
        pass


def preencher_campo(page, locator, valor):
    locator.click()
    time.sleep(0.03)
    page.keyboard.press("Control+a")
    time.sleep(0.02)
    page.keyboard.press("Delete")
    time.sleep(0.02)
    locator.fill(str(valor))
    time.sleep(0.03)


def preencher_campo_rapido(page, locator, valor):
    try:
        locator.focus(timeout=4000)
    except Exception:
        locator.click(timeout=6000)
    locator.fill(str(valor), timeout=8000)


def _fill_by_label_candidates(scope, labels, valor, page):
    if valor is None or str(valor).strip() == "":
        return False
    for lb in labels:
        loc = scope.get_by_label(lb, exact=False).first
        try:
            loc.wait_for(state="visible", timeout=2500)
            preencher_campo(page, loc, str(valor).strip())
            return True
        except Exception:
            continue
    return False


def _fill_by_placeholder(scope, page, patterns, valor):
    if not (valor or "").strip():
        return False
    for ph in patterns:
        try:
            loc = scope.get_by_placeholder(ph).first
            loc.wait_for(state="visible", timeout=2000)
            preencher_campo(page, loc, str(valor).strip())
            return True
        except Exception:
            continue
    return False


def _fill_by_css(scope, page, css: str, valor: str) -> bool:
    if not (valor or "").strip():
        return False
    try:
        loc = scope.locator(css).first
        loc.wait_for(state="visible", timeout=3500)
        preencher_campo(page, loc, str(valor).strip())
        return True
    except Exception:
        return False


def _fill_apos_texto(scope, page, trechos: tuple[str, ...], valor: str) -> bool:
    """Clica no input seguinte a um rotulo de texto (Bubble costuma nao ter <label>)."""
    if not (valor or "").strip():
        return False
    for trecho in trechos:
        try:
            rotulo = scope.get_by_text(
                re.compile(r"^\s*" + re.escape(trecho) + r"\s*\*?$", re.I)
            ).first
            rotulo.wait_for(state="visible", timeout=2200)
            loc = rotulo.locator(
                "xpath=following::input[not(@type='file') and not(@type='hidden')"
                " and not(@type='checkbox') and not(@type='radio')][1]"
            )
            loc.wait_for(state="visible", timeout=2200)
            preencher_campo(page, loc, str(valor).strip())
            return True
        except Exception:
            pass
        # fallback: contains (rotulo pode ter * ou texto extra)
        try:
            rotulo = scope.get_by_text(re.compile(re.escape(trecho), re.I)).first
            rotulo.wait_for(state="visible", timeout=1800)
            loc = rotulo.locator(
                "xpath=following::input[not(@type='file') and not(@type='hidden')"
                " and not(@type='checkbox') and not(@type='radio')][1]"
            )
            loc.wait_for(state="visible", timeout=1800)
            preencher_campo(page, loc, str(valor).strip())
            return True
        except Exception:
            continue
    return False


def _preencher_campo_repasse(scope, page, *, labels, placeholders, css, trechos, valor, nome):
    if not (valor or "").strip():
        return False
    scopes = []
    if scope is not None:
        scopes.append(scope)
    scopes.append(page)
    for sc in scopes:
        if css and _fill_by_css(sc, page, css, valor):
            print("    {} (CSS): {}".format(nome, valor[:60]))
            return True
        if _fill_by_label_candidates(sc, labels, valor, page):
            print("    {} (label): {}".format(nome, valor[:60]))
            return True
        if placeholders and _fill_by_placeholder(sc, page, placeholders, valor):
            print("    {} (placeholder): {}".format(nome, valor[:60]))
            return True
        if trechos and _fill_apos_texto(sc, page, trechos, valor):
            print("    {} (texto): {}".format(nome, valor[:60]))
            return True
    return False


def _caminho_arquivo(valor, pasta_base: Path | None = None):
    if not valor:
        return None
    p = Path(str(valor).strip().strip('"'))
    if p.is_file():
        return p.resolve()
    if pasta_base and not p.is_absolute():
        base = Path(pasta_base)
        cand = (base / p).resolve()
        if cand.is_file():
            return cand
        # Nome exato em qualquer subpasta
        nome = p.name
        for hit in base.rglob(nome):
            if hit.is_file():
                return hit.resolve()
        # Planilha com nome longo vs PDF curto na pasta:
        # "Repasse 01-2023 - 128 Duodécimo....pdf" → "Repasse 01-2023.pdf"
        m = re.match(
            r"^(Repasse\s+\d{2}-\d{4})\b",
            nome,
            flags=re.IGNORECASE,
        )
        if m:
            prefixo = m.group(1)
            ano_m = re.search(r"(\d{4})", prefixo)
            candidatos: list[Path] = []
            if ano_m:
                pasta_ano = base / "Repasses" / ano_m.group(1)
                if pasta_ano.is_dir():
                    candidatos.extend(pasta_ano.glob(prefixo + "*.pdf"))
            if not candidatos:
                candidatos.extend(base.rglob(prefixo + "*.pdf"))
            if candidatos:
                # Preferir o mais parecido (mesmo nome completo se existir)
                for c in candidatos:
                    if c.name.lower() == nome.lower():
                        return c.resolve()
                return sorted(candidatos, key=lambda x: len(x.name))[0].resolve()
    return None


def parse_mes_ano(valor: str) -> tuple[int, int] | None:
    txt = str(valor or "").strip()
    m = re.match(r"^\s*(\d{1,2})\s*/\s*(\d{4})\s*$", txt)
    if m:
        mes, ano = int(m.group(1)), int(m.group(2))
        if 1 <= mes <= 12:
            return mes, ano
    m = re.search(r"\b(0?[1-9]|1[0-2])[\s/\-]((?:20|19)\d{2})\b", txt)
    if m:
        return int(m.group(1)), int(m.group(2))
    return None


def formatar_mes_ano(valor: str) -> str:
    ma = parse_mes_ano(valor)
    if ma:
        return "{:02d}/{}".format(ma[0], ma[1])
    return str(valor or "").strip()


def formatar_data(valor: str, mes_ano: str = "") -> str:
    """Usa a data do registro (dia do repasse/documento). Nao inventa ultimo dia."""
    txt = str(valor or "").strip()
    m = re.match(r"^\s*(\d{1,2})/(\d{1,2})/(\d{4})\s*$", txt)
    if m:
        return "{:02d}/{:02d}/{}".format(int(m.group(1)), int(m.group(2)), int(m.group(3)))
    m = re.search(r"\b(\d{1,2})\s*[/\-.]\s*(\d{1,2})\s*[/\-.]\s*(\d{4})\b", txt)
    if m:
        return "{:02d}/{:02d}/{}".format(int(m.group(1)), int(m.group(2)), int(m.group(3)))
    return txt


def formatar_valor_ui(valor: str) -> str:
    """Garante formato tipo 1.234,56 (sem R$)."""
    txt = str(valor or "").strip()
    if not txt:
        return ""
    txt = re.sub(r"^\s*R\$\s*", "", txt, flags=re.I).strip()
    m = re.search(r"(\d{1,3}(?:\.\d{3})+,\d{2}|\d+,\d{2}|\d+[.,]\d{2}|\d+)", txt)
    if not m:
        return txt
    bruto = m.group(1)
    if "," in bruto:
        return bruto
    if "." in bruto and bruto.count(".") == 1 and len(bruto.split(".")[-1]) <= 2:
        reais, cents = bruto.split(".")
        return "{},{:0<2}".format(reais, cents)
    try:
        n = float(bruto.replace(".", "").replace(",", "."))
        return "{:,.2f}".format(n).replace(",", "X").replace(".", ",").replace("X", ".")
    except ValueError:
        return bruto


def registro_vazio():
    return {k: "" for k in CAMPOS}


# ---------------------------------------------------------------------
# Fila a partir de Repasses.xlsx
# ---------------------------------------------------------------------

def _resolver_pasta_base() -> Path:
    return Path(PASTA_BASE)


def _achar_planilha(pasta: Path) -> Path | None:
    for nome in ("Repasses.xlsx", "Repasses.csv", "repasses.xlsx", "repasses.csv"):
        p = pasta / nome
        if p.is_file():
            return p
    # planilha dentro de subpasta
    for p in pasta.glob("**/Repasses.xlsx"):
        return p
    return None


def _mapear_linha_planilha(headers: list[str], row: list) -> dict:
    item = registro_vazio()
    for i, h in enumerate(headers):
        if i >= len(row):
            break
        val = "" if row[i] is None else str(row[i]).strip()
        n = normalizar(h)
        if not n or not val:
            continue
        if "link" in n or n == "url":
            item["link"] = val
        elif "mes" in n and "ano" in n:
            item["mes_ano"] = val
        elif n in ("mes", "mês", "competencia", "competência"):
            item["mes_ano"] = val
        elif "data" in n:
            item["data"] = val
        elif "previsto" in n:
            item["valor_previsto"] = val
        elif "realizado" in n or n == "valor":
            if not item["valor_realizado"]:
                item["valor_realizado"] = val
        elif "descricao" in n or "descrição" in n:
            item["descricao"] = val
        elif "arquivo" in n or n.endswith("pdf"):
            item["arquivo"] = val
    return item


def ler_planilha_repasses(caminho: Path, pasta_base: Path) -> list[dict]:
    path = Path(caminho)
    if not path.is_file():
        return []
    itens: list[dict] = []

    if path.suffix.lower() in (".xlsx", ".xlsm"):
        from openpyxl import load_workbook

        wb = load_workbook(path, data_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        wb.close()
        if not rows:
            return []
        headers = [str(h or "") for h in rows[0]]
        for i, row in enumerate(rows[1:], start=2):
            if not row or all(c is None or str(c).strip() == "" for c in row):
                continue
            item = _mapear_linha_planilha(headers, list(row))
            if not any(item.get(k) for k in ("mes_ano", "data", "valor_previsto", "valor_realizado", "descricao")):
                continue
            # resolve PDF
            arq = _caminho_arquivo(item.get("arquivo"), pasta_base)
            if arq:
                item["arquivo"] = str(arq)
            item["linha"] = i
            itens.append(item)
    else:
        with path.open("r", encoding="utf-8-sig", newline="") as f:
            sample = f.read(2048)
            f.seek(0)
            delim = ";" if sample.count(";") >= sample.count(",") else ","
            reader = csv.reader(f, delimiter=delim)
            rows = list(reader)
        if not rows:
            return []
        headers = rows[0]
        for i, row in enumerate(rows[1:], start=2):
            item = _mapear_linha_planilha(headers, row)
            if not any(item.values()):
                continue
            arq = _caminho_arquivo(item.get("arquivo"), pasta_base)
            if arq:
                item["arquivo"] = str(arq)
            item["linha"] = i
            itens.append(item)

    # Normaliza campos para o modal
    for it in itens:
        it["mes_ano"] = formatar_mes_ano(it.get("mes_ano") or it.get("data") or "")
        it["data"] = formatar_data(it.get("data") or "", it.get("mes_ano") or "")
        it["valor_previsto"] = formatar_valor_ui(it.get("valor_previsto") or "")
        it["valor_realizado"] = formatar_valor_ui(it.get("valor_realizado") or "")
        if it["valor_previsto"] and not it["valor_realizado"]:
            it["valor_realizado"] = it["valor_previsto"]
        elif it["valor_realizado"] and not it["valor_previsto"]:
            it["valor_previsto"] = it["valor_realizado"]

    print("[INFO] {} repasse(s) na planilha {}".format(len(itens), path))
    return itens


def montar_fila() -> list[dict]:
    if isinstance(REGISTRO_UNICO, dict) and any(
        str(REGISTRO_UNICO.get(k) or "").strip() for k in CAMPOS
    ):
        item = registro_vazio()
        for k in CAMPOS:
            item[k] = str(REGISTRO_UNICO.get(k) or "").strip()
        item["mes_ano"] = formatar_mes_ano(item.get("mes_ano") or item.get("data") or "")
        item["data"] = formatar_data(item.get("data") or "", item.get("mes_ano") or "")
        item["valor_previsto"] = formatar_valor_ui(item.get("valor_previsto") or "")
        item["valor_realizado"] = formatar_valor_ui(item.get("valor_realizado") or "")
        if item["valor_previsto"] and not item["valor_realizado"]:
            item["valor_realizado"] = item["valor_previsto"]
        elif item["valor_realizado"] and not item["valor_previsto"]:
            item["valor_previsto"] = item["valor_realizado"]
        item["linha"] = 1
        print("[INFO] Fila: 1 repasse (registro unico).")
        return [item]

    pasta = _resolver_pasta_base()
    planilha = _achar_planilha(pasta)
    if not planilha:
        print("[AVISO] Repasses.xlsx nao encontrado em {}".format(pasta))
        return []
    return ler_planilha_repasses(planilha, pasta)


# ---------------------------------------------------------------------
# Login / navegacao
# ---------------------------------------------------------------------

def credenciais_portal_configuradas():
    return bool((PORTAL_USUARIO or "").strip() and (PORTAL_SENHA or "").strip())


def navegar_para_url(page, url, etiqueta, pausa_apos_carregar=0.5):
    print("[INFO] Carregando {}...".format(etiqueta))
    ultimo = None
    for wt in ("domcontentloaded", "load"):
        try:
            page.goto(url, wait_until=wt, timeout=120000)
            ultimo = None
            break
        except Exception as e:
            ultimo = e
    if ultimo:
        raise ultimo
    time.sleep(pausa_apos_carregar)


def aguardar_barra_carregamento_topo(page, timeout_s=None, etiqueta=""):
    if timeout_s is None:
        timeout_s = TIMEOUT_LOADER_TOPO_S
    limite = time.monotonic() + float(timeout_s)
    viu = False
    js = """
        () => {
            function ativa(el) {
                if (!el) return false;
                var s = window.getComputedStyle(el);
                if (s.display === 'none' || parseFloat(s.opacity) < 0.08) return false;
                var r = el.getBoundingClientRect();
                if (r.top > 22 || r.height > 28) return false;
                return r.width > 12;
            }
            if (ativa(document.querySelector('#nprogress .bar'))) return true;
            if (ativa(document.querySelector('.turbo-progress-bar'))) return true;
            return false;
        }
    """
    while time.monotonic() < limite:
        try:
            ativo = page.evaluate(js)
        except Exception:
            ativo = False
        if ativo:
            viu = True
            time.sleep(0.14)
            continue
        return
    if viu:
        raise TimeoutError("Barra de progresso ativa apos {}s [{}].".format(timeout_s, etiqueta))


def _resolver_escopo_login(page):
    ultimo = None
    for _ in range(50):
        ordem = [page]
        try:
            for fr in page.frames:
                if fr != page.main_frame:
                    ordem.append(fr)
        except Exception:
            pass
        for scope in ordem:
            try:
                scope.locator("input[type='password']").first.wait_for(
                    state="visible", timeout=300
                )
                return scope
            except Exception as e:
                ultimo = e
        time.sleep(0.08)
    raise TimeoutError("Formulario de login nao pronto: {}".format(ultimo))


def login_automatico_portal(page):
    usuario = PORTAL_USUARIO.strip()
    senha = PORTAL_SENHA.strip()
    scope = _resolver_escopo_login(page)
    preenchido = False
    for rx in (
        re.compile(r"informe seu e-?\s*mail\s*:?", re.I),
        re.compile(r"seu e-?\s*mail", re.I),
    ):
        try:
            loc = scope.get_by_label(rx).first
            loc.wait_for(state="visible", timeout=3500)
            preencher_campo_rapido(page, loc, usuario)
            preenchido = True
            break
        except Exception:
            continue
    if not preenchido:
        for sel in (
            "input[type='email']",
            "input[autocomplete='username']",
            "input[type='text']",
        ):
            try:
                loc = scope.locator(sel).first
                loc.wait_for(state="visible", timeout=2500)
                preencher_campo_rapido(page, loc, usuario)
                preenchido = True
                break
            except Exception:
                continue
    if not preenchido:
        raise RuntimeError("Campo de e-mail/usuario nao encontrado no login.")

    campo = scope.locator("input[type='password']").first
    campo.wait_for(state="visible", timeout=10000)
    preencher_campo_rapido(page, campo, senha)
    time.sleep(0.07)
    for texto in (PORTAL_LOGIN_BOTAO, "Entrar", "Login"):
        try:
            b = scope.get_by_role("button", name=re.compile(texto, re.I)).first
            b.wait_for(state="visible", timeout=3500)
            b.click()
            return
        except Exception:
            continue
    page.keyboard.press("Enter")


def aguardar_login_usuario(page, pular_enter=False):
    navegar_para_url(page, URL_LOGIN, "login", pausa_apos_carregar=0.12)
    print("[INFO] Pagina de login: {}".format(URL_LOGIN))
    if credenciais_portal_configuradas():
        try:
            login_automatico_portal(page)
            print("[INFO] Login enviado.")
            time.sleep(0.55)
        except Exception as e:
            print("[AVISO] Login automatico falhou: {}".format(str(e)[:120]))
    if pular_enter:
        time.sleep(1.25)
    else:
        try:
            input("\n[INFO] Quando estiver logado, Enter aqui...\n>>> ")
        except EOFError:
            time.sleep(8)


def garantir_pagina_portal(page, url_alvo, etiqueta_log="Repasse"):
    if not url_portal_ativa(url_alvo):
        raise ValueError("URL do portal de Repasse vazia.")
    try:
        page.goto(url_alvo, wait_until="domcontentloaded", timeout=120000)
    except Exception:
        page.goto(url_alvo, wait_until="load", timeout=120000)
    time.sleep(0.25)
    try:
        page.locator("button:has-text('Criar Publicação')").wait_for(
            state="visible", timeout=20000
        )
    except Exception:
        pass
    aguardar_barra_carregamento_topo(page, etiqueta=etiqueta_log)


def criar_navegador_e_login(pular_enter_pos_login=False):
    verificar_playwright_instalado()
    pw = sync_playwright().start()
    opera = resolver_caminho_opera()
    launch_kwargs = {"headless": HEADLESS}
    if opera:
        launch_kwargs["executable_path"] = str(opera)
        print("[INFO] Usando Opera: {}".format(opera))
    else:
        print("[INFO] Opera nao encontrado — Chromium do Playwright.")
    try:
        browser = pw.chromium.launch(**launch_kwargs)
    except Exception as e:
        print("[ERRO] Falha ao iniciar navegador: {}".format(e))
        pw.stop()
        sys.exit(1)
    page = browser.new_context().new_page()
    if ABRIR_LOGIN_ANTES_DO_PORTAL:
        aguardar_login_usuario(page, pular_enter=pular_enter_pos_login)
    return pw, browser, page


# ---------------------------------------------------------------------
# Modal Repasse
# ---------------------------------------------------------------------

def _loc_modal_titulo(page):
    return page.locator("text=/{}/i".format(MODAL_TITULO_REGEX)).first


def _limpar_overlays_portal(page) -> None:
    """Fecha modais/greyout que bloqueiam 'Criar Publicação'."""
    _fechar_criar_documento_se_aberto(page)
    fechar_dialogo_anexar_se_aberto(page)
    try:
        fechar_modal(page)
    except Exception:
        pass
    try:
        page.keyboard.press("Escape")
        time.sleep(0.2)
        page.keyboard.press("Escape")
        time.sleep(0.15)
    except Exception:
        pass
    try:
        page.evaluate(
            """
            () => {
                document.querySelectorAll('.greyout, .overlay, .modal-backdrop')
                    .forEach(el => {
                        try { el.style.pointerEvents = 'none'; el.style.display = 'none'; }
                        catch (e) {}
                    });
            }
            """
        )
    except Exception:
        pass


def abrir_modal(page):
    _limpar_overlays_portal(page)
    criar_btn = page.locator("button:has-text('Criar Publicação')").first
    criar_btn.wait_for(state="visible", timeout=15000)
    criar_btn.scroll_into_view_if_needed()
    time.sleep(0.06)
    try:
        criar_btn.click(timeout=8000)
    except Exception:
        _limpar_overlays_portal(page)
        criar_btn.click(force=True, timeout=8000)
    time.sleep(0.28)
    try:
        _loc_modal_titulo(page).wait_for(state="visible", timeout=10000)
    except Exception:
        page.locator("button:has-text('Publicar')").first.wait_for(
            state="visible", timeout=10000
        )
    time.sleep(0.12)


def _modal_bubble_repasse(page):
    try:
        page.evaluate(
            """
            () => {
                document.querySelectorAll('[data-cr2-repasse-modal-marker]').forEach(function (el) {
                    el.removeAttribute('data-cr2-repasse-modal-marker');
                });
                var pubs = Array.from(
                    document.querySelectorAll('button, div[role="button"], .bubble-element.Button')
                ).filter(function (b) {
                    var t = ((b.innerText || b.textContent || '') + '').trim();
                    return (t === 'Publicar' || t.indexOf('Publicar') >= 0) && t.length < 48;
                });
                for (var i = pubs.length - 1; i >= 0; i--) {
                    var node = pubs[i];
                    var depth = 0;
                    while (node && depth < 28) {
                        depth++;
                        node = node.parentElement;
                        if (!node || !node.querySelectorAll) continue;
                        var txt = (node.innerText || '').slice(0, 6000);
                        var tem =
                            /Cadastrar\\s+Repasse/i.test(txt) ||
                            /Repasse/i.test(txt) && /Valor\\s+Previsto/i.test(txt);
                        if (!tem) continue;
                        if (txt.indexOf('Publicar') < 0) continue;
                        node.setAttribute('data-cr2-repasse-modal-marker', '1');
                        return true;
                    }
                }
                return false;
            }
            """
        )
        root = page.locator('[data-cr2-repasse-modal-marker="1"]').first
        root.wait_for(state="visible", timeout=8000)
        return root
    except Exception:
        return None


def fechar_modal(page):
    try:
        root = _modal_bubble_repasse(page)
        if root is not None:
            root.locator("button:has-text('Fechar')").first.click(timeout=4000)
            time.sleep(0.25)
    except Exception:
        try:
            page.keyboard.press("Escape")
            time.sleep(0.22)
        except Exception:
            pass
    try:
        _loc_modal_titulo(page).wait_for(state="hidden", timeout=8000)
    except Exception:
        pass


def _restaurar_inputs_file(page):
    try:
        page.evaluate(
            """
            () => {
                document.querySelectorAll('input[type=file][data-cr2-file-style]').forEach(function (el) {
                    var raw = el.getAttribute('data-cr2-file-style');
                    if (raw) {
                        try {
                            var s = JSON.parse(raw);
                            el.style.display = s.display || '';
                            el.style.opacity = s.opacity || '';
                            el.style.visibility = s.visibility || '';
                            el.style.position = s.position || '';
                            el.style.top = s.top || '';
                            el.style.left = s.left || '';
                            el.style.zIndex = s.zIndex || '';
                        } catch (e) {}
                        el.removeAttribute('data-cr2-file-style');
                    }
                });
            }
            """
        )
    except Exception:
        pass


def _revelar_input_file(page, alvo):
    try:
        if hasattr(alvo, "evaluate"):
            alvo.evaluate(
                """
                (el) => {
                    if (!el || el.tagName !== 'INPUT') return;
                    if (!el.getAttribute('data-cr2-file-style')) {
                        el.setAttribute('data-cr2-file-style', JSON.stringify({
                            display: el.style.display,
                            opacity: el.style.opacity,
                            visibility: el.style.visibility,
                            position: el.style.position,
                            top: el.style.top,
                            left: el.style.left,
                            zIndex: el.style.zIndex
                        }));
                    }
                    el.style.position = 'fixed';
                    el.style.top = '8px';
                    el.style.left = '8px';
                    el.style.opacity = '0.01';
                    el.style.zIndex = '2147483647';
                    el.style.width = '1px';
                    el.style.height = '1px';
                }
                """
            )
    except Exception:
        pass


def _aguardar_confirmacao_upload(page, modal_root, path: Path):
    nome = path.name.lower()
    for _ in range(MAX_TENTATIVAS_POLL_UPLOAD):
        try:
            areas = (
                modal_root.locator(".file-input-text")
                if modal_root is not None
                else page.locator(".file-input-text")
            )
            for i in range(areas.count()):
                txt = areas.nth(i).inner_text().strip().lower()
                if nome in txt or (path.stem.lower()[:20] in txt and "clique aqui" not in txt):
                    time.sleep(PAUSA_APOS_CONFIRMAR_UPLOAD)
                    return
        except Exception:
            pass
        time.sleep(PAUSA_POLL_UPLOAD_UI)
    time.sleep(PAUSA_APOS_CONFIRMAR_UPLOAD)


def upload_arquivo(page, modal_root, caminho) -> bool:
    path = _caminho_arquivo(caminho, _resolver_pasta_base())
    if path is None:
        return False
    scope = modal_root if modal_root is not None else page
    alvo = None
    for lb in LABELS_ARQUIVO:
        try:
            loc = scope.get_by_text(re.compile(r"^\s*{}\s*$".format(re.escape(lb)), re.I)).first
            loc.wait_for(state="visible", timeout=1500)
            handle = loc.element_handle(timeout=1000)
            if handle:
                fh = handle.evaluate_handle(
                    """
                    (el) => {
                        var p = el.parentElement;
                        for (var i = 0; i < 10 && p; i++) {
                            var near = p.querySelector('input[type=file]');
                            if (near) return near;
                            p = p.parentElement;
                        }
                        return null;
                    }
                    """
                )
                el = fh.as_element() if fh else None
                if el:
                    alvo = el
                    break
        except Exception:
            continue
    if alvo is None:
        try:
            files = scope.locator("input[type=file]")
            if files.count() >= 1:
                alvo = files.first
        except Exception:
            pass
    if alvo is None:
        print("    [AVISO] Campo Arquivo nao encontrado (seguindo sem upload).")
        return False
    try:
        _revelar_input_file(page, alvo)
        time.sleep(0.05)
        alvo.set_input_files(str(path))
        time.sleep(PAUSA_APOS_ANEXAR)
        print("    Upload: {}".format(path.name))
        _aguardar_confirmacao_upload(page, modal_root, path)
        _restaurar_inputs_file(page)
        return True
    except Exception as e:
        _restaurar_inputs_file(page)
        print("    [Upload] Falhou: {}".format(str(e)[:80]))
        return False


def preencher_modal_repasse(page, item: dict):
    """Só os campos de informação — arquivo vai depois, no fluxo Anexar."""
    modal_root = _modal_bubble_repasse(page)
    scope = modal_root if modal_root is not None else page

    mes_ano = formatar_mes_ano(item.get("mes_ano") or "")
    data = formatar_data(item.get("data") or "", mes_ano)
    prev = formatar_valor_ui(item.get("valor_previsto") or "")
    real = formatar_valor_ui(item.get("valor_realizado") or "")
    if prev and not real:
        real = prev
    elif real and not prev:
        prev = real
    # Descrição do repasse: padrão "Repasse" (não usa link)
    desc = DESCRICAO_PADRAO_REPASSE

    ok_mes = _preencher_campo_repasse(
        scope,
        page,
        labels=LABELS_MES_ANO,
        placeholders=(
            re.compile(r"01/2024"),
            re.compile(r"mm\s*/\s*aaaa", re.I),
            re.compile(r"m[eê]s", re.I),
        ),
        css=(
            "input.bubble-element.Input[placeholder*='01/2024'], "
            "input.bubble-element.Input[placeholder*='01/20'], "
            "input[placeholder*='01/2024'], "
            "input[placeholder*='MM/AAAA'], "
            "input[placeholder*='mm/aaaa']"
        ),
        trechos=("Mês e Ano", "Mes e Ano", "Competência", "Competencia"),
        valor=mes_ano,
        nome="Mes e Ano",
    )
    if not ok_mes:
        raise RuntimeError("Campo 'Mes e Ano' nao encontrado.")
    time.sleep(PAUSA_ENTRE_CAMPOS)

    ok_data = _preencher_campo_repasse(
        scope,
        page,
        labels=LABELS_DATA,
        placeholders=(
            re.compile(r"01/01/2024"),
            re.compile(r"dd\s*/\s*mm", re.I),
            re.compile(r"\d{2}/\d{2}/\d{4}"),
        ),
        css=(
            "input.bubble-element.Input[placeholder*='01/01'], "
            "input.bubble-element.Input[placeholder*='01/01/2024'], "
            "input[placeholder*='01/01/2024'], "
            "input[placeholder*='dd/mm'], "
            "input[placeholder*='DD/MM']"
        ),
        trechos=("Data",),
        valor=data,
        nome="Data",
    )
    if not ok_data:
        raise RuntimeError("Campo 'Data' nao encontrado.")
    time.sleep(PAUSA_ENTRE_CAMPOS)

    if not prev:
        raise RuntimeError("Valor Previsto/Realizado vazio na planilha — rode a extracao com OCR.")
    ok_prev = _preencher_campo_repasse(
        scope,
        page,
        labels=LABELS_VALOR_PREVISTO,
        placeholders=(re.compile(r"previsto", re.I),),
        css="input[placeholder*='Previsto'], input[placeholder*='previsto']",
        trechos=("Valor Previsto", "Previsto"),
        valor=prev,
        nome="Valor Previsto",
    )
    if not ok_prev:
        raise RuntimeError("Campo 'Valor Previsto' nao encontrado.")
    time.sleep(PAUSA_ENTRE_CAMPOS)

    ok_real = _preencher_campo_repasse(
        scope,
        page,
        labels=LABELS_VALOR_REALIZADO,
        placeholders=(re.compile(r"realizado", re.I),),
        css="input[placeholder*='Realizado'], input[placeholder*='realizado']",
        trechos=("Valor Realizado", "Realizado"),
        valor=real,
        nome="Valor Realizado",
    )
    if not ok_real:
        raise RuntimeError("Campo 'Valor Realizado' nao encontrado.")
    time.sleep(PAUSA_ENTRE_CAMPOS)

    _preencher_campo_repasse(
        scope,
        page,
        labels=LABELS_DESCRICAO,
        placeholders=(re.compile(r"descri", re.I),),
        css="textarea, input[placeholder*='Descri']",
        trechos=("Descrição", "Descricao"),
        valor=desc,
        nome="Descricao",
    )
    time.sleep(PAUSA_ENTRE_CAMPOS)
    print("    Descricao = {!r} (padrao; sem link; arquivo depois)".format(desc))

    _restaurar_inputs_file(page)
    return modal_root


def _botao_anexar_visivel(page):
    try:
        if page.get_by_text(re.compile(r"anexar documentos a esta publica", re.I)).count():
            loc = page.get_by_text(re.compile(r"anexar documentos a esta publica", re.I)).first
            if loc.is_visible(timeout=400):
                btn = page.locator("button:has-text('Anexar')").first
                if btn.is_visible(timeout=400):
                    return btn
    except Exception:
        pass
    try:
        btn = page.locator("button:has-text('Anexar')").first
        if btn.is_visible(timeout=400):
            return btn
    except Exception:
        pass
    return None


def _parece_erro_real(texto: str) -> bool:
    if not texto:
        return False
    limpo = _TEXTO_LABEL_FORM_RX.sub(" ", texto)
    return bool(_TEXTO_ERRO_APOS_PUBLICAR_RX.search(limpo))


def aguardar_resultado_apos_publicar(page, modal_root) -> str:
    """
    Apos Publicar no repasse: espera diálogo Anexar, sucesso ou fechamento do modal.
    Retorna: 'anexar' | 'ok'
    """
    titulo_loc = _loc_modal_titulo(page)
    fim = time.monotonic() + TIMEOUT_RESULTADO_PUBLICACAO_S
    ultimo = ""
    viu_loader = False
    ok_desde = None
    while time.monotonic() < fim:
        btn_anexar = _botao_anexar_visivel(page)
        if btn_anexar is not None:
            print("    Dialogo 'Anexar documentos' apareceu.")
            return "anexar"

        try:
            loader = page.evaluate(
                """
                () => {
                    function ativa(el) {
                        if (!el) return false;
                        var s = window.getComputedStyle(el);
                        if (s.display === 'none' || parseFloat(s.opacity) < 0.08) return false;
                        return el.getBoundingClientRect().width > 12;
                    }
                    return !!(ativa(document.querySelector('#nprogress .bar'))
                        || ativa(document.querySelector('.turbo-progress-bar')));
                }
                """
            )
        except Exception:
            loader = False
        if loader:
            viu_loader = True
            ok_desde = None
            time.sleep(0.12)
            continue
        if viu_loader and ok_desde is None:
            ok_desde = time.monotonic()

        try:
            visivel = titulo_loc.is_visible(timeout=300)
        except Exception:
            visivel = False
        if not visivel:
            # Modal cadastro fechou — ainda pode abrir o Anexar em seguida
            time.sleep(0.18)
            btn_anexar = _botao_anexar_visivel(page)
            if btn_anexar is not None:
                print("    Dialogo 'Anexar documentos' apareceu.")
                return "anexar"
            try:
                if titulo_loc.is_visible(timeout=300):
                    continue
            except Exception:
                pass
            print("    Modal fechou — publicacao aceita.")
            return "ok"

        try:
            ultimo = (
                modal_root.inner_text(timeout=800)
                if modal_root is not None
                else page.locator("body").inner_text(timeout=800)
            )
        except Exception:
            ultimo = ""

        if _parece_erro_real(ultimo or ""):
            raise RuntimeError(
                "Resposta apos Publicar: {}".format(
                    (ultimo or "").replace("\n", " ").strip()[:260]
                )
            )
        if _TEXTO_SUCESSO_MODAL_RX.search(ultimo or ""):
            # Sucesso no cadastro — ainda aguarda o diálogo Anexar
            time.sleep(0.2)
            btn_anexar = _botao_anexar_visivel(page)
            if btn_anexar is not None:
                return "anexar"
            print("    Sucesso detectado.")
            return "ok"
        if viu_loader and ok_desde and (time.monotonic() - ok_desde) >= 2.2:
            btn_anexar = _botao_anexar_visivel(page)
            if btn_anexar is not None:
                return "anexar"
            print("    Envio terminou sem mensagem — assumindo OK.")
            return "ok"
        time.sleep(0.2)

    btn_anexar = _botao_anexar_visivel(page)
    if btn_anexar is not None:
        return "anexar"
    if _parece_erro_real(ultimo or ""):
        raise TimeoutError(
            "Sem confirmacao apos Publicar ({}s). {}".format(
                TIMEOUT_RESULTADO_PUBLICACAO_S,
                (ultimo or "").replace("\n", " ").strip()[:180],
            )
        )
    print(
        "    [AVISO] Modal aberto apos {}s sem erro — seguindo.".format(
            TIMEOUT_RESULTADO_PUBLICACAO_S
        )
    )
    return "ok"


def clicar_publicar(page) -> str:
    _restaurar_inputs_file(page)
    time.sleep(0.04)
    modal_root = _modal_bubble_repasse(page)
    btn = (
        modal_root.locator("button:has-text('Publicar')").first
        if modal_root is not None
        else page.locator("button:has-text('Publicar')").first
    )
    btn.wait_for(state="visible", timeout=15000)
    btn.scroll_into_view_if_needed()
    limite = time.monotonic() + TIMEOUT_PUBLICAR_HABILITADO_S
    while time.monotonic() < limite:
        try:
            if btn.is_enabled():
                break
        except Exception:
            pass
        time.sleep(0.18)
    else:
        salvar_screenshot(page, "TIMEOUT_PUBLICAR_REPASSE")
        raise TimeoutError("Botao Publicar desabilitado por demais tempo.")
    aguardar_barra_carregamento_topo(page, etiqueta="antes de Publicar")
    time.sleep(0.05)
    _restaurar_inputs_file(page)
    try:
        box = btn.bounding_box()
        if box:
            page.mouse.click(box["x"] + box["width"] / 2, box["y"] + box["height"] / 2)
        else:
            btn.click(timeout=15000)
    except Exception:
        btn.click(force=True, timeout=15000)
    print("    Clicou em Publicar.")
    time.sleep(0.1)
    status = aguardar_resultado_apos_publicar(page, modal_root)
    time.sleep(PAUSA_APOS_CLICAR_PUBLICAR)
    return status


def _aguardar_modal_texto(page, padrao: str, timeout_s: float):
    fim = time.monotonic() + timeout_s
    rx = re.compile(padrao, re.I)
    while time.monotonic() < fim:
        try:
            loc = page.get_by_text(rx).first
            if loc.is_visible(timeout=250):
                return loc
        except Exception:
            pass
        time.sleep(0.12)
    return None


def upload_em_criar_documento(page, caminho) -> Path:
    """Sobe o PDF clicando na area tracejada do Passo 1."""
    path = _caminho_arquivo(caminho, _resolver_pasta_base())
    if path is None:
        raise RuntimeError("Arquivo PDF nao encontrado para anexar: {}".format(caminho))
    if not path.is_file():
        raise RuntimeError("Arquivo PDF inexistente: {}".format(path))

    if _aguardar_modal_texto(page, r"Criar\s+Documento", TIMEOUT_CRIAR_DOCUMENTO_S) is None:
        salvar_screenshot(page, "SEM_CRIAR_DOCUMENTO")
        raise TimeoutError("Modal 'Criar Documento' nao apareceu.")

    time.sleep(0.4)
    print("    Upload na zona tracejada: {}".format(path.name))

    def _passo2_ok() -> bool:
        try:
            if page.get_by_text(re.compile(r"Passo\s*2", re.I)).first.is_visible(timeout=300):
                return True
        except Exception:
            pass
        stem = path.stem[:24]
        try:
            if page.get_by_text(re.compile(re.escape(stem), re.I)).first.is_visible(timeout=300):
                return True
        except Exception:
            pass
        return False

    def _aguardar_passo2(timeout_s: float = 28) -> bool:
        fim = time.monotonic() + timeout_s
        while time.monotonic() < fim:
            if _passo2_ok():
                time.sleep(PAUSA_APOS_UPLOAD_DOC)
                return True
            time.sleep(0.25)
        return False

    def _clicar_zona_tracejada() -> None:
        """Clica no centro da area 'Clique aqui para fazer upload…'."""
        rx = re.compile(r"Clique aqui para fazer upload", re.I)
        texto = page.get_by_text(rx).first
        texto.wait_for(state="visible", timeout=10000)
        # Preferir o div pai da zona (maior que o texto)
        try:
            zona = page.locator("div").filter(has_text=rx).filter(
                has_not=page.get_by_text(re.compile(r"Adicionar Documento|Passo\s*2|Finalizar", re.I))
            ).last
            box = zona.bounding_box()
            if box and box["width"] > 180 and box["height"] > 40:
                page.mouse.click(
                    box["x"] + box["width"] / 2,
                    box["y"] + box["height"] / 2,
                )
                print("    Clicou na zona tracejada (centro).")
                return
        except Exception:
            pass
        # Fallback: clique no proprio texto
        texto.click(timeout=5000)
        print("    Clicou no texto da zona de upload.")

    def _input_perto_da_zona():
        """Acha input[type=file] ligado a zona de upload."""
        return page.evaluate(
            """
            () => {
                const nodes = Array.from(document.querySelectorAll('div, span, p'));
                let alvo = null;
                for (const el of nodes) {
                    const t = (el.innerText || '').trim();
                    if (/Clique aqui para fazer upload/i.test(t) && t.length < 120) {
                        alvo = el;
                        break;
                    }
                }
                if (!alvo) return null;
                let n = alvo;
                for (let i = 0; i < 12 && n; i++) {
                    const inp = n.querySelector && n.querySelector('input[type=file]');
                    if (inp) {
                        // marca
                        inp.setAttribute('data-cr2-upload-zona', '1');
                        return true;
                    }
                    n = n.parentElement;
                }
                return false;
            }
            """
        )

    ok = False

    # 1) PRIORIDADE: clicar na zona tracejada → file chooser
    try:
        with page.expect_file_chooser(timeout=10000) as fc_info:
            _clicar_zona_tracejada()
        chooser = fc_info.value
        chooser.set_files(str(path))
        print("    Arquivo enviado pela zona tracejada.")
        ok = _aguardar_passo2(28)
    except Exception as e:
        print("    [Upload] zona+chooser: {}".format(str(e)[:100]))

    # 2) Zona: marca input interno e set_input_files
    if not ok:
        try:
            _clicar_zona_tracejada()
            time.sleep(0.25)
            _input_perto_da_zona()
            page.evaluate(
                """
                () => {
                    document.querySelectorAll('input[type=file]').forEach(function (el) {
                        el.style.display = 'block';
                        el.style.opacity = '1';
                        el.style.visibility = 'visible';
                        el.style.position = 'fixed';
                        el.style.top = '4px';
                        el.style.left = '4px';
                        el.style.zIndex = '2147483647';
                        el.style.width = '4px';
                        el.style.height = '4px';
                    });
                }
                """
            )
            time.sleep(0.1)
            alvo = page.locator("input[type=file][data-cr2-upload-zona='1']").first
            if alvo.count() < 1:
                alvo = page.locator("input[type=file]").last
            alvo.set_input_files(str(path), timeout=8000)
            print("    Arquivo via input da zona.")
            ok = _aguardar_passo2(22)
        except Exception as e:
            print("    [Upload] zona+input: {}".format(str(e)[:100]))
        finally:
            _restaurar_inputs_file(page)

    # 3) Fallback: botao Adicionar Documento Unico
    if not ok:
        try:
            with page.expect_file_chooser(timeout=8000) as fc_info:
                page.get_by_text(re.compile(r"Adicionar Documento", re.I)).first.click(
                    timeout=5000
                )
            fc_info.value.set_files(str(path))
            print("    Arquivo via 'Adicionar Documento'.")
            ok = _aguardar_passo2(22)
        except Exception as e:
            print("    [Upload] botao adicionar: {}".format(str(e)[:100]))

    if not ok:
        _restaurar_inputs_file(page)
        salvar_screenshot(page, "FALHA_UPLOAD_ZONA")
        raise RuntimeError(
            "Nao subiu o PDF na zona 'Clique aqui para fazer upload…' (Passo 1). "
            "Arquivo: {}".format(path)
        )

    _restaurar_inputs_file(page)
    print("    PDF na lista (Passo 2).")
    return path


def _fechar_criar_documento_se_aberto(page) -> None:
    """Fecha modal Criar Documento (evita greyout bloquear o proximo item)."""
    try:
        if not page.get_by_text(re.compile(r"Criar\s+Documento", re.I)).first.is_visible(
            timeout=400
        ):
            return
    except Exception:
        return
    for texto in ("Fechar", "Cancelar", "X"):
        try:
            btn = page.locator("button:has-text('{0}')".format(texto)).first
            if btn.is_visible(timeout=400):
                btn.click(timeout=4000, force=True)
                time.sleep(0.35)
                return
        except Exception:
            continue
    try:
        page.keyboard.press("Escape")
        time.sleep(0.25)
    except Exception:
        pass


def _clicar_lapis_editar_arquivo(page, path: Path) -> None:
    """Clica no botao Editar (lapis) na linha do PDF — Passo 2.

    Independente de cor. Estrategia:
      1) Acha o menor no com o nome do PDF
      2) Sobe na arvore e lista .clickable-element pequenos da mesma linha
      3) Ordena: semantica editar → meio da fileira → demais (lixeira por ultimo)
      4) Clica cada candidato (Playwright + eventos Bubble) ate abrir o modal
    """
    stem = path.stem
    trecho = stem[:36]
    _aguardar_modal_texto(page, r"Passo\s*2", 20)
    try:
        page.get_by_text(re.compile(re.escape(trecho[:24]), re.I)).first.wait_for(
            state="visible", timeout=12000
        )
    except Exception:
        pass
    time.sleep(0.55)

    def _marcar_candidatos():
        return page.evaluate(
            r"""
            (trecho) => {
                document.querySelectorAll('[data-cr2-lapis]').forEach(el => {
                    el.removeAttribute('data-cr2-lapis');
                });

                function visivel(el) {
                    if (!el) return false;
                    const s = window.getComputedStyle(el);
                    if (s.display === 'none' || s.visibility === 'hidden'
                        || parseFloat(s.opacity) < 0.05) return false;
                    const r = el.getBoundingClientRect();
                    return r.width > 6 && r.height > 6;
                }
                function tamanhoBotao(el) {
                    const r = el.getBoundingClientRect();
                    // Botoes de icone Bubble costumam ser 24–96px (as vezes o hitbox e maior)
                    return r.width >= 14 && r.width <= 120 && r.height >= 14 && r.height <= 120;
                }
                function blob(el) {
                    return [
                        el.getAttribute('aria-label') || '',
                        el.getAttribute('title') || '',
                        el.getAttribute('data-tip') || '',
                        typeof el.className === 'string' ? el.className : '',
                        el.innerHTML || '',
                    ].join(' ').toLowerCase();
                }
                function kind(el) {
                    const t = blob(el);
                    if (/trash|lixeira|delete|excluir|remover|rubbish|fa-trash|ion-trash/.test(t))
                        return 'trash';
                    if (/pencil|editar|edit\b|lapis|lápis|fa-pencil|fa-edit|ion-pencil|ion-compose|create-outline/.test(t))
                        return 'edit';
                    if (/check|confirma|fa-check|ion-checkmark|done\b/.test(t))
                        return 'check';
                    return 'unknown';
                }
                function textoBotaoGrande(el) {
                    const t = (el.innerText || '').replace(/\s+/g, ' ').trim();
                    return t.length > 0 && /finalizar|fechar|adicionar|cancelar|passo/i.test(t);
                }

                // Menor no visivel que contenha o trecho do PDF (evita container enorme)
                let alvoTxt = null;
                let melhorArea = Infinity;
                const nodes = Array.from(document.querySelectorAll('div, span, p, a, label'));
                for (const el of nodes) {
                    const t = (el.innerText || '').replace(/\s+/g, ' ').trim();
                    if (!t || t.length > 240) continue;
                    if (t.indexOf(trecho) < 0) continue;
                    if (t.indexOf('Passo') >= 0 || /clique aqui/i.test(t)) continue;
                    if (!visivel(el)) continue;
                    const r = el.getBoundingClientRect();
                    const area = r.width * r.height;
                    // Preferir nos "folha" (poucos filhos com o mesmo texto)
                    const filhosComTexto = Array.from(el.children || []).filter(c =>
                        ((c.innerText || '').indexOf(trecho) >= 0)
                    ).length;
                    const score = area + (filhosComTexto ? 50000 : 0);
                    if (score < melhorArea) {
                        melhorArea = score;
                        alvoTxt = el;
                    }
                }
                if (!alvoTxt) {
                    const curto = (trecho.split(' - ')[0] || trecho.slice(0, 18));
                    for (const el of nodes) {
                        const t = (el.innerText || '').replace(/\s+/g, ' ').trim();
                        if (!t || t.length > 240) continue;
                        if (t.indexOf(curto) >= 0 && /\.pdf/i.test(t) && visivel(el)) {
                            alvoTxt = el;
                            break;
                        }
                    }
                }
                if (!alvoTxt) return { count: 0, vias: [], debug: 'sem-texto-pdf' };

                const txtBox = alvoTxt.getBoundingClientRect();
                const yMid = txtBox.y + txtBox.height / 2;
                let row = alvoTxt;
                let acoes = [];
                for (let up = 0; up < 16 && row; up++) {
                    const raw = Array.from(
                        row.querySelectorAll('.clickable-element, button, [role=button]')
                    ).filter(el => visivel(el) && tamanhoBotao(el) && !textoBotaoGrande(el));

                    const byX = new Map();
                    for (const el of raw) {
                        const r = el.getBoundingClientRect();
                        // Mesma faixa vertical da linha do arquivo (tolerante)
                        if (Math.abs((r.y + r.height / 2) - yMid) > 56) continue;
                        const key = Math.round(r.x / 10);
                        const prev = byX.get(key);
                        // Preferir o proprio clickable-element
                        if (!prev || el.classList.contains('clickable-element')) {
                            byX.set(key, el);
                        }
                    }
                    acoes = Array.from(byX.values()).sort(
                        (a, b) => a.getBoundingClientRect().x - b.getBoundingClientRect().x
                    );
                    if (acoes.length >= 2) break;
                    row = row.parentElement;
                }

                // Fallback: qualquer clickable pequeno no modal Criar Documento perto do PDF
                if (acoes.length < 2) {
                    const modalHint = Array.from(document.querySelectorAll('div')).find(el => {
                        const t = (el.innerText || '');
                        return /Criar\s+Documento/i.test(t) && /Passo\s*2/i.test(t)
                            && el.getBoundingClientRect().height > 120;
                    });
                    const escopo = modalHint || document.body;
                    const raw = Array.from(
                        escopo.querySelectorAll('.clickable-element, button, [role=button]')
                    ).filter(el => visivel(el) && tamanhoBotao(el) && !textoBotaoGrande(el));
                    const byX = new Map();
                    for (const el of raw) {
                        const r = el.getBoundingClientRect();
                        if (Math.abs((r.y + r.height / 2) - yMid) > 70) continue;
                        byX.set(Math.round(r.x / 10), el);
                    }
                    acoes = Array.from(byX.values()).sort(
                        (a, b) => a.getBoundingClientRect().x - b.getBoundingClientRect().x
                    );
                }

                if (!acoes.length) {
                    return {
                        count: 0,
                        vias: [],
                        debug: 'sem-botoes y=' + Math.round(yMid),
                    };
                }

                const ranked = [];
                const seen = new Set();
                function push(el, via) {
                    if (!el || seen.has(el)) return;
                    seen.add(el);
                    ranked.push({ el, via });
                }
                for (const el of acoes) {
                    if (kind(el) === 'edit') push(el, 'icon-edit');
                }
                if (acoes.length >= 3) push(acoes[1], 'row-middle');
                if (acoes.length === 2) {
                    const nonTrash = acoes.filter(el => kind(el) !== 'trash');
                    if (nonTrash.length) push(nonTrash[0], 'non-trash');
                    else push(acoes[0], 'row-first');
                }
                if (acoes.length === 1) push(acoes[0], 'only-one');
                for (const el of acoes) {
                    if (kind(el) !== 'trash') push(el, 'scan');
                }
                for (const el of acoes) push(el, 'all');

                ranked.forEach((item, i) => {
                    item.el.setAttribute('data-cr2-lapis', String(i + 1));
                });
                return {
                    count: ranked.length,
                    vias: ranked.map(r => r.via),
                    debug: 'n=' + acoes.length + ' kinds=' + acoes.map(kind).join(','),
                };
            }
            """,
            trecho,
        )

    def _clicar_marcado(idx: int) -> bool:
        sel = "[data-cr2-lapis='{0}']".format(idx)
        loc = page.locator(sel).first
        try:
            loc.wait_for(state="visible", timeout=2000)
        except Exception:
            return False
        # 1) Clique Playwright normal
        try:
            loc.click(timeout=3500)
            return True
        except Exception:
            pass
        # 2) Force (greyout / overlay)
        try:
            loc.click(force=True, timeout=3000)
            return True
        except Exception:
            pass
        # 3) Mouse no centro + eventos Bubble
        try:
            box = loc.bounding_box()
            if box:
                x = box["x"] + box["width"] / 2
                y = box["y"] + box["height"] / 2
                page.mouse.move(x, y)
                page.mouse.down()
                page.mouse.up()
                page.evaluate(
                    """
                    (i) => {
                        const el = document.querySelector('[data-cr2-lapis="' + i + '"]');
                        if (!el) return;
                        for (const type of ['pointerdown', 'mousedown', 'mouseup',
                                            'pointerup', 'click']) {
                            el.dispatchEvent(new MouseEvent(type, {
                                bubbles: true, cancelable: true, view: window
                            }));
                        }
                    }
                    """,
                    idx,
                )
                return True
        except Exception:
            pass
        return False

    aberto = False
    ultimo_debug = ""
    for tentativa in range(1, 5):
        info = None
        try:
            info = _marcar_candidatos()
        except Exception as exc:
            ultimo_debug = "eval:{0}".format(exc)
            info = None
        if not info or not info.get("count"):
            ultimo_debug = (info or {}).get("debug") or ultimo_debug or "count=0"
            print("    [AVISO] Lápis: nenhum candidato ({0}).".format(ultimo_debug))
            time.sleep(0.45)
            continue

        print(
            "    Lápis: {0} candidato(s) [{1}]".format(
                info["count"], info.get("debug") or ""
            )
        )
        vias = info.get("vias") or []
        for idx in range(1, int(info["count"]) + 1):
            if not _clicar_marcado(idx):
                continue
            via = vias[idx - 1] if idx - 1 < len(vias) else "?"
            print(
                "    Clicou no lapis (via={0}, cand={1}/{2}, tentativa {3}).".format(
                    via, idx, info["count"], tentativa
                )
            )
            time.sleep(0.65)
            if _aguardar_modal_texto(page, r"Editar\s+Documento", 6) is not None:
                aberto = True
                break
        if aberto:
            break
        time.sleep(0.4)

    if not aberto:
        # Fallback absoluto: todos os clickable pequenos ao lado do nome do PDF
        try:
            loc_nome = page.get_by_text(
                re.compile(re.escape(trecho[:20]), re.I)
            ).first
            box = loc_nome.bounding_box()
            if box:
                # Varre pontos a direita do nome (icones da linha)
                xs = [
                    box["x"] + box["width"] + 28,
                    box["x"] + box["width"] + 56,
                    box["x"] + box["width"] + 84,
                    box["x"] + box["width"] + 112,
                ]
                y = box["y"] + box["height"] / 2
                for x in xs:
                    try:
                        page.mouse.click(x, y)
                    except Exception:
                        continue
                    print("    Clicou no lapis (fallback coord x=+{0}).".format(
                        int(x - box["x"] - box["width"])
                    ))
                    time.sleep(0.6)
                    if _aguardar_modal_texto(page, r"Editar\s+Documento", 4) is not None:
                        aberto = True
                        break
        except Exception:
            pass

    if not aberto:
        try:
            row = (
                page.locator("div")
                .filter(has_text=re.compile(re.escape(trecho[:20]), re.I))
                .filter(has_text=re.compile(r"\.pdf", re.I))
                .last
            )
            btns = row.locator(".clickable-element")
            n = min(btns.count(), 8)
            ordem = list(range(n))
            if 1 in ordem:
                ordem = [1] + [i for i in ordem if i != 1]
            for i in ordem:
                try:
                    btns.nth(i).click(force=True, timeout=3000)
                except Exception:
                    continue
                time.sleep(0.55)
                if _aguardar_modal_texto(page, r"Editar\s+Documento", 4) is not None:
                    aberto = True
                    print("    Clicou no lapis (fallback PW idx={0}).".format(i))
                    break
        except Exception:
            pass

    if not aberto:
        salvar_screenshot(page, "SEM_LAPIS_EDITAR")
        _fechar_criar_documento_se_aberto(page)
        raise RuntimeError(
            "Nao abriu 'Editar Documento' apos clicar no botao Editar (lapis) do PDF."
            + (" ({0})".format(ultimo_debug) if ultimo_debug else "")
        )
    time.sleep(0.25)


def _sair_do_campo_focado(page) -> None:
    """Bubble so habilita 'Editar' depois que o input perde o foco (clicar fora)."""
    try:
        page.keyboard.press("Tab")
    except Exception:
        pass
    time.sleep(PAUSA_SAIR_CAMPO)
    try:
        # Clica no titulo do modal — fora do input de data
        titulo = page.get_by_text(re.compile(r"Editar\s+Documento", re.I)).first
        if titulo.is_visible(timeout=500):
            titulo.click(timeout=1500)
            time.sleep(PAUSA_SAIR_CAMPO)
            return
    except Exception:
        pass
    try:
        page.evaluate(
            """
            () => {
                const a = document.activeElement;
                if (a && typeof a.blur === 'function') a.blur();
                const t = Array.from(document.querySelectorAll('div, span'))
                    .find(el => /Editar\\s+Documento/i.test((el.innerText || '').trim())
                        && (el.innerText || '').trim().length < 40);
                if (t) t.click();
            }
            """
        )
    except Exception:
        pass
    time.sleep(PAUSA_SAIR_CAMPO)


def _preencher_apos_rotulo_rapido(page, rotulos: tuple[str, ...], valor: str) -> bool:
    """Preenche o input logo apos o rotulo — timeouts curtos (modal Editar Documento)."""
    if not (valor or "").strip():
        return False
    valor = str(valor).strip()
    for rotulo in rotulos:
        try:
            lab = page.get_by_text(
                re.compile(r"^\s*" + re.escape(rotulo) + r"\s*\*?$", re.I)
            ).first
            lab.wait_for(state="visible", timeout=900)
            loc = lab.locator(
                "xpath=following::input[not(@type='file') and not(@type='hidden')"
                " and not(@type='checkbox') and not(@type='radio')][1]"
            )
            loc.wait_for(state="visible", timeout=900)
            loc.click(timeout=1500)
            time.sleep(0.08)
            loc.fill(valor, timeout=2500)
            time.sleep(0.1)
            return True
        except Exception:
            continue
    # Fallback: placeholder tipico da data de publicacao
    try:
        loc = page.get_by_placeholder(re.compile(r"01/01/2024|dd\s*/\s*mm", re.I)).first
        loc.wait_for(state="visible", timeout=700)
        loc.click(timeout=1200)
        time.sleep(0.08)
        loc.fill(valor, timeout=2500)
        time.sleep(0.1)
        return True
    except Exception:
        return False


def _preencher_editar_documento(page, item: dict, path: Path) -> None:
    if _aguardar_modal_texto(page, r"Editar\s+Documento", TIMEOUT_EDITAR_DOCUMENTO_S) is None:
        salvar_screenshot(page, "SEM_EDITAR_DOCUMENTO")
        raise TimeoutError("Modal 'Editar Documento' nao apareceu.")

    mes_ano = formatar_mes_ano(item.get("mes_ano") or "")
    data_pub = formatar_data(item.get("data") or "", mes_ano)
    if not data_pub:
        raise RuntimeError("Data de Publicacao vazia — preencha a Data na planilha.")

    desc_doc = path.stem.strip() or DESCRICAO_PADRAO_REPASSE

    # Descricao (ja costuma vir do PDF) — so ajusta se necessario, bem rapido
    _preencher_apos_rotulo_rapido(
        page,
        ("Descrição do Documento", "Descricao do Documento"),
        desc_doc,
    )

    ok_data = _preencher_apos_rotulo_rapido(
        page,
        ("Data de Publicação", "Data de Publicacao", "Data Publicação"),
        data_pub,
    )
    if not ok_data:
        raise RuntimeError("Campo 'Data de Publicacao' nao encontrado.")
    print("    Editar Documento: data={} desc={!r}".format(data_pub, desc_doc[:50]))

    # Clica fora da data → espera 1s → Editar (Bubble so libera o botao assim)
    _sair_do_campo_focado(page)
    print("    Aguardando 1s apos sair da data...")
    time.sleep(PAUSA_APOS_DATA_DOC_S)

    btn = page.locator("button:has-text('Editar')").last
    btn.wait_for(state="visible", timeout=8000)
    limite = time.monotonic() + 10
    while time.monotonic() < limite:
        try:
            if btn.is_enabled():
                break
        except Exception:
            pass
        time.sleep(0.15)
    else:
        salvar_screenshot(page, "TIMEOUT_EDITAR_DOC_DISABLED")
        raise TimeoutError("Botao Editar ficou desabilitado apos preencher a data.")
    btn.click(timeout=10000)
    print("    Clicou em Editar (salvar documento).")
    time.sleep(PAUSA_APOS_EDITAR_DOC)


def _clicar_finalizar_documento(page) -> None:
    _aguardar_modal_texto(page, r"Passo\s*3|Finalizar", 15)
    btn = page.locator("button:has-text('Finalizar')").first
    btn.wait_for(state="visible", timeout=15000)
    limite = time.monotonic() + 40
    while time.monotonic() < limite:
        try:
            if btn.is_enabled():
                break
        except Exception:
            pass
        time.sleep(0.18)
    else:
        salvar_screenshot(page, "TIMEOUT_FINALIZAR_DOC")
        raise TimeoutError("Botao Finalizar desabilitado por demais tempo.")
    btn.click(timeout=15000)
    print("    Clicou em Finalizar.")
    time.sleep(PAUSA_APOS_FINALIZAR)
    # Espera modal Criar Documento sumir
    fim = time.monotonic() + 18
    while time.monotonic() < fim:
        try:
            if not page.get_by_text(re.compile(r"Criar\s+Documento", re.I)).first.is_visible(timeout=200):
                break
        except Exception:
            break
        time.sleep(0.15)


def anexar_documento_apos_publicar(page, item: dict) -> None:
    """Anexar → upload → lapis → Data de Publicacao → Editar → Finalizar."""
    arquivo = item.get("arquivo")
    if not arquivo:
        raise RuntimeError("Sem arquivo PDF na fila para anexar apos Publicar.")

    btn = _botao_anexar_visivel(page)
    if btn is None:
        # Espera um pouco mais
        fim = time.monotonic() + TIMEOUT_DIALOGO_ANEXAR_S
        while time.monotonic() < fim and btn is None:
            btn = _botao_anexar_visivel(page)
            if btn is None:
                time.sleep(0.15)
    if btn is None:
        salvar_screenshot(page, "SEM_DIALOGO_ANEXAR")
        raise TimeoutError("Dialogo 'Deseja anexar documentos…?' nao apareceu.")

    btn.click(timeout=15000)
    print("    Clicou em Anexar.")
    time.sleep(PAUSA_APOS_ANEXAR_DIALOGO)

    path = upload_em_criar_documento(page, arquivo)
    try:
        _clicar_lapis_editar_arquivo(page, path)
        _preencher_editar_documento(page, item, path)
        _clicar_finalizar_documento(page)
    except Exception:
        _fechar_criar_documento_se_aberto(page)
        raise


def fechar_dialogo_anexar_se_aberto(page) -> None:
    try:
        btn = _botao_anexar_visivel(page)
        if btn is None:
            return
        fechar = page.locator("button:has-text('Fechar')").first
        if fechar.is_visible(timeout=800):
            fechar.click(timeout=5000)
            print("    Fechou dialogo Anexar (sem arquivo).")
    except Exception:
        pass


def publicar_um(page, item, idx, total):
    _abortar_se_cancelado()
    rotulo = item.get("mes_ano") or DESCRICAO_PADRAO_REPASSE
    print(
        "[-> REPASSE] [{}/{}] {} | desc={} | prev={} real={}".format(
            idx,
            total,
            item.get("mes_ano"),
            DESCRICAO_PADRAO_REPASSE,
            item.get("valor_previsto"),
            item.get("valor_realizado"),
        )
    )
    abrir_modal(page)
    preencher_modal_repasse(page, item)
    status = clicar_publicar(page)

    if item.get("arquivo"):
        if status != "anexar":
            # Tenta ainda pegar o dialogo
            time.sleep(0.2)
        anexar_documento_apos_publicar(page, item)
    else:
        fechar_dialogo_anexar_se_aberto(page)
        print("    [AVISO] Sem PDF — publicou so as informacoes.")

    try:
        if _loc_modal_titulo(page).is_visible():
            fechar_modal(page)
    except Exception:
        try:
            fechar_modal(page)
        except Exception:
            pass
    print("    [OK] Concluido ({})".format(rotulo))


# ---------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------

def parse_args(argv=None):
    p = argparse.ArgumentParser(description="Publicacao de Repasse — portal CR2")
    p.add_argument("--test", action="store_true")
    p.add_argument("--yes", action="store_true")
    p.add_argument("--headless", action="store_true")
    p.add_argument("--pasta", type=str, default="", help="Pasta com Repasses.xlsx")
    return p.parse_args(argv)


def main(argv=None):
    args = parse_args(argv)
    global HEADLESS, MODO_TESTE, PASTA_BASE
    if args.headless:
        HEADLESS = True
    if args.test:
        MODO_TESTE = True
    if args.pasta:
        PASTA_BASE = Path(args.pasta)

    if not url_portal_ativa(URL_PORTAL_REPASSE):
        raise ValueError(
            "URL_PORTAL_REPASSE vazia. Informe a URL admin do modulo Repasse."
        )

    fila = montar_fila()
    if not fila:
        raise ValueError(
            "Nenhum repasse na fila. Rode a Extração de Repasses e informe a pasta "
            "com Repasses.xlsx (ex.: C:\\Downloads\\repasses)."
        )
    if MODO_TESTE:
        fila = fila[:1]
        print("[INFO] Modo teste: 1 repasse.")

    print("=" * 60)
    print("  PUBLICACAO DE REPASSE — portal CR2")
    print("  Total: {}".format(len(fila)))
    print("  Pasta: {}".format(PASTA_BASE))
    print("  URL: {}".format(URL_PORTAL_REPASSE))
    print("  Ritmo: apos cada item, {}s".format(PAUSA_ENTRE_ITENS))
    print("=" * 60)

    pw = browser = page = None
    ok = erros = 0
    try:
        pw, browser, page = criar_navegador_e_login(pular_enter_pos_login=args.yes)
        garantir_pagina_portal(page, URL_PORTAL_REPASSE, "Repasse")
        for i, item in enumerate(fila, 1):
            try:
                publicar_um(page, item, i, len(fila))
                ok += 1
                if i < len(fila):
                    print("    Proximo em {:.0f}s...".format(PAUSA_ENTRE_ITENS))
                    time.sleep(PAUSA_ENTRE_ITENS)
            except Cancelado:
                raise
            except Exception as e:
                erros += 1
                print("    [ERRO] {}".format(e))
                salvar_screenshot(page, "repasse_erro_{}".format(i))
                try:
                    _limpar_overlays_portal(page)
                except Exception:
                    try:
                        fechar_modal(page)
                    except Exception:
                        pass
                if i < len(fila):
                    time.sleep(PAUSA_ENTRE_ITENS)
    finally:
        if browser:
            try:
                browser.close()
            except Exception:
                pass
        if pw:
            try:
                pw.stop()
            except Exception:
                pass

    print("")
    print("=" * 60)
    print("  CONCLUIDO! OK: {} | Erros: {} | Total: {}".format(ok, erros, len(fila)))
    print("=" * 60)
    if erros:
        raise RuntimeError("Publicacao de repasse com {} erro(s).".format(erros))


if __name__ == "__main__":
    try:
        main()
    except Cancelado:
        print("[AVISO] Interrompido.")
        sys.exit(2)
