#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
==============================================================================
 BAIXADOR + EXTRATOR DE LICITAÇÕES — Portais CR2
==============================================================================

SCRIPT ÚNICO. Faz tudo:
  1) Coleta as licitações (API REST do WordPress; fallback: raspagem HTML)
  2) Baixa os anexos em pastas por licitação
  3) Lê o texto dos documentos (OCR opcional; cache em .ocr.txt)
  4) Extrai valores, datas e situação (regras + IA local Ollama opcional)
  5) Gera Licitacoes_preenchida.xlsx (intermediária) e as planilhas oficiais
     de upload: subirLicitacoes.xlsx + subirDocumentosLicitacoes.xlsx
  6) Lê os contratos/aditivos e gera Contratos/subirContratos.xlsx

Genérico: a entidade vem da URL de --listagem (qualquer portal CR2).

------------------------------------------------------------------------------
 REQUISITOS
------------------------------------------------------------------------------
  pip install requests beautifulsoup4 openpyxl pdfplumber

  OCR (opcional): Tesseract + Poppler (ou ocrmypdf + Ghostscript)
  IA local (opcional): Ollama com modelo llama3.2:3b (ou --modelo-ia)

------------------------------------------------------------------------------
 USO
------------------------------------------------------------------------------
  python script.py --ocr --refinar-ia
  python script.py --so-planilha --ocr --refinar-ia
  python script.py --listagem https://OUTRA.pa.gov.br/c/licitacoes/ --ocr
  python script.py --ignorar-ssl
==============================================================================
"""

import argparse
import os
import re
import shutil
import subprocess
import sys
import time
import unicodedata
from concurrent.futures import ThreadPoolExecutor, as_completed
from datetime import datetime
from pathlib import Path
from typing import Any
from urllib.parse import urljoin, urlparse, unquote

import requests
from bs4 import BeautifulSoup

# Console Windows legado pode não aceitar UTF-8; evita crash em prints
try:
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")
except Exception:
    pass


class Cancelado(Exception):
    """Fila interrompida pelo usuario (centro-automacoes)."""


def pedido_cancelado():
    return False


def _abortar_se_cancelado():
    if pedido_cancelado():
        print("  [AVISO] Fila cancelada pelo usuario.")
        raise Cancelado()


# ============================================================================
# CONFIGURAÇÃO
# ============================================================================
# Defaults neutros — o painel / CLI deve informar listagem e pasta.
LISTAGEM_PADRAO = ""
SAIDA_PADRAO    = r"C:\Downloads\Licitacoes"
PLANILHA_SAIDA  = "Licitacoes_preenchida.xlsx"
# Dict mutável: o main() só faz .clear()/.update() — sem `global` (evita
# "assigned to before global declaration" se alguém reordenar o código).
ULTIMO_RESULTADO_UPLOAD: dict = {}
SUBCATEGORIAS   = ["licitacoes-fracassadas", "licitacoes-desertas"]
MAX_PAGINAS     = 300     # trava de segurança na varredura de /page/N/

# ----------------------------------------------------------------------------
# FILTRO DE ANOS — vazio = todos. Use --anos 2023,2024 ou o campo do painel.
# Com filtro, a listagem (mais recente → mais antiga) PARA ao achar um ano
# anterior ao menor pedido (ex.: pediu 2023 → para no primeiro 2022).
# ----------------------------------------------------------------------------
ANOS_FILTRO = []

# ----------------------------------------------------------------------------
# RENOMEAÇÃO PELO TÍTULO INTERNO — quando o nome do anexo é genérico
# ("Download", "documento", "anexo1"...), o script lê o título dentro do
# PDF (via texto nativo ou OCR) e renomeia o arquivo por ele.
# ----------------------------------------------------------------------------
RENOMEAR_POR_TITULO = True

# ----------------------------------------------------------------------------
# MOTOR DE OCR — edite aqui (ou use --motor-ocr na linha de comando).
#   "tesseract" | "auto" = Tesseract (rápido; padrão)
#   "paddleocr" = opcional, mais pesado
# A IA (Ollama) SEMPRE confirma as informações (não só valores).
# ----------------------------------------------------------------------------
MOTOR_OCR = "tesseract"
# Limite de páginas quando o PDF é escaneado (nativo ainda pode ler mais)
OCR_MAX_PAGINAS_PRIOR = 6
OCR_MAX_PAGINAS_OUTRO = 2

EXT_DOCS = {
    ".pdf", ".doc", ".docx", ".xls", ".xlsx", ".ppt", ".pptx",
    ".zip", ".rar", ".7z", ".txt", ".csv", ".odt", ".ods", ".p7s", ".xml",
}
EXT_TEXTAVEIS = {".pdf"}

HEADERS = {"User-Agent": ("Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
                          "AppleWebKit/537.36 (KHTML, like Gecko) "
                          "Chrome/124.0 Safari/537.36")}
PAUSA, TENTATIVAS, TIMEOUT = 0.5, 3, 90
# Conexões paralelas ao baixar anexos da mesma licitação (painel: OPTO_DOWNLOAD_WORKERS).
DOWNLOAD_WORKERS = 4


def _erro_rede_temporario(exc: BaseException) -> bool:
    """DNS/conexão instável no VPS — vale retry."""
    texto = str(exc).lower()
    nomes = (
        "nameresolutionerror",
        "temporary failure in name resolution",
        "failed to resolve",
        "getaddrinfo failed",
        "nodename nor servname",
        "name or service not known",
        "connection reset",
        "connection aborted",
        "connection refused",
        "timed out",
        "timeout",
        "temporarily unavailable",
        "network is unreachable",
    )
    return any(n in texto for n in nomes)


def http_get(sessao, url, *, timeout=None, tentativas=None, **kwargs):
    """GET com retry em falha temporária de DNS/rede (comum em VPS)."""
    tentativas = int(tentativas or TENTATIVAS)
    timeout = timeout if timeout is not None else TIMEOUT
    ultimo = None
    for t in range(1, tentativas + 1):
        _abortar_se_cancelado()
        try:
            return sessao.get(url, timeout=timeout, **kwargs)
        except requests.RequestException as e:
            ultimo = e
            if t >= tentativas or not _erro_rede_temporario(e):
                raise
            espera = min(2.5 * t, 12)
            print(
                f"  ! Rede/DNS instável ({t}/{tentativas}) em {urlparse(url).netloc}: "
                f"nova tentativa em {espera:.0f}s…"
            )
            time.sleep(espera)
    if ultimo:
        raise ultimo
    raise RuntimeError("Falha HTTP sem detalhes")

MODALIDADES_RE = re.compile(
    r"(preg[aã]o|inexigibilidade|dispensa|convite|tomada\s+de\s+pre[çc]os?|"
    r"concorr[eê]ncia|chamamento|credenciamento|leil[aã]o|concurso|"
    r"ades[aã]o|registro\s+de\s+pre[çc]os|rdc|"
    r"n[aã]o\s+houve\s+ades|comunicado\s+da\s+comiss[aã]o)", re.IGNORECASE)

COLUNAS = ["Modalidade", "Número", "Ano", "Objeto", "Data de Publicação",
           "Data de Abertura", "Valor Estimado", "Situação da Licitação",
           "Valor Homologado"]


# ============================================================================
# PARTE 1 — EXTRATOR DE CAMPOS  (valores, datas, situação)
# ============================================================================
def sem_acento(txt):
    """Remove acentos preservando o comprimento (1 char -> 1 char)."""
    return "".join(c for c in unicodedata.normalize("NFKD", txt)
                   if not unicodedata.combining(c))


def normaliza(txt):
    """minúsculo, sem acento, espaços colapsados — só para TESTES DE PERTENCIMENTO
    (nunca use as posições deste texto para fatiar o original)."""
    return re.sub(r"\s+", " ", sem_acento(txt).lower())


def normaliza_com_mapa(txt):
    """Normaliza colapsando espaços, MAS devolve também um mapa de posições:
    mapa[i] = índice, no texto ORIGINAL, do caractere norm[i].
    Isso permite achar um rótulo no texto normalizado e fatiar o original
    na posição correta — corrige o desalinhamento causado pelo colapso."""
    base = sem_acento(txt).lower()          # preserva comprimento
    norm_chars, mapa = [], []
    prev_space = False
    for i, ch in enumerate(base):
        if ch.isspace():
            if prev_space:
                continue
            norm_chars.append(" ")
            mapa.append(i)
            prev_space = True
        else:
            norm_chars.append(ch)
            mapa.append(i)
            prev_space = False
    return "".join(norm_chars), mapa


# --- Moeda brasileira: 1.234.567,89 ---
RE_MOEDA = re.compile(
    r"(?<![\d.,])(\d{1,3}(?:\.\d{3})+,\d{2}|\d+,\d{2})(?![\d])")
# Com prefixo R$ aceitamos também valores SEM centavos ("R$ 1.500.000"),
# forma comum em avisos e ratificações. O prefixo evita capturar números
# soltos (datas, nº de processo). Aceita "RS" — OCR frequentemente lê o
# cifrão como a letra S.
RE_MOEDA_RS = re.compile(
    r"r\s*[\$s]\s*(\d{1,3}(?:\.\d{3})+(?:,\d{2})?|\d+,\d{2})", re.IGNORECASE)


def primeiro_valor(trecho):
    """Primeiro valor monetário do trecho, tentando o padrão completo e o
    padrão prefixado por R$. Retorna float ou None."""
    m1 = RE_MOEDA.search(trecho)
    m2 = RE_MOEDA_RS.search(trecho)
    cands = []
    if m1:
        cands.append((m1.start(), m1.group(1)))
    if m2:
        cands.append((m2.start(), m2.group(1)))
    if not cands:
        return None
    cands.sort()                     # o que aparece primeiro após o rótulo
    txt = cands[0][1]
    if "," not in txt:               # sem centavos: só aceito se veio de R$
        txt += ",00"
    return para_float(txt)


def para_float(txt_moeda):
    if not txt_moeda:
        return None
    try:
        return round(float(txt_moeda.strip().replace(".", "").replace(",", ".")), 2)
    except ValueError:
        return None


ROTULOS_ESTIMADO = [
    "valor total estimado", "valor global estimado", "valor estimado",
    "valor total de referencia", "valor de referencia", "valor referencia",
    "valor maximo", "valor total maximo", "preco maximo", "preco estimado",
    "valor total previsto", "importancia estimada", "valor orcado",
    # ampliação (vocabulário real dos documentos):
    "total estimado", "estimado em", "estimada em", "orcado em",
    "custo estimado", "valor previsto", "preco de referencia",
    "valor medio estimado", "media dos precos", "valor da estimativa",
    "estimativa de preco", "estimativa de custo", "valor de mercado",
    "totalizando a estimativa", "cujo valor estimado",
    "despesa estimada", "valor da despesa estimada", "estimado no valor",
    "valor global maximo", "valor teto", "limite maximo",
]
ROTULOS_HOMOLOGADO = [
    "valor homologado", "valor adjudicado", "valor total homologado",
    "valor global", "valor total do contrato", "valor do contrato",
    "valor contratado", "valor da contratacao", "valor total adjudicado",
    "homologo o", "adjudico o", "no valor global de", "no valor total de",
    # ampliação (vocabulário real dos documentos):
    "perfazendo o valor", "perfazendo um total", "perfazendo a importancia",
    "importancia de", "pela importancia de", "no valor de",
    "pelo valor de", "pelo valor global", "com o valor de",
    "vencedora com o valor", "melhor proposta no valor",
    "proposta no valor", "proposta vencedora no valor",
    "valor final", "valor arrematado", "totalizando",
    "valor total de", "valor total", "montante de", "no montante de",
    "valor global do contrato", "valor anual", "valor mensal estimado do contrato",
    "contratado por", "adjudicado por", "homologado no valor",
    "adjudicado no valor", "valor da proposta", "valor negociado",
    "apos negociacao", "lance final", "melhor lance", "lance vencedor",
    "quantia de", "pela quantia de", "importe de", "no importe de",
    "soma de", "pela soma de", "vencedora do certame com",
]
REFORCO_TOTAL = ["total", "global", "geral"]

DOCS_ESTIMADO = [
    "termo de referencia", "termo referencia", "pesquisa de mercado",
    "pesquisa de preco", "analise de preco", "mapa de preco",
    "aviso de licitacao", "aviso", "edital",
]
DOCS_HOMOLOGADO = [
    "homologacao", "adjudicacao", "termo de adjudicacao",
    "ata sessao", "ata de sessao", "ata de julgamento", "ata julgamento",
    "resultado", "contrato administrativo", "contrato",
]


def _janela_valor(texto_orig, rotulos, janela=200):
    """Procura cada rótulo (em texto normalizado) e captura o 1º valor
    monetário logo após ele NO TEXTO ORIGINAL, usando o mapa de posições.
    Retorna lista de tuplas (valor_float, rotulo, tem_reforco_total, trecho)."""
    tnorm, mapa = normaliza_com_mapa(texto_orig)
    achados = []
    for rot in rotulos:
        inicio = 0
        while True:
            pos = tnorm.find(rot, inicio)
            if pos == -1:
                break
            inicio = pos + len(rot)
            # posição do FIM do rótulo, mapeada de volta ao texto original
            fim_norm = pos + len(rot)
            orig_fim = mapa[fim_norm] if fim_norm < len(mapa) else len(texto_orig)
            orig_ini = mapa[pos]
            trecho = texto_orig[orig_fim: orig_fim + janela]
            val = primeiro_valor(trecho)
            if val and val > 0:
                contexto = texto_orig[orig_ini: orig_fim + janela]
                reforco = any(r in normaliza(contexto) for r in REFORCO_TOTAL)
                achados.append((val, rot, reforco,
                                re.sub(r"\s+", " ", contexto).strip()[:120]))
    return achados


def extrair_valor(texto, tipo):
    rotulos = ROTULOS_ESTIMADO if tipo == "estimado" else ROTULOS_HOMOLOGADO
    achados = _janela_valor(texto, rotulos)
    if not achados:
        return None
    achados.sort(key=lambda x: (x[2], x[0]), reverse=True)
    melhor = achados[0]
    vistos, candidatos = set(), []
    for v, rot, ref, tr in achados:
        if v not in vistos:
            vistos.add(v)
            candidatos.append({"valor": v, "rotulo": rot, "trecho": tr})
    return {"valor": melhor[0], "rotulo": melhor[1], "reforco": melhor[2],
            "trecho": melhor[3], "candidatos": candidatos}


MESES = {"janeiro":1,"fevereiro":2,"marco":3,"abril":4,"maio":5,"junho":6,
         "julho":7,"agosto":8,"setembro":9,"outubro":10,"novembro":11,"dezembro":12}
RE_DATA_NUM  = re.compile(r"\b(\d{1,2})[/.\-](\d{1,2})[/.\-](\d{4})\b")
RE_DATA_NUM2 = re.compile(r"\b(\d{1,2})[/.\-](\d{1,2})[/.\-](\d{2})\b(?![\d/.\-])")
RE_DATA_EXT  = re.compile(r"\b(\d{1,2})\s+de\s+([a-z]+)\s+de\s+(\d{4})\b")
ROTULOS_ABERTURA = [
    "data de abertura", "abertura da sessao", "abertura das propostas",
    "sessao publica", "data da sessao", "recebimento das propostas",
    "abertura", "realizar-se-a", "sera realizada", "data e horario",
    # ampliação (vocabulário real dos avisos/editais):
    "ocorrera no dia", "ocorrera em", "marcada para", "prevista para",
    "designada para", "agendada para", "no dia", "do dia",
    "as propostas serao recebidas", "inicio da sessao", "data da abertura",
    "data limite", "credenciamento", "disputa de lances",
    "sessao de disputa", "abertura dos envelopes", "entrega dos envelopes",
    "recebimento dos envelopes", "protocolo dos envelopes",
    "data do certame", "data da licitacao", "realizacao do certame",
]

RE_HORA_ADIANTE = re.compile(r"^.{0,20}?\b\d{1,2}\s*[:h]\s*\d{0,2}\b")


def _tem_hora_apos(trecho_norm_apos):
    """True se logo após a data há um horário ('as 09:00', '09h', '9 h 30') —
    assinatura típica da data de SESSÃO, e não de outras datas do documento."""
    return bool(RE_HORA_ADIANTE.search(trecho_norm_apos))


def data_valida(d, mth, y):
    """Valida calendário real (evita 31/02 vindo de OCR ruim)."""
    try:
        datetime(y, mth, d)
        return True
    except ValueError:
        return False


def _datas_no_trecho(trecho_norm):
    """Todas as datas válidas num trecho normalizado, na ordem, como
    (data, tem_hora) — tem_hora=True quando um horário segue a data."""
    achadas = []
    def _add(m, d, mth, y):
        if data_valida(d, mth, y):
            hora = _tem_hora_apos(trecho_norm[m.end(): m.end() + 25])
            achadas.append((m.start(), (d, mth, y), hora))
    for m in RE_DATA_NUM.finditer(trecho_norm):
        _add(m, int(m[1]), int(m[2]), int(m[3]))
    for m in RE_DATA_EXT.finditer(trecho_norm):
        if m[2] in MESES:
            _add(m, int(m[1]), MESES[m[2]], int(m[3]))
    for m in RE_DATA_NUM2.finditer(trecho_norm):
        y2 = int(m[3])
        _add(m, int(m[1]), int(m[2]), 2000 + y2 if y2 <= 79 else 1900 + y2)
    achadas.sort(key=lambda x: x[0])
    return [(d, h) for _, d, h in achadas]


def extrair_data_abertura(texto):
    """Busca de datas: feita inteiramente no texto normalizado (datas
    numéricas sobrevivem à normalização), sem fatiar o original.
    1) tenta cada rótulo e pega a 1ª data válida na janela após ele;
    2) sem rótulo, retorna None (o chamador pode usar o fallback)."""
    tnorm, _ = normaliza_com_mapa(texto)
    for rot in ROTULOS_ABERTURA:
        pos = tnorm.find(rot)
        if pos == -1:
            continue
        datas = _datas_no_trecho(tnorm[pos: pos + 220])
        if datas:
            return datas[0][0]
    return None


def datas_do_documento(texto):
    """Todas as datas válidas do documento, na ordem em que aparecem, como
    (data, tem_hora). Usada como fallback quando nenhum rótulo casa."""
    tnorm, _ = normaliza_com_mapa(texto)
    return _datas_no_trecho(tnorm)


def eh_contratacao_direta(modalidade):
    """Dispensa, Inexigibilidade, Contratação Direta, Adesão e Carona não
    têm sessão pública de abertura de propostas (Lei 8.666 arts. 24/25;
    Lei 14.133 arts. 74/75): o rito de conclusão é a RATIFICAÇÃO."""
    m = normaliza(modalidade or "")
    return ("dispensa" in m or "inexigibilidade" in m
            or "contratacao direta" in m or "adesao" in m or "carona" in m)


# ============================================================================
# MODALIDADES — regras de reconhecimento (ordem: específica → genérica)
# Cada regra: (termos que DEVEM aparecer, nome padrão).
# ============================================================================
_REGRAS_MODALIDADE = [
    # --- compostas de Registro de Preços (SRP) primeiro ---
    (["pregao", "eletronico", "srp"],       "Registro de Preços Originário de Pregão Eletrônico"),
    (["pregao", "eletronico", "registro de precos"],
                                            "Registro de Preços Originário de Pregão Eletrônico"),
    (["pregao", "presencial", "srp"],       "Registro de Preços Originário de Pregão Presencial"),
    (["pregao", "presencial", "registro de precos"],
                                            "Registro de Preços Originário de Pregão Presencial"),
    (["chamamento", "registro de precos"],  "Registro de Preços Originário de Chamamento Público"),
    (["chamada publica", "registro de precos"],
                                            "Registro de Preços Originário de Chamamento Público"),
    # --- adesão / carona ---
    (["adesao"],                            "Adesão a Ata de Registro de Preço"),
    (["carona"],                            "Carona"),
    # --- contratação direta e afins ---
    (["dispensa"],                          "Dispensa de Licitação"),
    (["inexigibilidade"],                   "Inexigibilidade de Licitação"),
    (["contratacao direta"],                "Contratação Direta"),
    # --- demais modalidades (lista oficial Front) ---
    (["dialogo competitivo"],               "Diálogo Competitivo"),
    (["credenciamento"],                    "Credenciamento"),
    (["chamada publica"],                   "Chamada Pública"),
    (["chamamento publico"],                "Chamada Pública"),
    (["chamamento"],                        "Chamada Pública"),
    # Concorrência eletrônica/presencial → Concorrência (CC)
    (["concorrencia"],                      "Concorrência"),
    (["concurso"],                          "Concurso"),
    (["carta convite"],                     "Convite"),
    (["convite"],                           "Convite"),
    (["leilao"],                            "Leilão"),
    (["tomada de preco"],                   "Tomada de Preços"),
    (["pregao", "eletronico"],              "Pregão Eletrônico"),
    (["pregao", "presencial"],              "Pregão Presencial"),
]


def modalidade_padrao(titulo):
    """Converte o texto de modalidade do título para o nome PADRÃO da
    biblioteca. Regras específicas primeiro (SRP antes de pregão simples).
    Pregão sem qualificação: decide pelo prefixo do número (PE/PP) e, na
    ausência, assume Pregão Presencial se o padrão 'PP' aparecer, senão
    Pregão Eletrônico (forma predominante desde 2020).
    Sem correspondência: devolve o texto original (para não perder dado)."""
    t = normaliza(titulo)
    for termos, padrao in _REGRAS_MODALIDADE:
        if all(term in t for term in termos):
            return padrao
    if "pregao" in t:
        # decide pelo prefixo do número: PPRP/PP -> presencial; PERP/PE -> eletrônico
        if re.search(r"\bpp(rp)?\b", t):
            return "Pregão Presencial"
        if re.search(r"\bpe(rp)?\b", t):
            return "Pregão Eletrônico"
        return "Pregão Eletrônico"
    mod, _ = split_modalidade_numero(titulo)
    return mod.strip().title() if mod else titulo.strip().title()


# ============================================================================
# TÍTULO INTERNO DO DOCUMENTO — para renomear anexos com nome genérico
# ============================================================================
# Padrões de título que os documentos oficiais trazem no cabeçalho.
# Avaliados em ordem; o primeiro que casar nas linhas iniciais vence.
_TITULOS_DOC = [
    "termo de homologacao", "termo de adjudicacao", "termo de ratificacao",
    "termo de referencia", "termo de dispensa", "termo de inexigibilidade",
    "ata de registro de precos", "ata de julgamento", "ata da sessao",
    "ata de sessao", "ata de abertura", "ata de recebimento",
    "aviso de licitacao", "aviso de resultado", "aviso de homologacao",
    "extrato de contrato", "extrato de dispensa", "extrato de inexigibilidade",
    "extrato de ata", "extrato de termo aditivo", "extrato",
    "contrato administrativo", "termo de contrato", "contrato",
    "termo aditivo", "apostilamento",
    "edital de licitacao", "edital de pregao", "edital de chamamento",
    "edital de credenciamento", "edital",
    "parecer juridico", "parecer tecnico", "parecer",
    "projeto basico", "planilha orcamentaria", "cronograma fisico",
    "memorial descritivo", "mapa de precos", "mapa comparativo",
    "pesquisa de precos", "pesquisa de mercado", "cotacao de precos",
    "ordem de servico", "ordem de compra", "portaria",
    "justificativa", "autorizacao", "solicitacao de despesa",
    "resultado de julgamento", "resultado do julgamento", "resultado",
    "recurso administrativo", "contrarrazoes", "impugnacao",
    "certidao", "declaracao", "despacho", "homologacao", "adjudicacao",
    "ratificacao",
]

# Nomes de link considerados genéricos (não descrevem o documento).
_NOMES_GENERICOS = re.compile(
    r"^(download|documento|anexo|arquivo|clique( aqui)?|abrir|visualizar|"
    r"pdf|doc(umento)? ?\d*|anexo ?\d*|arquivo ?\d*|sem_nome|untitled|"
    r"scan(ned)?[\w\- ]*|img[\w\- ]*|image[\w\- ]*|digitalizad[oa][\w\- ]*|"
    r"\d+|[\d\-_ .]+)$",
    re.IGNORECASE)


def nome_eh_generico(nome_sem_ext):
    """True se o nome do arquivo não descreve o conteúdo."""
    n = nome_sem_ext.strip()
    return bool(_NOMES_GENERICOS.match(n)) or len(n) <= 3


def titulo_interno(texto, max_linhas=40):
    """Encontra o título real do documento nas primeiras linhas do texto
    (nativo ou OCR). Prefere a linha que contém um padrão conhecido de
    documento oficial; retorna o texto da linha limpo, ou ''. """
    if not texto:
        return ""
    linhas = [ln.strip() for ln in texto.splitlines() if ln.strip()][:max_linhas]
    melhor, melhor_rank = "", -1
    for ln in linhas:
        if len(ln) > 90:            # títulos são curtos
            continue
        n = normaliza(ln)
        for idx, padrao in enumerate(_TITULOS_DOC):
            if padrao in n:
                # padrões no INÍCIO da lista são mais específicos:
                # rank maior = mais específico
                rank = len(_TITULOS_DOC) - idx
                if rank > melhor_rank:
                    melhor, melhor_rank = ln, rank
                break               # primeiro padrão (mais específico) da linha
    if not melhor:
        return ""
    # limpeza: colapsa espaços, remove pontuação de borda
    t = re.sub(r"\s+", " ", melhor).strip(" .:-–—_|")
    return t[:100]


def inferir_situacao(titulo, nomes_docs, modalidade=""):
    """Deduz a situação APENAS do título e dos NOMES dos documentos.
    (Não usa o texto interno: editais normais contêm frases-padrão como
    'caso a licitação seja considerada deserta...', que geravam falso
    positivo.) Aplica o rito correto por modalidade: contratação direta
    conclui por ratificação; certames concorrenciais, por homologação."""
    t = normaliza(titulo)
    docs = normaliza(" ".join(nomes_docs))
    if "deserta" in t or "deserta" in docs:
        return "Deserta"
    if "fracassada" in t or "fracassada" in docs:
        return "Fracassada"
    if "revogacao" in docs or "revogada" in docs or "revogado" in docs:
        return "Revogada"
    if "anulacao" in docs or "anulada" in docs or "anulado" in docs:
        return "Anulada"
    if eh_contratacao_direta(modalidade):
        # Rito da contratação direta: ratificação (e o contrato/extrato
        # publicado é evidência de conclusão do processo).
        if ("ratificacao" in docs or "ratificado" in docs or "ratificada" in docs
                or "extrato" in docs or "contrato" in docs):
            return "Ratificada"
        return ""
    if "homologacao" in docs or "homologado" in docs or "homologada" in docs:
        return "Homologada"
    if "adjudicacao" in docs or "adjudicado" in docs or "adjudicada" in docs:
        return "Adjudicada"
    return ""


# ============================================================================
# PARTE 2 — NOMES DE ARQUIVO / PASTA / TÍTULO
# ============================================================================
ILEGAL     = re.compile(r'[:\*\?"<>|\r\n\t]')
MARCADOR_N = re.compile(r"\bN\s*[ºoO°\.]+\s*", re.IGNORECASE)
RE_ANO     = re.compile(r"(?:19|20)\d{2}")


def limpa_nome(texto, maxlen=180):
    if not texto:
        return "sem_nome"
    texto = texto.strip().replace("/", "-").replace("\\", "-")
    texto = ILEGAL.sub("", texto)
    texto = re.sub(r"\s+", " ", texto).strip(" .")
    return texto[:maxlen].strip(" .") or "sem_nome"


def split_modalidade_numero(titulo):
    base = titulo.split("(")[0].strip()
    m = MARCADOR_N.search(base)
    if m:
        return base[:m.start()].strip(" -–—"), base[m.end():].strip(" -–—")
    # Sem "Nº": "PREGÃO ELETRÔNICO 9/2023-029" / "CONVITE 1/2023-006"
    m2 = re.search(
        r"(?<!\d)(\d{1,4}\s*/\s*(?:19|20)\d{2}(?:\s*[-–—]\s*[\w.]+)?)\s*$",
        base,
    )
    if m2:
        return base[: m2.start()].strip(" -–—"), m2.group(1).strip()
    m3 = re.search(
        r"(?<!\d)(\d{1,4}\s*/\s*(?:19|20)\d{2}(?:\s*[-–—]\s*[\w.]+)?)",
        base,
    )
    if m3:
        return base[: m3.start()].strip(" -–—"), m3.group(1).strip()
    return base.strip(), ""


def extrai_ano(numero):
    """Ano da licitação a partir do número/título curto.

    Prioridade:
      1) ano logo após a barra (ex.: '9/2023-200402' → 2023; '202301/2024' → 2024)
      2) última ocorrência 'solta' de 19xx/20xx (não colada no meio de mais dígitos)

    Evita falso ano em sufixos tipo '-200402' (dia/mês/seq), que quebrava o
    filtro de anos (lia 2004 e parava a varredura cedo demais).
    """
    s = (numero or "").strip()
    if not s:
        return ""
    m = re.search(r"/\s*((?:19|20)\d{2})\b", s)
    if m:
        return m.group(1)
    achados = re.findall(r"(?<!\d)((?:19|20)\d{2})(?!\d)", s)
    return achados[-1] if achados else ""


def ano_do_titulo(titulo):
    """Ano extraído do título da licitação (número), ou '' se não houver."""
    _mod, numero = split_modalidade_numero(titulo or "")
    return extrai_ano(numero) or extrai_ano(titulo or "")


def ano_de_data_pub(data_pub):
    """Ano (YYYY) a partir da data de publicação (datetime, dd/mm/aaaa ou ISO)."""
    if data_pub is None or data_pub == "":
        return ""
    if isinstance(data_pub, datetime):
        return str(data_pub.year)
    s = str(data_pub).strip()
    m = re.match(r"(\d{1,2})/(\d{1,2})/((?:19|20)\d{2})$", s)
    if m:
        return m.group(3)
    m = re.match(r"((?:19|20)\d{2})-\d{2}-\d{2}", s)
    if m:
        return m.group(1)
    m = re.search(r"(?<!\d)((?:19|20)\d{2})(?!\d)", s)
    return m.group(1) if m else ""


def mes_de_data_pub(data_pub):
    """Mês (1–12) a partir da data de publicação, ou 0 se desconhecido."""
    if data_pub is None or data_pub == "":
        return 0
    if isinstance(data_pub, datetime):
        return int(data_pub.month)
    s = str(data_pub).strip()
    m = re.match(r"(\d{1,2})/(\d{1,2})/((?:19|20)\d{2})$", s)
    if m:
        try:
            mes = int(m.group(2))
            return mes if 1 <= mes <= 12 else 0
        except ValueError:
            return 0
    m = re.match(r"((?:19|20)\d{2})-(\d{2})-\d{2}", s)
    if m:
        try:
            mes = int(m.group(2))
            return mes if 1 <= mes <= 12 else 0
        except ValueError:
            return 0
    return 0


def chave_ano_mes_licitacao(lic):
    """
    Chave (ano, mês) para amostragem mensal.
    Prefere data de publicação; senão ano do título e mês 0 (sem mês).
    """
    data_pub = (lic or {}).get("data_pub")
    ano = ano_de_data_pub(data_pub)
    mes = mes_de_data_pub(data_pub)
    if not ano:
        ano = ano_do_titulo((lic or {}).get("titulo") or "") or "0000"
    try:
        ano_i = int(ano)
    except (TypeError, ValueError):
        ano_i = 0
    return (ano_i, int(mes or 0))


def amostrar_mensal_diversificada(licitacoes, por_mes=5):
    """
    Mantém até `por_mes` licitações por (ano, mês), priorizando modalidades
    diferentes (round-robin entre modalidades do mês).

    Retorna (selecionadas, restantes).
    Sem data/mês conhecido → bucket (ano, 0), ainda limitado a `por_mes`.
    """
    from collections import defaultdict, deque

    por_mes = int(por_mes) if por_mes else 5
    if por_mes < 1:
        por_mes = 5
    if not licitacoes:
        return [], []

    buckets = defaultdict(list)
    for lic in licitacoes:
        buckets[chave_ano_mes_licitacao(lic)].append(lic)

    selecionadas = []
    restantes = []
    for chave in sorted(buckets.keys()):
        itens = buckets[chave]
        por_mod = defaultdict(deque)
        for lic in itens:
            mod = modalidade_padrao((lic or {}).get("titulo") or "") or "Outros"
            por_mod[mod].append(lic)
        # Modalidades com mais itens primeiro; empate alfabético
        mods = sorted(por_mod.keys(), key=lambda m: (-len(por_mod[m]), m))
        escolhidas = []
        while len(escolhidas) < por_mes and any(por_mod[m] for m in mods):
            progrediu = False
            for m in mods:
                if len(escolhidas) >= por_mes:
                    break
                if por_mod[m]:
                    escolhidas.append(por_mod[m].popleft())
                    progrediu = True
            if not progrediu:
                break
        sobra = []
        for m in mods:
            sobra.extend(list(por_mod[m]))
        selecionadas.extend(escolhidas)
        restantes.extend(sobra)
    return selecionadas, restantes


def _anexos_como_pares(lic):
    """Normaliza lic['anexos'] para lista de (nome, url)."""
    pares = []
    for item in (lic or {}).get("anexos") or []:
        if isinstance(item, (list, tuple)) and len(item) >= 2:
            pares.append((str(item[0] or ""), str(item[1] or "")))
        elif isinstance(item, dict):
            pares.append((
                str(item.get("nome") or item.get("texto") or ""),
                str(item.get("url") or item.get("link") or ""),
            ))
        elif isinstance(item, str):
            pares.append(("", item))
    return pares


def filtrar_licitacoes_docs_leves(licitacoes):
    """
    Prioriza processos rápidos de extrair:
      1) rejeita se TODOS os anexos forem só contrato firmado e/ou aditivo;
         (minuta de contrato / contrato social NÃO contam como exclusão)
      2) exige ao menos um doc útil (DFD, TR, edital, homologação…);
      3) ordena por menos anexos e, em empate, mais docs-chave (DFD dá bônus).

    Retorna (selecionadas, rejeitadas).
    """
    from ia_local.classificar_docs import (
        TIPOS_EXCLUSAO_DOCS_LEVES,
        TIPOS_SCORE_DOCS_LEVES,
        TIPOS_UTEIS_DOCS_LEVES,
        classificar,
    )

    selecionadas = []
    rejeitadas = []
    score_peso = {t: len(TIPOS_SCORE_DOCS_LEVES) - i
                  for i, t in enumerate(TIPOS_SCORE_DOCS_LEVES)}

    for lic in licitacoes or []:
        pares = _anexos_como_pares(lic)
        qtd = len(pares)
        tipos = set()
        for nome, url in pares:
            meta = classificar(nome, url)
            tipos.add(meta.get("tipo") or "outro")

        lic_out = dict(lic)
        lic_out["_qtd_anexos"] = qtd
        lic_out["_tipos_anexos"] = sorted(tipos)

        # Pula pastas que são só contrato/aditivo (sem edital/TR/etc.)
        if qtd > 0 and tipos and tipos.issubset(TIPOS_EXCLUSAO_DOCS_LEVES):
            lic_out["_filtro_motivo"] = "só contrato/aditivo"
            rejeitadas.append(lic_out)
            continue
        if not (tipos & TIPOS_UTEIS_DOCS_LEVES):
            lic_out["_filtro_motivo"] = "sem doc útil (TR/edital/homologação/DFD)"
            rejeitadas.append(lic_out)
            continue

        score = sum(score_peso.get(t, 0) for t in tipos)
        if "dfd" in tipos:
            score += 20
        # Penaliza um pouco se já tiver contrato firmado (ainda processa, mas depois)
        if "contrato" in tipos or "aditivo" in tipos:
            score -= 5
        lic_out["_score_docs"] = score
        selecionadas.append(lic_out)

    selecionadas.sort(
        key=lambda x: (
            int(x.get("_qtd_anexos") or 0),
            -int(x.get("_score_docs") or 0),
            str(x.get("titulo") or ""),
        )
    )
    return selecionadas, rejeitadas


def salvar_planilha_nao_migradas(pasta_saida, licitacoes, nome_arquivo=""):
    """
    Planilha de controle das licitações NÃO processadas na amostra.
    Colunas focadas em link para migração posterior.
    """
    import openpyxl
    from openpyxl.styles import Font

    pasta_saida = pasta_saida or "."
    os.makedirs(pasta_saida, exist_ok=True)
    nome = (nome_arquivo or "").strip() or "Nao_migradas_links.xlsx"
    if not nome.lower().endswith(".xlsx"):
        nome += ".xlsx"
    caminho = os.path.join(pasta_saida, nome)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Não migradas"
    headers = (
        "Link", "Título", "Data publicação", "Modalidade", "Status",
        "Motivo", "Qtd anexos",
    )
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True)

    for i, lic in enumerate(licitacoes or [], 2):
        titulo = (lic or {}).get("titulo") or ""
        link = (lic or {}).get("link") or ""
        data_pub = (lic or {}).get("data_pub") or ""
        if hasattr(data_pub, "strftime"):
            try:
                data_pub = data_pub.strftime("%d/%m/%Y")
            except Exception:
                data_pub = str(data_pub)
        modalidade = modalidade_padrao(titulo) if titulo else ""
        motivo = (lic or {}).get("_filtro_motivo") or "Não migrada"
        qtd = (lic or {}).get("_qtd_anexos")
        if qtd is None:
            qtd = len(_anexos_como_pares(lic))
        ws.cell(row=i, column=1, value=link)
        ws.cell(row=i, column=2, value=titulo)
        ws.cell(row=i, column=3, value=str(data_pub))
        ws.cell(row=i, column=4, value=modalidade)
        ws.cell(row=i, column=5, value="Não migrada")
        ws.cell(row=i, column=6, value=motivo)
        ws.cell(row=i, column=7, value=qtd)

    ws.column_dimensions["A"].width = 60
    ws.column_dimensions["B"].width = 55
    ws.column_dimensions["C"].width = 16
    ws.column_dimensions["D"].width = 28
    ws.column_dimensions["E"].width = 14
    ws.column_dimensions["F"].width = 18
    ws.column_dimensions["G"].width = 12
    wb.save(caminho)
    return os.path.abspath(caminho)


def _anos_filtro_set(anos_filtro):
    return {str(a).strip() for a in (anos_filtro or []) if str(a).strip()}


def _ano_minimo_filtro(anos_set):
    vals = []
    for a in anos_set or ():
        try:
            vals.append(int(a))
        except (TypeError, ValueError):
            pass
    return min(vals) if vals else None


def _ano_int_ou_none(ano_str):
    try:
        return int(ano_str) if ano_str else None
    except (TypeError, ValueError):
        return None


def decidir_anos_vs_filtro(ano_titulo, ano_pub, anos_set, ano_min):
    """
    Filtro de anos considerando NÚMERO (título) e/ou DATA DE PUBLICAÇÃO.

      pegar  — título OU publicação no filtro; ou nenhum ano conhecido
      pular  — anos conhecidos fora do filtro, mas ainda >= mínimo (continua)
      parar  — referência de ordenação (pub se houver, senão título) < mínimo

    Sem filtro → sempre pegar.
    """
    if not anos_set or ano_min is None:
        return "pegar"
    yt = _ano_int_ou_none(ano_titulo)
    yp = _ano_int_ou_none(ano_pub)
    if (yt is not None and str(yt) in anos_set) or (
        yp is not None and str(yp) in anos_set
    ):
        return "pegar"
    if yt is None and yp is None:
        return "pegar"
    # Ordenação da listagem costuma ser por data de publicação
    ref = yp if yp is not None else yt
    if ref is not None and ref < ano_min:
        return "parar"
    return "pular"


def decidir_ano_vs_filtro(ano_str, anos_set, ano_min):
    """Compat: só ano do título/número (sem data de publicação)."""
    return decidir_anos_vs_filtro(ano_str, "", anos_set, ano_min)


def extrai_objeto(titulo):
    m = re.search(r"\((.*)\)\s*$", titulo.strip(), re.DOTALL)
    return re.sub(r"\s+", " ", m.group(1)).strip() if m else ""


def nome_pasta(titulo):
    mod, num = split_modalidade_numero(titulo)
    return limpa_nome(f"{mod} {num}") if num else limpa_nome(mod or titulo)


def nome_arquivo_bruto(texto_link, url):
    """Nome como nas versões anteriores (sem capitalização) — usado para
    reconhecer e migrar arquivos baixados por rodadas antigas."""
    ext = os.path.splitext(urlparse(url).path)[1].lower() or ".pdf"
    base = limpa_nome(texto_link) or limpa_nome(
        unquote(os.path.basename(urlparse(url).path)))
    return base if base.lower().endswith(ext) else f"{base}{ext}"


def nome_arquivo(texto_link, url):
    return capitaliza_nome_arquivo(nome_arquivo_bruto(texto_link, url))


def variante_numerada(arquivo, i):
    raiz, ext = os.path.splitext(arquivo)
    return f"{raiz} ({i}){ext}"


# ----------------------------------------------------------------------------
# Capitalização de nomes de arquivo (padrão de título em português)
# ----------------------------------------------------------------------------
# Conectivos ficam minúsculos (exceto como primeira palavra).
_CONECTIVOS = {
    "de", "da", "do", "das", "dos", "e", "em", "no", "na", "nos", "nas",
    "a", "o", "as", "os", "ao", "aos", "para", "por", "com", "sem",
    "sob", "sobre", "que",
}
# Siglas comuns em documentos de licitação: preservadas em MAIÚSCULO.
_SIGLAS = {
    "srp", "arp", "rp", "cpl", "cmp", "pmg", "pe", "pp", "pprp", "perp",
    "tp", "cc", "con", "rdc", "cnpj", "cpf", "me", "epp", "ltda", "eireli",
    "fme", "fms", "fmas", "fundeb", "semed", "semus", "sead", "pgm",
    "ad", "cr", "ca", "cd", "cv", "cp", "dc", "dl", "in", "ll",
    "rpcp", "rppe", "rppp",
}


def capitaliza_nome_arquivo(nome):
    """'TERMO DE HOMOLOGAÇÃO.PDF' -> 'Termo de Homologação.pdf'.
    Regras: 1ª letra de cada palavra maiúscula; conectivos minúsculos
    (exceto no início); siglas conhecidas em maiúsculo; palavras com
    dígitos/ordinais (059.2023, Nº, 1º) ficam como estão; extensão
    minúscula."""
    raiz, ext = os.path.splitext(nome)
    saida = []
    for i, p in enumerate(raiz.split(" ")):
        if not p:
            continue
        if any(ch.isdigit() for ch in p) or "º" in p or "°" in p:
            saida.append(p)                    # números, ordinais, códigos
            continue
        pl = p.lower()
        if pl in _SIGLAS:
            saida.append(pl.upper())
        elif i > 0 and pl in _CONECTIVOS:
            saida.append(pl)
        else:
            saida.append(pl[:1].upper() + pl[1:])
    return " ".join(saida) + ext.lower()


def caminho_unico_arquivo(pasta, novo_nome, arquivo_atual):
    """Caminho livre para renomear `arquivo_atual` como `novo_nome` dentro
    de `pasta`. Se já existir OUTRO arquivo com esse nome, numera; se o
    destino for o próprio arquivo, devolve-o inalterado."""
    destino = os.path.join(pasta, novo_nome)
    if os.path.abspath(destino) == os.path.abspath(arquivo_atual):
        return arquivo_atual
    if not os.path.exists(destino):
        return destino
    i = 2
    while True:
        cand = os.path.join(pasta, variante_numerada(novo_nome, i))
        if os.path.abspath(cand) == os.path.abspath(arquivo_atual):
            return arquivo_atual
        if not os.path.exists(cand):
            return cand
        i += 1


def eh_anexo(url):
    ext = os.path.splitext(urlparse(url).path)[1].lower()
    if ext not in EXT_DOCS:
        return False
    return "/wp-content/themes/" not in url and "/wp-content/plugins/" not in url


def eh_artefato_ocr(nome):
    """True para subprodutos do OCR (.ocr.pdf / .ocr.txt) e temporários."""
    n = nome.lower()
    return n.endswith(".ocr.pdf") or n.endswith(".ocr.txt") or n.endswith(".part")


# ============================================================================
# PARTE 3 — COLETA (API REST + fallback HTML com varredura de páginas)
# ============================================================================
def extrair_anexos(html, url_base):
    soup = BeautifulSoup(html, "html.parser")
    # Preferimos o contêiner do post; se não houver, o documento inteiro
    # (eh_anexo já filtra imagens de tema/plugin). NÃO usamos seletores de
    # page-builder (Elementor tem dezenas de containers e select_one pegava
    # o errado, zerando os anexos).
    conteudo = soup.select_one(
        ".entry-content, .post-content, .single-content, article, main")
    if conteudo is not None:
        achados = _links_anexo(conteudo, url_base)
        if achados:
            return achados
    return _links_anexo(soup, url_base)


def _links_anexo(no, url_base):
    vistos, anexos = set(), []
    for a in no.find_all("a", href=True):
        href = urljoin(url_base, a["href"].strip())
        if eh_anexo(href) and href not in vistos:
            vistos.add(href)
            anexos.append((a.get_text(strip=True) or "", href))
    return anexos


def descobrir_categoria_id(sessao, base, slug):
    r = http_get(
        sessao,
        f"{base}/wp-json/wp/v2/categories",
        params={"slug": slug},
        timeout=TIMEOUT,
        tentativas=5,
    )
    r.raise_for_status()
    dados = r.json()
    return dados[0]["id"] if isinstance(dados, list) and dados else None


def coletar_posts_api(sessao, base, cat_id, anos_filtro=None):
    anos_set = _anos_filtro_set(anos_filtro)
    ano_min = _ano_minimo_filtro(anos_set)
    posts, pagina = [], 1
    parou_ano = False
    while pagina <= MAX_PAGINAS and not parou_ano:
        r = http_get(
            sessao,
            f"{base}/wp-json/wp/v2/posts",
            timeout=TIMEOUT,
            tentativas=5,
            params={
                "categories": cat_id,
                "per_page": 100,
                "page": pagina,
                "_fields": "title,link,content,date",
                "orderby": "date",
                "order": "desc",
            },
        )
        if r.status_code == 400:      # passou da última página
            break
        r.raise_for_status()
        lote = r.json()
        if not lote:
            break
        for p in lote:
            titulo = BeautifulSoup(
                p["title"]["rendered"], "html.parser"
            ).get_text(strip=True)
            ano_t = ano_do_titulo(titulo)
            ano_p = ano_de_data_pub(p.get("date") or "")
            acao = decidir_anos_vs_filtro(ano_t, ano_p, anos_set, ano_min)
            if acao == "parar":
                print(
                    "    · scanner parou: ano título={0} pub={1} "
                    "(filtro desde {2})".format(
                        ano_t or "?", ano_p or "?", ano_min
                    )
                )
                parou_ano = True
                break
            if acao == "pular":
                continue
            posts.append({
                "titulo": titulo,
                "link": p["link"], "content": p["content"]["rendered"],
                "date": p.get("date", "")})
        if parou_ano:
            break
        total_hdr = r.headers.get("X-WP-TotalPages")
        if total_hdr:
            try:
                if pagina >= int(total_hdr):
                    break
            except ValueError:
                pass
        pagina += 1
        time.sleep(PAUSA)
    return posts


def coletar_via_api(sessao, base, slugs, anos_filtro=None):
    try:
        res = []
        for slug in slugs:
            cid = descobrir_categoria_id(sessao, base, slug)
            if not cid:
                print(f"  · categoria '{slug}' não encontrada na API.")
                continue
            posts = coletar_posts_api(sessao, base, cid, anos_filtro=anos_filtro)
            print(f"  · '{slug}': {len(posts)} post(s) via API"
                  + (" (com corte por ano)" if anos_filtro else "") + ".")
            for p in posts:
                data_pub = ""
                if p["date"]:
                    try:
                        data_pub = datetime.fromisoformat(p["date"]).strftime("%d/%m/%Y")
                    except Exception:
                        pass
                res.append({"titulo": p["titulo"], "link": p["link"],
                            "data_pub": data_pub,
                            "anexos": extrair_anexos(p["content"], p["link"])})
        return res or None
    except Exception as e:
        print(f"  ! API REST indisponível ({e}). Fallback HTML.")
        return None


def raiz_categoria(url_listagem):
    """Canonicaliza a URL de listagem removendo /page/N/ do fim.
    Assim, se o usuário passar .../licitacoes/page/3/, varremos DESDE a
    página 1 — senão as páginas não linkadas a partir da 3 ficariam de fora."""
    u = re.sub(r"/page/\d+/?$", "/", url_listagem.rstrip("/") + "/")
    return u if u.endswith("/") else u + "/"


def slug_categoria_da_listagem(url_listagem):
    """Slug WordPress da categoria a partir da URL de listagem.
    Ex.: …/c/licitacoes/ → licitacoes"""
    raiz = raiz_categoria(url_listagem)
    seg = [s for s in urlparse(raiz).path.split("/") if s]
    if len(seg) >= 2 and seg[0] == "c":
        return seg[1]
    if len(seg) == 1 and seg[0] not in (
        "c",
        "author",
        "tag",
        "category",
        "wp-content",
        "wp-json",
    ):
        return seg[0]
    return None


def _html_tem_carregar_mais(soup):
    """Portais CR2 com botão Carregar Mais (paginação AJAX — /page/2/ repete)."""
    if soup.find(class_=re.compile(r"load-button", re.I)):
        return True
    for a in soup.find_all("a", href=True):
        if re.search(r"carregar\s+mais", a.get_text(" ", strip=True), re.I):
            return True
    return False


def _complementar_posts_via_api(sessao, posts, raiz, anos_filtro=None):
    """Quando o HTML para cedo, busca o restante na API REST do WordPress."""
    slug = slug_categoria_da_listagem(raiz)
    if not slug:
        return posts
    base = f"{urlparse(raiz).scheme}://{urlparse(raiz).netloc}"
    try:
        cid = descobrir_categoria_id(sessao, base, slug)
        if not cid:
            return posts
        lote = coletar_posts_api(sessao, base, cid, anos_filtro=anos_filtro)
    except Exception as e:
        print(f"    ! API complementar indisponível: {e}")
        return posts
    antes = len(posts)
    for p in lote:
        href = p["link"]
        if href in posts:
            continue
        titulo = p["titulo"]
        ano_p = ano_de_data_pub(p.get("date") or "")
        posts[href] = (titulo, ano_p)
    extra = len(posts) - antes
    if extra:
        print(
            "    · API complementou +{0} licitação(ões) "
            "(HTML/Carregar Mais parou cedo)".format(extra)
        )
    return posts


def _ano_pub_perto_do_link(a_tag):
    """Tenta achar ano de publicação perto do link (time/datetime ou cabeçalho)."""
    nos = []
    if a_tag is not None:
        nos.append(a_tag.parent)
        for nome in ("article", "li", "div", "tr"):
            p = a_tag.find_parent(nome)
            if p is not None:
                nos.append(p)
    for parent in nos:
        if parent is None:
            continue
        for t in parent.find_all("time"):
            dt = (t.get("datetime") or "").strip()
            if re.match(r"(?:19|20)\d{2}", dt):
                return dt[:4]
            txt = t.get_text(" ", strip=True)
            m = re.search(r"(?<!\d)((?:19|20)\d{2})(?!\d)", txt)
            if m:
                return m.group(1)
        # Cabeçalhos de mês no arquivo: "maio, 2023"
        for prev in parent.find_all_previous(
            ["h2", "h3", "h4", "strong", "b"], limit=6
        ):
            txt = prev.get_text(" ", strip=True)
            if re.search(
                r"(jan|fev|mar|abr|mai|jun|jul|ago|set|out|nov|dez|janeiro|"
                r"fevereiro|mar[cç]o|abril|maio|junho|julho|agosto|"
                r"setembro|outubro|novembro|dezembro)",
                txt,
                re.I,
            ):
                m = re.search(r"(?<!\d)((?:19|20)\d{2})(?!\d)", txt)
                if m:
                    return m.group(1)
    return ""


def filtra_posts_da_pagina(soup, url, dominio):
    """Retorna {href: (titulo, ano_pub_hint)}."""
    posts = {}
    for a in soup.find_all("a", href=True):
        href = urljoin(url, a["href"].split("#")[0].strip())
        titulo = a.get_text(strip=True)
        pp = urlparse(href)
        if pp.netloc != dominio:
            continue
        seg = [s for s in pp.path.split("/") if s]
        if len(seg) != 1 or seg[0] in (
            "c","author","tag","category","wp-content","wp-json",
            "admin","webmail","mapa-do-site","o-municipio","o-governo",
            "lgpd","covid-19"):
            continue
        if titulo and MODALIDADES_RE.search(titulo):
            posts[href] = (titulo, _ano_pub_perto_do_link(a))
    return posts


def coletar_posts_html(sessao, url_listagem, anos_filtro=None):
    """Varre a listagem página a página SEQUENCIALMENTE (/, /page/2/, /page/3/,
    ...) até receber 404, página sem novos posts, ou (com filtro de anos) o
    primeiro post com ano anterior ao mínimo pedido.

    O filtro usa ano do NÚMERO e/ou da DATA DE PUBLICAÇÃO (quando visível).
    """
    anos_set = _anos_filtro_set(anos_filtro)
    ano_min = _ano_minimo_filtro(anos_set)
    raiz = raiz_categoria(url_listagem)
    dominio = urlparse(raiz).netloc
    posts = {}  # href -> (titulo, ano_pub_hint)
    n = 1
    paginacao_quebrada = False
    tem_load_more = False
    while n <= MAX_PAGINAS:
        url = raiz if n == 1 else f"{raiz}page/{n}/"
        try:
            r = http_get(sessao, url, timeout=TIMEOUT, tentativas=5)
            if r.status_code == 404:
                break
            r.raise_for_status()
            r.encoding = r.apparent_encoding or "utf-8"
            soup = BeautifulSoup(r.text, "html.parser")
            if n == 1:
                tem_load_more = _html_tem_carregar_mais(soup)
        except Exception as e:
            host = urlparse(url).netloc
            if _erro_rede_temporario(e):
                print(
                    f"  ! Erro ao listar {url}: DNS/rede do servidor não resolveu "
                    f"'{host}'. Confira DNS do VPS (ex.: 1.1.1.1 / 8.8.8.8) e tente de novo.\n"
                    f"    Detalhe: {e}"
                )
            else:
                print(f"  ! Erro ao listar {url}: {e}")
            break
        novos = filtra_posts_da_pagina(soup, url, dominio)
        inedito = {k: v for k, v in novos.items() if k not in posts}
        if n > 1 and not inedito:
            paginacao_quebrada = True
            break            # página repetida/vazia (comum com Carregar Mais)

        if anos_set and ano_min is not None:
            manter = {}
            parar = False
            ano_parada = ""
            for href, (titulo, ano_pub_hint) in inedito.items():
                ano_t = ano_do_titulo(titulo)
                acao = decidir_anos_vs_filtro(
                    ano_t, ano_pub_hint, anos_set, ano_min
                )
                if acao == "parar":
                    # Não interrompe no meio da página: a ordem no HTML pode
                    # misturar itens; só evita ir para /page/N+1 depois.
                    parar = True
                    if not ano_parada:
                        ano_parada = ano_pub_hint or ano_t
                    continue
                if acao == "pegar":
                    manter[href] = (titulo, ano_pub_hint)
            posts.update(manter)
            if parar:
                print(
                    "    · scanner parou após a página {0}: achou ano {1} "
                    "(filtro desde {2})".format(n, ano_parada or "?", ano_min)
                )
                break
        else:
            posts.update(inedito)

        n += 1
        time.sleep(PAUSA)

    if paginacao_quebrada or tem_load_more:
        if paginacao_quebrada:
            print(
                "    · listagem HTML repetiu posts em /page/{0}/ "
                "(portal usa Carregar Mais)".format(n)
            )
        elif tem_load_more:
            print(
                "    · portal com botão Carregar Mais — "
                "complementando via API REST…"
            )
        posts = _complementar_posts_via_api(sessao, posts, raiz, anos_filtro)

    # Compat com callers: lista de (href, titulo)
    return [(h, t) for h, (t, _ano) in posts.items()]


RE_META_PUB = re.compile(
    r'property=["\']article:published_time["\']\s+content=["\'](\d{4})-(\d{2})-(\d{2})',
    re.IGNORECASE)
RE_TIME_TAG = re.compile(
    r'<time[^>]+datetime=["\'](\d{4})-(\d{2})-(\d{2})', re.IGNORECASE)


def data_pub_do_html(html):
    """Data de publicação do post extraída das metatags do WordPress
    (article:published_time) ou da tag <time>. Formato dd/mm/aaaa ou ''. """
    for rx in (RE_META_PUB, RE_TIME_TAG):
        m = rx.search(html)
        if m:
            y, mth, d = int(m[1]), int(m[2]), int(m[3])
            if data_valida(d, mth, y):
                return f"{d:02d}/{mth:02d}/{y}"
    return ""


def coletar_via_html(sessao, listagens, anos_filtro=None):
    res = []
    anos_set = _anos_filtro_set(anos_filtro)
    ano_min = _ano_minimo_filtro(anos_set)
    for url_listagem in listagens:
        print(f"  · raspando {raiz_categoria(url_listagem)} (varre /page/N/)")
        posts = coletar_posts_html(
            sessao, url_listagem, anos_filtro=anos_filtro
        )
        print(
            f"    {len(posts)} licitação(ões) no filtro — abrindo posts p/ anexos..."
        )
        for i, (url_post, titulo) in enumerate(posts, 1):
            if i == 1 or i % 20 == 0 or i == len(posts):
                print(f"    · anexos {i}/{len(posts)}…")
            time.sleep(PAUSA)
            try:
                r = sessao.get(url_post, timeout=TIMEOUT)
                r.raise_for_status()
                r.encoding = r.apparent_encoding or "utf-8"
                anexos = extrair_anexos(r.text, url_post)
                data_pub = data_pub_do_html(r.text)
            except Exception as e:
                print(f"    ! Erro em {url_post}: {e}")
                anexos, data_pub = [], ""
            # Confirma filtro com título + data real de publicação do post
            if anos_set and ano_min is not None:
                acao = decidir_anos_vs_filtro(
                    ano_do_titulo(titulo),
                    ano_de_data_pub(data_pub),
                    anos_set,
                    ano_min,
                )
                if acao != "pegar":
                    continue
            res.append({"titulo": titulo, "link": url_post,
                        "data_pub": data_pub, "anexos": anexos})
    return res


# ---------------------------------------------------------------------------
# Planilha Google / Excel como fonte (links na coluna Documentos)
# ---------------------------------------------------------------------------

_RE_DRIVE_SHEET_ID = re.compile(r"/spreadsheets/d/([a-zA-Z0-9_-]+)")
_RE_DRIVE_FOLDER_ID = re.compile(
    r"drive\.google\.com/(?:drive/)?folders/([a-zA-Z0-9_-]+)",
    re.I,
)
_RE_DRIVE_FILE_ID = re.compile(
    r"drive\.google\.com/(?:file/d/|open\?id=|uc\?(?:[^#]*&)?id=)([a-zA-Z0-9_-]+)",
    re.I,
)
_RE_DRIVE_URL_ANY = re.compile(
    r"https?://(?:drive|docs)\.google\.com/[^\s\"'<>]+",
    re.I,
)
_EXT_DOC_OK = (
    ".pdf", ".doc", ".docx", ".xls", ".xlsx", ".ppt", ".pptx",
    ".odt", ".ods", ".zip", ".rar", ".7z",
)


def _extrair_id_google_sheets(url: str) -> str | None:
    m = _RE_DRIVE_SHEET_ID.search(url or "")
    return m.group(1) if m else None


def eh_url_google_drive(url: str) -> bool:
    u = (url or "").lower()
    return "drive.google.com" in u or (
        "docs.google.com" in u and "/document" not in u and "/spreadsheets" not in u
    )


def _id_pasta_drive(url: str) -> str | None:
    m = _RE_DRIVE_FOLDER_ID.search(url or "")
    return m.group(1) if m else None


def _id_arquivo_drive(url: str) -> str | None:
    if _id_pasta_drive(url):
        return None
    m = _RE_DRIVE_FILE_ID.search(url or "")
    return m.group(1) if m else None


def url_download_drive(file_id: str) -> str:
    return "https://drive.google.com/uc?export=download&id={0}".format(file_id)


def _listar_pasta_drive_embedded(sessao, folder_id: str) -> list[tuple[str, str]]:
    """
    Lista arquivos públicos de uma pasta Drive via embeddedfolderview.
    Retorna [(nome, url_download), ...].
    """
    folder_id = (folder_id or "").strip()
    if not folder_id:
        return []
    url = "https://drive.google.com/embeddedfolderview?id={0}#list".format(folder_id)
    try:
        r = sessao.get(url, timeout=TIMEOUT)
        r.raise_for_status()
        r.encoding = r.apparent_encoding or "utf-8"
        html = r.text or ""
    except Exception as e:
        print("    ! Drive pasta {0}: {1}".format(folder_id[:12], e))
        return []

    from bs4 import BeautifulSoup

    soup = BeautifulSoup(html, "html.parser")
    out: list[tuple[str, str]] = []
    vistos: set[str] = set()
    for a in soup.find_all("a", href=True):
        href = a.get("href") or ""
        m = re.search(r"/file/d/([a-zA-Z0-9_-]+)", href)
        if not m:
            continue
        fid = m.group(1)
        if fid in vistos:
            continue
        vistos.add(fid)
        nome = (a.get_text(strip=True) or "").strip() or "documento.pdf"
        # Ignora pastas (sem extensão típica e sem mime no nome) — só arquivos
        low = nome.lower()
        if not any(low.endswith(ext) for ext in _EXT_DOC_OK):
            # Ainda assim aceita se Drive listou como arquivo; força .pdf só se sem ponto
            if "." not in nome:
                nome = nome + ".pdf"
        out.append((nome, url_download_drive(fid)))

    # Subpastas: links /folders/ no embedded (raro) — 1 nível
    for a in soup.find_all("a", href=True):
        href = a.get("href") or ""
        m = re.search(r"/folders/([a-zA-Z0-9_-]+)", href)
        if not m:
            continue
        sub = m.group(1)
        if sub == folder_id or sub in vistos:
            continue
        vistos.add(sub)
        for nome, dl in _listar_pasta_drive_embedded(sessao, sub):
            out.append((nome, dl))
    return out


def _listar_pasta_drive_data_id(sessao, folder_id: str) -> list[tuple[str, str]]:
    """Fallback: página normal da pasta (data-id=…)."""
    url = "https://drive.google.com/drive/folders/{0}".format(folder_id)
    try:
        r = sessao.get(url, timeout=TIMEOUT)
        r.raise_for_status()
        html = r.text or ""
    except Exception:
        return []
    ids = sorted(set(re.findall(r'data-id="([a-zA-Z0-9_-]{20,})"', html)))
    # O primeiro data-id costuma ser a própria pasta — filtramos depois via download
    out = []
    for fid in ids:
        if fid == folder_id:
            continue
        out.append(("documento-{0}.pdf".format(fid[:8]), url_download_drive(fid)))
    return out


def anexos_google_drive(sessao, url: str) -> list[tuple[str, str]]:
    """
    Converte link Drive (pasta ou arquivo) em lista de anexos
    no formato [(nome, url_download), ...] usado pelo baixador.
    """
    url = (url or "").strip()
    if not url:
        return []

    folder_id = _id_pasta_drive(url)
    if folder_id:
        anexos = _listar_pasta_drive_embedded(sessao, folder_id)
        if not anexos:
            anexos = _listar_pasta_drive_data_id(sessao, folder_id)
        return anexos

    file_id = _id_arquivo_drive(url)
    if not file_id:
        # open?id= sem /file/d/ — tenta pasta, senão arquivo
        m = re.search(r"[?&]id=([a-zA-Z0-9_-]+)", url)
        if m:
            cand = m.group(1)
            anexos = _listar_pasta_drive_embedded(sessao, cand)
            if anexos:
                return anexos
            file_id = cand

    if file_id:
        nome = "documento.pdf"
        # tenta obter nome pela página view
        try:
            r = sessao.get(
                "https://drive.google.com/file/d/{0}/view".format(file_id),
                timeout=TIMEOUT,
            )
            if r.ok:
                mt = re.search(
                    r'<meta\s+property="og:title"\s+content="([^"]+)"',
                    r.text or "",
                    re.I,
                )
                if mt and mt.group(1).strip():
                    nome = mt.group(1).strip()
        except Exception:
            pass
        if "." not in nome:
            nome = nome + ".pdf"
        return [(nome, url_download_drive(file_id))]
    return []


def _link_documentos_na_linha(cells: list[str], idx: dict[str, int]) -> str:
    """
    Preferência: coluna Documentos. Se vazia/deslocada, pega o 1º link Drive
    (ou http de documento) em qualquer célula da linha.
    """
    url = _cel(cells, idx, "documentos")
    if url.startswith("http") and (
        eh_url_google_drive(url) or eh_anexo(url) or "/wp-content/" in url.lower()
    ):
        return url.split()[0].strip()

    # Célula Documentos com URL no meio do texto
    if url:
        m = _RE_DRIVE_URL_ANY.search(url)
        if m:
            return m.group(0).rstrip(").,;")

    candidatos: list[str] = []
    for c in cells:
        s = str(c or "").strip()
        if not s or "http" not in s.lower():
            continue
        for m in _RE_DRIVE_URL_ANY.finditer(s):
            candidatos.append(m.group(0).rstrip(").,;"))
        if not candidatos and s.startswith("http"):
            # URL solta (não Drive) — só se parecer documento/página
            low = s.lower()
            if any(x in low for x in ("drive.google", ".pdf", "wp-content", "/licit")):
                candidatos.append(s.split()[0].strip())
    if not candidatos:
        return url if url.startswith("http") else ""
    # Prefere pasta Drive, depois arquivo Drive, depois demais
    for u in candidatos:
        if _id_pasta_drive(u):
            return u
    for u in candidatos:
        if eh_url_google_drive(u):
            return u
    return candidatos[0]


def baixar_planilha_google(url_ou_path: str, destino: Path) -> Path | None:
    """Baixa Google Sheets (export xlsx) ou usa arquivo local."""
    url_ou_path = (url_ou_path or "").strip().strip('"')
    if not url_ou_path:
        return None
    local = Path(url_ou_path)
    if local.is_file() and local.suffix.lower() in (".xlsx", ".xlsm", ".xls", ".csv"):
        return local.resolve()
    file_id = _extrair_id_google_sheets(url_ou_path)
    if not file_id:
        print("  ! ID da planilha Google não encontrado no link.")
        return None
    destino.parent.mkdir(parents=True, exist_ok=True)
    export_url = (
        "https://docs.google.com/spreadsheets/d/{0}/export?format=xlsx".format(
            file_id
        )
    )
    try:
        r = requests.get(export_url, headers=HEADERS, timeout=120)
        r.raise_for_status()
        if r.content[:2] != b"PK":
            print(
                "  ! Planilha não baixou como Excel. "
                "Compartilhe como «Qualquer pessoa com o link pode ver»."
            )
            return None
        destino.write_bytes(r.content)
        print("  · Planilha baixada: {0}".format(destino))
        return destino.resolve()
    except Exception as e:
        print("  ! Falha ao baixar planilha Google: {0}".format(e))
        return None


def _norm_col_planilha(txt: str) -> str:
    return normaliza(txt or "").replace(" ", "")


def _indice_colunas_planilha(linha_cab: list) -> dict[str, int]:
    """Mapeia cabeçalhos → índice de coluna."""
    idx: dict[str, int] = {}
    for i, raw in enumerate(linha_cab):
        n = _norm_col_planilha(str(raw or ""))
        if not n:
            continue
        if "modalidade" in n:
            idx.setdefault("modalidade", i)
        elif n in ("numero", "n", "no") or "numero" in n or n.startswith("n"):
            if "ordem" not in n and "cronolog" not in n:
                idx.setdefault("numero", i)
        elif "objeto" in n:
            idx.setdefault("objeto", i)
        elif "public" in n:
            idx.setdefault("publicacao", i)
        elif "abertura" in n:
            idx.setdefault("abertura", i)
        elif "situ" in n:
            idx.setdefault("situacao", i)
        elif "homolog" in n or "valor" in n:
            idx.setdefault("valor_homologado", i)
        elif "document" in n or n in ("link", "url", "site"):
            idx.setdefault("documentos", i)
    return idx


def _cel(row: list, idx: dict, chave: str) -> str:
    i = idx.get(chave)
    if i is None or i >= len(row):
        return ""
    return str(row[i] or "").strip()


def _parse_data_planilha(valor: str):
    v = (valor or "").strip()
    if not v:
        return ""
    for fmt in ("%d/%m/%Y", "%d-%m-%Y", "%Y-%m-%d"):
        try:
            return datetime.strptime(v, fmt)
        except ValueError:
            continue
    return v


def ler_linhas_planilha_fonte(caminho: Path) -> list[dict[str, Any]]:
    """Lê linhas da planilha-fonte (Google Sheets exportado ou xlsx local)."""
    from openpyxl import load_workbook

    wb = load_workbook(caminho, read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    wb.close()
    if not rows:
        return []
    idx = _indice_colunas_planilha([str(c or "") for c in rows[0]])
    # Sem coluna Documentos ainda tenta achar links Drive nas linhas
    out: list[dict[str, Any]] = []
    for row in rows[1:]:
        if not row:
            continue
        cells = [str(c) if c is not None else "" for c in row]
        url_doc = _link_documentos_na_linha(cells, idx)
        if not url_doc.startswith("http"):
            continue
        mod = _cel(cells, idx, "modalidade")
        num = _cel(cells, idx, "numero")
        obj = _cel(cells, idx, "objeto")
        data_pub = _cel(cells, idx, "publicacao")
        if data_pub.startswith("http") or eh_url_google_drive(data_pub):
            data_pub = ""
        elif re.search(r"finaliz|publicad|realizad|fracass", data_pub or "", re.I):
            # Coluna deslocada: situação caiu em Publicação
            data_pub = ""
        titulo = " ".join(
            x for x in (mod, num, ("({0})".format(obj) if obj else "")) if x
        ).strip() or url_doc
        out.append(
            {
                "titulo": titulo,
                "link": url_doc,
                "data_pub": data_pub,
                "anexos": [],
                "dados_planilha": {
                    "modalidade": mod,
                    "numero": num,
                    "objeto": obj,
                    "publicacao": _cel(cells, idx, "publicacao"),
                    "abertura": _cel(cells, idx, "abertura"),
                    "situacao": _cel(cells, idx, "situacao"),
                    "valor_homologado": _cel(cells, idx, "valor_homologado"),
                    "documentos": url_doc,
                },
            }
        )
    if not out and "documentos" not in idx:
        print("  ! Coluna «Documentos» (ou Link/URL) não encontrada e nenhum link Drive nas linhas.")
    return out


def coletar_via_planilha_fonte(
    sessao, url_ou_path: str, pasta_cache: Path, anos_filtro=None
) -> list:
    """Usa planilha (Google ou local): coluna Documentos → anexos (Drive ou HTML)."""
    path = baixar_planilha_google(
        url_ou_path, pasta_cache / "_planilha_fonte.xlsx"
    )
    if not path:
        return []
    linhas = ler_linhas_planilha_fonte(path)
    print("  · {0} linha(s) com link na planilha-fonte".format(len(linhas)))

    anos_set = _anos_filtro_set(anos_filtro)
    ano_min = _ano_minimo_filtro(anos_set)
    if anos_set and ano_min is not None:
        filtradas = []
        puladas = 0
        for lic in linhas:
            num_pl = ((lic.get("dados_planilha") or {}).get("numero") or "").strip()
            ano_t = ano_do_titulo(lic.get("titulo") or "") or extrai_ano(num_pl)
            ano_p = ano_de_data_pub(lic.get("data_pub"))
            acao = decidir_anos_vs_filtro(ano_t, ano_p, anos_set, ano_min)
            if acao != "pegar":
                puladas += 1
                continue
            filtradas.append(lic)
        if puladas:
            print(
                "  · filtro anos {0}: {1} linha(s) mantida(s), {2} fora".format(
                    ",".join(sorted(anos_set)), len(filtradas), puladas
                )
            )
        linhas = filtradas

    res = []
    for i, lic in enumerate(linhas, 1):
        url_post = lic["link"]
        if i == 1 or i % 10 == 0 or i == len(linhas):
            print("    · anexos planilha {0}/{1}…".format(i, len(linhas)))
        time.sleep(PAUSA)
        try:
            if eh_url_google_drive(url_post):
                anexos = anexos_google_drive(sessao, url_post)
                if not anexos:
                    print(
                        "    ! Drive sem arquivos públicos: {0}".format(
                            url_post[:72]
                        )
                    )
            else:
                r = sessao.get(url_post, timeout=TIMEOUT)
                r.raise_for_status()
                r.encoding = r.apparent_encoding or "utf-8"
                anexos = extrair_anexos(r.text, url_post)
                if not lic.get("data_pub"):
                    lic["data_pub"] = data_pub_do_html(r.text)
        except Exception as e:
            print("    ! Erro em {0}: {1}".format(url_post, e))
            anexos = []
        lic["anexos"] = anexos
        res.append(lic)
    return res

# ============================================================================
# PARTE 4 — OCR + LEITURA DE TEXTO
# ============================================================================
_OCR_BACKEND = None
_OCR_PATHS_OK = False


def _configurar_caminhos_ocr():
    """Localiza Tesseract/Poppler no Windows (INSTALAR.bat / winget)."""
    global _OCR_PATHS_OK
    if _OCR_PATHS_OK:
        return
    _OCR_PATHS_OK = True
    candidatos_tess = [
        os.environ.get("TESSERACT_CMD") or "",
        r"C:\Program Files\Tesseract-OCR\tesseract.exe",
        r"C:\Program Files (x86)\Tesseract-OCR\tesseract.exe",
        shutil.which("tesseract") or "",
    ]
    # WinGet packages
    local = os.environ.get("LOCALAPPDATA") or ""
    if local:
        winget = os.path.join(local, "Microsoft", "WinGet", "Packages")
        if os.path.isdir(winget):
            for nome in os.listdir(winget):
                if "Tesseract" in nome:
                    cand = os.path.join(winget, nome, "tesseract.exe")
                    if os.path.isfile(cand):
                        candidatos_tess.append(cand)
                if "Poppler" in nome:
                    for sub in ("Library\\bin", "bin"):
                        pdir = os.path.join(winget, nome, sub)
                        if os.path.isfile(os.path.join(pdir, "pdftoppm.exe")):
                            os.environ["PATH"] = pdir + os.pathsep + os.environ.get("PATH", "")
    tess = next((c for c in candidatos_tess if c and os.path.isfile(c)), None)
    if tess:
        try:
            import pytesseract
            pytesseract.pytesseract.tesseract_cmd = tess
            tess_dir = os.path.dirname(tess)
            if tess_dir not in os.environ.get("PATH", ""):
                os.environ["PATH"] = tess_dir + os.pathsep + os.environ.get("PATH", "")
        except Exception:
            pass
    for pop in (
        r"C:\Program Files\poppler\Library\bin",
        r"C:\Program Files\poppler\bin",
    ):
        if os.path.isfile(os.path.join(pop, "pdftoppm.exe")):
            os.environ["PATH"] = pop + os.pathsep + os.environ.get("PATH", "")
            break


def detectar_backend_ocr():
    global _OCR_BACKEND
    if _OCR_BACKEND is not None:
        return _OCR_BACKEND
    _configurar_caminhos_ocr()
    if shutil.which("ocrmypdf"):
        _OCR_BACKEND = "ocrmypdf"
    else:
        try:
            import pytesseract, pdf2image  # noqa
            # Confirma que o binário existe
            try:
                pytesseract.get_tesseract_version()
                _OCR_BACKEND = "pytesseract"
            except Exception:
                _OCR_BACKEND = "nenhum"
        except Exception:
            _OCR_BACKEND = "nenhum"
    return _OCR_BACKEND


def ler_texto_pdf(caminho, max_paginas=None):
    try:
        import pdfplumber
        with pdfplumber.open(caminho) as pdf:
            pages = pdf.pages
            if max_paginas is not None and max_paginas > 0:
                pages = pages[: int(max_paginas)]
            return "\n".join((p.extract_text() or "") for p in pages)
    except Exception:
        return ""


def caminho_sidecar(caminho_pdf):
    return caminho_pdf[:-4] + ".ocr.txt"


def ocr_para_texto(caminho, idioma="por", motor="auto", max_paginas=None):
    """Roda OCR e devolve o texto (cache .ocr.txt).

    auto = Tesseract primeiro; Paddle so se fraco. max_paginas limita o OCR.
    """
    txt_path = caminho_sidecar(caminho)
    motor = (motor or MOTOR_OCR or "auto").strip().lower() or "auto"
    if motor in ("easyocr", "docling", "surya"):
        motor = "auto"

    if motor in ("auto", "paddleocr", "tesseract"):
        try:
            raiz = Path(__file__).resolve().parent.parent
            if str(raiz) not in sys.path:
                sys.path.insert(0, str(raiz))
            from _comum.ocr_multi import ocr_pdf

            texto = ocr_pdf(caminho, motor=motor, max_paginas=max_paginas)
            if texto and texto.strip():
                _grava_sidecar(txt_path, texto)
                return texto
        except Exception as e:
            print(f"        (ocr_multi falhou: {e})")

    texto = _ocr_tesseract(caminho, idioma, max_paginas=max_paginas)
    _grava_sidecar(txt_path, texto)
    return texto


def _grava_sidecar(txt_path, texto):
    if not texto:
        return
    try:
        with open(txt_path, "w", encoding="utf-8") as f:
            f.write(texto)
    except Exception:
        pass


def _ocr_tesseract(caminho, idioma="por", max_paginas=None):
    """OCR via Tesseract (ocrmypdf preferido; pytesseract como alternativa)."""
    backend = detectar_backend_ocr()
    max_p = max_paginas if max_paginas and int(max_paginas) > 0 else OCR_MAX_PAGINAS_PRIOR
    max_p = max(1, int(max_p))
    if backend == "ocrmypdf":
        try:
            saida = caminho[:-4] + ".ocr.pdf"
            txt_path_s = caminho_sidecar(caminho)
            subprocess.run(
                ["ocrmypdf", "-l", idioma, "--force-ocr",
                 "--deskew", "--rotate-pages", "--oversample", "200",
                 "--pages", "1-%d" % max_p,
                 "--optimize", "1", "--sidecar", txt_path_s, caminho, saida],
                capture_output=True, text=True, timeout=300)
            if os.path.exists(txt_path_s):
                with open(txt_path_s, encoding="utf-8", errors="ignore") as f:
                    return f.read()
        except Exception as ee:
            print(f"        (ocrmypdf falhou: {ee})")
        return ""
    if backend == "pytesseract":
        try:
            import pytesseract
            from pdf2image import convert_from_path
            return "\n".join(
                pytesseract.image_to_string(img, lang=idioma, config="--oem 1 --psm 6")
                for img in convert_from_path(caminho, dpi=180, last_page=max_p))
        except Exception as ee:
            print(f"        (pytesseract falhou: {ee})")
        return ""
    return ""


def ocr_instalado():
    """True se Tesseract (binário) ou PaddleOCR estão disponíveis."""
    if detectar_backend_ocr() != "nenhum":
        return True
    try:
        import importlib.util
        return bool(importlib.util.find_spec("paddleocr"))
    except Exception:
        return False


def obter_texto(caminho, usar_ocr, idioma="por", min_chars=40, motor="auto",
                max_paginas=None, max_chars=None):
    """Retorna (texto, origem in {'nativo','ocr','ocr-cache','vazio'})."""
    if os.path.splitext(caminho)[1].lower() not in EXT_TEXTAVEIS:
        return "", "vazio"
    texto = ler_texto_pdf(caminho, max_paginas=max_paginas)
    if max_chars and texto and len(texto) > max_chars:
        texto = texto[:max_chars]
    if len(texto.strip()) >= min_chars:
        return texto, "nativo"
    # cache de OCR de rodadas anteriores
    txt_path = caminho_sidecar(caminho)
    if os.path.exists(txt_path):
        try:
            with open(txt_path, encoding="utf-8", errors="ignore") as f:
                cache = f.read()
            if max_chars and cache and len(cache) > max_chars:
                cache = cache[:max_chars]
            if len(cache.strip()) >= min_chars:
                return cache, "ocr-cache"
        except Exception:
            pass
    if usar_ocr and ocr_instalado():
        # OCR nunca lê o PDF inteiro — limita páginas (IA confirma valores)
        ocr_pag = max_paginas
        if ocr_pag is None or int(ocr_pag) <= 0:
            ocr_pag = OCR_MAX_PAGINAS_PRIOR
        else:
            ocr_pag = min(int(ocr_pag), OCR_MAX_PAGINAS_PRIOR)
        texto_ocr = ocr_para_texto(
            caminho, idioma, motor, max_paginas=ocr_pag
        )
        if max_chars and texto_ocr and len(texto_ocr) > max_chars:
            texto_ocr = texto_ocr[:max_chars]
        if len(texto_ocr.strip()) >= min_chars:
            return texto_ocr, "ocr"
    return texto, ("nativo" if texto.strip() else "vazio")


# ============================================================================
# PARTE 5 — DOWNLOAD
# ============================================================================
def baixar_arquivo(sessao, url, destino):
    """Baixa anexo; trata confirmação anti-vírus do Google Drive em arquivos grandes."""
    for t in range(1, TENTATIVAS + 1):
        try:
            with sessao.get(url, stream=True, timeout=TIMEOUT) as r:
                r.raise_for_status()
                ctype = (r.headers.get("Content-Type") or "").lower()
                # Drive às vezes devolve HTML pedindo confirm=
                if "text/html" in ctype and "drive.google" in (url or "").lower():
                    trecho = b"".join(list(r.iter_content(65536))[:8])
                    html = trecho.decode("utf-8", errors="ignore")
                    m = re.search(
                        r"confirm=([0-9A-Za-z_]+)",
                        html,
                    ) or re.search(
                        r'name="confirm"\s+value="([^"]+)"',
                        html,
                    )
                    fid = _id_arquivo_drive(url) or ""
                    if not fid:
                        mm = re.search(r"[?&]id=([a-zA-Z0-9_-]+)", url or "")
                        fid = mm.group(1) if mm else ""
                    if m and fid:
                        url2 = (
                            "https://drive.google.com/uc?export=download"
                            "&confirm={0}&id={1}".format(m.group(1), fid)
                        )
                        with sessao.get(url2, stream=True, timeout=TIMEOUT) as r2:
                            r2.raise_for_status()
                            os.makedirs(os.path.dirname(destino), exist_ok=True)
                            tmp = destino + ".part"
                            with open(tmp, "wb") as f:
                                for chunk in r2.iter_content(65536):
                                    if chunk:
                                        f.write(chunk)
                            os.replace(tmp, destino)
                        return True
                    print(
                        "        tentativa {0}/{1}: Drive pediu confirmação "
                        "e não foi possível ler o token".format(t, TENTATIVAS)
                    )
                    time.sleep(1.5 * t)
                    continue

                os.makedirs(os.path.dirname(destino), exist_ok=True)
                tmp = destino + ".part"
                with open(tmp, "wb") as f:
                    for chunk in r.iter_content(65536):
                        if chunk:
                            f.write(chunk)
                # HTML salvo por engano (login/bloqueio)
                if os.path.getsize(tmp) < 5000:
                    with open(tmp, "rb") as fh:
                        head = fh.read(200).lstrip().lower()
                    if head.startswith(b"<!doctype") or head.startswith(b"<html"):
                        os.remove(tmp)
                        raise RuntimeError("resposta HTML em vez de arquivo")
                os.replace(tmp, destino)
            return True
        except Exception as e:
            print(f"        tentativa {t}/{TENTATIVAS}: {e}")
            time.sleep(1.5 * t)
    return False


def _nova_sessao_download(sessao_base):
    """Session própria por thread (requests.Session não é thread-safe)."""
    s = requests.Session()
    try:
        s.headers.update(dict(sessao_base.headers))
    except Exception:
        s.headers.update(HEADERS)
    s.verify = getattr(sessao_base, "verify", True)
    return s


def baixar_anexos_da_licitacao(sessao, anexos, pasta, *, so_planilha=False):
    """
    Prepara nomes únicos e baixa anexos em série ou em paralelo (DOWNLOAD_WORKERS).
    Retorna lista de caminhos locais prontos.
    """
    arquivos_locais = []
    nomes_nesta_execucao = set()
    pendentes = []  # (arq, url, destino)

    for texto_link, url_arq in anexos or []:
        _abortar_se_cancelado()
        arq = nome_arquivo(texto_link, url_arq)
        if arq in nomes_nesta_execucao:
            i = 2
            while variante_numerada(arq, i) in nomes_nesta_execucao:
                i += 1
            arq = variante_numerada(arq, i)
        nomes_nesta_execucao.add(arq)
        destino = os.path.join(pasta, arq)

        if not os.path.exists(destino):
            legado = os.path.join(pasta, nome_arquivo_bruto(texto_link, url_arq))
            if legado != destino and os.path.exists(legado):
                try:
                    os.replace(legado, destino)
                    sc = caminho_sidecar(legado)
                    if os.path.exists(sc):
                        os.replace(sc, caminho_sidecar(destino))
                    print(f"    [REN ] {os.path.basename(legado)} -> {arq}")
                except OSError:
                    destino = legado

        if os.path.exists(destino):
            arquivos_locais.append(destino)
            continue
        if so_planilha:
            continue
        pendentes.append((arq, url_arq, destino))

    if not pendentes:
        return arquivos_locais

    workers = max(1, min(12, int(DOWNLOAD_WORKERS or 1)))

    def _um(item):
        arq, url_arq, destino = item
        _abortar_se_cancelado()
        print(f"    [DOWN] {arq}")
        s = _nova_sessao_download(sessao)
        try:
            ok = baixar_arquivo(s, url_arq, destino)
        finally:
            try:
                s.close()
            except Exception:
                pass
        if workers <= 1:
            time.sleep(PAUSA)
        return destino if ok else None

    if workers <= 1:
        for item in pendentes:
            caminho = _um(item)
            if caminho:
                arquivos_locais.append(caminho)
        return arquivos_locais

    print(f"    · downloads paralelos: {workers} conexões ({len(pendentes)} arquivo(s))")
    with ThreadPoolExecutor(max_workers=workers) as pool:
        futures = [pool.submit(_um, item) for item in pendentes]
        for fut in as_completed(futures):
            _abortar_se_cancelado()
            try:
                caminho = fut.result()
            except Cancelado:
                for f in futures:
                    f.cancel()
                raise
            except Exception as e:
                print(f"        erro no download paralelo: {e}")
                caminho = None
            if caminho:
                arquivos_locais.append(caminho)
    time.sleep(min(PAUSA, 0.25))
    return arquivos_locais


# ============================================================================
# PARTE 6 — EXTRAÇÃO DE VALORES POR LICITAÇÃO
# ============================================================================
def doc_serve_para(nome_arq, tipo):
    n = normaliza(nome_arq)
    alvos = DOCS_ESTIMADO if tipo == "estimado" else DOCS_HOMOLOGADO
    return any(a in n for a in alvos)


# Valor mínimo plausível para o TOTAL de uma contratação pública (valores
# menores encontrados perto de rótulos costumam ser multas, taxas ou itens).
VALOR_MINIMO_PLAUSIVEL = 100.00

# Tolerância da regra "homologado <= estimado": nos certames concorrenciais
# a disputa reduz o preço; homologado muito acima do estimado indica erro de
# extração (Lei 10.520 e Lei 14.133 desclassificam proposta acima do teto).
TOLERANCIA_HOMOLOGADO = 1.10


def _candidatos_do_tipo(arquivos_texto, tipo):
    """Todos os candidatos de valor do tipo, com metadados de confiança."""
    candidatos = []
    for nome, texto in arquivos_texto:
        if not texto:
            continue
        prioridade = 1 if doc_serve_para(nome, tipo) else 0
        achado = extrair_valor(texto, tipo)
        if not achado:
            continue
        for c in achado["candidatos"]:
            if c["valor"] < VALOR_MINIMO_PLAUSIVEL:
                continue
            candidatos.append({
                "valor": c["valor"], "rotulo": c["rotulo"], "doc": nome,
                "trecho": c.get("trecho", ""),
                "prioridade": prioridade,
                "reforco": any(r in c["rotulo"] for r in REFORCO_TOTAL)
                           or achado["reforco"] and c is achado["candidatos"][0],
                # Confirmação por extenso: "R$ X (… reais)" logo após o valor
                # é a redação padrão dos atos oficiais — evidência forte de
                # que o número capturado é mesmo um valor da contratação.
                "extenso": "reais" in normaliza(c.get("trecho", "")),
            })
    return candidatos


def _confianca(cand, contagem_docs):
    """Pontuação de confiança de um candidato de valor:
      +2 se o MESMO valor aparece em 2+ documentos distintos (validação cruzada)
      +1 se veio do documento correto para o tipo
      +1 se o rótulo indica total/global
      +1 se o valor vem seguido do extenso '(… reais)' — redação oficial
    A pontuação ORDENA os candidatos (o melhor é preenchido); não é veto."""
    pts = 0
    if contagem_docs.get(cand["valor"], 0) >= 2:
        pts += 2
    if cand["prioridade"] == 1:
        pts += 1
    if cand["reforco"]:
        pts += 1
    if cand.get("extenso"):
        pts += 1
    return pts


def _pontuar(candidatos):
    """Retorna lista [(pts, cand)] ordenada do melhor para o pior."""
    if not candidatos:
        return []
    docs_por_valor = {}
    for c in candidatos:
        docs_por_valor.setdefault(c["valor"], set()).add(c["doc"])
    contagem = {v: len(d) for v, d in docs_por_valor.items()}
    pontuados = [(_confianca(c, contagem), c) for c in candidatos]
    pontuados.sort(key=lambda x: (x[0], x[1]["prioridade"], x[1]["reforco"],
                                  x[1]["valor"]), reverse=True)
    return pontuados


def _melhor(candidatos):
    """Melhor candidato disponível (sempre preenche se houver algum)."""
    pontuados = _pontuar(candidatos)
    return pontuados[0][1] if pontuados else None


def _pts_de(cand, candidatos):
    for pts, c in _pontuar(candidatos):
        if c is cand:
            return pts
    return 0


def _resumo_candidatos(candidatos, limite=6):
    """String compacta 'doc: R$ valor (pts)' dos melhores candidatos."""
    partes = []
    for pts, c in _pontuar(candidatos)[:limite]:
        partes.append(f"{c['doc']}: {c['valor']:,.2f} ({pts}pt)".replace(
            ",", "@").replace(".", ",").replace("@", "."))
    return " | ".join(partes)


def _item_auditoria(campo, escolhido, candidatos, motivo=""):
    """Monta um registro de auditoria para um campo de valor."""
    if escolhido is None:
        return {"campo": campo, "valor": None, "doc": "—",
                "rotulo": motivo or "nenhum candidato encontrado",
                "trecho": "", "pts": "",
                "outros": _resumo_candidatos(candidatos)}
    return {"campo": campo, "valor": escolhido["valor"],
            "doc": escolhido["doc"], "rotulo": escolhido["rotulo"],
            "trecho": escolhido.get("trecho", ""),
            "pts": _pts_de(escolhido, candidatos),
            "outros": _resumo_candidatos(candidatos)}


def extrair_valores_da_licitacao(arquivos_texto, modalidade="",
                                 data_pub=None, ano=""):
    """Extrai valores e data de abertura. Preenche SEMPRE o melhor candidato
    disponível; os critérios de coerência da legislação ordenam a escolha e
    corrigem pares incoerentes (não vetam o preenchimento). Devolve também a
    auditoria de cada campo (documento, rótulo, trecho, pontos)."""
    res = {"valor_estimado": None, "valor_homologado": None,
           "data_abertura": None, "auditoria": []}

    cand_est = _candidatos_do_tipo(arquivos_texto, "estimado")
    cand_hom = _candidatos_do_tipo(arquivos_texto, "homologado")
    esc_est = _melhor(cand_est)
    esc_hom = _melhor(cand_hom)
    motivo_est = motivo_hom = ""

    # --- Coerência legal: homologado não deve superar o estimado ---
    if esc_est and esc_hom and esc_hom["valor"] > esc_est["valor"] * TOLERANCIA_HOMOLOGADO:
        coerentes = [c for c in cand_hom
                     if c["valor"] <= esc_est["valor"] * TOLERANCIA_HOMOLOGADO]
        alt = _melhor(coerentes)
        if alt:
            esc_hom = alt
            motivo_hom = "recombinado p/ coerência (homologado <= estimado)"
        else:
            coerentes_e = [c for c in cand_est
                           if esc_hom["valor"] <= c["valor"] * TOLERANCIA_HOMOLOGADO]
            alt_e = _melhor(coerentes_e)
            if alt_e:
                esc_est = alt_e
                motivo_est = "recombinado p/ coerência (homologado <= estimado)"
            else:
                if _pts_de(esc_hom, cand_hom) > _pts_de(esc_est, cand_est):
                    esc_est = None
                    motivo_est = "descartado: incoerente com homologado (sem alternativa)"
                else:
                    esc_hom = None
                    motivo_hom = "descartado: incoerente com estimado (sem alternativa)"

    res["valor_estimado"] = esc_est["valor"] if esc_est else None
    res["valor_homologado"] = esc_hom["valor"] if esc_hom else None

    it_e = _item_auditoria("Valor Estimado", esc_est, cand_est, motivo_est)
    it_h = _item_auditoria("Valor Homologado", esc_hom, cand_hom, motivo_hom)
    if motivo_est and esc_est:
        it_e["rotulo"] += f"  [{motivo_est}]"
    if motivo_hom and esc_hom:
        it_h["rotulo"] += f"  [{motivo_hom}]"
    res["auditoria"].append(it_e)
    res["auditoria"].append(it_h)

    # --- Data de abertura ---
    # Contratação direta (dispensa/inexigibilidade) NÃO tem sessão de
    # abertura de propostas: o campo fica vazio por definição legal.
    if eh_contratacao_direta(modalidade):
        res["auditoria"].append({
            "campo": "Data de Abertura", "valor": None, "doc": "—",
            "rotulo": "contratação direta: não há sessão de abertura (vazio correto)",
            "trecho": "", "pts": "", "outros": ""})
    else:
        d, doc_d, metodo = _achar_data_abertura(arquivos_texto, data_pub, ano)
        res["data_abertura"] = d
        if d:
            res["auditoria"].append({
                "campo": "Data de Abertura",
                "valor": f"{d[0]:02d}/{d[1]:02d}/{d[2]}",
                "doc": doc_d, "rotulo": metodo, "trecho": "", "pts": "",
                "outros": ""})
        else:
            res["auditoria"].append({
                "campo": "Data de Abertura", "valor": None, "doc": "—",
                "rotulo": "nenhuma data coerente encontrada",
                "trecho": "", "pts": "", "outros": ""})
    return res


def _achar_data_abertura(arquivos_texto, data_pub, ano):
    """Estratégia em camadas para a data de abertura. Retorna
    (data, documento, método):
      1) rótulo de abertura no AVISO/EDITAL;
      2) rótulo de abertura em qualquer documento;
      3) fallback: data com HORÁRIO adjacente coerente (aviso/edital primeiro)
         — 'dd/mm/aaaa às 09:00' é a assinatura da data de sessão;
      4) fallback final: primeira data coerente de qualquer documento."""
    docs_prior = [(n, t) for n, t in arquivos_texto
                  if t and any(k in normaliza(n) for k in ("aviso", "edital"))]
    docs_resto = [(n, t) for n, t in arquivos_texto
                  if t and (n, t) not in docs_prior]

    # 1 e 2: por rótulo
    for grupo, tag in ((docs_prior, "rótulo no aviso/edital"),
                       (docs_resto, "rótulo no documento")):
        for nome, texto in grupo:
            d = extrair_data_abertura(texto)
            if d and _data_abertura_coerente(d, data_pub, ano):
                return d, nome, tag
    # 3: data com horário adjacente
    for grupo in (docs_prior, docs_resto):
        for nome, texto in grupo:
            for d, tem_hora in datas_do_documento(texto):
                if tem_hora and _data_abertura_coerente(d, data_pub, ano):
                    return d, nome, "data seguida de horário (fallback)"
    # 4: primeira data coerente
    for grupo in (docs_prior, docs_resto):
        for nome, texto in grupo:
            for d, _h in datas_do_documento(texto):
                if _data_abertura_coerente(d, data_pub, ano):
                    return d, nome, "primeira data coerente (fallback)"
    return None, "", ""


def _data_abertura_coerente(d, data_pub, ano):
    """Sanidade da data de abertura:
      - nunca ANTERIOR à publicação (o edital é publicado antes da sessão);
      - no máximo ~1 ano após a publicação;
      - no ano do certame ou no seguinte (processos publicados em dezembro
        podem abrir em janeiro)."""
    dia, mes, a = d
    try:
        dt = datetime(a, mes, dia)
    except ValueError:
        return False
    if isinstance(data_pub, datetime):
        if dt < data_pub:
            return False
        if (dt - data_pub).days > 400:
            return False
    if ano:
        try:
            ano_i = int(ano)
            if not (ano_i <= a <= ano_i + 1):
                return False
        except ValueError:
            pass
    return True


# ============================================================================
# PARTE 7 — PLANILHA
# ============================================================================
def ultima_linha_com_dados(ws, colunas_idx):
    """Última linha que tem conteúdo real nas colunas mapeadas. (max_row do
    openpyxl conta formatação fantasma — no modelo, vai até a linha 992.)"""
    ultima = 1
    for r in range(2, ws.max_row + 1):
        if any(ws.cell(row=r, column=c).value not in (None, "")
               for c in colunas_idx):
            ultima = r
    return ultima


def preencher_planilha(linhas, auditoria, modelo, saida):
    import openpyxl
    from openpyxl.styles import Font
    from copy import copy

    if modelo and os.path.exists(modelo):
        wb = openpyxl.load_workbook(modelo)
        ws = wb.active
        header = {}
        for c in range(1, ws.max_column + 1):
            v = ws.cell(row=1, column=c).value
            if v:
                header[str(v).strip()] = c
        ref = ws.cell(row=1, column=1).font
        estilo = Font(name=ref.name or "Arial", size=ref.size or 10, bold=False)
        linha_ini = ultima_linha_com_dados(ws, header.values()) + 1
    else:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Licitações"
        for i, nome in enumerate(COLUNAS, 1):
            ws.cell(row=1, column=i, value=nome).font = Font(
                name="Arial", size=10, bold=True)
        header = {nome: i for i, nome in enumerate(COLUNAS, 1)}
        estilo = Font(name="Arial", size=10, bold=False)
        linha_ini = 2

    fmt = {"Data de Publicação": "dd/mm/yyyy", "Data de Abertura": "dd/mm/yyyy",
           "Valor Estimado": "#,##0.00", "Valor Homologado": "#,##0.00"}

    for r_off, dados in enumerate(linhas):
        r = linha_ini + r_off
        for nome_col, valor in dados.items():
            if nome_col not in header:
                continue
            cell = ws.cell(row=r, column=header[nome_col], value=valor)
            cell.font = copy(estilo)
            if nome_col in fmt:
                cell.number_format = fmt[nome_col]

    # ----- Aba Auditoria: origem de cada dado extraído -----
    if "Auditoria" in wb.sheetnames:
        del wb["Auditoria"]
    aud = wb.create_sheet("Auditoria")
    cab = ["Licitação", "Campo", "Valor preenchido", "Documento de origem",
           "Rótulo / método / motivo", "Trecho do documento",
           "Confiança (pts)", "Outros candidatos (doc: valor (pts))"]
    for i, h in enumerate(cab, 1):
        aud.cell(row=1, column=i, value=h).font = Font(
            name="Arial", size=10, bold=True)
    ra = 2
    for item in auditoria:
        aud.cell(row=ra, column=1, value=item["licitacao"])
        aud.cell(row=ra, column=2, value=item["campo"])
        cel_v = aud.cell(row=ra, column=3,
                         value=item["valor"] if item["valor"] is not None
                         else "(em branco)")
        if isinstance(item["valor"], (int, float)):
            cel_v.number_format = "#,##0.00"
        aud.cell(row=ra, column=4, value=item["doc"])
        aud.cell(row=ra, column=5, value=item["rotulo"])
        aud.cell(row=ra, column=6, value=item["trecho"])
        aud.cell(row=ra, column=7, value=item["pts"])
        aud.cell(row=ra, column=8, value=item["outros"])
        for c in range(1, 9):
            aud.cell(row=ra, column=c).font = Font(name="Arial", size=10)
        ra += 1
    for col, w in {"A": 42, "B": 17, "C": 16, "D": 34, "E": 42,
                   "F": 60, "G": 13, "H": 55}.items():
        aud.column_dimensions[col].width = w

    wb.save(saida)


# ============================================================================
# PRIORIDADE DE DOCS + IA LOCAL (Ollama)
# ============================================================================
def _log(msg, *args):
    """Print imediato (flush) para o painel não parecer parado."""
    if args:
        msg = msg.format(*args)
    print(msg, flush=True)


def _barra(atual, total, largura=18):
    if total <= 0:
        return "[" + ("#" * largura) + "]"
    frac = max(0.0, min(1.0, float(atual) / float(total)))
    cheios = int(round(frac * largura))
    return "[" + ("#" * cheios) + ("-" * (largura - cheios)) + "]"


def _pct(atual, total):
    if total <= 0:
        return 0
    return int(100 * atual / total)


def _garantir_path_script():
    d = os.path.dirname(os.path.abspath(__file__))
    if d not in sys.path:
        sys.path.insert(0, d)
    return d


def situacao_para_front(situacao):
    """Mapeia rótulos internos (Homologada, Deserta…) para o vocabulário Front."""
    if not situacao:
        return "Em andamento"
    s = str(situacao).strip()
    try:
        _garantir_path_script()
        from gestor_regras.config_front import MAPA_SITUACAO, SITUACOES
    except ImportError:
        return s
    if s in SITUACOES:
        return s
    return MAPA_SITUACAO.get(normaliza(s), s)


def numero_com_sigla_front(numero, modalidade):
    """Aplica sigla Front sem alterar dígitos/códigos do número."""
    try:
        _garantir_path_script()
        from ia_local import numero_com_sigla
        return numero_com_sigla(numero or "", modalidade or "")
    except ImportError:
        return (numero or "").strip()


def padronizar_linha_para_todas_planilhas(linha):
    """
    Deixa Modalidade/Número/Ano iguais em todas as planilhas
    (Licitacoes_preenchida + subirLicitacoes + subirDocumentos).

    Usa as mesmas regras do Front (inclui Registro de Preços → RPPP/RPPE).
    Números/códigos do nome ficam iguais; só a categoria (sigla) muda.
    """
    if not isinstance(linha, dict):
        return linha
    try:
        _garantir_path_script()
        from gestor_regras.upload import registro_de_linha_planilha
        from gestor_regras.front import linha_front
        lf = linha_front(registro_de_linha_planilha(linha))
    except Exception:
        num = numero_com_sigla_front(
            linha.get("Número") or linha.get("Numero") or "",
            linha.get("Modalidade") or "",
        )
        if num:
            linha["Número"] = num
        return linha

    if lf.get("modalidade"):
        linha["Modalidade"] = lf["modalidade"]
    if lf.get("numero"):
        linha["Número"] = lf["numero"]
    if lf.get("ano"):
        linha["Ano"] = lf["ano"]
    return linha


def _precisa_ia(linha):
    """True se falta campo que a IA pode confirmar/preencher."""
    for k in (
        "Número", "Objeto", "Situação da Licitação",
        "Data de Publicação", "Data de Abertura",
        "Valor Estimado", "Valor Homologado",
    ):
        v = linha.get(k)
        if v is None or str(v).strip() in ("", "Não informado"):
            return True
    return False


def _ler_docs_priorizados(arquivos_locais, usar_ocr, idioma_ocr, motor_ocr,
                          renomear, pasta):
    """
    Classifica anexos; Edital, DFD, TR e Termo de Homologação são lidos
    INTEIROS para a IA de valores. Devolve (arquivos_texto, cabecalhos, nomes).
    """
    _garantir_path_script()
    from ia_local import classificar, limites_leitura, selecionar_para_leitura
    try:
        from ia_local.classificar_docs import TIPOS_OBRIGATORIOS_VALORES
    except ImportError:
        # processo do painel pode ter módulo antigo em cache
        import importlib
        import ia_local.classificar_docs as _cd
        importlib.reload(_cd)
        TIPOS_OBRIGATORIOS_VALORES = getattr(
            _cd,
            "TIPOS_OBRIGATORIOS_VALORES",
            ("dfd", "edital", "termo_referencia", "orcamento", "homologacao", "contrato"),
        )

    metas = []
    for fp in arquivos_locais:
        nome = os.path.basename(fp)
        meta = classificar(nome, fp)
        meta["caminho"] = fp
        meta["nome"] = nome
        meta["url"] = fp  # selecionar_para_leitura usa url p/ .pdf
        metas.append(meta)

    metas.sort(key=lambda x: (x.get("prioridade", 99), x.get("nome", "").lower()))
    escolhidos = selecionar_para_leitura(metas, max_pdfs=8, so_pdf=True)
    escolhidos_paths = {os.path.abspath(e["caminho"]) for e in escolhidos}

    # Garante Edital + DFD (e demais obrigatórios de valor) mesmo se a seleção falhar
    for m in metas:
        if m.get("tipo") in TIPOS_OBRIGATORIOS_VALORES:
            escolhidos_paths.add(os.path.abspath(m["caminho"]))

    # Garante limites também nos não escolhidos (leitura curta p/ rename/situação)
    por_path = {os.path.abspath(m["caminho"]): m for m in metas}

    arquivos_texto = []
    cabecalhos = []
    novos_paths = []

    for fp in arquivos_locais:
        _abortar_se_cancelado()
        abs_fp = os.path.abspath(fp)
        meta = por_path.get(abs_fp) or classificar(os.path.basename(fp), fp)
        tipo = meta.get("tipo") or "outro"
        max_pag, max_chars = limites_leitura(tipo)
        eh_escolhido = abs_fp in escolhidos_paths
        # não-priorizados: leitura curta só p/ rename / nomes — SEM OCR
        usar_ocr_doc = bool(usar_ocr) and eh_escolhido
        if not eh_escolhido:
            if max_pag is None:
                max_pag = 3
            else:
                max_pag = min(max_pag, 3)
            max_chars = min(max_chars, 3500)
        elif tipo in ("edital", "dfd", "termo_referencia", "homologacao"):
            # Nativo: pode ler bastante; OCR (se precisar) fica limitado
            if max_pag is None:
                max_pag = OCR_MAX_PAGINAS_PRIOR
            else:
                max_pag = min(max_pag, OCR_MAX_PAGINAS_PRIOR)
            _log(
                "        lendo {0} ({1}, até {2} pág.)…",
                meta.get("rotulo") or tipo,
                os.path.basename(fp)[:50],
                max_pag,
            )
        elif eh_escolhido and max_pag is None:
            max_pag = OCR_MAX_PAGINAS_PRIOR

        texto, origem = obter_texto(
            fp, usar_ocr_doc, idioma_ocr, motor=motor_ocr,
            max_paginas=max_pag, max_chars=max_chars,
        )
        if origem == "ocr":
            print(f"        (OCR leve) {os.path.basename(fp)}")

        nome_atual = os.path.basename(fp)
        raiz_atual, ext = os.path.splitext(nome_atual)
        if renomear and texto and nome_eh_generico(raiz_atual):
            tit = titulo_interno(texto)
            if tit:
                novo = capitaliza_nome_arquivo(limpa_nome(tit) + ext.lower())
                novo_fp = caminho_unico_arquivo(pasta, novo, fp)
                if novo_fp != fp:
                    try:
                        os.replace(fp, novo_fp)
                        sc_velho = caminho_sidecar(fp)
                        if os.path.exists(sc_velho):
                            os.replace(sc_velho, caminho_sidecar(novo_fp))
                        print(f"        [REN ] {nome_atual} -> "
                              f"{os.path.basename(novo_fp)}")
                        fp = novo_fp
                        meta = classificar(os.path.basename(fp), fp)
                        meta["caminho"] = fp
                        meta["nome"] = os.path.basename(fp)
                        tipo = meta.get("tipo") or tipo
                    except OSError as e:
                        print(f"        (renomear falhou: {e})")

        nome_final = os.path.basename(fp)
        arquivos_texto.append((nome_final, texto))
        novos_paths.append(fp)

        doc = {
            "nome": nome_final,
            "tipo": tipo,
            "rotulo": meta.get("rotulo") or tipo,
            "texto": texto or "",
            "prioritario": bool(meta.get("prioritario")),
        }
        if eh_escolhido or meta.get("prioritario") or tipo in TIPOS_OBRIGATORIOS_VALORES:
            cabecalhos.append(doc)
        elif tipo in ("contrato", "aceite_adesao", "homologacao", "ata"):
            cabecalhos.append(doc)

    # ordena cabecalhos: DFD/Edital/TR/Homologação primeiro (valores)
    peso = {
        "dfd": 0, "edital": 1, "termo_referencia": 2, "homologacao": 3,
        "orcamento": 4, "etp": 5, "contrato": 6,
    }
    cabecalhos.sort(key=lambda d: (peso.get(d.get("tipo"), 40), d.get("nome", "")))

    # log dos prioritários lidos
    tipos_lidos = [
        "{0}({1}c)".format(c["tipo"], len(c.get("texto") or ""))
        for c in cabecalhos if c.get("texto")
    ]
    if tipos_lidos:
        _log("        docs p/ IA/valores: {0}", ", ".join(tipos_lidos[:8]))

    return arquivos_texto, cabecalhos, [os.path.basename(p) for p in novos_paths]


def _aplicar_valores_prioritarios(linha, cabecalhos, arquivos_texto,
                                  modalidade, data_pub, ano):
    """Valores via regras_valores (docs tipados); fallback na extração antiga."""
    _garantir_path_script()
    from ia_local import extrair_valores_dos_docs

    vals_prio = extrair_valores_dos_docs(cabecalhos)
    auditoria = []

    if vals_prio.get("valor_estimado"):
        try:
            linha["Valor Estimado"] = float(vals_prio["valor_estimado"])
        except ValueError:
            linha["Valor Estimado"] = vals_prio["valor_estimado"]
        meta = vals_prio.get("valor_estimado_meta") or {}
        auditoria.append({
            "campo": "Valor Estimado", "valor": vals_prio["valor_estimado"],
            "origem": "prioritario", "doc": meta.get("doc", ""),
            "rotulo": meta.get("rotulo", ""), "trecho": meta.get("trecho", ""),
            "pts": "", "outros": "",
        })
    if vals_prio.get("valor_homologado"):
        try:
            linha["Valor Homologado"] = float(vals_prio["valor_homologado"])
        except ValueError:
            linha["Valor Homologado"] = vals_prio["valor_homologado"]
        meta = vals_prio.get("valor_homologado_meta") or {}
        auditoria.append({
            "campo": "Valor Homologado", "valor": vals_prio["valor_homologado"],
            "origem": "prioritario", "doc": meta.get("doc", ""),
            "rotulo": meta.get("rotulo", ""), "trecho": meta.get("trecho", ""),
            "pts": "", "outros": "",
        })

    # data de abertura + fallback de valores pela lógica antiga
    vals = extrair_valores_da_licitacao(
        arquivos_texto, modalidade=modalidade,
        data_pub=data_pub if isinstance(data_pub, datetime) else None,
        ano=ano)
    if not linha.get("Valor Estimado") and vals["valor_estimado"] is not None:
        linha["Valor Estimado"] = vals["valor_estimado"]
    if not linha.get("Valor Homologado") and vals["valor_homologado"] is not None:
        linha["Valor Homologado"] = vals["valor_homologado"]
    if vals["data_abertura"]:
        d, mth, y = vals["data_abertura"]
        linha["Data de Abertura"] = datetime(y, mth, d)
    # Evita duplicar Valor Estimado/Homologado se já veio dos docs prioritários
    campos_ja = {a.get("campo") for a in auditoria}
    for item in vals.get("auditoria") or []:
        if item.get("campo") in campos_ja and item.get("campo") in (
            "Valor Estimado", "Valor Homologado",
        ):
            continue
        auditoria.append(item)
    return auditoria


def _item_aud_simples(campo, valor, doc, rotulo, trecho=""):
    return {
        "campo": campo,
        "valor": valor if valor not in ("", None) else None,
        "doc": doc or "—",
        "rotulo": rotulo or "",
        "trecho": trecho or "",
        "pts": "",
        "outros": "",
    }


def _refinar_com_ollama(titulo, linha, cabecalhos, modalidade,
                        modelo, ollama_url, pasta_cache, so_se_faltar):
    """Chama Ollama para CONFIRMAR as informações; nunca quebra o job se offline."""
    import threading

    if so_se_faltar and not _precisa_ia(linha):
        _log("        etapa: IA — pulada (campos já preenchidos)")
        return None

    _garantir_path_script()
    from ia_local import ErroIA, ollama_disponivel, refinar
    from pathlib import Path

    if not ollama_disponivel(ollama_url):
        _log("        ! Ollama offline — seguindo só com regras ({0})",
             ollama_url)
        return None

    def _fmt_data(v):
        if isinstance(v, datetime):
            return v.strftime("%d/%m/%Y")
        return str(v or "").strip()

    try:
        from ia_local.regras_titulo import numero_sem_categoria
        num_bruto = numero_sem_categoria(linha.get("Número") or "")
    except Exception:
        num_bruto = re.sub(r"-([A-Za-z]+)$", "", str(linha.get("Número") or ""))

    leitura_local = {
        "numero": linha.get("Número") or "",
        "numero_bruto": num_bruto,
        "ano": str(linha.get("Ano") or ""),
        "objeto": linha.get("Objeto") or "",
        "situacao": linha.get("Situação da Licitação") or "",
        "data_publicacao": _fmt_data(linha.get("Data de Publicação")),
        "data_abertura": _fmt_data(linha.get("Data de Abertura")),
        "valor_estimado": (
            f"{linha['Valor Estimado']:.2f}"
            if isinstance(linha.get("Valor Estimado"), (int, float))
            else str(linha.get("Valor Estimado") or "")
        ),
        "valor_homologado": (
            f"{linha['Valor Homologado']:.2f}"
            if isinstance(linha.get("Valor Homologado"), (int, float))
            else str(linha.get("Valor Homologado") or "")
        ),
        "modalidade": modalidade,
    }

    stop = threading.Event()

    def _heartbeat():
        t0 = time.time()
        while not stop.wait(15):
            _log("        IA: ainda confirmando... {0}s (aguarde)",
                 int(time.time() - t0))

    th = threading.Thread(target=_heartbeat, daemon=True)
    try:
        _log(
            "        etapa: IA — confirmando informações com Ollama ({0})…",
            modelo,
        )
        th.start()
        out = refinar(
            titulo, leitura_local, cabecalhos,
            provedor="ollama", modelo=modelo, ollama_url=ollama_url,
            pasta_cache=Path(pasta_cache) if pasta_cache else None,
            usar_cache=True,
        )
    except ErroIA as e:
        _log("        ! IA: {0}", e)
        return None
    except Exception as e:
        _log("        ! IA falhou: {0}: {1}", type(e).__name__, e)
        return None
    finally:
        stop.set()

    if out.get("cache"):
        _log("        IA: cache hit (resposta instantânea)")
    mudancas = out.get("mudancas") or []
    if mudancas:
        _log("        IA: " + "; ".join(mudancas[:6]))
    else:
        _log("        IA: confirmou a leitura local (sem mudanças)")

    # funde só o que veio validado + monta auditoria da IA
    aud_ia = []
    origem_ia = out.get("origem") or "ia_local"
    if out.get("cache"):
        origem_ia = "ia_cache"

    def _aud_ia(campo, valor, trecho="", motivo=""):
        aud_ia.append(_item_aud_simples(
            campo, valor,
            doc=origem_ia,
            rotulo=motivo or "confirmação Ollama",
            trecho=trecho,
        ))

    if out.get("numero"):
        ant = linha.get("Número")
        try:
            from ia_local.regras_titulo import numero_pos_confirmacao
            linha["Número"] = numero_pos_confirmacao(
                out["numero"], ant or "", modalidade
            )
        except Exception:
            linha["Número"] = numero_com_sigla_front(out["numero"], modalidade)
        if out.get("ano"):
            linha["Ano"] = out["ano"]
        _aud_ia(
            "Número", linha["Número"],
            trecho=out.get("trecho_numero") or "",
            motivo=(
                "IA alterou (antes: {0})".format(ant or "—")
                if str(ant or "") != str(linha["Número"])
                else "IA confirmou"
            ),
        )
    if out.get("objeto"):
        ant = (linha.get("Objeto") or "")[:60]
        linha["Objeto"] = out["objeto"]
        _aud_ia(
            "Objeto", out["objeto"],
            trecho=out.get("trecho_objeto") or "",
            motivo=(
                "IA alterou (antes: {0})".format(ant or "—")
                if (ant or "") != (out["objeto"] or "")[:60]
                else "IA confirmou"
            ),
        )
    if out.get("situacao"):
        ant = linha.get("Situação da Licitação")
        linha["Situação da Licitação"] = situacao_para_front(out["situacao"])
        _aud_ia(
            "Situação da Licitação", linha["Situação da Licitação"],
            trecho=out.get("motivo_situacao") or "",
            motivo=(
                "IA alterou (antes: {0})".format(ant or "—")
                if ant != linha["Situação da Licitação"]
                else "IA confirmou"
            ),
        )

    def _aplicar_data(campo_linha, campo_ia, chave_trecho):
        raw = (out.get(campo_ia) or "").strip()
        if not raw or raw == "Não informado":
            return
        try:
            dt = datetime.strptime(raw, "%d/%m/%Y")
        except ValueError:
            return
        ant = linha.get(campo_linha)
        linha[campo_linha] = dt
        _aud_ia(
            campo_linha, dt.strftime("%d/%m/%Y"),
            trecho=out.get(chave_trecho) or "",
            motivo=(
                "IA alterou (antes: {0})".format(_fmt_data(ant) or "—")
                if _fmt_data(ant) != dt.strftime("%d/%m/%Y")
                else "IA confirmou"
            ),
        )

    _aplicar_data(
        "Data de Publicação", "data_publicacao", "trecho_data_publicacao"
    )
    _aplicar_data(
        "Data de Abertura", "data_abertura", "trecho_data_abertura"
    )

    for campo_linha, campo_ia, chave_trecho in (
        ("Valor Estimado", "valor_estimado", "trecho_valor_estimado"),
        ("Valor Homologado", "valor_homologado", "trecho_valor_homologado"),
    ):
        v = out.get(campo_ia) or ""
        if not v:
            continue
        ant = linha.get(campo_linha)
        try:
            linha[campo_linha] = float(v)
            valor_aud = float(v)
        except ValueError:
            linha[campo_linha] = v
            valor_aud = v
        _aud_ia(
            campo_linha, valor_aud,
            trecho=out.get(chave_trecho) or "",
            motivo=(
                "IA alterou (antes: {0})".format(
                    ant if ant not in ("", None) else "—"
                )
                if str(ant) != str(valor_aud)
                else "IA confirmou"
            ),
        )
    out["auditoria"] = aud_ia
    return out


# ============================================================================
# MAIN
# ============================================================================
def main():
    ap = argparse.ArgumentParser(
        description="Baixa licitações de portais CR2 e preenche planilha.")
    ap.add_argument(
        "--planilha-fonte",
        default="",
        help="URL Google Sheets ou caminho .xlsx já preenchido. "
             "Usa coluna Documentos (links) para baixar PDFs; "
             "demais colunas viram dados da planilha de saída.",
    )
    ap.add_argument("--listagem", default=LISTAGEM_PADRAO,
                    help="URL da listagem de licitações.")
    ap.add_argument("--saida", default=SAIDA_PADRAO, help="Pasta dos downloads.")
    ap.add_argument("--planilha-modelo", default="Front.xlsx",
                    help="Planilha-modelo com os cabeçalhos (padrão: Front.xlsx).")
    ap.add_argument("--planilha-saida", default="",
                    help="Caminho da planilha gerada. Vazio = "
                         f"{PLANILHA_SAIDA} DENTRO da pasta de downloads (--saida).")
    ap.add_argument("--incluir-subcategorias", action="store_true")
    ap.add_argument("--so-html", action="store_true", help="Ignora a API REST.")
    ap.add_argument("--so-planilha", action="store_true",
                    help="Não rebaixa anexos; usa os já salvos em --saida.")
    ap.add_argument("--sem-extracao", action="store_true",
                    help="Só baixa; não lê documentos nem preenche planilha.")
    ap.add_argument("--ocr", action="store_true",
                    help="Usa OCR nos PDFs sem camada de texto.")
    ap.add_argument("--idioma-ocr", default="por")
    ap.add_argument("--motor-ocr", default="",
                    choices=["", "auto", "tesseract", "easyocr", "paddleocr", "docling", "surya"],
                    help="Motor de OCR: tesseract (padrão/rápido), auto (=tesseract) "
                         "ou paddleocr. easyocr/docling/surya viram tesseract.")
    ap.add_argument("--anos", default="",
                    help="Anos a extrair, separados por vírgula (ex.: 2023,2024). "
                         "Com filtro, para o scanner ao achar ano anterior ao mínimo "
                         "(ex.: 2023 → para no 2022). Vazio = todos.")
    ap.add_argument("--sem-renomear", action="store_true",
                    help="Não renomeia anexos pelo título interno do documento.")
    ap.add_argument("--ignorar-ssl", action="store_true",
                    help="Ignora erros de certificado SSL (portais com cert quebrado).")
    ap.add_argument(
        "--link-pasta-base",
        default="",
        help="URL base (Drive/SharePoint) para montar LinkDaPasta. "
             "Vazio = usa caminho local absoluto da pasta de anexos.",
    )
    ap.add_argument(
        "--refinar-ia", action="store_true",
        help="Confirma com Ollama: número, objeto, situação, datas e valores.",
    )
    ap.add_argument(
        "--modelo-ia", default="llama3.2:3b",
        help="Modelo Ollama (padrão: llama3.2:3b).",
    )
    ap.add_argument(
        "--ollama-url", default="http://127.0.0.1:11434",
        help="URL do Ollama (padrão: http://127.0.0.1:11434).",
    )
    ap.add_argument(
        "--ia-sempre", action="store_true",
        help="Com --refinar-ia, confirma todas as informações mesmo se "
             "as regras já preencheram (número, objeto, situação, datas, valores).",
    )
    ap.add_argument(
        "--limite", type=int, default=0,
        help="Processa no máximo N licitações (0 = todas). Útil para testes.",
    )
    ap.add_argument(
        "--amostra-mensal", action="store_true",
        help="Após coletar, mantém só N licitações por mês "
             "(diversificando modalidades). Ver --amostra-por-mes.",
    )
    ap.add_argument(
        "--amostra-por-mes", type=int, default=5,
        help="Com --amostra-mensal: quantas licitações por mês (padrão: 5).",
    )
    ap.add_argument(
        "--priorizar-docs-leves", action="store_true",
        help="Pula pastas só de contrato/aditivo; prioriza docs úteis "
             "(TR/edital/homologação/DFD) e menos anexos.",
    )
    ap.add_argument(
        "--download-workers", type=int, default=0,
        help="Downloads paralelos por licitação (1–12). "
             "0 = usa DOWNLOAD_WORKERS do ambiente/painel (padrão 4).",
    )
    args = ap.parse_args()
    # Garante dict mutável (runner antigo do painel às vezes setava None).
    _ur = globals().get("ULTIMO_RESULTADO_UPLOAD")
    if not isinstance(_ur, dict):
        globals()["ULTIMO_RESULTADO_UPLOAD"] = {}
    else:
        _ur.clear()

    if not (args.planilha_fonte or "").strip() and not (args.listagem or "").strip():
        ap.error("Informe --listagem ou --planilha-fonte (Google Sheets / xlsx).")

    p = urlparse(args.listagem) if (args.listagem or "").strip() else None
    base = "{0}://{1}".format(p.scheme, p.netloc) if p and p.scheme and p.netloc else ""

    # Filtro de anos efetivo: linha de comando > configuração do topo
    anos_filtro = ([a.strip() for a in args.anos.split(",") if a.strip()]
                   if args.anos.strip() else list(ANOS_FILTRO))
    renomear = RENOMEAR_POR_TITULO and not args.sem_renomear
    global DOWNLOAD_WORKERS
    if getattr(args, "download_workers", 0) and int(args.download_workers) > 0:
        DOWNLOAD_WORKERS = max(1, min(12, int(args.download_workers)))
    else:
        DOWNLOAD_WORKERS = max(1, min(12, int(DOWNLOAD_WORKERS or 4)))

    os.makedirs(args.saida, exist_ok=True)

    # Planilha: por padrão, DENTRO da pasta de downloads (junto dos arquivos)
    if not args.planilha_saida.strip():
        args.planilha_saida = os.path.join(args.saida, PLANILHA_SAIDA)
    sessao = requests.Session()
    sessao.headers.update(HEADERS)
    if args.ignorar_ssl:
        sessao.verify = False
        try:
            import urllib3
            urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)
        except Exception:
            pass

    if args.ocr:
        motor_ocr = args.motor_ocr or MOTOR_OCR
        tess = detectar_backend_ocr()
        try:
            import importlib.util
            paddle_ok = bool(importlib.util.find_spec("paddleocr"))
        except Exception:
            paddle_ok = False
        print(f"  · OCR: motor '{motor_ocr}' | Tesseract: "
              f"{tess if tess != 'nenhum' else 'NÃO instalado'} | PaddleOCR: "
              f"{'instalado' if paddle_ok else 'não instalado'}")
        if not ocr_instalado():
            print("    ! Tesseract não encontrado — OCR será ignorado "
                  "(instale Tesseract-OCR; Auto usa só Tesseract)")
    else:
        motor_ocr = args.motor_ocr or MOTOR_OCR

    print("=" * 66)
    if (args.planilha_fonte or "").strip():
        print("Fonte   : planilha ({0})".format(args.planilha_fonte[:72]))
    else:
        print("Entidade : {0}".format(p.netloc if p else "?"))
        print("Listagem : {0}".format(raiz_categoria(args.listagem)))
    print(f"Downloads: {args.saida}")
    print(f"Planilha : {args.planilha_saida}  (modelo: {args.planilha_modelo})")
    print(f"Anos     : {', '.join(anos_filtro) if anos_filtro else 'todos'}")
    print(f"Renomear : {'pelo título interno dos documentos' if renomear else 'não'}")
    print(f"Paralelo : {DOWNLOAD_WORKERS} download(s) por licitação")
    if getattr(args, "amostra_mensal", False):
        print(
            "Amostra  : até {0} por mês (modalidades diversificadas)".format(
                getattr(args, "amostra_por_mes", 5) or 5
            )
        )
    if getattr(args, "priorizar_docs_leves", False):
        print(
            "Filtro   : docs leves (pula só contrato/aditivo puro; "
            "prioriza TR/edital/homologação/DFD; menos anexos)"
        )
    if args.limite and args.limite > 0:
        print(f"Limite   : {args.limite} licitação(ões)")
    if args.refinar_ia:
        # Sempre confirma todas as informações (não só valores / não só gaps)
        print(
            f"IA       : Ollama / {args.modelo_ia} @ {args.ollama_url}"
            " (confirmando número, objeto, situação, datas e valores)"
        )
    else:
        print("IA       : desligada")
    print("=" * 66 + "\n")

    licitacoes = None
    if (args.planilha_fonte or "").strip():
        print("► Coletando via planilha-fonte (links Documentos)...")
        cache_dir = Path(args.saida) / "_cache"
        licitacoes = coletar_via_planilha_fonte(
            sessao, args.planilha_fonte.strip(), cache_dir, anos_filtro=anos_filtro
        )
    elif not args.so_html:
        print("► Coletando via API REST...")
        slugs = ["licitacoes"] + (SUBCATEGORIAS if args.incluir_subcategorias else [])
        licitacoes = coletar_via_api(sessao, base, slugs, anos_filtro=anos_filtro)
    if licitacoes is None and not (args.planilha_fonte or "").strip():
        print("► Coletando via HTML (varredura de páginas)...")
        listagens = [args.listagem] + (
            [f"{base}/c/licitacoes/{s}/" for s in SUBCATEGORIAS]
            if args.incluir_subcategorias else [])
        licitacoes = coletar_via_html(
            sessao, listagens, anos_filtro=anos_filtro
        )
    nao_migradas_acum = []
    if getattr(args, "priorizar_docs_leves", False):
        antes = len(licitacoes or [])
        selecionadas, rejeitadas = filtrar_licitacoes_docs_leves(licitacoes or [])
        licitacoes = selecionadas
        print(
            "\n► Filtro docs leves: {0} → {1} licitação(ões) "
            "(pula só contrato/aditivo puro; prioriza docs úteis; menos anexos).".format(
                antes, len(licitacoes)
            )
        )
        if rejeitadas:
            from collections import Counter
            motivos = Counter(
                (r.get("_filtro_motivo") or "?") for r in rejeitadas
            )
            print(
                "  · Rejeitadas no filtro: {0} ({1})".format(
                    len(rejeitadas),
                    ", ".join("{0}={1}".format(k, v) for k, v in motivos.most_common()),
                )
            )
            nao_migradas_acum.extend(rejeitadas)
        else:
            print("  · Nenhuma licitação rejeitada pelo filtro de docs.")
    if getattr(args, "amostra_mensal", False):
        antes = len(licitacoes or [])
        por_mes = getattr(args, "amostra_por_mes", 5) or 5
        selecionadas, restantes = amostrar_mensal_diversificada(
            licitacoes or [], por_mes=por_mes
        )
        licitacoes = selecionadas
        print(
            "\n► Amostra mensal: {0} → {1} licitação(ões) "
            "(até {2}/mês, modalidades diversificadas).".format(
                antes, len(licitacoes), por_mes
            )
        )
        if restantes:
            for r in restantes:
                item = dict(r)
                item.setdefault("_filtro_motivo", "fora da amostra mensal")
                nao_migradas_acum.append(item)
            print("  · Fora da amostra: {0}".format(len(restantes)))
        else:
            print("  · Nenhuma licitação restante para controle de não migradas.")
    if nao_migradas_acum:
        planilha_nao = salvar_planilha_nao_migradas(args.saida, nao_migradas_acum)
        print(
            "  · Não migradas (só links): {0} → {1}".format(
                len(nao_migradas_acum), planilha_nao
            )
        )
        ULTIMO_RESULTADO_UPLOAD.update({
            "planilha_nao_migradas": planilha_nao,
            "nao_migradas": len(nao_migradas_acum),
        })
    if args.limite and args.limite > 0:
        licitacoes = licitacoes[: args.limite]
    print(f"\n► {len(licitacoes)} licitação(ões) a processar.\n")

    linhas_planilha, auditoria_geral = [], []
    itens_upload = []  # {linha, pasta, titulo} para subir*.xlsx

    puladas_ano = 0
    cancelado = False
    total_lic = len(licitacoes)
    try:
        for idx, lic in enumerate(licitacoes, 1):
            _abortar_se_cancelado()
            titulo = lic["titulo"]

            modalidade_bruta, numero = split_modalidade_numero(titulo)
            modalidade = modalidade_padrao(titulo)      # nome padronizado
            ano = (
                extrai_ano(numero)
                or ano_do_titulo(titulo)
                or extrai_ano(
                    ((lic.get("dados_planilha") or {}).get("numero") or "")
                )
            )
            objeto = extrai_objeto(titulo)

            # --- FILTRO DE ANOS: número do título E/OU data de publicação.
            # Sem nenhum dos dois → mantém (não perde por falha de parsing).
            if anos_filtro:
                ano_pub_chk = ano_de_data_pub(lic.get("data_pub"))
                anos_set_chk = _anos_filtro_set(anos_filtro)
                ano_min_chk = _ano_minimo_filtro(anos_set_chk)
                acao_ano = decidir_anos_vs_filtro(
                    ano, ano_pub_chk, anos_set_chk, ano_min_chk
                )
                if acao_ano != "pegar":
                    puladas_ano += 1
                    _log(
                        "  · [{0}/{1}] pulada (ano título={2} pub={3} "
                        "fora do filtro)",
                        idx, total_lic, ano or "?", ano_pub_chk or "?",
                    )
                    continue
                if not ano and not ano_pub_chk:
                    _log(
                        "  ! [{0}/{1}] sem ano (título/pub) — mantida: {2}",
                        idx, total_lic, titulo[:55],
                    )

            pasta = os.path.join(args.saida, nome_pasta(titulo))
            os.makedirs(pasta, exist_ok=True)

            barra = _barra(idx - 1, total_lic)
            _log("")
            _log("── [{0}/{1} · {2}%] {3} {4}",
                 idx, total_lic, _pct(idx - 1, total_lic), barra, titulo[:55])
            _log("    etapa: baixar anexos ({0} link(s))...",
                 len(lic.get("anexos") or []))

            arquivos_locais = baixar_anexos_da_licitacao(
                sessao,
                lic.get("anexos") or [],
                pasta,
                so_planilha=bool(args.so_planilha),
            )

            if args.so_planilha:
                for f in os.listdir(pasta):
                    fp = os.path.join(pasta, f)
                    if (os.path.isfile(fp) and fp not in arquivos_locais
                            and not eh_artefato_ocr(f)
                            and os.path.splitext(f)[1].lower() in EXT_DOCS):
                        arquivos_locais.append(fp)

            # Converte a data de publicação ANTES da extração, para que sirva
            # de referência na validação da data de abertura.
            data_pub = lic["data_pub"]
            if isinstance(data_pub, str) and data_pub:
                try:
                    data_pub = datetime.strptime(data_pub, "%d/%m/%Y")
                except Exception:
                    pass

            dp = lic.get("dados_planilha") or {}
            usar_dados_planilha = bool(dp)

            if usar_dados_planilha:
                modalidade = modalidade_padrao(
                    dp.get("modalidade") or modalidade
                ) or (dp.get("modalidade") or modalidade)
                numero = dp.get("numero") or numero
                ano = extrai_ano(numero) or ano
                objeto = dp.get("objeto") or objeto or titulo
                pub_dp = _parse_data_planilha(dp.get("publicacao"))
                if pub_dp:
                    data_pub = pub_dp
                linha = {
                    "Modalidade": modalidade,
                    "Número": numero,
                    "Ano": ano,
                    "Objeto": objeto,
                    "Data de Publicação": data_pub,
                    "Data de Abertura": _parse_data_planilha(dp.get("abertura")),
                    "Valor Estimado": "",
                    "Situação da Licitação": dp.get("situacao") or "Em andamento",
                    "Valor Homologado": dp.get("valor_homologado") or "",
                }
            else:
                linha = {
                    "Modalidade": modalidade,
                    "Número": numero,
                    "Ano": ano,
                    "Objeto": objeto or titulo,
                    "Data de Publicação": data_pub,
                    "Data de Abertura": "",
                    "Valor Estimado": "",
                    "Situação da Licitação": "Em andamento",
                    "Valor Homologado": "",
                }

            if not args.sem_extracao and not usar_dados_planilha:
                _log("    etapa: ler documentos prioritários...")
                arquivos_texto, cabecalhos, nomes_docs = _ler_docs_priorizados(
                    arquivos_locais, args.ocr, args.idioma_ocr, motor_ocr,
                    renomear, pasta,
                )
                _log("    etapa: extrair valores / situação...")
                for item in _aplicar_valores_prioritarios(
                    linha, cabecalhos, arquivos_texto,
                    modalidade, data_pub, ano,
                ):
                    item["licitacao"] = titulo
                    auditoria_geral.append(item)

                sit = situacao_para_front(
                    inferir_situacao(titulo, nomes_docs, modalidade=modalidade)
                )
                linha["Situação da Licitação"] = sit
                auditoria_geral.append({
                    **_item_aud_simples(
                        "Situação da Licitação", sit,
                        doc="título + nomes dos anexos",
                        rotulo="inferir_situacao (regras)",
                        trecho=(titulo or "")[:120],
                    ),
                    "licitacao": titulo,
                })

                # Número com sigla (ex.: 9/2023-007-CMVX-RPPP) — só troca categoria
                num_final = numero_com_sigla_front(
                    linha.get("Número") or numero, modalidade
                )
                linha["Número"] = num_final
                auditoria_geral.append({
                    **_item_aud_simples(
                        "Número", num_final,
                        doc="título da listagem",
                        rotulo="split_modalidade_numero + sigla Front",
                        trecho=(titulo or "")[:120],
                    ),
                    "licitacao": titulo,
                })
                auditoria_geral.append({
                    **_item_aud_simples(
                        "Modalidade", modalidade,
                        doc="título da listagem",
                        rotulo="modalidade_padrao (regras)",
                        trecho=(titulo or "")[:120],
                    ),
                    "licitacao": titulo,
                })
                auditoria_geral.append({
                    **_item_aud_simples(
                        "Objeto", linha.get("Objeto") or "",
                        doc="título da listagem",
                        rotulo="extrai_objeto (texto entre parênteses)",
                        trecho=(titulo or "")[:160],
                    ),
                    "licitacao": titulo,
                })

                est = linha.get("Valor Estimado")
                hom = linha.get("Valor Homologado")
                _log(
                    "    leitura: nº {0} | sit. {1} | est. {2} | hom. {3}",
                    linha.get("Número") or "—",
                    linha.get("Situação da Licitação") or "—",
                    est if est not in ("", None) else "—",
                    hom if hom not in ("", None) else "—",
                )

                if args.refinar_ia:
                    pasta_cache = os.path.join(
                        os.path.dirname(os.path.abspath(__file__)), "cache_ia"
                    )
                    # Sempre confirma (número, objeto, situação, datas, valores)
                    out_ia = _refinar_com_ollama(
                        titulo, linha, cabecalhos, modalidade,
                        modelo=args.modelo_ia,
                        ollama_url=args.ollama_url,
                        pasta_cache=pasta_cache,
                        so_se_faltar=False,
                    )
                    if out_ia:
                        for item in out_ia.get("auditoria") or []:
                            item["licitacao"] = titulo
                            auditoria_geral.append(item)

                # Mesmo Número/Modalidade em preenchida + subir* (+ contratos)
                ant_num = linha.get("Número")
                padronizar_linha_para_todas_planilhas(linha)
                if str(linha.get("Número") or "") != str(ant_num or ""):
                    auditoria_geral.append({
                        **_item_aud_simples(
                            "Número", linha.get("Número"),
                            doc="padronização Front (todas as planilhas)",
                            rotulo="padronizar_linha_para_todas_planilhas",
                            trecho=(titulo or "")[:120],
                        ),
                        "licitacao": titulo,
                    })
                    modalidade = linha.get("Modalidade") or modalidade

            elif not args.sem_extracao and usar_dados_planilha:
                num_final = numero_com_sigla_front(
                    linha.get("Número") or numero, modalidade
                )
                linha["Número"] = num_final
                padronizar_linha_para_todas_planilhas(linha)
                modalidade = linha.get("Modalidade") or modalidade
                _log(
                    "    planilha-fonte: nº {0} | sit. {1} | hom. {2}",
                    linha.get("Número") or "—",
                    linha.get("Situação da Licitação") or "—",
                    linha.get("Valor Homologado") or "—",
                )

            linhas_planilha.append(linha)
            itens_upload.append({
                "linha": dict(linha),
                "pasta": os.path.abspath(pasta),
                "titulo": titulo,
            })
            _log("    ✓ [{0}/{1} · {2}%] concluída {3}",
                 idx, total_lic, _pct(idx, total_lic), _barra(idx, total_lic))
    except Cancelado:
        cancelado = True

    if anos_filtro and puladas_ano:
        print(f"\n  · {puladas_ano} licitação(ões) fora dos anos "
              f"{', '.join(anos_filtro)} — puladas.")

    if not args.sem_extracao and linhas_planilha:
        print(f"\n► Preenchendo planilha ({len(linhas_planilha)} linhas)...")
        preencher_planilha(linhas_planilha, auditoria_geral,
                           args.planilha_modelo, args.planilha_saida)
        print(f"  ✓ {args.planilha_saida}")
        print("  · Aba 'Auditoria': origem de cada campo (documento + trecho)")
        ULTIMO_RESULTADO_UPLOAD.update({
            "planilha_preenchida": os.path.abspath(args.planilha_saida),
            "planilha_auditoria": os.path.abspath(args.planilha_saida),
        })

        # Planilhas oficiais de upload (Front)
        try:
            from gestor_regras import gerar_planilhas_upload
        except ImportError:
            # execução via load_module do painel: pasta do script no path
            _dir_script = os.path.dirname(os.path.abspath(__file__))
            if _dir_script not in sys.path:
                sys.path.insert(0, _dir_script)
            from gestor_regras import gerar_planilhas_upload

        print("\n► Planilhas oficiais (licitação + contratos)...")

        def _ler_texto_contrato(caminho):
            """Leitor para a planilha de contratos — mesmo OCR do restante.

            max_paginas=None de propósito: o texto NATIVO tem que vir inteiro,
            porque a cláusula do valor global costuma ficar depois da página 6.
            O OCR continua limitado (obter_texto corta em OCR_MAX_PAGINAS_PRIOR).
            """
            try:
                texto, _origem = obter_texto(
                    caminho,
                    args.ocr,
                    idioma=args.idioma_ocr,
                    motor=args.motor_ocr,
                    max_paginas=None,
                    max_chars=400000,
                )
                return texto or ""
            except Cancelado:
                raise
            except Exception as e:
                print("    (falha ao ler %s: %s)" % (os.path.basename(caminho), e))
                return ""

        resultado_upload = gerar_planilhas_upload(
            itens_upload,
            args.saida,
            link_pasta_base=(args.link_pasta_base or "").strip(),
            ler_texto=_ler_texto_contrato,
        )
        print(
            "\n  Resumo — Prontas: {0}  |  Pendentes: {1}".format(
                resultado_upload["prontas"],
                resultado_upload["pendentes"],
            )
        )
        print("  ✓ Licitação: {0}".format(resultado_upload["planilha_licitacoes"]))
        print("  ✓ Documentos: {0}".format(resultado_upload["planilha_documentos"]))
        if resultado_upload.get("planilha_contratos"):
            print("  ✓ Contratos: {0}  ({1} linha(s))".format(
                resultado_upload["planilha_contratos"],
                resultado_upload.get("contratos_linhas", 0)))
        if resultado_upload.get("contratos_relatorio"):
            print("  · Origem dos campos: {0}".format(
                resultado_upload["contratos_relatorio"]))
        if resultado_upload.get("contratos_movidos"):
            print("  · Contratos separados: {0} arquivo(s)".format(
                resultado_upload["contratos_movidos"]))
            for msg in (resultado_upload.get("logs_contratos") or [])[:8]:
                print("  · {0}".format(msg))
        if resultado_upload["pendentes"]:
            print("  · Pendentes: {0}".format(resultado_upload["pendentes_relatorio"]))
            for msg in (resultado_upload.get("logs_move") or [])[:10]:
                print("  · Pasta pendente: {0}".format(msg))
            if len(resultado_upload.get("logs_move") or []) > 10:
                print("  · … (+{0} pastas movidas)".format(
                    len(resultado_upload["logs_move"]) - 10))
        for msg in (resultado_upload.get("logs_link") or [])[:5]:
            print("  · {0}".format(msg))
        if len(resultado_upload.get("logs_link") or []) > 5:
            print("  · … (+{0} pastas)".format(
                len(resultado_upload["logs_link"]) - 5))

        # expõe caminhos para o runner do painel
        ULTIMO_RESULTADO_UPLOAD.update(resultado_upload or {})
        ULTIMO_RESULTADO_UPLOAD["planilha_preenchida"] = os.path.abspath(
            args.planilha_saida
        )
        ULTIMO_RESULTADO_UPLOAD["planilha_auditoria"] = (
            ULTIMO_RESULTADO_UPLOAD["planilha_preenchida"]
        )  # aba Auditoria dentro do mesmo arquivo

    print("\n" + "=" * 66)
    print("CANCELADO." if cancelado else "Concluído.")
    if not args.sem_extracao:
        print("  1) Licitação: subirLicitacoes.xlsx + subirDocumentosLicitacoes.xlsx")
        print("  2) Contratos: pasta Contratos/ + subirContratos.xlsx")
        print("  Veja também a aba 'Auditoria' e a pasta PENDENTES/ se houver faltas.")
    print("=" * 66)
    if cancelado:
        raise Cancelado()


if __name__ == "__main__":
    main()
