# -*- coding: utf-8 -*-
"""Gera subirContratos.xlsx a partir dos PDFs de contrato/aditivo.

Roda depois da separação (Contratos/<licitação>/). Fontes de cada linha:

  - contratos  -> Contratos/<licitação>/ (movidos na etapa anterior)
  - aditivos   -> pasta da própria licitação (regra 13: aditivo não é contrato,
                  então o arquivo NÃO é movido — só entra na planilha)
  - fiscal     -> do próprio contrato ou da portaria de designação da pasta
"""

from __future__ import annotations

import os
from pathlib import Path
from typing import Any, Callable

from .campos_contrato import alertas_contrato, falta_para_o_portal, linha_contrato
from .config_front import (
    ABA_CONTRATOS,
    ARQ_MODELO_CONTRATOS,
    CAMPOS_CONTRATO,
    ROTULOS_CONTRATO,
)
from .contratos import (
    eh_arquivo_contrato,
    eh_arquivo_portaria_fiscal,
    nome_pasta_contrato,
)
from .campos_contrato import eh_aditivo

PASTA_CONTRATOS = "Contratos"
ARQ_RELATORIO = "_RELATORIO_CONTRATOS.txt"
EXT_LEGIVEIS = (".pdf",)

_DIR = Path(__file__).resolve().parent.parent
MODELOS_DIR = _DIR / "modelos"


def _ler_texto_padrao(caminho: str) -> str:
    """Leitura nativa (sem OCR) — usada quando o chamador não passa leitor."""
    try:
        import pdfplumber

        with pdfplumber.open(caminho) as pdf:
            return "\n".join((p.extract_text() or "") for p in pdf.pages[:12])
    except Exception:
        return ""


def _arquivos(pasta: str) -> list[str]:
    try:
        nomes = sorted(os.listdir(pasta))
    except OSError:
        return []
    saida = []
    for nome in nomes:
        caminho = os.path.join(pasta, nome)
        if os.path.isfile(caminho) and nome.lower().endswith(EXT_LEGIVEIS):
            saida.append(caminho)
    return saida


def _link_documento(
    caminho: str, sub: str, pasta_saida: str, link_pasta_base: str = ""
) -> str:
    """Valor da coluna `documento`.

    Padrão: caminho RELATIVO à pasta de saída, com barra normal. É o que a
    publicação consome — publicacao-repasses._caminho_arquivo() resolve
    relativo à pasta base (ou acha pelo nome) e faz set_input_files().
    Caminho absoluto não serve: na VPS ele aponta para
    /opt/opto-automacoes/data/users/... e morre quando o ZIP é baixado.
    """
    base = (link_pasta_base or "").strip().rstrip("/")
    if base:
        return "%s/%s/%s" % (base, sub, os.path.basename(caminho))
    try:
        rel = os.path.relpath(os.path.abspath(caminho), os.path.abspath(pasta_saida))
    except ValueError:  # outra unidade de disco no Windows
        return os.path.abspath(caminho)
    if rel.startswith(".."):  # fora da pasta de saída: relativo não ajuda
        return os.path.abspath(caminho)
    return rel.replace(os.sep, "/")


def _achar_aba(wb, preferida: str):
    if preferida in wb.sheetnames:
        return wb[preferida]
    return wb[wb.sheetnames[0]]


def _limpar_dados(ws) -> None:
    if ws.max_row and ws.max_row > 1:
        ws.delete_rows(2, ws.max_row - 1)
    for i, (_chave, rotulo) in enumerate(CAMPOS_CONTRATO, 1):
        if not ws.cell(1, i).value:
            ws.cell(1, i, value=rotulo)


def coletar_linhas_contratos(
    prontas: list[tuple[dict, str, str]],
    pasta_saida: str,
    *,
    ler_texto: Callable[[str], str] | None = None,
    link_pasta_base: str = "",
    log: Callable[[str], None] | None = None,
) -> tuple[list[dict[str, Any]], list[tuple[str, list[str], list[str]]]]:
    """Devolve (linhas_prontas, problemas) lendo contratos e aditivos."""
    leitor = ler_texto or _ler_texto_padrao
    fala = log or (lambda _m: None)

    linhas: list[dict[str, Any]] = []
    problemas: list[tuple[str, list[str], list[str]]] = []

    for lf, pasta_lic, titulo in prontas:
        sub = nome_pasta_contrato(lf)
        dir_contratos = os.path.join(os.path.abspath(pasta_saida), PASTA_CONTRATOS, sub)
        origem = (lf.get("numero") or "").strip()

        candidatos: list[tuple[str, str]] = []  # (caminho, sub_para_link)
        texto_portaria = ""

        for caminho in _arquivos(dir_contratos):
            nome = os.path.basename(caminho)
            if eh_arquivo_portaria_fiscal(nome):
                if not texto_portaria:
                    texto_portaria = leitor(caminho)
                continue
            if eh_arquivo_contrato(nome):
                candidatos.append((caminho, sub))

        # Aditivos ficam na pasta da licitação (regra 13) — entram só na planilha
        for caminho in _arquivos(pasta_lic):
            nome = os.path.basename(caminho)
            if eh_aditivo(nome):
                candidatos.append(
                    (caminho, os.path.basename(os.path.abspath(pasta_lic).rstrip("\\/")))
                )

        if not candidatos:
            continue

        for caminho, sub_link in candidatos:
            nome = os.path.basename(caminho)
            texto = leitor(caminho)
            linha = linha_contrato(
                nome,
                texto,
                licitacao_origem=origem,
                documento=_link_documento(
                    caminho, sub_link, pasta_saida, link_pasta_base
                ),
                texto_portaria=texto_portaria,
            )
            faltas = falta_para_o_portal(linha)
            if faltas:
                problemas.append(("%s — %s" % (titulo, nome), faltas, []))
                fala("  · %s: falta %s" % (nome, "; ".join(faltas)))
                continue
            alertas = alertas_contrato(linha)
            linhas.append(linha)
            if alertas:
                problemas.append(("%s — %s (aceito)" % (titulo, nome), [], alertas))
            fala(
                "  · %s | %s %s | %s"
                % (
                    nome,
                    linha["tipo_contrato"],
                    linha["numero"],
                    linha["nome_razao_social"] or "(sem contratada)",
                )
            )

    return linhas, problemas


def gerar_planilha_contratos(
    prontas: list[tuple[dict, str, str]],
    pasta_saida: str,
    *,
    ler_texto: Callable[[str], str] | None = None,
    link_pasta_base: str = "",
    modelos_dir: str | Path | None = None,
) -> dict[str, Any]:
    """Escreve subirContratos.xlsx + relatório. Não cria arquivo se não houver linha."""
    import openpyxl

    linhas, problemas = coletar_linhas_contratos(
        prontas,
        pasta_saida,
        ler_texto=ler_texto,
        link_pasta_base=link_pasta_base,
        log=lambda m: print(m),
    )

    contratos_dir = os.path.join(os.path.abspath(pasta_saida), PASTA_CONTRATOS)
    resultado: dict[str, Any] = {
        "planilha_contratos": "",
        "contratos_linhas": len(linhas),
        "contratos_problemas": len(problemas),
        "contratos_relatorio": "",
    }
    if not linhas and not problemas:
        return resultado

    os.makedirs(contratos_dir, exist_ok=True)

    if linhas:
        modelos = Path(modelos_dir) if modelos_dir else MODELOS_DIR
        modelo = modelos / ARQ_MODELO_CONTRATOS
        if not modelo.is_file():
            raise FileNotFoundError("Modelo não encontrado: %s" % modelo)
        wb = openpyxl.load_workbook(modelo)
        ws = _achar_aba(wb, ABA_CONTRATOS)
        _limpar_dados(ws)
        for r, linha in enumerate(linhas, start=2):
            for c, (chave, _rotulo) in enumerate(CAMPOS_CONTRATO, start=1):
                ws.cell(row=r, column=c, value=linha.get(chave, ""))
        saida = os.path.join(contratos_dir, ARQ_MODELO_CONTRATOS)
        wb.save(saida)
        resultado["planilha_contratos"] = saida

    rel = os.path.join(contratos_dir, ARQ_RELATORIO)
    with open(rel, "w", encoding="utf-8") as fh:
        fh.write("CONTRATOS — origem dos campos de %s\n" % ARQ_MODELO_CONTRATOS)
        fh.write("=" * 70 + "\n\n")
        fh.write("Linhas na planilha: %d\n" % len(linhas))
        fh.write("Documentos com pendência/alerta: %d\n\n" % len(problemas))
        fh.write("Campos obrigatórios: %s\n\n" % ", ".join(
            ROTULOS_CONTRATO[c] for c, _r in CAMPOS_CONTRATO
            if c in ("ano", "tipo_contrato", "numero", "objeto", "nome_razao_social")
        ))
        if not problemas:
            fh.write("(nenhuma pendência)\n")
        for titulo, faltas, alertas in problemas:
            fh.write("- %s\n" % titulo)
            if faltas:
                fh.write("  FORA da planilha — falta: %s\n" % "; ".join(faltas))
            for a in alertas:
                fh.write("  · %s\n" % a)
            fh.write("\n")
    resultado["contratos_relatorio"] = rel
    return resultado
