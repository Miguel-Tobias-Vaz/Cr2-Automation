# -*- coding: utf-8 -*-
"""Gera subirLicitacoes.xlsx e subirDocumentosLicitacoes.xlsx a partir dos modelos.

Etapa final: separa PDFs de contrato/portaria em Contratos/<licitação>/
(sem preencher planilha de contratos).
"""

from __future__ import annotations

import os
import re
import shutil
from pathlib import Path
from typing import Any

from .config_front import (
    ABA_DOCUMENTOS,
    ABA_LICITACOES,
    ARQ_MODELO_DOCUMENTOS,
    ARQ_MODELO_LICITACOES,
    CAMPOS_FRONT,
)
from .contratos import separar_contratos_da_pasta
from .front import alertas_licitacao, falta_para_o_front, linha_front

PASTA_CONTRATOS = "Contratos"
PASTA_PENDENTES = "PENDENTES"

_DIR = Path(__file__).resolve().parent.parent
MODELOS_DIR = _DIR / "modelos"


def _achar_aba(wb, preferida: str):
    if preferida in wb.sheetnames:
        return wb[preferida]
    return wb[wb.sheetnames[0]]


def _limpar_dados(ws, n_cols: int) -> None:
    if ws.max_row and ws.max_row > 1:
        ws.delete_rows(2, ws.max_row - 1)
    for i, (_chave, rotulo) in enumerate(CAMPOS_FRONT, 1):
        if i <= n_cols and not ws.cell(1, i).value:
            ws.cell(1, i, value=rotulo)


def _link_pasta(pasta_abs: str, link_pasta_base: str = "") -> tuple[str, str]:
    pasta_abs = os.path.abspath(pasta_abs) if pasta_abs else ""
    base = (link_pasta_base or "").strip().rstrip("/")
    if base and pasta_abs:
        nome = os.path.basename(pasta_abs.rstrip("\\/"))
        url = "%s/%s" % (base, nome)
        return url, "URL montada (%s)" % url
    return pasta_abs, "caminho local (%s)" % pasta_abs


def _nome_pasta_seguro(nome: str) -> str:
    nome = (nome or "").strip() or "licitacao"
    nome = re.sub(r'[<>:"/\\|?*\x00-\x1f]', "_", nome)
    nome = re.sub(r"\s+", " ", nome).strip(" .")
    return nome[:120] or "licitacao"


def _destino_unico(destino: str) -> str:
    if not os.path.exists(destino):
        return destino
    base, n = destino, 2
    while os.path.exists("%s_%d" % (base, n)):
        n += 1
    return "%s_%d" % (base, n)


def _mover_para_pendentes(pasta: str, pendentes_dir: str, titulo: str) -> tuple[str, str]:
    """
    Move a pasta da licitação incompleta para PENDENTES/<nome>.
    Retorna (destino, mensagem_log). Se não houver pasta, só registra.
    """
    pasta = os.path.abspath(pasta) if pasta else ""
    if not pasta or not os.path.isdir(pasta):
        return "", "pasta não encontrada — só registrada no relatório"

    pendentes_abs = os.path.abspath(pendentes_dir)
    try:
        if os.path.commonpath([pasta, pendentes_abs]) == pendentes_abs:
            return pasta, "já em PENDENTES (%s)" % pasta
    except ValueError:
        pass

    nome = _nome_pasta_seguro(os.path.basename(pasta.rstrip("\\/")) or titulo)
    destino = _destino_unico(os.path.join(pendentes_dir, nome))
    shutil.move(pasta, destino)
    return destino, "movida para %s" % destino


def registro_de_linha_planilha(linha: dict[str, Any]) -> dict[str, Any]:
    """Converte a linha do script CR2 (chaves em português) para registro interno."""
    return {
        "modalidade": linha.get("Modalidade") or "",
        "numero": linha.get("Número") or linha.get("Numero") or "",
        "ano": str(linha.get("Ano") or "").strip(),
        "objeto": linha.get("Objeto") or "",
        "data_publicacao": linha.get("Data de Publicação") or "",
        "data_abertura": linha.get("Data de Abertura") or "",
        "valor_estimado": linha.get("Valor Estimado") if linha.get("Valor Estimado") != "" else "",
        "situacao": linha.get("Situação da Licitação") or "",
        "valor_homologado": linha.get("Valor Homologado") if linha.get("Valor Homologado") != "" else "",
    }


def _garantir_numero_igual_nas_planilhas(lf: dict[str, Any], linha_bruta: dict[str, Any]) -> dict[str, Any]:
    """
    Garante que o Número do Front preserve o do nome (só troca categoria)
    e fique idêntico ao que foi/será gravado nas outras planilhas.
    """
    from ia_local.regras_titulo import numero_com_sigla

    bruto = (
        (linha_bruta.get("Número") or linha_bruta.get("Numero") or "")
        or (lf.get("numero") or "")
    )
    modalidade = lf.get("modalidade") or linha_bruta.get("Modalidade") or ""
    num = numero_com_sigla(str(bruto), str(modalidade))
    if num:
        lf["numero"] = num
    return lf


def gerar_planilhas_upload(
    itens: list[dict[str, Any]],
    pasta_saida: str,
    *,
    link_pasta_base: str = "",
    modelos_dir: str | Path | None = None,
) -> dict[str, Any]:
    """
    1) Classifica prontas/pendentes → subirLicitacoes + subirDocumentos (+ PENDENTES)
    2) Separa contratos/portarias → Contratos/<licitação>/
    """
    import openpyxl

    modelos = Path(modelos_dir) if modelos_dir else MODELOS_DIR
    os.makedirs(pasta_saida, exist_ok=True)
    pendentes_dir = os.path.join(pasta_saida, PASTA_PENDENTES)
    os.makedirs(pendentes_dir, exist_ok=True)

    prontas: list[tuple[dict, str, str]] = []
    pendentes: list[tuple[str, list[str], list[str], str, str]] = []
    alertas_ok: list[tuple[str, list[str]]] = []
    logs_move: list[str] = []

    print("=" * 66)
    print("  ETAPA 1/2 — LICITAÇÃO")
    print("  Classificar prontas/pendentes e gerar planilhas de upload")
    print("=" * 66)

    for item in itens:
        linha_bruta = item.get("linha") or {}
        pasta = item.get("pasta") or ""
        titulo = item.get("titulo") or str(linha_bruta.get("Número") or "licitacao")
        reg = registro_de_linha_planilha(linha_bruta)
        faltas = falta_para_o_front(reg)
        alerts = alertas_licitacao(reg)
        if faltas:
            destino, log_move = _mover_para_pendentes(pasta, pendentes_dir, titulo)
            logs_move.append("%s -> %s" % (titulo, log_move))
            pendentes.append((titulo, faltas, alerts, destino, log_move))
            continue
        lf = linha_front(reg)
        lf = _garantir_numero_igual_nas_planilhas(lf, linha_bruta)
        prontas.append((lf, pasta, titulo))
        if alerts:
            alertas_ok.append((titulo, alerts))

    print(
        "  · Prontas: {0}  |  Pendentes: {1}".format(len(prontas), len(pendentes))
    )

    modelo_lic = modelos / ARQ_MODELO_LICITACOES
    if not modelo_lic.is_file():
        raise FileNotFoundError("Modelo não encontrado: %s" % modelo_lic)
    wb_lic = openpyxl.load_workbook(modelo_lic)
    ws_lic = _achar_aba(wb_lic, ABA_LICITACOES)
    _limpar_dados(ws_lic, 9)
    for r, (lf, _pasta, _tit) in enumerate(prontas, start=2):
        for c, (chave, _rotulo) in enumerate(CAMPOS_FRONT, start=1):
            ws_lic.cell(row=r, column=c, value=lf.get(chave, ""))
    saida_lic = os.path.join(pasta_saida, ARQ_MODELO_LICITACOES)
    wb_lic.save(saida_lic)
    print("  ✓ {0}".format(saida_lic))

    modelo_doc = modelos / ARQ_MODELO_DOCUMENTOS
    if not modelo_doc.is_file():
        raise FileNotFoundError("Modelo não encontrado: %s" % modelo_doc)
    wb_doc = openpyxl.load_workbook(modelo_doc)
    ws_doc = _achar_aba(wb_doc, ABA_DOCUMENTOS)
    if ws_doc.max_row and ws_doc.max_row > 1:
        ws_doc.delete_rows(2, ws_doc.max_row - 1)
    if not ws_doc.cell(1, 1).value:
        ws_doc.cell(1, 1, value="LinkDaPasta")
        ws_doc.cell(1, 2, value="Modalidade")
        ws_doc.cell(1, 3, value="Numero")

    logs_link: list[str] = []
    for r, (lf, pasta, _tit) in enumerate(prontas, start=2):
        link, como = _link_pasta(pasta, link_pasta_base)
        logs_link.append("LinkDaPasta = %s" % como)
        ws_doc.cell(row=r, column=1, value=link)
        ws_doc.cell(row=r, column=2, value=lf.get("modalidade", ""))
        ws_doc.cell(row=r, column=3, value=lf.get("numero", ""))
    saida_doc = os.path.join(pasta_saida, ARQ_MODELO_DOCUMENTOS)
    wb_doc.save(saida_doc)
    print("  ✓ {0}".format(saida_doc))

    rel = os.path.join(pendentes_dir, "_RELATORIO.txt")
    with open(rel, "w", encoding="utf-8") as fh:
        fh.write("PENDENTES — licitações que NÃO entraram nas planilhas de upload\n")
        fh.write("A pasta de anexos de cada uma foi movida para esta pasta.\n")
        fh.write("=" * 70 + "\n\n")
        if not pendentes:
            fh.write("(nenhuma)\n")
        for titulo, faltas, alerts, destino, log_move in pendentes:
            fh.write("- %s\n" % titulo)
            fh.write("  Falta: %s\n" % "; ".join(faltas))
            if alerts:
                fh.write("  Alertas: %s\n" % "; ".join(alerts))
            fh.write("  Pasta: %s\n" % (destino or log_move))
            fh.write("\n")
        if alertas_ok:
            fh.write("\nALERTAS em linhas ACEITAS (preenchimento por regra)\n")
            fh.write("-" * 70 + "\n")
            for titulo, alerts in alertas_ok:
                fh.write("- %s\n" % titulo)
                for a in alerts:
                    fh.write("  · %s\n" % a)
                fh.write("\n")
    if pendentes:
        print("  · Pendentes: {0} → {1}".format(len(pendentes), rel))

    print("  ✓ ETAPA 1/2 concluída — planilhas de licitação prontas.")

    print("")
    print("=" * 66)
    print("  ETAPA 2/2 — CONTRATOS")
    print("  Separar PDFs → Contratos/<licitação>/")
    print("=" * 66)

    logs_contratos: list[str] = []
    n_contratos = 0
    for lf, pasta, titulo in prontas:
        movidos = separar_contratos_da_pasta(pasta, pasta_saida, lf)
        if movidos:
            n_contratos += len(movidos)
            pasta_ctr = os.path.dirname(movidos[0])
            msg = "%s: %d arquivo(s) → %s" % (titulo, len(movidos), pasta_ctr)
            logs_contratos.append(msg)
            print("  · {0}".format(msg))

    if n_contratos:
        print("  · Separados: {0} arquivo(s) de contrato/portaria.".format(n_contratos))
    else:
        print("  · Nenhum contrato/portaria encontrado nas pastas prontas.")
    print("  ✓ ETAPA 2/2 concluída — contratos separados.")

    contratos_dir = os.path.join(pasta_saida, PASTA_CONTRATOS)
    return {
        "planilha_licitacoes": saida_lic,
        "planilha_documentos": saida_doc,
        "pendentes_relatorio": rel,
        "pasta_contratos": contratos_dir if n_contratos else "",
        "prontas": len(prontas),
        "pendentes": len(pendentes),
        "contratos_movidos": n_contratos,
        "logs_link": logs_link,
        "logs_move": logs_move,
        "logs_contratos": logs_contratos,
        "alertas_ok": alertas_ok,
        "pendentes_itens": pendentes,
    }
