# -*- coding: utf-8 -*-
"""Gera contratos.csv / contratos.xlsx (+ aba Auditoria) a partir de Contratos/."""

from __future__ import annotations

import csv
import os
import re
from typing import Any, Callable

from .config_front import (
    ARQ_CONTRATOS_CSV,
    ARQ_CONTRATOS_XLSX,
    CAMPOS_CONTRATO,
    CHAVES_CONTRATO,
)
from .extrair_contrato import extrair_contrato, normalizar
from .ia_contrato import refinar_contrato_ia

EXT_OK = {".pdf", ".txt"}

# Campos que entram na planilha de auditoria (documento fica sempre vazio)
CAMPOS_AUDITORIA = [k for k, _ in CAMPOS_CONTRATO]


def licitacao_de_nome_pasta(nome: str) -> str:
    """003-2025-RPPE -> 003/2025-RPPE"""
    nome = (nome or "").strip()
    m = re.match(r"^(\d{1,10})-(\d{4})(?:-(.+))?$", nome)
    if not m:
        return nome.replace("-", "/", 1) if "-" in nome else nome
    base = "%s/%s" % (m.group(1), m.group(2))
    if m.group(3):
        return "%s-%s" % (base, m.group(3))
    return base


def _listar_arquivos(pasta: str) -> list[str]:
    try:
        nomes = os.listdir(pasta)
    except OSError:
        return []
    out = []
    for nome in sorted(nomes):
        fp = os.path.join(pasta, nome)
        if not os.path.isfile(fp):
            continue
        if os.path.splitext(nome)[1].lower() not in EXT_OK:
            continue
        out.append(fp)
    return out


def _trecho_no_doc(texto: str, valor: str, raio: int = 50) -> str:
    """Trecho literal ao redor do valor no documento."""
    if not texto or not valor:
        return ""
    n = normalizar(texto)
    alvo = normalizar(valor)
    if len(alvo) < 3:
        return ""
    # tenta valor completo; senão primeiros 40 chars
    for cand in (alvo, alvo[:40], re.sub(r"\D", "", alvo)[:14]):
        if len(cand) < 3:
            continue
        pos = n.find(cand)
        if pos < 0:
            continue
        ini = max(0, pos - raio)
        fim = min(len(n), pos + len(cand) + raio)
        return re.sub(r"\s+", " ", n[ini:fim]).strip()[:160]
    return ""


def _origem_campo(
    chave: str,
    valor: str,
    docs: list[dict[str, str]],
    *,
    pasta_nome: str = "",
) -> tuple[str, str, str]:
    """
    Retorna (documento, método, trecho) para auditoria.
    """
    valor = (valor or "").strip()

    if chave == "documento":
        return "—", "sempre vazia (regra 11 — link Drive)", ""
    if chave == "tipoContrato":
        return "—", "fixo (Contrato)", ""
    if chave == "licitacaoOrigem":
        return pasta_nome or "pasta Contratos/", "origem da licitação (pasta)", ""

    if not valor:
        return "—", "em branco", ""

    # Aguardando informação / CNPJ placeholder (regra 14)
    if valor in ("Aguardando informação", "00.000.000/0000-00"):
        return "—", "regra 14 (sem razão social/CNPJ nos docs)", ""

    melhor_doc = ""
    melhor_trecho = ""
    for d in docs:
        trecho = _trecho_no_doc(d.get("texto") or "", valor)
        if trecho:
            melhor_doc = d.get("nome") or ""
            melhor_trecho = trecho
            break

    if melhor_doc:
        return melhor_doc, "regras (trecho no documento)", melhor_trecho

    nomes = "; ".join(d.get("nome") or "" for d in docs if d.get("nome"))
    return nomes or "—", "regras (sem trecho localizado)", ""


def montar_auditoria_contrato(
    reg: dict[str, Any],
    docs: list[dict[str, str]],
    *,
    antes_ia: dict[str, Any] | None = None,
    usou_ia: bool = False,
    pasta_nome: str = "",
) -> list[dict[str, Any]]:
    """Uma linha de auditoria por campo preenchido."""
    itens: list[dict[str, Any]] = []
    contrato_id = (
        (reg.get("numero") or "").strip()
        or (reg.get("licitacaoOrigem") or "").strip()
        or pasta_nome
        or "?"
    )
    licitacao = (reg.get("licitacaoOrigem") or "").strip() or pasta_nome

    for chave, rotulo in CAMPOS_CONTRATO:
        valor = "" if chave == "documento" else (reg.get(chave) or "")
        doc, metodo, trecho = _origem_campo(
            chave, str(valor), docs, pasta_nome=pasta_nome,
        )

        if usou_ia and antes_ia is not None:
            ant = (antes_ia.get(chave) or "").strip()
            novo = str(valor).strip()
            if chave == "documento":
                pass
            elif ant != novo:
                metodo = "IA (alterou)"
                if ant:
                    metodo = "IA (alterou; antes: %s)" % (ant[:80],)
                # tenta trecho do valor novo
                for d in docs:
                    t2 = _trecho_no_doc(d.get("texto") or "", novo)
                    if t2:
                        doc = d.get("nome") or doc
                        trecho = t2
                        break
            elif novo and usou_ia:
                if metodo.startswith("regras"):
                    metodo = "regras + IA (confirmou)"

        # vigência assumida
        if chave == "dataVigenciaFIM" and reg.get("vigencia_assumida") == "sim":
            metodo = (metodo + "; " if metodo else "") + "vigência assumida (+365 dias)"

        itens.append({
            "licitacao": licitacao,
            "contrato": contrato_id,
            "campo": rotulo,
            "valor": valor if valor != "" else None,
            "doc": doc,
            "rotulo": metodo,
            "trecho": trecho,
            "pts": "IA" if "IA (alterou)" in metodo else ("ok" if valor else "—"),
            "outros": "; ".join(
                d.get("nome") or "" for d in docs if d.get("nome")
            ),
        })
    return itens


def processar_pasta_contrato(
    pasta: str,
    *,
    licitacao_origem: str = "",
    ler_texto: Callable[..., tuple[str, str]] | None = None,
    usar_ocr: bool = False,
    idioma_ocr: str = "por",
    motor_ocr: str = "auto",
    usar_ia: bool = True,
    modelo_ia: str = "llama3.2:3b",
    ollama_url: str = "http://127.0.0.1:11434",
    log: Callable[[str], None] | None = None,
) -> dict[str, Any]:
    """
    Lê TODOS os PDFs da pasta do contrato, extrai por regras e confirma com IA.
    Anexa reg['_auditoria'] com origem de cada campo.
    """
    def _log(msg: str) -> None:
        if log:
            log(msg)

    pasta = os.path.abspath(pasta)
    pasta_nome = os.path.basename(pasta)
    origem = licitacao_origem or licitacao_de_nome_pasta(pasta_nome)
    arquivos = _listar_arquivos(pasta)
    docs: list[dict[str, str]] = []
    textos: list[str] = []

    for fp in arquivos:
        nome = os.path.basename(fp)
        texto, origem_txt = "", "vazio"
        if ler_texto:
            try:
                texto, origem_txt = ler_texto(
                    fp, usar_ocr, idioma_ocr, motor=motor_ocr,
                    max_paginas=None, max_chars=200_000,
                )
            except TypeError:
                texto, origem_txt = ler_texto(fp, usar_ocr, idioma_ocr)
            except Exception as exc:
                _log("  · falha ao ler {0}: {1}".format(nome, exc))
                continue
        elif os.path.splitext(fp)[1].lower() == ".txt":
            try:
                with open(fp, encoding="utf-8", errors="ignore") as fh:
                    texto = fh.read()
                origem_txt = "txt"
            except OSError:
                continue
        if not (texto or "").strip():
            _log("  · sem texto: {0}".format(nome))
            continue
        _log("  · lido {0} ({1}, {2}c)".format(nome, origem_txt, len(texto)))
        docs.append({"nome": nome, "texto": texto})
        textos.append(texto)

    texto_tudo = "\n\n===== DOC =====\n\n".join(textos)
    reg = extrair_contrato(
        texto_tudo,
        licitacao_origem=origem,
        arquivo="; ".join(d["nome"] for d in docs),
    )
    antes_ia = {k: reg.get(k, "") for k in CHAVES_CONTRATO}
    usou_ia = False

    if usar_ia and docs:
        _log("  · IA confirmando com {0} documento(s)…".format(len(docs)))
        reg = refinar_contrato_ia(
            reg, docs, modelo=modelo_ia, ollama_url=ollama_url,
        )
        reg["documento"] = ""
        usou_ia = True

    reg["_auditoria"] = montar_auditoria_contrato(
        reg,
        docs,
        antes_ia=antes_ia if usou_ia else None,
        usou_ia=usou_ia,
        pasta_nome=pasta_nome,
    )
    reg["_docs_lidos"] = [d.get("nome") or "" for d in docs]
    return reg


def varrer_contratos(
    pasta_contratos: str,
    *,
    origem_por_pasta: dict[str, str] | None = None,
    **kwargs: Any,
) -> list[dict[str, Any]]:
    """Percorre Contratos/<subpasta>/ e devolve linhas da planilha."""
    pasta_contratos = os.path.abspath(pasta_contratos)
    origem_por_pasta = origem_por_pasta or {}
    if not os.path.isdir(pasta_contratos):
        return []

    linhas: list[dict[str, Any]] = []
    log = kwargs.get("log")
    for nome in sorted(os.listdir(pasta_contratos)):
        sub = os.path.join(pasta_contratos, nome)
        if not os.path.isdir(sub):
            continue
        if log:
            log("► Contrato pasta: {0}".format(nome))
        origem = origem_por_pasta.get(os.path.abspath(sub)) or origem_por_pasta.get(sub)
        if not origem:
            origem = licitacao_de_nome_pasta(nome)
        reg = processar_pasta_contrato(sub, licitacao_origem=origem, **kwargs)
        linhas.append(reg)
    return linhas


def _escrever_aba_auditoria(wb, auditoria: list[dict[str, Any]]) -> None:
    from openpyxl.styles import Font

    if "Auditoria" in wb.sheetnames:
        del wb["Auditoria"]
    aud = wb.create_sheet("Auditoria")
    cab = [
        "Licitação origem",
        "Contrato",
        "Campo",
        "Valor preenchido",
        "Documento de origem",
        "Método / motivo",
        "Trecho do documento",
        "Status",
        "Docs lidos na pasta",
    ]
    for i, h in enumerate(cab, 1):
        aud.cell(row=1, column=i, value=h).font = Font(
            name="Arial", size=10, bold=True,
        )
    ra = 2
    for item in auditoria:
        aud.cell(row=ra, column=1, value=item.get("licitacao") or "")
        aud.cell(row=ra, column=2, value=item.get("contrato") or "")
        aud.cell(row=ra, column=3, value=item.get("campo") or "")
        cel_v = aud.cell(
            row=ra, column=4,
            value=item["valor"] if item.get("valor") is not None else "(em branco)",
        )
        aud.cell(row=ra, column=5, value=item.get("doc") or "")
        aud.cell(row=ra, column=6, value=item.get("rotulo") or "")
        aud.cell(row=ra, column=7, value=item.get("trecho") or "")
        aud.cell(row=ra, column=8, value=item.get("pts") or "")
        aud.cell(row=ra, column=9, value=item.get("outros") or "")
        for c in range(1, 10):
            aud.cell(row=ra, column=c).font = Font(name="Arial", size=10)
        ra += 1
    for col, w in {
        "A": 18, "B": 14, "C": 18, "D": 28, "E": 32,
        "F": 42, "G": 55, "H": 10, "I": 40,
    }.items():
        aud.column_dimensions[col].width = w


def salvar_planilha_contratos(
    linhas: list[dict[str, Any]],
    pasta_saida: str,
) -> dict[str, str]:
    """Grava contratos.csv e contratos.xlsx (com aba Auditoria). Retorna caminhos."""
    os.makedirs(pasta_saida, exist_ok=True)
    csv_path = os.path.join(pasta_saida, ARQ_CONTRATOS_CSV)
    xlsx_path = os.path.join(pasta_saida, ARQ_CONTRATOS_XLSX)

    auditoria: list[dict[str, Any]] = []
    for lin in linhas:
        auditoria.extend(lin.get("_auditoria") or [])

    with open(csv_path, "w", encoding="utf-8-sig", newline="") as fh:
        w = csv.DictWriter(fh, fieldnames=CHAVES_CONTRATO, extrasaction="ignore")
        w.writeheader()
        for lin in linhas:
            row = {k: (lin.get(k) or "") for k in CHAVES_CONTRATO}
            row["documento"] = ""
            w.writerow(row)

    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font

        wb = Workbook()
        ws = wb.active
        ws.title = "Contratos"
        for c, (_k, rotulo) in enumerate(CAMPOS_CONTRATO, start=1):
            ws.cell(1, c, value=rotulo).font = Font(name="Arial", size=10, bold=True)
        for r, lin in enumerate(linhas, start=2):
            for c, (chave, _rot) in enumerate(CAMPOS_CONTRATO, start=1):
                val = "" if chave == "documento" else (lin.get(chave) or "")
                ws.cell(r, c, value=val).font = Font(name="Arial", size=10)

        _escrever_aba_auditoria(wb, auditoria)
        wb.save(xlsx_path)
    except Exception:
        xlsx_path = ""

    return {
        "csv": csv_path,
        "xlsx": xlsx_path,
        "auditoria_linhas": str(len(auditoria)),
    }
