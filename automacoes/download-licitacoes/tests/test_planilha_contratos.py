# -*- coding: utf-8 -*-
import os
import shutil
import tempfile
import unittest

import openpyxl

from gestor_regras.config_front import ABA_CONTRATOS, CAMPOS_CONTRATO
from gestor_regras.upload_contratos import (
    coletar_linhas_contratos,
    gerar_planilha_contratos,
)
from tests.test_campos_contrato import ADITIVO, CONTRATO, PORTARIA


class BaseContratos(unittest.TestCase):
    """Monta saida/Contratos/009-2025-PE/ + pasta da licitação com aditivo."""

    def setUp(self):
        self.tmp = tempfile.mkdtemp(prefix="ctr_")
        self.saida = os.path.join(self.tmp, "saida")
        self.pasta_lic = os.path.join(self.saida, "PE 009-2025")
        self.dir_ctr = os.path.join(self.saida, "Contratos", "009-2025-PE")
        os.makedirs(self.pasta_lic)
        os.makedirs(self.dir_ctr)

        self.textos = {}
        self._criar(self.dir_ctr, "Contrato 003-2025.pdf", CONTRATO)
        self._criar(self.dir_ctr, "Portaria Fiscal 45-2025.pdf", PORTARIA)
        self._criar(self.pasta_lic, "2 Termo Aditivo 003-2025.pdf", ADITIVO)
        self._criar(self.pasta_lic, "Edital 009-2025.pdf", "edital qualquer")

        self.lf = {"numero": "009/2025-PE", "modalidade": "Pregão Eletrônico"}
        self.prontas = [(self.lf, self.pasta_lic, "PE 009/2025")]

    def tearDown(self):
        shutil.rmtree(self.tmp, ignore_errors=True)

    def _criar(self, pasta, nome, texto):
        caminho = os.path.join(pasta, nome)
        with open(caminho, "wb") as fh:
            fh.write(b"%PDF-1.4 fake")
        self.textos[os.path.abspath(caminho)] = texto

    def ler(self, caminho):
        return self.textos.get(os.path.abspath(caminho), "")


class TestColeta(BaseContratos):
    def test_gera_uma_linha_por_contrato_e_aditivo(self):
        linhas, _problemas = coletar_linhas_contratos(
            self.prontas, self.saida, ler_texto=self.ler
        )
        tipos = sorted(l["tipo_contrato"] for l in linhas)
        self.assertEqual(tipos, ["Aditivo 02", "Contrato"])

    def test_portaria_e_edital_nao_viram_linha(self):
        linhas, _p = coletar_linhas_contratos(
            self.prontas, self.saida, ler_texto=self.ler
        )
        docs = [os.path.basename(l["documento"]) for l in linhas]
        self.assertNotIn("Portaria Fiscal 45-2025.pdf", docs)
        self.assertNotIn("Edital 009-2025.pdf", docs)

    def test_licitacao_origem_vem_do_numero_da_licitacao(self):
        linhas, _p = coletar_linhas_contratos(
            self.prontas, self.saida, ler_texto=self.ler
        )
        self.assertTrue(all(l["licitacao_origem"] == "009/2025-PE" for l in linhas))

    def test_fiscal_do_contrato_prevalece_sobre_a_portaria(self):
        linhas, _p = coletar_linhas_contratos(
            self.prontas, self.saida, ler_texto=self.ler
        )
        contrato = [l for l in linhas if l["tipo_contrato"] == "Contrato"][0]
        self.assertEqual(contrato["fiscal_contrato"], "JOAO CARLOS PEREIRA")

    def test_aditivo_usa_fiscal_da_portaria(self):
        linhas, _p = coletar_linhas_contratos(
            self.prontas, self.saida, ler_texto=self.ler
        )
        aditivo = [l for l in linhas if l["tipo_contrato"] == "Aditivo 02"][0]
        self.assertEqual(aditivo["fiscal_contrato"], "MARIA APARECIDA SOUZA")

    def test_documento_recebe_url_quando_ha_link_base(self):
        linhas, _p = coletar_linhas_contratos(
            self.prontas,
            self.saida,
            ler_texto=self.ler,
            link_pasta_base="https://exemplo.com/arquivos",
        )
        contrato = [l for l in linhas if l["tipo_contrato"] == "Contrato"][0]
        self.assertEqual(
            contrato["documento"],
            "https://exemplo.com/arquivos/009-2025-PE/Contrato 003-2025.pdf",
        )

    def test_documento_e_relativo_a_pasta_de_saida(self):
        # caminho absoluto morre quando o ZIP sai da VPS para o PC do usuário;
        # relativo é o que publicacao-repasses._caminho_arquivo() resolve
        linhas, _p = coletar_linhas_contratos(
            self.prontas, self.saida, ler_texto=self.ler
        )
        contrato = [l for l in linhas if l["tipo_contrato"] == "Contrato"][0]
        doc = contrato["documento"]
        self.assertEqual(doc, "Contratos/009-2025-PE/Contrato 003-2025.pdf")
        self.assertFalse(os.path.isabs(doc))
        self.assertNotIn("\\", doc)
        # e o relativo realmente aponta para o arquivo
        self.assertTrue(os.path.isfile(os.path.join(self.saida, doc)))

    def test_documento_do_aditivo_e_relativo_tambem(self):
        linhas, _p = coletar_linhas_contratos(
            self.prontas, self.saida, ler_texto=self.ler
        )
        aditivo = [l for l in linhas if l["tipo_contrato"] == "Aditivo 02"][0]
        self.assertEqual(
            aditivo["documento"], "PE 009-2025/2 Termo Aditivo 003-2025.pdf"
        )
        self.assertTrue(os.path.isfile(os.path.join(self.saida, aditivo["documento"])))

    def test_documento_sem_campo_obrigatorio_fica_de_fora(self):
        self._criar(self.dir_ctr, "Contrato vazio.pdf", "papel em branco")
        linhas, problemas = coletar_linhas_contratos(
            self.prontas, self.saida, ler_texto=self.ler
        )
        docs = [os.path.basename(l["documento"]) for l in linhas]
        self.assertNotIn("Contrato vazio.pdf", docs)
        self.assertTrue(any("Contrato vazio.pdf" in p[0] for p in problemas))


class TestPlanilha(BaseContratos):
    def test_planilha_tem_cabecalho_do_portal(self):
        res = gerar_planilha_contratos(self.prontas, self.saida, ler_texto=self.ler)
        wb = openpyxl.load_workbook(res["planilha_contratos"])
        ws = wb[ABA_CONTRATOS]
        header = [c.value for c in ws[1]]
        self.assertEqual(header, [rotulo for _chave, rotulo in CAMPOS_CONTRATO])

    def test_planilha_grava_as_linhas_na_ordem_das_colunas(self):
        res = gerar_planilha_contratos(self.prontas, self.saida, ler_texto=self.ler)
        wb = openpyxl.load_workbook(res["planilha_contratos"])
        ws = wb[ABA_CONTRATOS]
        self.assertEqual(ws.max_row, 3)  # cabeçalho + contrato + aditivo
        linha = {
            rotulo: ws.cell(2, i).value
            for i, (_c, rotulo) in enumerate(CAMPOS_CONTRATO, start=1)
        }
        self.assertEqual(linha["ano"], "2025")
        self.assertEqual(linha["numero"], "003/2025")
        self.assertEqual(linha["tipoContrato"], "Contrato")
        self.assertEqual(linha["cpfCnpj"], "12.345.678/0001-99")
        self.assertEqual(linha["valor"], "254300.50")
        self.assertEqual(linha["dataVigenciaIN"], "02/01/2025")
        self.assertEqual(linha["dataVigenciaFIM"], "31/12/2025")

    def test_salva_dentro_da_pasta_contratos(self):
        res = gerar_planilha_contratos(self.prontas, self.saida, ler_texto=self.ler)
        self.assertEqual(
            os.path.dirname(res["planilha_contratos"]),
            os.path.join(self.saida, "Contratos"),
        )
        self.assertEqual(res["contratos_linhas"], 2)

    def test_relatorio_lista_pendencias(self):
        self._criar(self.dir_ctr, "Contrato ruim.pdf", "nada")
        res = gerar_planilha_contratos(self.prontas, self.saida, ler_texto=self.ler)
        with open(res["contratos_relatorio"], encoding="utf-8") as fh:
            texto = fh.read()
        self.assertIn("Contrato ruim.pdf", texto)
        self.assertIn("FORA da planilha", texto)

    def test_sem_contrato_nenhum_nao_cria_planilha(self):
        vazio = os.path.join(self.tmp, "saida2")
        pasta = os.path.join(vazio, "PE 001-2025")
        os.makedirs(pasta)
        res = gerar_planilha_contratos(
            [({"numero": "001/2025-PE"}, pasta, "PE 001/2025")],
            vazio,
            ler_texto=lambda _c: "",
        )
        self.assertEqual(res["planilha_contratos"], "")
        self.assertEqual(res["contratos_linhas"], 0)


if __name__ == "__main__":
    unittest.main()
