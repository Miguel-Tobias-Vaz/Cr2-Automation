# -*- coding: utf-8 -*-
"""Testes básicos da extração de contratos."""
from __future__ import annotations

import sys
import tempfile
import unittest
from pathlib import Path

ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT))

from gestor_regras.extrair_contrato import extrair_contrato  # noqa: E402
from gestor_regras.planilha_contratos import (  # noqa: E402
    licitacao_de_nome_pasta,
    salvar_planilha_contratos,
)


TEXTO = """
CONTRATO ADMINISTRATIVO Nº 055/2026
CONTRATANTE: Prefeitura Municipal de Exemplo, CNPJ 11.111.111/0001-11
CONTRATADA: EMPRESA ALPHA SERVICOS LTDA, inscrita no CNPJ 22.222.222/0001-22,
com sede na Rua X.
OBJETO: Contratação de empresa para prestação de serviços de limpeza urbana.
VALOR GLOBAL: R$ 68.069,00
Vigência de 01/03/2026 a 01/03/2027.
Fiscal do contrato: JOAO DA SILVA SANTOS
"""


class TestContrato(unittest.TestCase):
    def test_extrai_campos(self):
        reg = extrair_contrato(TEXTO, licitacao_origem="003/2025-PE")
        self.assertEqual(reg["licitacaoOrigem"], "003/2025-PE")
        self.assertEqual(reg["numero"], "055/2026")
        self.assertEqual(reg["ano"], "2026")
        self.assertIn("ALPHA", reg["nomeRazaoSocial"].upper())
        self.assertEqual(reg["cpfCnpj"], "22.222.222/0001-22")
        self.assertEqual(reg["valor"], "68069,00")
        self.assertEqual(reg["dataVigenciaIN"], "01/03/2026")
        self.assertEqual(reg["dataVigenciaFIM"], "01/03/2027")
        self.assertIn("JOAO", reg["fiscalContrato"].upper())
        self.assertEqual(reg["documento"], "")
        self.assertNotIn("11.111.111", reg["cpfCnpj"])

    def test_nome_pasta(self):
        self.assertEqual(licitacao_de_nome_pasta("003-2025-RPPE"), "003/2025-RPPE")

    def test_salva_csv(self):
        reg = extrair_contrato(TEXTO, licitacao_origem="003/2025-PE")
        from gestor_regras.planilha_contratos import montar_auditoria_contrato
        reg["_auditoria"] = montar_auditoria_contrato(
            reg,
            [{"nome": "contrato.pdf", "texto": TEXTO}],
            pasta_nome="003-2025-PE",
        )
        with tempfile.TemporaryDirectory() as tmp:
            paths = salvar_planilha_contratos([reg], tmp)
            self.assertTrue(Path(paths["csv"]).is_file())
            self.assertTrue(Path(paths["xlsx"]).is_file())
            txt = Path(paths["csv"]).read_text(encoding="utf-8-sig")
            self.assertIn("licitacaoOrigem", txt)
            self.assertIn("68069,00", txt)
            # aba Auditoria no xlsx
            from openpyxl import load_workbook
            wb = load_workbook(paths["xlsx"])
            self.assertIn("Auditoria", wb.sheetnames)
            self.assertGreater(wb["Auditoria"].max_row, 1)


if __name__ == "__main__":
    unittest.main()
