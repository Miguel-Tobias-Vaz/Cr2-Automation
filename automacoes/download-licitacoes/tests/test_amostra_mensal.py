# -*- coding: utf-8 -*-
"""Amostra: N por mês diversificando modalidades."""

from __future__ import annotations

import importlib.util
from pathlib import Path

ROOT = Path(__file__).resolve().parents[1]
MOD_PATH = ROOT / "script.py"


def _load():
    spec = importlib.util.spec_from_file_location("download_licitacoes_script", MOD_PATH)
    mod = importlib.util.module_from_spec(spec)
    assert spec.loader is not None
    spec.loader.exec_module(mod)
    return mod


def test_amostra_mensal_diversifica_modalidades():
    mod = _load()
    itens = [
        {"titulo": "Pregão Eletrônico Nº 01/2023 (...)", "data_pub": "05/03/2023", "link": "u1"},
        {"titulo": "Pregão Eletrônico Nº 02/2023 (...)", "data_pub": "10/03/2023", "link": "u2"},
        {"titulo": "Pregão Eletrônico Nº 03/2023 (...)", "data_pub": "12/03/2023", "link": "u3"},
        {"titulo": "Dispensa de Licitação Nº 04/2023 (...)", "data_pub": "15/03/2023", "link": "u4"},
        {"titulo": "Inexigibilidade Nº 05/2023 (...)", "data_pub": "18/03/2023", "link": "u5"},
        {"titulo": "Concorrência Nº 06/2023 (...)", "data_pub": "20/03/2023", "link": "u6"},
        {"titulo": "Tomada de Preços Nº 07/2023 (...)", "data_pub": "22/03/2023", "link": "u7"},
        {"titulo": "Pregão Eletrônico Nº 08/2023 (...)", "data_pub": "25/03/2023", "link": "u8"},
        {"titulo": "Pregão Eletrônico Nº 09/2023 (...)", "data_pub": "02/04/2023", "link": "u9"},
        {"titulo": "Dispensa de Licitação Nº 10/2023 (...)", "data_pub": "04/04/2023", "link": "u10"},
    ]
    sel, rest = mod.amostrar_mensal_diversificada(itens, por_mes=5)
    mar = [x for x in sel if (x.get("data_pub") or "").endswith("/03/2023")]
    abr = [x for x in sel if (x.get("data_pub") or "").endswith("/04/2023")]
    assert len(mar) == 5
    assert len(abr) == 2
    assert len(sel) + len(rest) == len(itens)
    mods_mar = {mod.modalidade_padrao(x["titulo"]) for x in mar}
    assert len(mods_mar) >= 4


def test_amostra_mensal_nao_ultrapassa_disponivel():
    mod = _load()
    itens = [
        {"titulo": "Pregão Eletrônico Nº 01/2022 (...)", "data_pub": "01/01/2022", "link": "a"},
        {"titulo": "Dispensa de Licitação Nº 02/2022 (...)", "data_pub": "02/01/2022", "link": "b"},
    ]
    sel, rest = mod.amostrar_mensal_diversificada(itens, por_mes=5)
    assert len(sel) == 2
    assert rest == []


def test_salvar_planilha_nao_migradas(tmp_path):
    mod = _load()
    itens = [
        {
            "titulo": "Pregão Eletrônico Nº 01/2023 (objeto)",
            "data_pub": "10/03/2023",
            "link": "https://exemplo.gov.br/lic/1",
        },
    ]
    path = mod.salvar_planilha_nao_migradas(str(tmp_path), itens)
    assert Path(path).is_file()
    from openpyxl import load_workbook

    wb = load_workbook(path)
    ws = wb.active
    assert ws.cell(1, 1).value == "Link"
    assert ws.cell(2, 1).value == "https://exemplo.gov.br/lic/1"
    assert ws.cell(2, 5).value == "Não migrada"
