# =====================================================================
#  Central CR2 — Estagiários, Terceirizados e Dívida Ativa (Playwright)
# =====================================================================
#
#  Le planilhas Excel (.xlsx) SEPARADAS no Drive (link vazio = tipo desligado):
#    PLANILHA_DRIVE_ESTAGIARIO | PLANILHA_DRIVE_TERCEIRIZADO | PLANILHA_DRIVE_DIVIDA
#  Pode preencher so 1, 2 ou as 3 — so as que tiverem link entram na fila.
#  Cada linha = uma publicacao via Criar Publicacao (1 a 1), igual ao RGF:
#    abrir modal → preencher → Publicar → proxima linha.
#  Coluna Arquivo = link Drive do anexo (opcional).
#
#  Uso:
#    python publicar_estagiario_terceirizado_divida.py
#    python publicar_estagiario_terceirizado_divida.py --test
#    python publicar_estagiario_terceirizado_divida.py --yes
#    python publicar_estagiario_terceirizado_divida.py --gerar-modelo
#
# =====================================================================

import csv
import json
import os
import re
import shutil
import subprocess
import sys
import time
import unicodedata
from pathlib import Path
from urllib.parse import parse_qs, urlparse

try:
    from playwright.sync_api import TimeoutError as PWTimeout
    from playwright.sync_api import sync_playwright
except ImportError:
    PWTimeout = None
    sync_playwright = None

try:
    import openpyxl
except ImportError:
    openpyxl = None

try:
    import urllib.request
    import urllib.error
except ImportError:
    urllib = None


# ---------------------------------------------------------------------
#  Bootstrap Playwright
# ---------------------------------------------------------------------


def _recarregar_playwright():
    global PWTimeout, sync_playwright
    try:
        from playwright.sync_api import TimeoutError as PWTimeout
        from playwright.sync_api import sync_playwright

        return True
    except ImportError:
        PWTimeout = None
        sync_playwright = None
        return False


def _mesmo_python(a, b):
    try:
        return Path(a).resolve() == Path(b).resolve()
    except Exception:
        return False


def _python_venv_do_projeto():
    return Path(__file__).resolve().parent / "venv" / "Scripts" / "python.exe"


def _pip_instalar(python_exe, pacote):
    print("[INFO] Instalando '{}' com: {}".format(pacote, python_exe))
    r = subprocess.run(
        [str(python_exe), "-m", "pip", "install", pacote],
        check=False,
    )
    return r.returncode == 0


def garantir_playwright_pronto():
    if sync_playwright is not None:
        return
    if _pip_instalar(sys.executable, "playwright") and _recarregar_playwright():
        print("[INFO] Playwright OK neste interpretador.")
        return
    venv_py = _python_venv_do_projeto()
    if not venv_py.is_file():
        print("[INFO] Criando venv em: {}".format(venv_py.parent.parent))
        subprocess.run(
            [sys.executable, "-m", "venv", str(venv_py.parent.parent)],
            check=True,
        )
    if venv_py.is_file():
        _pip_instalar(venv_py, "playwright")
        if not _mesmo_python(sys.executable, venv_py):
            print("[INFO] Reiniciando o script com o Python do venv do projeto...")
            os.execv(str(venv_py), [str(venv_py)] + sys.argv)
        if _recarregar_playwright():
            print("[INFO] Playwright OK no venv do projeto.")
            return


def garantir_openpyxl():
    global openpyxl
    if openpyxl is not None:
        return True
    if _pip_instalar(sys.executable, "openpyxl"):
        try:
            import openpyxl as _ox

            openpyxl = _ox
            return True
        except ImportError:
            pass
    return False


# ---------------------------------------------------------------------
#  CONFIG — edite aqui
# ---------------------------------------------------------------------

PASTA_PROJETO = Path(__file__).resolve().parent
PASTA_DATA = PASTA_PROJETO / "data"
PASTA_DOWNLOADS = PASTA_DATA / "downloads"
PASTA_PLANILHAS = PASTA_DATA / "planilhas"
PASTA_SCREENSHOTS = PASTA_PROJETO / "runtime" / "screenshots"

# Links das planilhas no Google Drive (compartilhadas: qualquer pessoa com o link)
# Link vazio = aquele tipo NAO roda. Pode preencher 1, 2 ou as 3.
# Ex.: https://docs.google.com/spreadsheets/d/ID_AQUI/edit?usp=sharing
PLANILHA_DRIVE_ESTAGIARIO = ""
PLANILHA_DRIVE_TERCEIRIZADO = ""
PLANILHA_DRIVE_DIVIDA = ""

# Cache local apos baixar do Drive (nao precisa editar)
PLANILHA_ESTAGIARIO = PASTA_PLANILHAS / "planilha_estagiarios.xlsx"
PLANILHA_TERCEIRIZADO = PASTA_PLANILHAS / "planilha_terceirizados.xlsx"
PLANILHA_DIVIDA = PASTA_PLANILHAS / "planilha_divida_ativa.xlsx"

# URL vazia = fila desligada
URL_PORTAL_ESTAGIARIO = "https://www.portalcr2.com.br/relacao-estagiarios/estagiarios-prefeitura-municipal-de-testes"
URL_PORTAL_TERCEIRIZADO = "https://www.portalcr2.com.br/relacao_prestadores/servicos-terceirizados-prefeitura-municipal-de-testes"
URL_PORTAL_DIVIDA = "https://www.portalcr2.com.br/divida-ativa/divida-prefeitura-municipal-de-testes"

URL_LOGIN = "https://www.portalcr2.com.br/?view=login"
ABRIR_LOGIN_ANTES_DO_PORTAL = True
HEADLESS = False

# Credenciais só pelo painel ou variáveis de ambiente (nunca no código).
PORTAL_USUARIO = os.environ.get("PORTAL_USUARIO", "").strip()
PORTAL_SENHA = os.environ.get("PORTAL_SENHA", "").strip()
PORTAL_LOGIN_BOTAO = "Acessar"
TIMEOUT_LOGIN_CONCLUIDO_S = 60

MODO_TESTE = False
PUBLICAR_DUPLO_BUBBLE = False

# Publicacao = mesmo padrao do RGF: Criar Publicacao → preencher → Publicar (1 a 1)
PAUSA_APOS_ANEXAR_PDF = 0.35
PAUSA_POLL_UPLOAD_UI = 0.2
MAX_TENTATIVAS_POLL_UPLOAD = 14
PAUSA_APOS_CONFIRMAR_UPLOAD = 0.45
PAUSA_APOS_CLICAR_PUBLICAR = 0.55

# Em lote: nao tira screenshot a cada linha (so em erro) — bem mais rapido
SALVAR_SCREENSHOTS_OK = False
# Reusa anexo ja baixado pelo mesmo ID do Drive
CACHE_ANEXOS_DRIVE = True

TIMEOUT_PUBLICAR_HABILITADO_S = 75
TIMEOUT_RESULTADO_PUBLICACAO_S = 40
TIMEOUT_LOADER_TOPO_S = 120
OPERA_EXE = None

MODAL_TITULO_REGEX_ESTAGIARIO = (
    r"Criar.*Rela[cç][aã]o\s+de\s+Estagi[aá]rios"
)
MODAL_TITULO_REGEX_TERCEIRIZADO = (
    r"Criar.*Rela[cç][aã]o\s+de\s+Prestadores\s+Terceirizados"
)
MODAL_TITULO_REGEX_DIVIDA = (
    r"Cadastrar.*D[ií]vida\s+Ativa\s+Municipal"
)

_TEXTO_ERRO_APOS_PUBLICAR_RX = re.compile(
    r"(erro|falha|inv[aá]lid|obrigat[oó]rio|n[aã]o\s+foi\s+poss[ií]vel|"
    r"tente\s+novamente|j[aá]\s+existe|duplicad|\bduplicat)",
    re.I,
)
_TEXTO_SUCESSO_MODAL_RX = re.compile(
    r"(publicad|salvo\s+com\s+sucesso|cadastrad|registrad|enviad\s+com\s+sucesso)",
    re.I,
)

# Nomes aceitos para abas / colunas (sem acento, minúsculo)
ABAS_PLANILHA = {
    "estagiario": ("estagiarios", "estagiario", "estagiários", "estagiário"),
    "terceirizado": (
        "terceirizados",
        "terceirizado",
        "prestadores",
        "prestadores terceirizados",
    ),
    "divida": (
        "divida ativa",
        "dívida ativa",
        "divida",
        "dívida",
        "divida_ativa",
    ),
}

COLUNAS_ESTAGIARIO = {
    "nome": ("nome",),
    "inicio": ("inicio do contrato", "início do contrato", "inicio", "início"),
    "fim": ("fim do contrato", "fim"),
    "lotacao": ("lotacao", "lotação"),
    "situacao": ("situacao", "situação"),
    "arquivo": ("arquivo", "link arquivo", "link drive", "drive", "url arquivo"),
}

COLUNAS_TERCEIRIZADO = {
    "mes_ano": ("mes e ano", "mês e ano", "mes/ano", "mês/ano"),
    "nome": ("nome completo", "nome"),
    "empresa": ("empresa",),
    "funcao": ("funcao", "função"),
    "arquivo": ("arquivo", "link arquivo", "link drive", "drive", "url arquivo"),
}

COLUNAS_DIVIDA = {
    "ano": ("ano",),
    "tipo": ("tipo",),
    "nome": ("nome", "nome / razao social", "nome / razão social", "razao social"),
    "valor": ("valor", "valor (r$)", "valor r$"),
    "link": ("link",),
    "arquivo": ("arquivo", "link arquivo", "link drive", "drive", "url arquivo"),
}


# ---------------------------------------------------------------------
#  Utilitarios
# ---------------------------------------------------------------------


def url_portal_ativa(url):
    return bool((url or "").strip())


def _fold_ascii(s):
    if not s:
        return ""
    nfd = unicodedata.normalize("NFD", str(s))
    return "".join(c for c in nfd if unicodedata.category(c) != "Mn").lower().strip()


def normalizar(texto):
    if texto is None:
        return ""
    texto = str(texto)
    subs = {
        "ç": "c",
        "ã": "a",
        "á": "a",
        "â": "a",
        "à": "a",
        "é": "e",
        "ê": "e",
        "í": "i",
        "ó": "o",
        "ô": "o",
        "õ": "o",
        "ú": "u",
        "ü": "u",
    }
    for orig, dest in subs.items():
        texto = texto.replace(orig, dest)
    return texto


def _safe_print(msg):
    """Evita crash de encoding no console Windows (cp1252)."""
    try:
        print(msg)
    except UnicodeEncodeError:
        enc = getattr(sys.stdout, "encoding", None) or "utf-8"
        print(str(msg).encode(enc, errors="replace").decode(enc, errors="replace"))


def _celula_str(v, modo=None):
    """modo: None | 'data' | 'mes_ano' | 'ano'."""
    if v is None:
        return ""
    # datetime/date vindos do Excel / Google Sheets
    try:
        from datetime import date, datetime

        if isinstance(v, datetime):
            if modo == "mes_ano":
                return "{:02d}/{}".format(v.month, v.year)
            if modo == "ano":
                return str(v.year)
            return "{:02d}/{:02d}/{}".format(v.day, v.month, v.year)
        if isinstance(v, date):
            if modo == "mes_ano":
                return "{:02d}/{}".format(v.month, v.year)
            if modo == "ano":
                return str(v.year)
            return "{:02d}/{:02d}/{}".format(v.day, v.month, v.year)
    except Exception:
        pass
    if isinstance(v, float) and v == int(v):
        return str(int(v))
    s = str(v).strip()
    # "2029-01-01 00:00:00" ou "2029-01-01"
    m = re.match(r"^(\d{4})-(\d{2})-(\d{2})(?:\s|$)", s)
    if m:
        y, mo, d = m.group(1), m.group(2), m.group(3)
        if modo == "mes_ano":
            return "{}/{}".format(mo, y)
        if modo == "ano":
            return y
        return "{}/{}/{}".format(d, mo, y)
    return s


def _mapear_colunas(headers, mapa_esperado):
    """headers: lista de strings da 1a linha. Retorna dict chave->indice."""
    folded = {_fold_ascii(h): i for i, h in enumerate(headers) if h}
    out = {}
    for chave, aliases in mapa_esperado.items():
        for alias in aliases:
            idx = folded.get(_fold_ascii(alias))
            if idx is not None:
                out[chave] = idx
                break
    return out


# ---------------------------------------------------------------------
#  Planilha modelo
# ---------------------------------------------------------------------


def gerar_planilha_modelo(pasta=None):
    """Gera 3 planilhas modelo (uma por tipo)."""
    if not garantir_openpyxl():
        print("[ERRO] openpyxl necessario para gerar o modelo. pip install openpyxl")
        sys.exit(1)
    pasta = Path(pasta) if pasta else PASTA_PLANILHAS
    pasta.mkdir(parents=True, exist_ok=True)

    modelos = [
        (
            pasta / "planilha_estagiarios.xlsx",
            "Estagiarios",
            [
                "Nome",
                "Início do Contrato",
                "Fim do Contrato",
                "Lotação",
                "Situação",
                "Arquivo",
            ],
            [
                "João da Silva",
                "01/01/2024",
                "31/12/2024",
                "Secretaria de Obras",
                "Ativo",
                "https://drive.google.com/file/d/COLE_ID_AQUI/view",
            ],
        ),
        (
            pasta / "planilha_terceirizados.xlsx",
            "Terceirizados",
            ["Mês e ano", "Nome Completo", "Empresa", "Função", "Arquivo"],
            [
                "01/2024",
                "Maria Souza",
                "Cidade Limpa LTDA",
                "Auxiliar de Limpeza",
                "https://drive.google.com/file/d/COLE_ID_AQUI/view",
            ],
        ),
        (
            pasta / "planilha_divida_ativa.xlsx",
            "Divida Ativa",
            ["Ano", "Tipo", "Nome", "Valor", "Link", "Arquivo"],
            [
                "2024",
                "",
                "José da Silva",
                "R$ 1.000,00",
                "www.prefeitura.com.br",
                "https://drive.google.com/file/d/COLE_ID_AQUI/view",
            ],
        ),
    ]

    gerados = []
    for caminho, titulo, headers, exemplo in modelos:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = titulo
        ws.append(headers)
        ws.append(exemplo)
        wb.save(str(caminho))
        gerados.append(caminho)
        print("[OK] Modelo: {}".format(caminho))

    print("     Suba cada arquivo no Drive (ou use Google Sheets) e cole o link no CONFIG.")
    return gerados



_SCHEMA_EXPORT = {
    "estagiario": (
        "Estagiarios",
        [
            "Linha original",
            "Motivo",
            "Nome",
            "Início do Contrato",
            "Fim do Contrato",
            "Lotação",
            "Situação",
            "Arquivo",
        ],
        ("linha", "motivo", "nome", "inicio", "fim", "lotacao", "situacao", "arquivo"),
    ),
    "terceirizado": (
        "Terceirizados",
        [
            "Linha original",
            "Motivo",
            "Mês e ano",
            "Nome Completo",
            "Empresa",
            "Função",
            "Arquivo",
        ],
        ("linha", "motivo", "mes_ano", "nome", "empresa", "funcao", "arquivo"),
    ),
    "divida": (
        "Divida Ativa",
        [
            "Linha original",
            "Motivo",
            "Ano",
            "Tipo",
            "Nome",
            "Valor",
            "Link",
            "Arquivo",
        ],
        ("linha", "motivo", "ano", "tipo", "nome", "valor", "link", "arquivo"),
    ),
}


def entrada_nao_publicada(it, motivo):
    """Monta registro completo para exportar / listar linhas nao publicadas."""
    dados = dict(it or {})
    dados["motivo"] = motivo
    return {
        "kind": dados.get("kind") or "?",
        "linha": dados.get("linha"),
        "nome": dados.get("nome") or "(sem nome)",
        "motivo": motivo,
        "dados": dados,
    }


def gerar_planilha_nao_publicadas(itens, destino=None):
    """
    Gera um .xlsx com as linhas que falharam/foram puladas.
    Uma aba por tipo (Estagiarios / Terceirizados / Divida Ativa),
    com colunas originais + Motivo + Linha original.
    """
    if not itens:
        return None
    if not garantir_openpyxl():
        raise RuntimeError("openpyxl necessario para exportar nao publicadas")
    PASTA_DOWNLOADS.mkdir(parents=True, exist_ok=True)
    ts = time.strftime("%Y%m%d_%H%M%S")
    destino = Path(destino) if destino else (
        PASTA_DOWNLOADS / "linhas_nao_publicadas_{}.xlsx".format(ts)
    )

    por_kind = {"estagiario": [], "terceirizado": [], "divida": []}
    outros = []
    for entry in itens:
        kind = (entry.get("kind") or "").strip()
        dados = dict(entry.get("dados") or {})
        if "motivo" not in dados:
            dados["motivo"] = entry.get("motivo") or ""
        if "linha" not in dados and entry.get("linha") is not None:
            dados["linha"] = entry.get("linha")
        if "nome" not in dados:
            dados["nome"] = entry.get("nome") or ""
        if kind in por_kind:
            por_kind[kind].append(dados)
        else:
            outros.append(dados)

    wb = openpyxl.Workbook()
    # remove aba padrao vazia no fim se criarmos outras
    primeira = True
    for kind, rows in por_kind.items():
        if not rows:
            continue
        titulo, headers, keys = _SCHEMA_EXPORT[kind]
        if primeira:
            ws = wb.active
            ws.title = titulo
            primeira = False
        else:
            ws = wb.create_sheet(titulo)
        ws.append(headers)
        for d in rows:
            ws.append(
                [
                    d.get(k) if d.get(k) is not None else ""
                    for k in keys
                ]
            )

    if outros:
        if primeira:
            ws = wb.active
            ws.title = "Outros"
            primeira = False
        else:
            ws = wb.create_sheet("Outros")
        ws.append(["Linha original", "Motivo", "Nome", "Kind"])
        for d in outros:
            ws.append(
                [
                    d.get("linha") or "",
                    d.get("motivo") or "",
                    d.get("nome") or "",
                    d.get("kind") or "",
                ]
            )

    if primeira:
        # so entradas sem kind reconhecido — ativa ja e Outros acima; se nada:
        ws = wb.active
        ws.title = "Nao_publicadas"
        ws.append(["Linha original", "Motivo", "Nome"])
        for entry in itens:
            ws.append(
                [
                    entry.get("linha") or "",
                    entry.get("motivo") or "",
                    entry.get("nome") or "",
                ]
            )

    wb.save(str(destino))
    print(
        "    Planilha de correcao: {} linha(s) -> {}".format(
            len(itens), destino.name
        )
    )
    return destino.resolve()


# ---------------------------------------------------------------------
#  Leitura da planilha
# ---------------------------------------------------------------------


def _resolver_aba(wb, tipo):
    """Usa aba pelo nome; se a planilha tiver so 1 aba, usa ela."""
    if len(wb.sheetnames) == 1:
        return wb[wb.sheetnames[0]]
    aliases = ABAS_PLANILHA[tipo]
    alias_set = {_fold_ascii(a) for a in aliases}
    for sheet in wb.sheetnames:
        if _fold_ascii(sheet) in alias_set:
            return wb[sheet]
    # fallback: primeira aba
    return wb[wb.sheetnames[0]] if wb.sheetnames else None


def _ler_linhas_xlsx(caminho, tipo, mapa_colunas, max_itens=None):
    """Le planilha. max_itens limita quantas linhas de dados retornar (modo teste)."""
    if not garantir_openpyxl():
        raise RuntimeError("openpyxl nao instalado — pip install openpyxl")
    wb = openpyxl.load_workbook(str(caminho), data_only=True, read_only=True)
    try:
        ws = _resolver_aba(wb, tipo)
        if ws is None:
            print(
                "[AVISO] Aba '{}' nao encontrada em {}. Abas: {}".format(
                    tipo, caminho.name, ", ".join(wb.sheetnames)
                )
            )
            return []

        rows_iter = ws.iter_rows(values_only=True)
        try:
            header_row = next(rows_iter)
        except StopIteration:
            return []
        headers = [_celula_str(c) for c in header_row]
        cols = _mapear_colunas(headers, mapa_colunas)
        faltando = [k for k in mapa_colunas if k not in cols]
        obrigatorios_header = {
            "estagiario": ("nome", "inicio", "fim", "situacao"),
            "terceirizado": ("mes_ano", "nome", "empresa", "funcao"),
            "divida": ("ano", "nome", "valor"),
        }
        for k in obrigatorios_header.get(tipo, ()):
            if k not in cols:
                print(
                    "[ERRO] Coluna obrigatoria ausente na aba '{}': {}. Cabecalhos: {}".format(
                        ws.title, k, headers
                    )
                )
                return []

        itens = []
        for n_linha, row in enumerate(rows_iter, start=2):
            if row is None or all(c is None or str(c).strip() == "" for c in row):
                continue

            def get(chave, modo=None, _row=row):
                idx = cols.get(chave)
                if idx is None or idx >= len(_row):
                    return ""
                return _celula_str(_row[idx], modo=modo)

            item = {"kind": tipo, "linha": n_linha, "aba": ws.title}
            for chave in mapa_colunas:
                modo = None
                if chave in ("inicio", "fim"):
                    modo = "data"
                elif chave == "mes_ano":
                    modo = "mes_ano"
                elif chave == "ano":
                    modo = "ano"
                item[chave] = get(chave, modo=modo)
            itens.append(item)
            if max_itens is not None and len(itens) >= int(max_itens):
                break

        if faltando:
            print(
                "[INFO] Colunas opcionais ausentes na aba '{}': {}".format(
                    ws.title, ", ".join(faltando)
                )
            )
        return itens
    finally:
        try:
            wb.close()
        except Exception:
            pass


def ler_fila_estagiario(caminho, retornar_puladas=False, max_ok=None):
    # Em modo teste, le um pouco a mais para achar a 1a valida
    max_scan = None if max_ok is None else max(int(max_ok) * 30, 50)
    itens = _ler_linhas_xlsx(
        caminho, "estagiario", COLUNAS_ESTAGIARIO, max_itens=max_scan
    )
    validos = []
    puladas = []
    for it in itens:
        faltam = [
            c
            for c, v in (
                ("Nome", it["nome"]),
                ("Início do Contrato", it["inicio"]),
                ("Fim do Contrato", it["fim"]),
                ("Situação", it["situacao"]),
            )
            if not v
        ]
        if faltam:
            motivo = "Campos obrigatorios em branco: {}".format(", ".join(faltam))
            print(
                "[AVISO] Estagiario linha {}: {} — pulando.".format(it["linha"], motivo)
            )
            puladas.append(entrada_nao_publicada(it, motivo))
            continue
        validos.append(it)
        if max_ok is not None and len(validos) >= int(max_ok):
            break
    if retornar_puladas:
        return validos, puladas
    return validos


def ler_fila_terceirizado(caminho, retornar_puladas=False, max_ok=None):
    max_scan = None if max_ok is None else max(int(max_ok) * 30, 50)
    itens = _ler_linhas_xlsx(
        caminho, "terceirizado", COLUNAS_TERCEIRIZADO, max_itens=max_scan
    )
    validos = []
    puladas = []
    for it in itens:
        faltam = [
            c
            for c, v in (
                ("Mês e ano", it["mes_ano"]),
                ("Nome Completo", it["nome"]),
                ("Empresa", it["empresa"]),
                ("Função", it["funcao"]),
            )
            if not v
        ]
        if faltam:
            motivo = "Campos obrigatorios em branco: {}".format(", ".join(faltam))
            print(
                "[AVISO] Terceirizado linha {}: {} — pulando.".format(
                    it["linha"], motivo
                )
            )
            puladas.append(entrada_nao_publicada(it, motivo))
            continue
        validos.append(it)
        if max_ok is not None and len(validos) >= int(max_ok):
            break
    if retornar_puladas:
        return validos, puladas
    return validos


def ler_fila_divida(caminho, retornar_puladas=False, max_ok=None):
    max_scan = None if max_ok is None else max(int(max_ok) * 30, 50)
    itens = _ler_linhas_xlsx(caminho, "divida", COLUNAS_DIVIDA, max_itens=max_scan)
    validos = []
    puladas = []
    for it in itens:
        faltam = [
            c
            for c, v in (
                ("Ano", it["ano"]),
                ("Nome", it["nome"]),
                ("Valor", it["valor"]),
            )
            if not v
        ]
        if faltam:
            motivo = "Campos obrigatorios em branco: {}".format(", ".join(faltam))
            print(
                "[AVISO] Divida Ativa linha {}: {} — pulando.".format(
                    it["linha"], motivo
                )
            )
            puladas.append(entrada_nao_publicada(it, motivo))
            continue
        validos.append(it)
        if max_ok is not None and len(validos) >= int(max_ok):
            break
    if retornar_puladas:
        return validos, puladas
    return validos


# ---------------------------------------------------------------------
#  Download Google Drive
# ---------------------------------------------------------------------


def extrair_id_drive(url):
    if not url:
        return None
    url = url.strip()
    if not url or url.upper().startswith("COLE_"):
        return None
    m = re.search(r"/spreadsheets/d/([a-zA-Z0-9_-]+)", url)
    if m:
        return m.group(1)
    m = re.search(r"/file/d/([a-zA-Z0-9_-]+)", url)
    if m:
        return m.group(1)
    m = re.search(r"[?&]id=([a-zA-Z0-9_-]+)", url)
    if m:
        return m.group(1)
    m = re.search(r"/open\?id=([a-zA-Z0-9_-]+)", url)
    if m:
        return m.group(1)
    # ID direto
    if re.fullmatch(r"[a-zA-Z0-9_-]{20,}", url):
        return url
    return None


def _eh_google_sheets(url):
    return "docs.google.com/spreadsheets" in (url or "").lower()


def baixar_planilha_drive(url_drive, destino=None):
    """
    Baixa a planilha do Drive/Sheets e salva como .xlsx local.
    Retorna Path do arquivo ou None.
    """
    url_drive = (url_drive or "").strip()
    if not url_drive:
        return None

    destino = Path(destino) if destino else PLANILHA_ESTAGIARIO
    destino.parent.mkdir(parents=True, exist_ok=True)
    file_id = extrair_id_drive(url_drive)
    if not file_id:
        print("[ERRO] Nao foi possivel extrair o ID do link da planilha.")
        return None

    print("[INFO] Baixando planilha do Drive...")

    # Google Sheets -> exportar XLSX direto
    if _eh_google_sheets(url_drive):
        export_url = (
            "https://docs.google.com/spreadsheets/d/{}/export?format=xlsx".format(
                file_id
            )
        )
        try:
            req = urllib.request.Request(
                export_url,
                headers={
                    "User-Agent": (
                        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36"
                    )
                },
            )
            with urllib.request.urlopen(req, timeout=120) as resp:
                data = resp.read()
            if data[:2] != b"PK":
                print(
                    "[ERRO] Download da planilha nao parece um Excel. "
                    "Confira se o link esta compartilhado (qualquer pessoa com o link)."
                )
                return None
            destino.write_bytes(data)
            print("[OK] Planilha salva em: {}".format(destino))
            return destino.resolve()
        except Exception as e:
            print("[ERRO] Falha ao exportar Google Sheets: {}".format(str(e)[:160]))
            return None

    # Arquivo .xlsx/.xls no Drive
    baixado = baixar_arquivo_drive(url_drive, destino.parent, nome_hint="planilha")
    if baixado is None:
        return None
    if baixado.suffix.lower() not in (".xlsx", ".xlsm", ".xls"):
        novo = destino if destino.suffix.lower() == ".xlsx" else destino.with_suffix(".xlsx")
        try:
            shutil.copy2(baixado, novo)
            print("[OK] Planilha copiada para: {}".format(novo))
            return novo.resolve()
        except Exception as e:
            print("[ERRO] Nao foi possivel copiar planilha: {}".format(e))
            return None
    if baixado.resolve() != destino.resolve():
        try:
            shutil.copy2(baixado, destino)
            print("[OK] Planilha salva em: {}".format(destino))
            return destino.resolve()
        except Exception:
            return baixado.resolve()
    return baixado.resolve()


def _nome_arquivo_content_disposition(header):
    if not header:
        return None
    m = re.search(r"filename\*=UTF-8''([^;]+)", header, re.I)
    if m:
        from urllib.parse import unquote

        return unquote(m.group(1)).strip().strip('"')
    m = re.search(r'filename="?([^";]+)"?', header, re.I)
    if m:
        return m.group(1).strip()
    return None


def baixar_arquivo_drive(link_ou_id, pasta_destino, nome_hint="arquivo"):
    """
    Baixa arquivo do Google Drive. Retorna Path local ou None se link vazio.
    Aceita URL completa ou ID.
    """
    file_id = extrair_id_drive(link_ou_id)
    if not file_id:
        if (link_ou_id or "").strip():
            # Caminho local?
            p = Path(link_ou_id.strip())
            if p.is_file():
                return p.resolve()
            print(
                "    [AVISO] Link Drive invalido (nao e URL nem caminho local): {}".format(
                    str(link_ou_id)[:80]
                )
            )
        return None

    pasta_destino = Path(pasta_destino)
    pasta_destino.mkdir(parents=True, exist_ok=True)

    # Cache: mesmo ID do Drive = mesmo arquivo (evita re-baixar 20s por linha)
    if CACHE_ANEXOS_DRIVE:
        for existente in pasta_destino.glob("*{}*".format(file_id[:12])):
            try:
                if existente.is_file() and existente.stat().st_size > 80:
                    print("    Anexo em cache: {}".format(existente.name))
                    return existente.resolve()
            except Exception:
                continue
        pasta_id = pasta_destino / "by_id"
        pasta_id.mkdir(parents=True, exist_ok=True)
        for existente in pasta_id.glob("{}.*".format(file_id)):
            try:
                if existente.is_file() and existente.stat().st_size > 80:
                    print("    Anexo em cache: {}".format(existente.name))
                    return existente.resolve()
            except Exception:
                continue

    url = "https://drive.google.com/uc?export=download&id={}".format(file_id)
    print("    Baixando Drive id={}...".format(file_id[:12]))

    try:
        import http.cookiejar

        jar = http.cookiejar.CookieJar()
        opener = urllib.request.build_opener(urllib.request.HTTPCookieProcessor(jar))
        req = urllib.request.Request(
            url,
            headers={
                "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36"
            },
        )
        with opener.open(req, timeout=120) as resp:
            content_type = (resp.headers.get("Content-Type") or "").lower()
            data = resp.read()
            cd = resp.headers.get("Content-Disposition")
            nome = _nome_arquivo_content_disposition(cd)

        # Aviso de virus scan (arquivos grandes)
        if b"confirm=" in data[:8000] or (
            "text/html" in content_type and b"download_warning" in data
        ):
            m = re.search(rb"confirm=([0-9A-Za-z_]+)", data)
            if not m:
                m = re.search(rb'name="confirm"\s+value="([^"]+)"', data)
            token = m.group(1).decode("ascii", "ignore") if m else "t"
            url2 = (
                "https://drive.google.com/uc?export=download&confirm={}&id={}".format(
                    token, file_id
                )
            )
            req2 = urllib.request.Request(
                url2,
                headers={
                    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36"
                },
            )
            with opener.open(req2, timeout=180) as resp2:
                data = resp2.read()
                cd = resp2.headers.get("Content-Disposition")
                nome = _nome_arquivo_content_disposition(cd) or nome

        if not nome:
            ext = ".pdf"
            if data[:4] == b"%PDF":
                ext = ".pdf"
            elif data[:2] == b"PK":
                ext = ".zip"
            nome = "{}_{}{}".format(normalizar(nome_hint)[:40], file_id[:8], ext)

        # Sanitiza nome (ASCII seguro para Windows)
        nome = normalizar(nome)
        nome = re.sub(r'[<>:"/\\|?*\x00-\x1f]', "_", nome)
        nome = "".join(c if ord(c) < 128 else "_" for c in nome).strip(" ._") or (
            "arquivo_{}".format(file_id[:8])
        )
        if "." not in nome:
            if data[:4] == b"%PDF":
                nome += ".pdf"
            elif data[:2] == b"PK":
                nome += ".xlsx"
        # Salva tambem por ID para cache estavel entre linhas
        pasta_id = pasta_destino / "by_id"
        pasta_id.mkdir(parents=True, exist_ok=True)
        ext = Path(nome).suffix or ".bin"
        destino_id = pasta_id / "{}{}".format(file_id, ext)
        if destino_id.is_file() and destino_id.stat().st_size > 80:
            print("    Anexo em cache: {}".format(destino_id.name))
            return destino_id.resolve()
        destino_id.write_bytes(data)
        if destino_id.stat().st_size < 80:
            _safe_print(
                "    [AVISO] Arquivo muito pequeno — Drive pode exigir permissao publica."
            )
        _safe_print("    Arquivo salvo: {}".format(destino_id.name))
        return destino_id.resolve()
    except Exception as e:
        _safe_print("    [ERRO] Falha ao baixar Drive: {}".format(repr(e)[:160]))
        return None


# ---------------------------------------------------------------------
#  Opera / navegador
# ---------------------------------------------------------------------


def _opera_via_program_files():
    local = os.environ.get("LOCALAPPDATA", "")
    prog = os.environ.get("PROGRAMFILES", "")
    prog_x86 = os.environ.get("PROGRAMFILES(X86)", "")
    candidatos = []
    base_local = Path(local) / "Programs"
    if base_local.is_dir():
        for folder in sorted(base_local.glob("Opera*")):
            if folder.is_dir():
                exe = folder / "opera.exe"
                if exe.is_file():
                    candidatos.append(exe.resolve())
    candidatos.extend(
        [
            Path(local) / "Programs" / "Opera" / "opera.exe",
            Path(local) / "Programs" / "Opera GX" / "opera.exe",
            Path(prog) / "Opera" / "opera.exe",
            Path(prog) / "Opera GX" / "opera.exe",
            Path(prog_x86) / "Opera" / "opera.exe",
            Path(prog_x86) / "Opera GX" / "opera.exe",
        ]
    )
    visto = set()
    for p in candidatos:
        try:
            if p.is_file():
                r = p.resolve()
                if r not in visto:
                    visto.add(r)
                    return r
        except Exception:
            continue
    return None


def _opera_via_registro_windows():
    if sys.platform != "win32":
        return None
    try:
        import winreg
    except ImportError:
        return None
    roots = (
        (winreg.HKEY_LOCAL_MACHINE, r"SOFTWARE\Microsoft\Windows\CurrentVersion\Uninstall"),
        (
            winreg.HKEY_LOCAL_MACHINE,
            r"SOFTWARE\WOW6432Node\Microsoft\Windows\CurrentVersion\Uninstall",
        ),
        (winreg.HKEY_CURRENT_USER, r"SOFTWARE\Microsoft\Windows\CurrentVersion\Uninstall"),
    )
    for hive, path in roots:
        try:
            key = winreg.OpenKey(hive, path)
        except OSError:
            continue
        i = 0
        while True:
            try:
                subname = winreg.EnumKey(key, i)
            except OSError:
                break
            i += 1
            try:
                sk = winreg.OpenKey(key, subname)
            except OSError:
                continue
            try:
                try:
                    disp = winreg.QueryValueEx(sk, "DisplayName")[0]
                except OSError:
                    continue
                if "Opera" not in str(disp):
                    continue
                try:
                    loc = winreg.QueryValueEx(sk, "InstallLocation")[0]
                    exe = Path(str(loc).strip().strip('"')) / "opera.exe"
                    if exe.is_file():
                        return exe.resolve()
                except OSError:
                    pass
            finally:
                try:
                    winreg.CloseKey(sk)
                except Exception:
                    pass
        try:
            winreg.CloseKey(key)
        except Exception:
            pass
    return None


def resolver_caminho_opera():
    if OPERA_EXE:
        p = Path(OPERA_EXE).expanduser()
        if p.is_file():
            return p.resolve()
        print("[ERRO] OPERA_EXE nao encontrado: {}".format(p))
        return None
    via_path = shutil.which("opera.exe")
    if via_path:
        p = Path(via_path)
        if p.is_file():
            return p.resolve()
    p = _opera_via_program_files()
    if p is not None:
        return p
    return _opera_via_registro_windows()


def salvar_screenshot(page, nome, forcar=False):
    if not forcar and not SALVAR_SCREENSHOTS_OK:
        # Em lote so salva se o nome indicar erro/timeout
        low = (nome or "").lower()
        if not (low.startswith("erro") or low.startswith("timeout")):
            return
    try:
        PASTA_SCREENSHOTS.mkdir(parents=True, exist_ok=True)
        caminho = PASTA_SCREENSHOTS / "{}.png".format(nome)
        page.screenshot(path=str(caminho), full_page=True)
    except Exception:
        pass


def preencher_campo(page, locator, valor):
    valor = str(valor)
    locator.click()
    time.sleep(0.08)
    page.keyboard.press("Control+a")
    time.sleep(0.05)
    page.keyboard.press("Delete")
    time.sleep(0.05)
    locator.fill(valor)
    time.sleep(0.06)
    # Bubble precisa de eventos input/change + blur
    try:
        locator.evaluate(
            """(el, v) => {
                el.focus();
                el.value = v;
                el.dispatchEvent(new Event('input', { bubbles: true }));
                el.dispatchEvent(new Event('change', { bubbles: true }));
                el.dispatchEvent(new Event('blur', { bubbles: true }));
            }""",
            valor,
        )
    except Exception:
        pass
    try:
        page.keyboard.press("Tab")
    except Exception:
        pass
    time.sleep(0.08)


def preencher_campo_rapido(page, locator, valor):
    try:
        locator.focus(timeout=4000)
    except Exception:
        locator.click(timeout=6000)
    locator.fill(str(valor), timeout=8000)


def fazer_upload(page, arquivo_path, modal_root=None):
    try:
        page.evaluate(
            """
            () => {
                document.querySelectorAll('input[type=file]').forEach(function (el) {
                    el.style.display = 'block';
                    el.style.opacity = '1';
                    el.style.visibility = 'visible';
                    el.style.position = 'fixed';
                    el.style.top = '0';
                    el.style.left = '0';
                    el.style.zIndex = '99999';
                });
            }
            """
        )
        time.sleep(0.1)
        if modal_root is not None:
            input_file = modal_root.locator("input[type=file]").first
        else:
            input_file = page.locator("input[type=file]").first
        input_file.wait_for(state="attached", timeout=5000)
        input_file.set_input_files(str(arquivo_path))
        time.sleep(PAUSA_APOS_ANEXAR_PDF)
        return True
    except Exception as e:
        _safe_print("    [Upload] Falhou: {}".format(str(e)[:80]))
        return False
    finally:
        # Remove overlay do input file — senao bloqueia o botao Publicar
        try:
            page.evaluate(
                """
                () => {
                    document.querySelectorAll('input[type=file]').forEach(function (el) {
                        el.style.display = 'none';
                        el.style.opacity = '0';
                        el.style.visibility = 'hidden';
                        el.style.position = '';
                        el.style.top = '';
                        el.style.left = '';
                        el.style.zIndex = '';
                        el.style.pointerEvents = 'none';
                    });
                }
                """
            )
        except Exception:
            pass


def confirmar_upload_ui(page, arquivo_path, modal_root=None):
    nome_lower = Path(arquivo_path).name.lower()
    print("    Aguardando confirmacao do upload...")
    for _ in range(MAX_TENTATIVAS_POLL_UPLOAD):
        try:
            if modal_root is not None:
                area = modal_root.locator(".file-input-text").first
            else:
                area = page.locator(".file-input-text").first
            txt = area.inner_text().strip()
            tl = txt.lower()
            if nome_lower and nome_lower in tl:
                print("    Upload confirmado (arquivo na UI).")
                return
            if len(txt) > 5 and "clique aqui" not in tl:
                print("    Upload confirmado: '{}'".format(txt[:50]))
                return
        except Exception:
            pass
        time.sleep(PAUSA_POLL_UPLOAD_UI)
    time.sleep(PAUSA_APOS_CONFIRMAR_UPLOAD)


# ---------------------------------------------------------------------
#  Modal Bubble
# ---------------------------------------------------------------------


def _regex_modal_titulo(modal_kind):
    if modal_kind == "estagiario":
        return MODAL_TITULO_REGEX_ESTAGIARIO
    if modal_kind == "terceirizado":
        return MODAL_TITULO_REGEX_TERCEIRIZADO
    if modal_kind == "divida":
        return MODAL_TITULO_REGEX_DIVIDA
    return MODAL_TITULO_REGEX_ESTAGIARIO


def _loc_modal_titulo(page, modal_kind):
    return page.locator("text=/{}/i".format(_regex_modal_titulo(modal_kind))).first


def abrir_modal(page, modal_kind):
    """Clica estritamente em Criar Publicacao (nunca Upload/Importar Planilha)."""
    rx = re.compile(r"^\s*Criar\s+Publica[cç][aã]o\s*$", re.I)
    candidatos = (
        page.get_by_role("button", name=rx),
        page.locator("button").filter(has_text=rx),
        page.locator("div.bubble-element.Button").filter(has_text=rx),
        page.locator("[role='button']").filter(has_text=rx),
    )
    clicou = False
    last_err = None
    for loc in candidatos:
        try:
            btn = loc.first
            btn.wait_for(state="visible", timeout=8000)
            texto = (btn.inner_text(timeout=1500) or "").strip()
            if re.search(r"planilha|upload|importar|exportar", texto, re.I):
                continue
            if not re.search(r"Criar\s+Publica", texto, re.I):
                continue
            btn.scroll_into_view_if_needed()
            time.sleep(0.06)
            try:
                btn.click(timeout=5000)
            except Exception:
                btn.click(force=True, timeout=5000)
            print("    Clique: Criar Publicacao ({})".format(texto[:40]))
            clicou = True
            break
        except Exception as e:
            last_err = e
            continue
    if not clicou:
        salvar_screenshot(page, "ERRO_BOTAO_CRIAR_PUBLICACAO")
        raise RuntimeError(
            "Botao 'Criar Publicacao' nao encontrado "
            "(evitado Upload/Importar Planilha). {}".format(
                str(last_err)[:100] if last_err else ""
            )
        )
    time.sleep(0.26)
    _loc_modal_titulo(page, modal_kind).wait_for(state="visible", timeout=10000)
    time.sleep(0.11)


def _limpar_marcador_modal_cr2(page):
    try:
        page.evaluate(
            """
            () => {
                document.querySelectorAll('[data-cr2-pub-modal-marker]').forEach(function (el) {
                    el.removeAttribute('data-cr2-pub-modal-marker');
                });
            }
            """
        )
    except Exception:
        pass


def _modal_bubble_publicacao(page, modal_kind):
    """Localiza o Group Bubble do modal pelo titulo + botao Publicar."""
    _limpar_marcador_modal_cr2(page)
    try:
        marcado = page.evaluate(
            """
            (kind) => {
                document.querySelectorAll('[data-cr2-pub-modal-marker]').forEach(function (el) {
                    el.removeAttribute('data-cr2-pub-modal-marker');
                });
                function texto(btn) {
                    return ((btn.innerText || btn.textContent || '') + '').trim();
                }
                var pubs = Array.from(
                    document.querySelectorAll('button, div[role="button"], .bubble-element.Button')
                ).filter(function (b) {
                    var t = texto(b);
                    return (t === 'Publicar' || t.indexOf('Publicar') >= 0) && t.length < 48;
                });
                for (var i = pubs.length - 1; i >= 0; i--) {
                    var node = pubs[i];
                    var depth = 0;
                    while (node && depth < 28) {
                        depth++;
                        node = node.parentElement;
                        if (!node || !node.querySelectorAll) continue;
                        var txt = (node.innerText || '').slice(0, 6000);
                        var files = node.querySelectorAll('input[type=file]');
                        if (kind === 'estagiario') {
                            if (txt.indexOf('Estagi') < 0 && txt.indexOf('estagi') < 0) continue;
                            if (txt.indexOf('Publicar') < 0) continue;
                            node.setAttribute('data-cr2-pub-modal-marker', '1');
                            return true;
                        }
                        if (kind === 'terceirizado') {
                            if (txt.indexOf('Terceiriz') < 0 &&
                                txt.indexOf('Prestadores') < 0) continue;
                            if (txt.indexOf('Publicar') < 0) continue;
                            node.setAttribute('data-cr2-pub-modal-marker', '1');
                            return true;
                        }
                        if (kind === 'divida') {
                            if (txt.indexOf('Dívida') < 0 &&
                                txt.indexOf('Divida') < 0 &&
                                txt.indexOf('dívida') < 0) continue;
                            if (txt.indexOf('Publicar') < 0) continue;
                            node.setAttribute('data-cr2-pub-modal-marker', '1');
                            return true;
                        }
                    }
                }
                return false;
            }
            """,
            modal_kind,
        )
        if marcado:
            root = page.locator('[data-cr2-pub-modal-marker="1"]').first
            root.wait_for(state="visible", timeout=12000)
            return root
    except Exception:
        pass

    try:
        root = (
            page.locator("div.bubble-element.Group")
            .filter(has=page.locator("button:has-text('Fechar')"))
            .filter(has=page.locator("button:has-text('Publicar')"))
            .first
        )
        root.wait_for(state="visible", timeout=12000)
        return root
    except Exception:
        return None


def fechar_modal(page, modal_kind):
    try:
        modal_root = _modal_bubble_publicacao(page, modal_kind)
        if modal_root is not None:
            fechar = modal_root.locator("button:has-text('Fechar')").first
        else:
            fechar = page.locator("button:has-text('Fechar')").first
        fechar.wait_for(state="visible", timeout=5000)
        fechar.click()
        time.sleep(0.25)
        _loc_modal_titulo(page, modal_kind).wait_for(state="hidden", timeout=8000)
        time.sleep(0.11)
    except Exception:
        try:
            page.keyboard.press("Escape")
            time.sleep(0.22)
        except Exception:
            pass


def _fill_by_label_candidates(scope, labels, valor, page):
    for lb in labels:
        loc = scope.get_by_label(lb, exact=False).first
        try:
            loc.wait_for(state="visible", timeout=2500)
            preencher_campo(page, loc, valor)
            return True
        except Exception:
            continue
    return False


def _listar_inputs_texto_modal(modal_root, page):
    """Inputs de texto editaveis do modal (exclui file/checkbox/hidden)."""
    scope = modal_root if modal_root is not None else page
    locs = scope.locator(
        "input.bubble-element.Input, input[type='input'], "
        "input[type='text'], input:not([type])"
    )
    out = []
    try:
        n = locs.count()
    except Exception:
        return out
    for i in range(n):
        loc = locs.nth(i)
        try:
            tipo = (loc.get_attribute("type") or "").lower()
            if tipo in ("file", "hidden", "checkbox", "radio", "password", "submit"):
                continue
            if not loc.is_visible():
                continue
            ph = loc.get_attribute("placeholder") or ""
            out.append((loc, ph))
        except Exception:
            continue
    return out


def _preencher_por_placeholder(scope, page, valor, *ph_keywords):
    """Tenta achar input cujo placeholder contenha alguma keyword."""
    for kw in ph_keywords:
        try:
            loc = scope.locator(
                "input[placeholder*='{}']".format(kw.replace("'", ""))
            ).first
            loc.wait_for(state="visible", timeout=2500)
            preencher_campo(page, loc, valor)
            return True
        except Exception:
            pass
        try:
            loc = scope.get_by_placeholder(re.compile(re.escape(kw), re.I)).first
            loc.wait_for(state="visible", timeout=2000)
            preencher_campo(page, loc, valor)
            return True
        except Exception:
            pass
    return False


def _preencher_input_por_rotulo(page, modal_root, labels, valor, log_nome, placeholders=None):
    if not valor and valor != 0:
        return False
    valor = str(valor)
    scope = modal_root if modal_root is not None else page
    if _fill_by_label_candidates(scope, labels, valor, page):
        _safe_print("    {} preenchido (label).".format(log_nome))
        return True
    if placeholders and _preencher_por_placeholder(scope, page, valor, *placeholders):
        _safe_print("    {} preenchido (placeholder).".format(log_nome))
        return True
    for lb in labels:
        try:
            # XPath sem depender de acento exato
            trecho = lb.split()[0] if lb else lb
            loc = scope.locator(
                "xpath=(//*[contains(translate(normalize-space(.),"
                "'ÁÀÂÃÄÉÈÊËÍÌÎÏÓÒÔÕÖÚÙÛÜÇáàâãäéèêëíìîïóòôõöúùûüç',"
                "'AAAAAEEEEIIIIOOOOOUUUUCaaaaaeeeeiiiiooooouuuuc'),"
                "'{}')])[1]"
                "/following::input[not(@type='file') and not(@type='hidden')"
                " and not(@type='checkbox')][1]".format(_fold_ascii(trecho).upper())
            ).first
            loc.wait_for(state="visible", timeout=3000)
            preencher_campo(page, loc, valor)
            _safe_print("    {} preenchido (xpath).".format(log_nome))
            return True
        except Exception:
            continue
    for lb in labels:
        try:
            loc = scope.locator(
                "xpath=(//*[contains(normalize-space(.),'{}')])[1]"
                "/following::input[not(@type='file') and not(@type='hidden')"
                " and not(@type='checkbox')][1]".format(lb)
            ).first
            loc.wait_for(state="visible", timeout=2500)
            preencher_campo(page, loc, valor)
            _safe_print("    {} preenchido (xpath literal).".format(log_nome))
            return True
        except Exception:
            continue
    _safe_print("    [AVISO] Campo {} nao encontrado.".format(log_nome))
    return False


def _preencher_inputs_por_ordem(page, modal_root, valores, log_prefix="campo"):
    """
    Preenche inputs de texto do modal na ordem visual.
    valores: lista de strings (None/'' = pular indice).
    """
    inputs = _listar_inputs_texto_modal(modal_root, page)
    if not inputs:
        _safe_print("    [AVISO] Nenhum input de texto no modal.")
        return False
    ok_algum = False
    for i, valor in enumerate(valores):
        if valor is None or valor == "":
            continue
        if i >= len(inputs):
            break
        loc, ph = inputs[i]
        try:
            preencher_campo(page, loc, str(valor))
            _safe_print(
                "    {}[{}] preenchido (ordem, ph={!r}).".format(
                    log_prefix, i, (ph or "")[:40]
                )
            )
            ok_algum = True
            time.sleep(0.05)
        except Exception as e:
            _safe_print(
                "    [AVISO] Falha ordem[{}]: {}".format(i, repr(e)[:80])
            )
    return ok_algum


def _selecionar_dropdown(page, modal_root, labels, valor, log_nome):
    if not valor:
        return False
    valor = str(valor)
    scope = modal_root if modal_root is not None else page

    def _apos_selecionar():
        try:
            page.keyboard.press("Tab")
        except Exception:
            pass
        time.sleep(0.12)

    # select nativo
    if modal_root is not None:
        try:
            selects = modal_root.locator("select")
            n = selects.count()
            for i in range(n):
                sel = selects.nth(i)
                try:
                    sel.wait_for(state="attached", timeout=2500)
                    sel.select_option(label=valor, timeout=3500)
                    try:
                        sel.evaluate(
                            """(el) => {
                                el.dispatchEvent(new Event('input', { bubbles: true }));
                                el.dispatchEvent(new Event('change', { bubbles: true }));
                            }"""
                        )
                    except Exception:
                        pass
                    _safe_print("    {} (select): {}".format(log_nome, valor))
                    _apos_selecionar()
                    return True
                except Exception:
                    try:
                        sel.select_option(
                            label=re.compile(re.escape(valor), re.I),
                            timeout=2500,
                        )
                        _safe_print(
                            "    {} (select regex): {}".format(log_nome, valor)
                        )
                        _apos_selecionar()
                        return True
                    except Exception:
                        continue
        except Exception:
            pass

    # Clique no dropdown visual Bubble (texto "Selecione" perto do rotulo)
    for lb in labels:
        try:
            bloco = scope.locator(
                "xpath=//*[contains(normalize-space(.),'{}')]"
                "/following::*[contains(.,'Selecione') or self::select][1]".format(lb)
            ).first
            bloco.click(timeout=3000)
            time.sleep(0.25)
            opt = page.get_by_text(valor, exact=True).last
            opt.wait_for(state="visible", timeout=4000)
            opt.click()
            _safe_print("    {} (clique lista): {}".format(log_nome, valor))
            _apos_selecionar()
            return True
        except Exception:
            continue

    for lb in labels:
        try:
            loc = scope.get_by_label(lb, exact=False).first
            loc.wait_for(state="visible", timeout=3000)
            tag = loc.evaluate("el => el.tagName")
            if tag == "SELECT":
                loc.select_option(label=valor)
                _safe_print("    {} (label select): {}".format(log_nome, valor))
                _apos_selecionar()
                return True
            loc.click()
            time.sleep(0.2)
            opt = page.get_by_text(valor, exact=True).last
            opt.wait_for(state="visible", timeout=3500)
            opt.click()
            _safe_print("    {} (lista): {}".format(log_nome, valor))
            _apos_selecionar()
            return True
        except Exception:
            continue
    _safe_print(
        "    [AVISO] Nao foi possivel selecionar {}: {!r}".format(log_nome, valor)
    )
    return False


# ---------------------------------------------------------------------
#  Preencher modais
# ---------------------------------------------------------------------


def preencher_modal_estagiario(page, item, arquivo_path):
    modal_root = _modal_bubble_publicacao(page, "estagiario")
    # Ordem tipica do modal: Nome, Inicio, Fim, Lotacao
    ok = _preencher_inputs_por_ordem(
        page,
        modal_root,
        [
            item["nome"],
            item["inicio"],
            item["fim"],
            item.get("lotacao") or "",
        ],
        log_prefix="Estagiario",
    )
    if not ok:
        ok = True
        ok &= _preencher_input_por_rotulo(
            page,
            modal_root,
            ("Nome",),
            item["nome"],
            "Nome",
            placeholders=("João", "Joao", "Silva"),
        )
        ok &= _preencher_input_por_rotulo(
            page,
            modal_root,
            ("Início do Contrato", "Inicio do Contrato"),
            item["inicio"],
            "Inicio do Contrato",
            placeholders=("01/01", "01/01/2024"),
        )
        ok &= _preencher_input_por_rotulo(
            page,
            modal_root,
            ("Fim do Contrato",),
            item["fim"],
            "Fim do Contrato",
            placeholders=("01/01", "01/01/2024"),
        )
        if item.get("lotacao"):
            _preencher_input_por_rotulo(
                page,
                modal_root,
                ("Lotação", "Lotacao"),
                item["lotacao"],
                "Lotacao",
                placeholders=("Secretaria", "Obras"),
            )
    _selecionar_dropdown(
        page,
        modal_root,
        ("Situação", "Situacao"),
        item["situacao"],
        "Situacao",
    )
    if arquivo_path is not None:
        if not fazer_upload(page, arquivo_path, modal_root=modal_root):
            raise Exception("Upload falhou")
        confirmar_upload_ui(page, arquivo_path, modal_root)
    else:
        _safe_print("    [INFO] Sem arquivo — publicando so com os campos.")
    if not ok:
        raise Exception("Campos obrigatorios de Estagiario nao preenchidos")
    time.sleep(PAUSA_APOS_CONFIRMAR_UPLOAD)


def preencher_modal_terceirizado(page, item, arquivo_path):
    modal_root = _modal_bubble_publicacao(page, "terceirizado")
    ok = _preencher_inputs_por_ordem(
        page,
        modal_root,
        [
            item["mes_ano"],
            item["nome"],
            item["empresa"],
            item["funcao"],
        ],
        log_prefix="Terceirizado",
    )
    if not ok:
        ok = True
        ok &= _preencher_input_por_rotulo(
            page,
            modal_root,
            ("Mês e ano", "Mes e ano", "Mês e Ano"),
            item["mes_ano"],
            "Mes e ano",
            placeholders=("01/2024", "01/20"),
        )
        ok &= _preencher_input_por_rotulo(
            page,
            modal_root,
            ("Nome Completo", "Nome"),
            item["nome"],
            "Nome Completo",
            placeholders=("João", "Joao", "Silva"),
        )
        ok &= _preencher_input_por_rotulo(
            page,
            modal_root,
            ("Empresa",),
            item["empresa"],
            "Empresa",
            placeholders=("Cidade Limpa", "LTDA"),
        )
        ok &= _preencher_input_por_rotulo(
            page,
            modal_root,
            ("Função", "Funcao"),
            item["funcao"],
            "Funcao",
            placeholders=("Auxiliar", "Limpeza"),
        )
    if arquivo_path is not None:
        if not fazer_upload(page, arquivo_path, modal_root=modal_root):
            raise Exception("Upload falhou")
        confirmar_upload_ui(page, arquivo_path, modal_root)
    else:
        _safe_print("    [INFO] Sem arquivo — publicando so com os campos.")
    if not ok:
        raise Exception("Campos obrigatorios de Terceirizado nao preenchidos")
    time.sleep(PAUSA_APOS_CONFIRMAR_UPLOAD)


def preencher_modal_divida(page, item, arquivo_path):
    modal_root = _modal_bubble_publicacao(page, "divida")
    # Ordem tipica: Ano, (Tipo=select), Nome, Valor, Link
    ok = _preencher_inputs_por_ordem(
        page,
        modal_root,
        [
            item["ano"],
            item["nome"],
            item["valor"],
            item.get("link") or "",
        ],
        log_prefix="Divida",
    )
    if not ok:
        ok = True
        ok &= _preencher_input_por_rotulo(
            page,
            modal_root,
            ("Ano",),
            item["ano"],
            "Ano",
            placeholders=("2024", "Ex.: 202"),
        )
        ok &= _preencher_input_por_rotulo(
            page,
            modal_root,
            ("Nome / Razão Social", "Nome / Razao Social", "Nome"),
            item["nome"],
            "Nome",
            placeholders=("José", "Jose", "Silva"),
        )
        ok &= _preencher_input_por_rotulo(
            page,
            modal_root,
            ("Valor (R$)", "Valor"),
            item["valor"],
            "Valor",
            placeholders=("R$", "1.000"),
        )
        if item.get("link"):
            _preencher_input_por_rotulo(
                page,
                modal_root,
                ("Link",),
                item["link"],
                "Link",
                placeholders=("www.", "prefeitura"),
            )
    if item.get("tipo"):
        _selecionar_dropdown(page, modal_root, ("Tipo",), item["tipo"], "Tipo")
    if arquivo_path is not None:
        if not fazer_upload(page, arquivo_path, modal_root=modal_root):
            raise Exception("Upload falhou")
        confirmar_upload_ui(page, arquivo_path, modal_root)
    else:
        _safe_print("    [INFO] Sem arquivo — publicando so com os campos.")
    if not ok:
        raise Exception("Campos obrigatorios de Divida Ativa nao preenchidos")
    time.sleep(PAUSA_APOS_CONFIRMAR_UPLOAD)


# ---------------------------------------------------------------------
#  Publicar
# ---------------------------------------------------------------------


def _inner_text_modal_publicacao(page, modal_kind, modal_root):
    if modal_root is not None:
        try:
            return modal_root.inner_text(timeout=2500)
        except Exception:
            pass
    try:
        alt = _modal_bubble_publicacao(page, modal_kind)
        if alt is not None:
            return alt.inner_text(timeout=2500)
    except Exception:
        pass
    try:
        return page.locator("body").inner_text(timeout=2000)
    except Exception:
        return ""


def aguardar_resultado_apos_publicar(page, modal_kind, modal_root):
    titulo_loc = _loc_modal_titulo(page, modal_kind)
    fim = time.monotonic() + TIMEOUT_RESULTADO_PUBLICACAO_S
    ultimo = ""
    while time.monotonic() < fim:
        try:
            visivel = titulo_loc.is_visible(timeout=400)
        except Exception:
            visivel = False
        if not visivel:
            time.sleep(0.28)
            try:
                if titulo_loc.is_visible(timeout=400):
                    continue
            except Exception:
                pass
            print("    Modal fechou — publicacao aceita.")
            return
        ultimo = _inner_text_modal_publicacao(page, modal_kind, modal_root)
        if _TEXTO_ERRO_APOS_PUBLICAR_RX.search(ultimo):
            raise RuntimeError(
                "Resposta no modal apos Publicar: {}".format(
                    ultimo.replace("\n", " ").strip()[:260]
                )
            )
        if _TEXTO_SUCESSO_MODAL_RX.search(ultimo):
            print("    Mensagem de sucesso detectada.")
            return
        time.sleep(0.42)
    salvar_screenshot(page, "TIMEOUT_APOS_PUBLICAR_{}".format(modal_kind))
    raise TimeoutError(
        "Modal ainda aberto apos Publicar (~{}s). Ultimo texto: {}".format(
            TIMEOUT_RESULTADO_PUBLICACAO_S,
            ultimo.replace("\n", " ").strip()[:200],
        )
    )


def _botao_publicar_habilitado(btn):
    """Bubble costuma 'desabilitar' por CSS/classe, sem disabled nativo."""
    try:
        if not btn.is_enabled():
            return False
    except Exception:
        return False
    try:
        info = btn.evaluate(
            """(el) => {
                var s = window.getComputedStyle(el);
                var cls = (el.className || '').toString().toLowerCase();
                var aria = (el.getAttribute('aria-disabled') || '').toLowerCase();
                var op = parseFloat(s.opacity || '1');
                var pe = s.pointerEvents;
                return {
                    aria: aria,
                    opacity: op,
                    pointerEvents: pe,
                    cls: cls,
                    disabledAttr: !!el.disabled
                };
            }"""
        )
        if info.get("disabledAttr"):
            return False
        if info.get("aria") in ("true", "1"):
            return False
        if info.get("opacity", 1) < 0.55:
            return False
        if info.get("pointerEvents") == "none":
            return False
        cls = info.get("cls") or ""
        if "disabled" in cls or "not-allowed" in cls or "inactive" in cls:
            return False
        return True
    except Exception:
        return True


def clicar_publicar(page, modal_kind):
    modal_root = _modal_bubble_publicacao(page, modal_kind)

    def _achar_botao():
        candidatos = []
        root = modal_root if modal_root is not None else page
        for sel in (
            "button:has-text('Publicar')",
            "div.bubble-element.Button:has-text('Publicar')",
            "[role='button']:has-text('Publicar')",
        ):
            try:
                loc = root.locator(sel).first
                if loc.count() > 0:
                    candidatos.append(loc)
            except Exception:
                continue
        return candidatos[0] if candidatos else page.locator(
            "button:has-text('Publicar')"
        ).first

    publicar_btn = _achar_botao()
    publicar_btn.wait_for(state="visible", timeout=15000)
    publicar_btn.scroll_into_view_if_needed()

    limite = time.monotonic() + TIMEOUT_PUBLICAR_HABILITADO_S
    while time.monotonic() < limite:
        try:
            if _botao_publicar_habilitado(publicar_btn):
                break
        except Exception:
            pass
        time.sleep(0.35)
    else:
        salvar_screenshot(page, "TIMEOUT_PUBLICAR_DESABILITADO_{}".format(modal_kind))
        raise TimeoutError(
            "Botao Publicar desabilitado > {}s — formulario incompleto?".format(
                TIMEOUT_PUBLICAR_HABILITADO_S
            )
        )

    aguardar_barra_carregamento_topo(
        page, etiqueta="antes de Publicar ({})".format(modal_kind)
    )
    time.sleep(0.6)

    clicou = False
    # 1) clique Playwright
    try:
        publicar_btn.click(timeout=15000)
        clicou = True
        _safe_print("    Clique Publicar (playwright).")
    except Exception as e:
        _safe_print("    [AVISO] Clique playwright falhou: {}".format(repr(e)[:80]))

    # 2) force
    if not clicou:
        try:
            publicar_btn.click(force=True, timeout=10000)
            clicou = True
            _safe_print("    Clique Publicar (force).")
        except Exception:
            pass

    # 3) JS click no elemento do Bubble
    try:
        page.evaluate(
            """() => {
                var nodes = Array.from(document.querySelectorAll(
                    'button, div.bubble-element.Button, [role="button"]'
                ));
                for (var i = nodes.length - 1; i >= 0; i--) {
                    var t = ((nodes[i].innerText || '') + '').trim();
                    if (t === 'Publicar' || (t.indexOf('Publicar') === 0 && t.length < 20)) {
                        nodes[i].click();
                        return true;
                    }
                }
                return false;
            }"""
        )
        _safe_print("    Clique Publicar (JS).")
    except Exception:
        pass

    time.sleep(0.5)

    # Reclique se modal continuar igual
    if modal_titulo_visivel(page, modal_kind):
        time.sleep(1.5)
        if modal_titulo_visivel(page, modal_kind):
            try:
                _safe_print("    Reclicando Publicar...")
                publicar_btn = _achar_botao()
                publicar_btn.click(force=True, timeout=8000)
                page.evaluate(
                    """() => {
                        var nodes = Array.from(document.querySelectorAll(
                            'button, div.bubble-element.Button, [role="button"]'
                        ));
                        for (var i = nodes.length - 1; i >= 0; i--) {
                            var t = ((nodes[i].innerText || '') + '').trim();
                            if (t === 'Publicar') { nodes[i].click(); return true; }
                        }
                        return false;
                    }"""
                )
            except Exception:
                pass

    aguardar_resultado_apos_publicar(page, modal_kind, modal_root)
    time.sleep(PAUSA_APOS_CLICAR_PUBLICAR)


def modal_titulo_visivel(page, modal_kind):
    try:
        return _loc_modal_titulo(page, modal_kind).is_visible()
    except Exception:
        return False


def _resolver_arquivo_item(item):
    link = (item.get("arquivo") or "").strip()
    if not link:
        return None
    hint = item.get("nome") or item.get("mes_ano") or "doc"
    return baixar_arquivo_drive(link, PASTA_DOWNLOADS, nome_hint=hint)


def publicar_um_item(page, item, modal_kind, preencher_fn):
    nome_base = normalizar(
        (item.get("nome") or "item").replace(" ", "_")
    )[:50]
    arquivo_path = _resolver_arquivo_item(item)

    def rodada(prefixo_shot):
        abrir_modal(page, modal_kind)
        preencher_fn(page, item, arquivo_path)
        salvar_screenshot(
            page, "{}_antes_{}_{}".format(prefixo_shot, modal_kind, nome_base)
        )
        clicar_publicar(page, modal_kind)
        salvar_screenshot(
            page, "{}_apos_{}_{}".format(prefixo_shot, modal_kind, nome_base)
        )

    if PUBLICAR_DUPLO_BUBBLE:
        print("    [1/2] Workaround duplo...")
        rodada("t1")
        fechar_modal(page, modal_kind)
        time.sleep(0.28)
        rodada("t2")
    else:
        rodada("pub")

    if modal_titulo_visivel(page, modal_kind):
        print("    Fechando modal apos publicacao...")
        fechar_modal(page, modal_kind)
    time.sleep(0.14)
    print("    Concluido.")


# ---------------------------------------------------------------------
#  Login / navegacao
# ---------------------------------------------------------------------


def credenciais_portal_configuradas():
    u = (PORTAL_USUARIO or "").strip()
    s = (PORTAL_SENHA or "").strip()
    return bool(u and s)


def _resolver_escopo_login(page):
    ultimo = None
    for _ in range(50):
        ordem = [page]
        try:
            mf = page.main_frame
            for fr in page.frames:
                if fr != mf:
                    ordem.append(fr)
        except Exception:
            ordem.extend([f for f in page.frames if f not in ordem])
        for scope in ordem:
            try:
                scope.locator("input[type='password']").first.wait_for(
                    state="visible", timeout=300
                )
                return scope
            except Exception as e:
                ultimo = e
                continue
        time.sleep(0.08)
    raise TimeoutError(
        "Formulario de login nao ficou pronto. Ultimo: {}".format(ultimo)
    )


def login_automatico_portal(page):
    """Preenche usuario/senha na tela CR2 (Bubble) e clica Acessar — igual ao RGF."""
    usuario = PORTAL_USUARIO.strip()
    senha = PORTAL_SENHA.strip()
    scope = _resolver_escopo_login(page)

    rotulos_email_cr2 = (
        re.compile(r"informe seu e-?\s*mail\s*:?", re.I),
        re.compile(r"^\s*informe seu e-?\s*mail", re.I),
        re.compile(r"seu e-?\s*mail", re.I),
    )
    rotulos_senha_cr2 = (
        re.compile(r"informe sua senha\s*:?", re.I),
        re.compile(r"^\s*informe sua senha", re.I),
        re.compile(r"sua senha", re.I),
    )

    preenchido_usuario = False
    for rx in rotulos_email_cr2:
        try:
            loc = scope.get_by_label(rx).first
            loc.wait_for(state="visible", timeout=3500)
            preencher_campo_rapido(page, loc, usuario)
            preenchido_usuario = True
            print("[INFO] Login: e-mail via rotulo Portal CR2.")
            break
        except Exception:
            continue

    if not preenchido_usuario:
        for label in (
            "Email",
            "E-mail",
            "Usuário",
            "Usuario",
            "Login",
            "User",
        ):
            try:
                loc = scope.get_by_label(
                    re.compile("^" + re.escape(label) + "$", re.I)
                ).first
                loc.wait_for(state="visible", timeout=2000)
                preencher_campo_rapido(page, loc, usuario)
                preenchido_usuario = True
                print("[INFO] Login: usuario via rotulo '{}'.".format(label))
                break
            except Exception:
                continue

    if not preenchido_usuario:
        for loc in (
            scope.get_by_placeholder(
                re.compile(r"gmail|meucontato|e-?\s*mail|@", re.I)
            ),
            scope.locator("input[type='email']"),
            scope.get_by_placeholder(
                re.compile(
                    r"e\s*-?\s*mail|usu[aá]rio|login|user|account",
                    re.I,
                )
            ),
            scope.locator("input[autocomplete='username']"),
            scope.locator("input[autocomplete='email']"),
            scope.locator("input[name*='email' i]"),
            scope.locator("input[type='text']"),
        ):
            try:
                alvo = loc.first
                alvo.wait_for(state="visible", timeout=4500)
                preencher_campo_rapido(page, alvo, usuario)
                preenchido_usuario = True
                print("[INFO] Login: usuario via placeholder/input.")
                break
            except Exception:
                continue

    if not preenchido_usuario:
        raise RuntimeError("Campo de usuario/email nao encontrado")

    preenchido_senha = False
    for rx in rotulos_senha_cr2:
        try:
            loc = scope.get_by_label(rx).first
            loc.wait_for(state="visible", timeout=4000)
            preencher_campo_rapido(page, loc, senha)
            preenchido_senha = True
            print("[INFO] Login: senha via rotulo Portal CR2.")
            break
        except Exception:
            continue
    if not preenchido_senha:
        campo_senha = scope.locator("input[type='password']").first
        campo_senha.wait_for(state="visible", timeout=10000)
        preencher_campo_rapido(page, campo_senha, senha)
        print("[INFO] Login: senha via input[type=password].")

    time.sleep(0.12)
    clicou = False
    txt_botao = (PORTAL_LOGIN_BOTAO or "").strip()
    if txt_botao:
        try:
            b = scope.get_by_role(
                "button", name=re.compile(re.escape(txt_botao), re.I)
            ).first
            b.wait_for(state="visible", timeout=6000)
            b.click(force=True)
            clicou = True
            print("[INFO] Login: clique em '{}' (config).".format(txt_botao))
        except Exception:
            try:
                b = scope.locator("div[role='button'], div.bubble-element.Button").filter(
                    has_text=re.compile(r"^\s*" + re.escape(txt_botao) + r"\s*$", re.I)
                ).first
                b.wait_for(state="visible", timeout=4000)
                b.click(force=True)
                clicou = True
                print("[INFO] Login: clique (div Bubble) '{}'.".format(txt_botao))
            except Exception:
                pass

    if not clicou:
        for rotulo in (
            "Acessar",
            "Entrar",
            "Login",
            "Sign in",
            "Continuar",
            "Log in",
            "Entrar na conta",
            "Acessar conta",
        ):
            try:
                loc_btn = scope.get_by_role(
                    "button", name=re.compile(re.escape(rotulo), re.I)
                ).first
                loc_btn.wait_for(state="visible", timeout=2200)
                loc_btn.click(force=True)
                clicou = True
                print("[INFO] Login: clique em '{}'.".format(rotulo))
                break
            except Exception:
                try:
                    alt = scope.locator(
                        "button:has-text('{0}'), div[role='button']:has-text('{0}'), "
                        "div.bubble-element.Button:has-text('{0}')".format(rotulo)
                    ).first
                    alt.wait_for(state="visible", timeout=1800)
                    alt.click(force=True)
                    clicou = True
                    print("[INFO] Login: clique (fallback) '{}'.".format(rotulo))
                    break
                except Exception:
                    continue

    if not clicou:
        sub = scope.locator("button[type='submit'], input[type='submit']").first
        sub.wait_for(state="visible", timeout=6000)
        sub.click(force=True)
        print("[INFO] Login: clique em submit generico.")


def aguardar_login_concluido(page, timeout_s=None):
    """Espera sair de view=login (sessao criada). Nao segue cego apos 1s."""
    if timeout_s is None:
        timeout_s = TIMEOUT_LOGIN_CONCLUIDO_S
    print("[INFO] Aguardando login concluir (sair de view=login)...")
    fim = time.monotonic() + float(timeout_s)
    while time.monotonic() < fim:
        try:
            url = (page.url or "").lower()
            if "view=login" not in url:
                try:
                    aguardar_barra_carregamento_topo(page, timeout_s=20, etiqueta="pos-login")
                except Exception:
                    pass
                print("[INFO] Login OK — URL: {}".format(page.url))
                return True
            try:
                txt = page.locator("body").inner_text(timeout=1200) or ""
            except Exception:
                txt = ""
            if re.search(
                r"(senha\s+(incorreta|inv[aá]lida)|credencia(is)?\s+inv[aá]lid|"
                r"e-?\s*mail\s+ou\s+senha|acesso\s+negado|usu[aá]rio\s+n[aã]o)",
                txt,
                re.I,
            ):
                salvar_screenshot(page, "ERRO_LOGIN_CREDENCIAIS")
                raise RuntimeError(
                    "Portal rejeitou o login (senha/usuario). Confira no front."
                )
        except RuntimeError:
            raise
        except Exception:
            pass
        time.sleep(0.45)
    salvar_screenshot(page, "ERRO_LOGIN_TIMEOUT")
    raise TimeoutError(
        "Login nao concluiu em {}s (ainda em view=login). "
        "Confira usuario/senha no front ou faca login manual no Opera.".format(
            int(timeout_s)
        )
    )


def aguardar_barra_carregamento_topo(page, timeout_s=None, etiqueta=""):
    if timeout_s is None:
        timeout_s = TIMEOUT_LOADER_TOPO_S
    tag = " [{}]".format(etiqueta) if etiqueta else ""
    limite = time.monotonic() + float(timeout_s)
    viu_barra = False
    js_topo = """
        () => {
            function barraTopoAtiva(el) {
                if (!el) return false;
                var s = window.getComputedStyle(el);
                if (s.display === 'none' || parseFloat(s.opacity) < 0.08) return false;
                var r = el.getBoundingClientRect();
                if (r.top > 22 || r.height > 28) return false;
                return r.width > 12;
            }
            var np = document.querySelector('#nprogress .bar');
            if (barraTopoAtiva(np)) return true;
            var turbo = document.querySelector('.turbo-progress-bar');
            if (barraTopoAtiva(turbo)) return true;
            return false;
        }
    """
    while time.monotonic() < limite:
        try:
            ativo = page.evaluate(js_topo)
        except Exception:
            ativo = False
        if ativo:
            if not viu_barra:
                print(
                    "[INFO] Aguardando barra de progresso no topo sumir{}...".format(
                        tag
                    )
                )
                viu_barra = True
            time.sleep(0.14)
            continue
        if viu_barra:
            print("[INFO] Barra de progresso concluida{}.".format(tag))
        return
    raise TimeoutError(
        "Barra de progresso ainda ativa apos {}s{}.".format(timeout_s, tag)
    )


def navegar_para_url(page, url, etiqueta, pausa_apos_carregar=1.0):
    print("[INFO] Carregando {}...".format(etiqueta))
    ultimo = None
    for wt in ("domcontentloaded", "load"):
        try:
            page.goto(url, wait_until=wt, timeout=120000)
            ultimo = None
            break
        except Exception as e:
            ultimo = e
    if ultimo:
        raise ultimo
    time.sleep(pausa_apos_carregar)


def garantir_pagina_portal(page, url_alvo, etiqueta_log):
    navegar_para_url(page, url_alvo, etiqueta_log, pausa_apos_carregar=0.52)
    path = (urlparse(url_alvo).path or "").strip("/")
    slug_esperado = path.split("/")[-1] if path else ""
    if slug_esperado and slug_esperado not in page.url:
        print("[AVISO] URL pode nao bater (slug {}). Forcando...".format(slug_esperado))
        try:
            page.evaluate("(u) => { window.location.href = u; }", url_alvo)
            page.wait_for_load_state("domcontentloaded", timeout=120000)
        except Exception:
            page.goto(url_alvo, wait_until="domcontentloaded", timeout=120000)
        time.sleep(0.22)
    print("[INFO] URL atual: {}".format(page.url))
    try:
        page.locator("button:has-text('Criar Publicação')").wait_for(
            state="visible", timeout=45000
        )
        print("[INFO] Pagina carregada (Criar Publicacao visivel).")
        aguardar_barra_carregamento_topo(page, etiqueta=etiqueta_log)
    except Exception as e:
        print(
            "[AVISO] Botao 'Criar Publicacao' nao apareceu. "
            "Confira se o login concluiu. ({})".format(str(e)[:120])
        )


def aguardar_login_usuario(page, pular_enter=False):
    navegar_para_url(page, URL_LOGIN, "login — {}".format(URL_LOGIN), 0.12)
    print("[INFO] Pagina de login aberta: {}".format(URL_LOGIN))
    if credenciais_portal_configuradas():
        print("[INFO] Tentando login automatico...")
        try:
            login_automatico_portal(page)
            print("[INFO] Formulario de login enviado.")
            time.sleep(0.4)
            aguardar_login_concluido(page)
        except Exception as e:
            print(
                "[AVISO] Login automatico falhou ({}).".format(str(e)[:200])
            )
            salvar_screenshot(page, "ERRO_LOGIN_AUTOMATICO")
            if pular_enter:
                print(
                    "[INFO] Faca login manual no Opera se a tela ainda estiver aberta. "
                    "Aguardando ate {}s...".format(TIMEOUT_LOGIN_CONCLUIDO_S)
                )
                try:
                    aguardar_login_concluido(page, timeout_s=TIMEOUT_LOGIN_CONCLUIDO_S)
                except Exception as e2:
                    raise RuntimeError(
                        "Login nao concluido: {} | {}".format(
                            str(e)[:120], str(e2)[:120]
                        )
                    )
    else:
        print(
            "[INFO] Preencha usuario/senha no front (ou PORTAL_* no script)."
        )
    if pular_enter:
        if "view=login" in (page.url or "").lower():
            print("[INFO] Ainda na tela de login — aguardando sessao...")
            aguardar_login_concluido(page)
        else:
            time.sleep(0.6)
    else:
        try:
            input(
                "[INFO] Quando estiver logado, Enter para continuar...\n>>> "
            )
        except EOFError:
            print(
                "[INFO] Sem terminal: aguardando login ate {}s...".format(
                    TIMEOUT_LOGIN_CONCLUIDO_S
                )
            )
            aguardar_login_concluido(page)


def verificar_playwright_instalado():
    if sync_playwright is not None:
        return
    garantir_playwright_pronto()
    if sync_playwright is not None:
        return
    print("\n[ERRO] Playwright indisponivel. pip install playwright")
    print("       Depois: python -m playwright install chromium")
    sys.exit(1)


def criar_navegador_e_login(pular_enter_pos_login=False):
    verificar_playwright_instalado()
    opera_path = resolver_caminho_opera()
    if opera_path is None:
        print("\n[ERRO] Opera nao encontrado. Instale ou defina OPERA_EXE.")
        sys.exit(1)
    pw = sync_playwright().start()
    print("\n[INFO] Abrindo Opera: {}".format(opera_path))
    try:
        browser = pw.chromium.launch(
            executable_path=str(opera_path),
            headless=HEADLESS,
        )
    except Exception as e:
        print("\n[ERRO] Falha ao iniciar Opera: {}".format(e))
        pw.stop()
        sys.exit(1)
    context = browser.new_context()
    page = context.new_page()
    if ABRIR_LOGIN_ANTES_DO_PORTAL:
        aguardar_login_usuario(page, pular_enter=pular_enter_pos_login)
    return pw, browser, page


# ---------------------------------------------------------------------
#  Orquestracao
# ---------------------------------------------------------------------


def publicar_filas(
    fila_estagiario,
    fila_terceirizado,
    fila_divida,
    pular_enter_pos_login=False,
    on_item=None,
):
    """
    Publica 1 a 1 no padrao RGF:
      Criar Publicacao → preencher → Publicar → proxima linha.
    Em erro, fecha o modal, registra a linha e segue.
    on_item(ok, item, kind, publicadas, erros) opcional (progresso/cancelamento).
    Retorna (ok, linhas_nao_publicadas).
    """
    verificar_playwright_instalado()
    pw = None
    browser = None
    ok = 0
    nao_publicadas = []
    url_corrente = None

    try:
        pw, browser, page = criar_navegador_e_login(
            pular_enter_pos_login=pular_enter_pos_login
        )

        def navegar_se_preciso(url, etiqueta):
            nonlocal url_corrente
            if not url_portal_ativa(url):
                return False
            if url_corrente != url:
                garantir_pagina_portal(page, url, etiqueta)
                url_corrente = url
            return True

        def processar(itens, url, kind, preencher_fn, etiqueta):
            nonlocal ok
            if not itens:
                return True
            if not navegar_se_preciso(url, etiqueta):
                print(
                    "[ERRO] URL vazia para {} — pulando {} item(ns).".format(
                        kind, len(itens)
                    )
                )
                for it in itens:
                    nao_publicadas.append(
                        entrada_nao_publicada(it, "URL portal vazia")
                    )
                return True
            for it in itens:
                if on_item is not None:
                    try:
                        if on_item(
                            ok=None,
                            item=it,
                            kind=kind,
                            publicadas=ok,
                            erros=len(nao_publicadas),
                            fase="antes",
                        ) is False:
                            print("[INFO] Publicacao interrompida (cancelamento).")
                            return False
                    except Exception:
                        pass
                # Cancela tambem se job_runtime pediu parada sem callback
                try:
                    import job_runtime as _jobrt

                    if _jobrt.pedido_cancelado():
                        print("[INFO] Publicacao interrompida (cancelamento).")
                        return False
                except Exception:
                    pass
                rotulo = it.get("nome") or "(sem nome)"
                print(
                    "\n[-> {}] linha {} — {}".format(kind.upper(), it["linha"], rotulo)
                )
                try:
                    publicar_um_item(page, it, kind, preencher_fn)
                    ok += 1
                    try:
                        import job_runtime as _jobrt

                        _jobrt.salvar_checkpoint_linha(
                            kind, it.get("linha"), publicadas_total=ok
                        )
                    except Exception:
                        pass
                    if on_item is not None:
                        try:
                            on_item(
                                ok=True,
                                item=it,
                                kind=kind,
                                publicadas=ok,
                                erros=len(nao_publicadas),
                                fase="ok",
                            )
                        except Exception:
                            pass
                except Exception as e:
                    motivo = str(e)[:160]
                    print("    [PULO] {}".format(motivo[:200]))
                    salvar_screenshot(
                        page,
                        "ERRO_{}_{}".format(
                            kind, normalizar(rotulo.replace(" ", "_"))[:40]
                        ),
                    )
                    try:
                        fechar_modal(page, kind)
                    except Exception:
                        pass
                    nao_publicadas.append(entrada_nao_publicada(it, motivo))
                    if on_item is not None:
                        try:
                            on_item(
                                ok=False,
                                item=it,
                                kind=kind,
                                publicadas=ok,
                                erros=len(nao_publicadas),
                                fase="erro",
                            )
                        except Exception:
                            pass
            return True

        if not processar(
            fila_estagiario,
            URL_PORTAL_ESTAGIARIO,
            "estagiario",
            preencher_modal_estagiario,
            "Estagiarios — {}".format(URL_PORTAL_ESTAGIARIO),
        ):
            print("[INFO] Filas restantes canceladas.")
        elif not processar(
            fila_terceirizado,
            URL_PORTAL_TERCEIRIZADO,
            "terceirizado",
            preencher_modal_terceirizado,
            "Terceirizados — {}".format(URL_PORTAL_TERCEIRIZADO),
        ):
            print("[INFO] Filas restantes canceladas.")
        elif not processar(
            fila_divida,
            URL_PORTAL_DIVIDA,
            "divida",
            preencher_modal_divida,
            "Divida Ativa — {}".format(URL_PORTAL_DIVIDA),
        ):
            print("[INFO] Fila cancelada.")

        print("\n" + "=" * 50)
        print(
            "  Resumo: {} ok | {} nao publicada(s)".format(ok, len(nao_publicadas))
        )
        if nao_publicadas:
            print("  Linhas nao publicadas:")
            for item in nao_publicadas[:30]:
                print(
                    "    - [{}] L{} {}: {}".format(
                        item.get("kind"),
                        item.get("linha"),
                        item.get("nome"),
                        item.get("motivo"),
                    )
                )
            if len(nao_publicadas) > 30:
                print("    ... +{} ".format(len(nao_publicadas) - 30))
        print("=" * 50)
        return ok, nao_publicadas
    finally:
        if browser is not None:
            try:
                browser.close()
            except Exception:
                pass
        if pw is not None:
            try:
                pw.stop()
            except Exception:
                pass


def _format_resumo(item):
    t = item.get("kind") or item.get("tipo")
    if t == "estagiario":
        return "nome={} inicio={} fim={} situacao={} arq={}".format(
            item["nome"],
            item["inicio"],
            item["fim"],
            item["situacao"],
            "sim" if item.get("arquivo") else "nao",
        )
    if t == "terceirizado":
        return "mes={} nome={} empresa={} funcao={} arq={}".format(
            item["mes_ano"],
            item["nome"],
            item["empresa"],
            item["funcao"],
            "sim" if item.get("arquivo") else "nao",
        )
    return "ano={} nome={} valor={} arq={}".format(
        item["ano"],
        item["nome"],
        item["valor"],
        "sim" if item.get("arquivo") else "nao",
    )




# ---------------------------------------------------------------------
#  Main
# ---------------------------------------------------------------------


if __name__ == "__main__":
    if sync_playwright is None and "--help" not in sys.argv and "-h" not in sys.argv:
        garantir_playwright_pronto()

    if "--help" in sys.argv or "-h" in sys.argv:
        print(
            "Uso: python publicar_estagiario_terceirizado_divida.py [opcoes]\n\n"
            "  Tres planilhas independentes no CONFIG (link vazio = tipo desligado):\n"
            "    PLANILHA_DRIVE_ESTAGIARIO\n"
            "    PLANILHA_DRIVE_TERCEIRIZADO\n"
            "    PLANILHA_DRIVE_DIVIDA\n"
            "  Pode preencher so 1, 2 ou as 3. Tambem URL_PORTAL_* de cada tipo.\n\n"
            "  --gerar-modelo       Cria 3 planilhas Excel modelo nesta pasta.\n"
            "  --test, -t           No max. 1 linha por fila ativa (Criar Publicacao).\n"
            "  --so-estagiario      Apenas Estagiarios.\n"
            "  --so-terceirizado    Apenas Terceirizados.\n"
            "  --so-divida          Apenas Divida Ativa.\n"
            "  --yes, -y            Pula confirmacoes + Enter pos-login.\n"
        )
        sys.exit(0)

    if "--gerar-modelo" in sys.argv:
        gerar_planilha_modelo()
        sys.exit(0)

    modo_teste = MODO_TESTE or "--test" in sys.argv or "-t" in sys.argv
    confirmar_automatico = "--yes" in sys.argv or "-y" in sys.argv
    only_est = "--so-estagiario" in sys.argv
    only_ter = "--so-terceirizado" in sys.argv
    only_div = "--so-divida" in sys.argv

    # Tipo ativo = tem planilha Drive + URL do portal (salvo --so-* forcar so um)
    quer_est = bool((PLANILHA_DRIVE_ESTAGIARIO or "").strip()) and url_portal_ativa(
        URL_PORTAL_ESTAGIARIO
    )
    quer_ter = bool((PLANILHA_DRIVE_TERCEIRIZADO or "").strip()) and url_portal_ativa(
        URL_PORTAL_TERCEIRIZADO
    )
    quer_div = bool((PLANILHA_DRIVE_DIVIDA or "").strip()) and url_portal_ativa(
        URL_PORTAL_DIVIDA
    )

    if only_est:
        quer_est, quer_ter, quer_div = True, False, False
    elif only_ter:
        quer_est, quer_ter, quer_div = False, True, False
    elif only_div:
        quer_est, quer_ter, quer_div = False, False, True

    print("=" * 50)
    print("  CR2 — Estagiarios + Terceirizados + Divida Ativa")
    print("=" * 50)

    def _preparar_planilha(url_drive, cache_local, rotulo):
        url_drive = (url_drive or "").strip()
        if not url_drive:
            return None
        path = baixar_planilha_drive(url_drive, cache_local)
        if path is None or not path.is_file():
            print(
                "[ERRO] Nao foi possivel baixar planilha de {}.\n"
                "       Confira o link e o compartilhamento "
                "(qualquer pessoa com o link).".format(rotulo)
            )
            sys.exit(1)
        return path

    planilha_est = None
    planilha_ter = None
    planilha_div = None

    if quer_est:
        if not (PLANILHA_DRIVE_ESTAGIARIO or "").strip():
            print("[ERRO] --so-estagiario exige PLANILHA_DRIVE_ESTAGIARIO no CONFIG.")
            sys.exit(1)
        if not url_portal_ativa(URL_PORTAL_ESTAGIARIO):
            print("[ERRO] URL_PORTAL_ESTAGIARIO vazia.")
            sys.exit(1)
        planilha_est = _preparar_planilha(
            PLANILHA_DRIVE_ESTAGIARIO, PLANILHA_ESTAGIARIO, "Estagiarios"
        )
        print("  [Estagiario] ATIVO — planilha + portal")
    else:
        motivo = []
        if not (PLANILHA_DRIVE_ESTAGIARIO or "").strip():
            motivo.append("planilha vazia")
        if not url_portal_ativa(URL_PORTAL_ESTAGIARIO):
            motivo.append("portal vazio")
        if only_ter or only_div:
            motivo = ["filtro --so-*"]
        print("  [Estagiario] pulado ({})".format(", ".join(motivo) or "?"))

    if quer_ter:
        if not (PLANILHA_DRIVE_TERCEIRIZADO or "").strip():
            print("[ERRO] --so-terceirizado exige PLANILHA_DRIVE_TERCEIRIZADO no CONFIG.")
            sys.exit(1)
        if not url_portal_ativa(URL_PORTAL_TERCEIRIZADO):
            print("[ERRO] URL_PORTAL_TERCEIRIZADO vazia.")
            sys.exit(1)
        planilha_ter = _preparar_planilha(
            PLANILHA_DRIVE_TERCEIRIZADO, PLANILHA_TERCEIRIZADO, "Terceirizados"
        )
        print("  [Terceirizado] ATIVO — planilha + portal")
    else:
        motivo = []
        if not (PLANILHA_DRIVE_TERCEIRIZADO or "").strip():
            motivo.append("planilha vazia")
        if not url_portal_ativa(URL_PORTAL_TERCEIRIZADO):
            motivo.append("portal vazio")
        if only_est or only_div:
            motivo = ["filtro --so-*"]
        print("  [Terceirizado] pulado ({})".format(", ".join(motivo) or "?"))

    if quer_div:
        if not (PLANILHA_DRIVE_DIVIDA or "").strip():
            print("[ERRO] --so-divida exige PLANILHA_DRIVE_DIVIDA no CONFIG.")
            sys.exit(1)
        if not url_portal_ativa(URL_PORTAL_DIVIDA):
            print("[ERRO] URL_PORTAL_DIVIDA vazia.")
            sys.exit(1)
        planilha_div = _preparar_planilha(
            PLANILHA_DRIVE_DIVIDA, PLANILHA_DIVIDA, "Divida Ativa"
        )
        print("  [Divida Ativa] ATIVO — planilha + portal")
    else:
        motivo = []
        if not (PLANILHA_DRIVE_DIVIDA or "").strip():
            motivo.append("planilha vazia")
        if not url_portal_ativa(URL_PORTAL_DIVIDA):
            motivo.append("portal vazio")
        if only_est or only_ter:
            motivo = ["filtro --so-*"]
        print("  [Divida Ativa] pulado ({})".format(", ".join(motivo) or "?"))

    fila_est = ler_fila_estagiario(planilha_est) if planilha_est else []
    fila_ter = ler_fila_terceirizado(planilha_ter) if planilha_ter else []
    fila_div = ler_fila_divida(planilha_div) if planilha_div else []

    if modo_teste:
        print("\n  *** MODO TESTE — ate 1 linha por fila (modal) ***\n")
        fila_est = fila_est[:1]
        fila_ter = fila_ter[:1]
        fila_div = fila_div[:1]

    total = len(fila_est) + len(fila_ter) + len(fila_div)
    if total == 0:
        print(
            "\n[INFO] Nenhuma linha valida na fila.\n"
            "       Cole o(s) link(s) PLANILHA_DRIVE_* no CONFIG e preencha os dados.\n"
            "       Modelos: python publicar_estagiario_terceirizado_divida.py --gerar-modelo\n"
        )
        sys.exit(0)

    print("\n{} item(ns) na fila (Criar Publicacao, 1 a 1):\n".format(total))
    if fila_est:
        print("--- Estagiarios ({}) ---".format(len(fila_est)))
        for it in fila_est:
            print("  L{} -> {}".format(it["linha"], _format_resumo(it)))
    if fila_ter:
        print("--- Terceirizados ({}) ---".format(len(fila_ter)))
        for it in fila_ter:
            print("  L{} -> {}".format(it["linha"], _format_resumo(it)))
    if fila_div:
        print("--- Divida Ativa ({}) ---".format(len(fila_div)))
        for it in fila_div:
            print("  L{} -> {}".format(it["linha"], _format_resumo(it)))

    print()
    if confirmar_automatico:
        resposta = "s"
        print("[INFO] Confirmacao automatica (--yes).\n")
    else:
        try:
            resposta = input("Prosseguir? (s/n): ").strip().lower()
        except EOFError:
            print("[INFO] Sem terminal. Use --yes. Encerrando.")
            sys.exit(0)
    if resposta != "s":
        sys.exit(0)

    if fila_est and not url_portal_ativa(URL_PORTAL_ESTAGIARIO):
        print("[ERRO] URL_PORTAL_ESTAGIARIO vazia.")
        sys.exit(1)
    if fila_ter and not url_portal_ativa(URL_PORTAL_TERCEIRIZADO):
        print("[ERRO] URL_PORTAL_TERCEIRIZADO vazia.")
        sys.exit(1)
    if fila_div and not url_portal_ativa(URL_PORTAL_DIVIDA):
        print("[ERRO] URL_PORTAL_DIVIDA vazia.")
        sys.exit(1)

    publicar_filas(
        fila_est,
        fila_ter,
        fila_div,
        pular_enter_pos_login=confirmar_automatico,
    )
