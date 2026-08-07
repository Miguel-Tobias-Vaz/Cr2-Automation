# -*- coding: utf-8 -*-
"""
Repasses — lê planilha (Google Sheets / Drive / .xlsx local), baixa documentos,
extrai dados (planilha + OCR) e gera:
  PASTA_BASE\\Repasses\\<ano>\\Repasse-MM-AAAA[-desc].pdf
  PASTA_BASE\\Repasses.xlsx
"""

from __future__ import annotations

import os
import re
import shutil
import sys
import tempfile
import urllib.request
from pathlib import Path
from urllib.parse import parse_qs, urlparse

import requests

# =============================================================
# CONFIG
# =============================================================

PASTA_BASE = r"C:\Downloads"
URL_PLANILHA = ""  # link Google Sheets/Drive ou caminho local .xlsx
USAR_OCR = True
REFINAR_IA = True  # Ollama: completa mes/ano, data, valores, descricao
MODELO_IA = "llama3.2:3b"
OLLAMA_URL = "http://127.0.0.1:11434"
# auto | paddleocr | tesseract
MOTOR_OCR = "auto"
ANOS_FILTRO: list[str] = []  # ex.: ["2023","2024"]; vazio = todos

HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
        "AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36"
    )
}
_DRIVE_UC = "https://drive.google.com/uc"


class Cancelado(Exception):
    """Fila interrompida pelo usuario (centro-automacoes)."""


def pedido_cancelado():
    return False


def _abortar_se_cancelado():
    if pedido_cancelado():
        print("[AVISO] Fila cancelada pelo usuario.")
        raise Cancelado()


# ---------------------------------------------------------------------------
# Planilha (Drive / Sheets / local)
# ---------------------------------------------------------------------------

def extrair_id_drive(url: str) -> str | None:
    if not url:
        return None
    url = url.strip()
    m = re.search(r"/spreadsheets/d/([a-zA-Z0-9_-]+)", url)
    if m:
        return m.group(1)
    m = re.search(r"/file/d/([a-zA-Z0-9_-]+)", url)
    if m:
        return m.group(1)
    m = re.search(r"[?&]id=([a-zA-Z0-9_-]+)", url)
    if m:
        return m.group(1)
    if re.fullmatch(r"[a-zA-Z0-9_-]{20,}", url):
        return url
    return None


def _eh_google_sheets(url: str) -> bool:
    return "docs.google.com/spreadsheets" in (url or "").lower()


def baixar_planilha(url_ou_path: str, destino: Path) -> Path | None:
    url_ou_path = (url_ou_path or "").strip().strip('"')
    if not url_ou_path:
        print("[ERRO] Informe o link ou caminho da planilha.")
        return None

    local = Path(url_ou_path)
    if local.is_file() and local.suffix.lower() in (".xlsx", ".xlsm", ".xls", ".csv"):
        print(f"[OK] Usando planilha local: {local}")
        return local.resolve()

    if not url_ou_path.startswith("http"):
        print(f"[ERRO] Arquivo nao encontrado: {url_ou_path}")
        return None

    destino.parent.mkdir(parents=True, exist_ok=True)
    file_id = extrair_id_drive(url_ou_path)
    if not file_id:
        print("[ERRO] Nao foi possivel extrair o ID do link da planilha.")
        return None

    print("[INFO] Baixando planilha...")
    if _eh_google_sheets(url_ou_path):
        export_url = (
            f"https://docs.google.com/spreadsheets/d/{file_id}/export?format=xlsx"
        )
        try:
            req = urllib.request.Request(export_url, headers=HEADERS)
            with urllib.request.urlopen(req, timeout=120) as resp:
                data = resp.read()
            if data[:2] != b"PK":
                print(
                    "[ERRO] Download nao parece Excel. "
                    "Confira se o link esta compartilhado (qualquer pessoa com o link)."
                )
                return None
            destino.write_bytes(data)
            print(f"[OK] Planilha salva em: {destino}")
            return destino.resolve()
        except Exception as e:
            print(f"[ERRO] Falha ao exportar Google Sheets: {str(e)[:160]}")
            return None

    # Arquivo no Drive
    try:
        baixar_arquivo_google_drive(file_id, str(destino))
        if destino.is_file() and destino.stat().st_size > 0:
            print(f"[OK] Planilha salva em: {destino}")
            return destino.resolve()
    except Exception as e:
        print(f"[ERRO] Falha ao baixar arquivo do Drive: {str(e)[:160]}")
    return None


def baixar_arquivo_google_drive(file_id: str, caminho_destino: str, timeout: int = 120):
    params = {"export": "download", "id": file_id}
    sessao = requests.Session()
    sessao.headers.update(HEADERS)
    r1 = sessao.get(_DRIVE_UC, params=params, timeout=timeout)
    r1.raise_for_status()
    token = None
    for chave, valor in r1.cookies.items():
        if chave.startswith("download_warning"):
            token = valor
            break
    if token is None and (
        r1.content[:5].startswith(b"%PDF") or r1.content[:2] == b"PK"
    ):
        with open(caminho_destino, "wb") as arquivo:
            arquivo.write(r1.content)
        return
    confirm = token
    if confirm is None:
        m = re.search(r"confirm=([\w-]+)", r1.text)
        if not m:
            raise ValueError(
                "Download do Drive bloqueado ou link sem permissao publica."
            )
        confirm = m.group(1)
    r2 = sessao.get(
        _DRIVE_UC,
        params={**params, "confirm": confirm},
        timeout=timeout,
        stream=True,
    )
    r2.raise_for_status()
    with open(caminho_destino, "wb") as arquivo:
        for parte in r2.iter_content(chunk_size=8192):
            if parte:
                arquivo.write(parte)


def baixar_documento(url: str, destino: Path) -> bool:
    url = (url or "").strip()
    if not url:
        return False
    destino.parent.mkdir(parents=True, exist_ok=True)
    file_id = extrair_id_drive(url)
    try:
        if file_id and ("drive.google.com" in url or "docs.google.com" in url):
            # Sheets como PDF não; arquivos sim
            if "/spreadsheets/" in url:
                print(f"  [AVISO] Link de planilha (nao PDF): {url[:80]}")
                return False
            baixar_arquivo_google_drive(file_id, str(destino))
        else:
            r = requests.get(url, headers=HEADERS, timeout=120, stream=True)
            r.raise_for_status()
            with open(destino, "wb") as f:
                for chunk in r.iter_content(8192):
                    if chunk:
                        f.write(chunk)
        if destino.is_file() and destino.stat().st_size > 100:
            # Se veio HTML (página de login), rejeita
            head = destino.read_bytes()[:200].lstrip()
            if head.startswith(b"<!DOCTYPE") or head.startswith(b"<html"):
                destino.unlink(missing_ok=True)
                print("  [AVISO] Download veio HTML (link privado?).")
                return False
            return True
    except Exception as e:
        print(f"  [ERRO] Download: {str(e)[:120]}")
    return False


# ---------------------------------------------------------------------------
# Modulo irmao (sempre recarrega do disco — painel mantem cache em sys.modules)
# ---------------------------------------------------------------------------

def _mod_extrair():
    import importlib
    import importlib.util

    caminho = Path(__file__).resolve().parent / "extrair_repasses.py"
    nome = "extrair_repasses"
    # tira cache velho (ex.: job anterior sem limpar_nome_arquivo)
    for key in list(sys.modules):
        if key == nome or key.endswith(".extrair_repasses"):
            sys.modules.pop(key, None)
    spec = importlib.util.spec_from_file_location(nome, str(caminho))
    if spec is None or spec.loader is None:
        raise ImportError("Nao foi possivel carregar {}".format(caminho))
    mod = importlib.util.module_from_spec(spec)
    sys.modules[nome] = mod
    spec.loader.exec_module(mod)
    return mod


# ---------------------------------------------------------------------------
# Leitura da planilha
# ---------------------------------------------------------------------------

def _ler_linhas_xlsx(path: Path) -> list[dict]:
    from openpyxl import load_workbook

    er = _mod_extrair()
    achar_url_na_celula = er.achar_url_na_celula
    celula_para_texto = er.celula_para_texto
    completar_mes_ano_e_data = er.completar_mes_ano_e_data
    mapear_cabecalhos = er.mapear_cabecalhos
    parse_data = er.parse_data
    parse_valor = er.parse_valor
    registro_vazio = er.registro_vazio

    wb = load_workbook(path, data_only=True)
    registros: list[dict] = []

    for ws in wb.worksheets:
        rows = list(ws.iter_rows())
        if not rows:
            continue
        # achar linha de cabeçalho (primeira com >=2 textos)
        header_idx = 0
        headers: list[str] = []
        for i, row in enumerate(rows[:15]):
            vals = [celula_para_texto(c.value) for c in row]
            nao_vazios = [v for v in vals if v]
            if len(nao_vazios) >= 2:
                headers = vals
                header_idx = i
                break
        if not headers:
            continue
        mapa = mapear_cabecalhos(headers)
        if not mapa:
            print(f"[AVISO] Aba '{ws.title}': cabecalhos nao reconhecidos: {headers[:8]}")
            continue
        print(f"[OK] Aba '{ws.title}': colunas {mapa}")

        for row in rows[header_idx + 1 :]:
            _abortar_se_cancelado()
            reg = registro_vazio()
            # hyperlinks openpyxl
            for campo, col_i in mapa.items():
                if col_i >= len(row):
                    continue
                cell = row[col_i]
                raw = cell.value
                hl = None
                if getattr(cell, "hyperlink", None) and cell.hyperlink:
                    hl = getattr(cell.hyperlink, "target", None) or str(cell.hyperlink)
                if campo == "link":
                    reg["link"] = achar_url_na_celula(raw, hl)
                elif campo == "mes_ano":
                    reg["mes_ano"] = celula_para_texto(raw)
                elif campo == "mes":
                    reg["mes"] = celula_para_texto(raw)
                elif campo == "ano":
                    reg["ano"] = celula_para_texto(raw)
                elif campo == "data_repasse":
                    reg["data_repasse"] = parse_data(raw)
                elif campo in ("valor_previsto", "valor_realizado"):
                    reg[campo] = parse_valor(raw)
                elif campo == "descricao":
                    reg["descricao"] = celula_para_texto(raw)
                else:
                    reg[campo] = celula_para_texto(raw)

            # Varre outras células da linha em busca de URL (se link vazio)
            if not reg.get("link"):
                for cell in row:
                    hl = None
                    if getattr(cell, "hyperlink", None) and cell.hyperlink:
                        hl = getattr(cell.hyperlink, "target", None) or str(
                            cell.hyperlink
                        )
                    u = achar_url_na_celula(cell.value, hl)
                    if u:
                        reg["link"] = u
                        break

            tem_dado = any(
                (reg.get(k) or "").strip()
                for k in (
                    "mes_ano",
                    "mes",
                    "ano",
                    "data_repasse",
                    "valor_previsto",
                    "valor_realizado",
                    "descricao",
                    "link",
                )
            )
            if tem_dado:
                # Junta Mês+Ano separados → MM/AAAA
                completar_mes_ano_e_data(reg)
                registros.append(reg)

    return registros


def _ler_linhas_csv(path: Path) -> list[dict]:
    import csv

    er = _mod_extrair()
    achar_url_na_celula = er.achar_url_na_celula
    completar_mes_ano_e_data = er.completar_mes_ano_e_data
    mapear_cabecalhos = er.mapear_cabecalhos
    parse_data = er.parse_data
    parse_valor = er.parse_valor
    registro_vazio = er.registro_vazio

    registros: list[dict] = []
    with path.open("r", encoding="utf-8-sig", newline="") as f:
        sample = f.read(4096)
        f.seek(0)
        try:
            dialect = csv.Sniffer().sniff(sample, delimiters=";,\t")
        except csv.Error:
            dialect = csv.excel
            dialect.delimiter = ";" if sample.count(";") >= sample.count(",") else ","
        reader = csv.reader(f, dialect)
        rows = list(reader)
    if not rows:
        return []
    headers = rows[0]
    mapa = mapear_cabecalhos(headers)
    for row in rows[1:]:
        reg = registro_vazio()
        for campo, col_i in mapa.items():
            if col_i >= len(row):
                continue
            raw = row[col_i]
            if campo == "link":
                reg["link"] = achar_url_na_celula(raw)
            elif campo == "mes_ano":
                reg["mes_ano"] = (raw or "").strip()
            elif campo == "mes":
                reg["mes"] = (raw or "").strip()
            elif campo == "ano":
                reg["ano"] = (raw or "").strip()
            elif campo == "data_repasse":
                reg["data_repasse"] = parse_data(raw)
            elif campo in ("valor_previsto", "valor_realizado"):
                reg[campo] = parse_valor(raw)
            else:
                reg[campo] = (raw or "").strip()
        if any(reg.get(k) for k in reg):
            completar_mes_ano_e_data(reg)
            registros.append(reg)
    return registros


# ---------------------------------------------------------------------------
# Texto / OCR
# ---------------------------------------------------------------------------

def ler_texto_pdf(caminho: Path) -> str:
    try:
        from _comum.ocr_multi import ler_texto_nativo

        return ler_texto_nativo(caminho)
    except Exception:
        try:
            import pdfplumber

            with pdfplumber.open(str(caminho)) as pdf:
                return "\n".join((p.extract_text() or "") for p in pdf.pages)
        except Exception:
            return ""


def obter_texto_documento(caminho: Path, usar_ocr: bool) -> str:
    try:
        auto = Path(__file__).resolve().parent.parent
        if str(auto) not in sys.path:
            sys.path.insert(0, str(auto))
        from _comum.ocr_multi import obter_texto_pdf

        texto, origem = obter_texto_pdf(
            caminho,
            usar_ocr=usar_ocr,
            motor=MOTOR_OCR or "auto",
            min_nativo=80,
            cache=True,
        )
        if origem.startswith("ocr"):
            print(f"  [OCR] origem={origem}")
        return texto or ""
    except Exception as e:
        print(f"  [AVISO] obter_texto falhou: {str(e)[:100]}")
        return ler_texto_pdf(caminho)


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main():
    # Garante import do módulo irmão (painel carrega via importlib)
    _dir = Path(__file__).resolve().parent
    if str(_dir) not in sys.path:
        sys.path.insert(0, str(_dir))
    er = _mod_extrair()
    completar_mes_ano_e_data = er.completar_mes_ano_e_data
    duplicar_valores = er.duplicar_valores
    escolher_valor_repasse = er.escolher_valor_repasse
    extrair_ano = er.extrair_ano
    extrair_com_ia = er.extrair_com_ia
    extrair_do_texto = er.extrair_do_texto
    limpar_nome_arquivo = er.limpar_nome_arquivo
    mesclar = er.mesclar
    nome_arquivo_final = er.nome_arquivo_final
    nome_arquivo_repasse = er.nome_arquivo_repasse
    normalizar_descricao = er.normalizar_descricao
    parse_mes_ano = er.parse_mes_ano
    salvar_planilha_repasses = er.salvar_planilha_repasses
    sanitizar_valores_reg = er.sanitizar_valores_reg
    valor_parece_lixo = er.valor_parece_lixo

    print("=" * 56)
    print(" REPASSES — planilha + documentos + OCR/IA")
    print("=" * 56)
    print(
        "  Saida: Repasses.xlsx (Mes e Ano | Data | Previsto | Realizado | Descricao | Arquivo)"
    )
    print("  OCR: {} | IA: {}".format(
        "ligado" if USAR_OCR else "desligado",
        "ligada ({})".format(MODELO_IA) if REFINAR_IA else "desligada",
    ))
    print("=" * 56)

    pasta = Path(PASTA_BASE)
    pasta.mkdir(parents=True, exist_ok=True)
    pasta_docs = pasta / "Repasses"
    pasta_docs.mkdir(parents=True, exist_ok=True)

    cache_planilha = pasta / "_cache_planilha_repasses.xlsx"
    path_planilha = baixar_planilha(URL_PLANILHA, cache_planilha)
    if not path_planilha:
        print("[ERRO] Sem planilha — abortando.")
        sys.exit(1)

    if path_planilha.suffix.lower() == ".csv":
        linhas = _ler_linhas_csv(path_planilha)
    else:
        linhas = _ler_linhas_xlsx(path_planilha)

    if not linhas:
        print("[ERRO] Nenhuma linha util encontrada na planilha.")
        sys.exit(1)

    print(f"[OK] {len(linhas)} linha(s) na planilha.")
    if ANOS_FILTRO:
        print(f"[INFO] Filtro de anos: {', '.join(ANOS_FILTRO)}")

    # Monta a fila só com o que entra no filtro — progresso [i/N] sobre N filtrado,
    # não sobre o total da planilha (ex.: 100 de 2023, não 200 linhas).
    fila: list[dict] = []
    pulados_ano = 0
    for reg in linhas:
        if not (reg.get("mes_ano") or "").strip():
            reg["mes_ano"] = parse_mes_ano("", reg.get("data_repasse") or "")
        completar_mes_ano_e_data(reg)
        if (reg.get("descricao") or "").strip():
            reg["descricao"] = normalizar_descricao(reg["descricao"])

        ano = extrair_ano(reg.get("mes_ano") or "", reg.get("data_repasse") or "")
        if ANOS_FILTRO:
            if ano and ano not in ANOS_FILTRO:
                pulados_ano += 1
                continue
            if not ano and not (reg.get("link") or "").strip():
                # sem ano e sem PDF — não dá para filtrar depois
                pulados_ano += 1
                continue
            # sem ano mas com link: entra na fila (OCR/IA pode revelar o ano)
        fila.append(reg)

    if ANOS_FILTRO:
        print(
            "[INFO] Fila: {0} de {1} (filtro {2}; {3} fora do ano)".format(
                len(fila),
                len(linhas),
                ", ".join(ANOS_FILTRO),
                pulados_ano,
            )
        )
        # Painel lê "fila: N" / "[i/N]" para a barra de progresso
        print("fila: {0}".format(len(fila)))
    else:
        print("fila: {0}".format(len(fila)))

    if not fila:
        print("[ERRO] Nenhuma linha na fila apos o filtro de anos.")
        sys.exit(1)

    saida: list[dict] = []
    total = len(fila)
    ok_docs = 0
    falha_docs = 0

    with tempfile.TemporaryDirectory(prefix="repasses_") as tmp:
        tmp_dir = Path(tmp)
        for i, reg in enumerate(fila, 1):
            _abortar_se_cancelado()
            print(f"[{i}/{total}] Processando...")

            ano = extrair_ano(reg.get("mes_ano") or "", reg.get("data_repasse") or "")
            # Redeclare filtro só para itens sem ano (confirmados após OCR abaixo)
            link = (reg.get("link") or "").strip()
            caminho_final: Path | None = None

            if link:
                tmp_pdf = tmp_dir / f"doc_{i}.pdf"
                if baixar_documento(link, tmp_pdf):
                    texto = obter_texto_documento(tmp_pdf, USAR_OCR)
                    extra = extrair_do_texto(texto, nome_arquivo=tmp_pdf.name)
                    reg = mesclar(reg, extra)
                    # IA completa o que OCR/planilha nao preencheram
                    faltando = any(
                        not (reg.get(k) or "").strip()
                        for k in (
                            "mes_ano",
                            "data_repasse",
                            "valor_previsto",
                            "valor_realizado",
                            "descricao",
                        )
                    )
                    if REFINAR_IA and (faltando or len((texto or "").strip()) > 40):
                        print("  [INFO] Extraindo campos com IA...")
                        ia = extrair_com_ia(
                            texto,
                            nome_arquivo=tmp_pdf.name,
                            modelo=MODELO_IA,
                            ollama_url=OLLAMA_URL,
                        )
                        reg = mesclar(reg, ia)
                    # Se ainda ficou placeholder (1.234,56), tenta de novo no texto
                    sanitizar_valores_reg(reg)
                    if (
                        valor_parece_lixo(reg.get("valor_previsto") or "")
                        or valor_parece_lixo(reg.get("valor_realizado") or "")
                        or (
                            not (reg.get("valor_previsto") or "").strip()
                            and not (reg.get("valor_realizado") or "").strip()
                        )
                    ):
                        alt = escolher_valor_repasse(texto)
                        if alt:
                            if not (reg.get("valor_realizado") or "").strip() or valor_parece_lixo(
                                reg.get("valor_realizado") or ""
                            ):
                                reg["valor_realizado"] = alt
                            if not (reg.get("valor_previsto") or "").strip() or valor_parece_lixo(
                                reg.get("valor_previsto") or ""
                            ):
                                reg["valor_previsto"] = alt
                    if (reg.get("descricao") or "").strip():
                        reg["descricao"] = normalizar_descricao(reg["descricao"])
                    completar_mes_ano_e_data(reg)
                    duplicar_valores(reg)
                    sanitizar_valores_reg(reg)
                    duplicar_valores(reg)
                    sanitizar_valores_reg(reg)
                    ano = extrair_ano(
                        reg.get("mes_ano") or "", reg.get("data_repasse") or ""
                    ) or "sem-ano"
                    if ANOS_FILTRO and ano != "sem-ano" and ano not in ANOS_FILTRO:
                        print(f"  (pulado apos OCR — ano {ano})")
                        continue
                    nome = nome_arquivo_repasse(reg)
                    dest_dir = pasta_docs / limpar_nome_arquivo(str(ano))
                    dest_dir.mkdir(parents=True, exist_ok=True)
                    caminho_final = dest_dir / nome
                    # evita sobrescrever
                    n = 2
                    while caminho_final.exists():
                        stem = Path(nome).stem
                        caminho_final = dest_dir / nome_arquivo_final(f"{stem} ({n})")
                        n += 1
                    shutil.copy2(tmp_pdf, caminho_final)
                    reg["arquivo"] = str(caminho_final.relative_to(pasta))
                    # Reforça competência pelo nome final (Repasse MM-AAAA)
                    ma_arq = er.mes_ano_do_arquivo(caminho_final.name)
                    if er.mes_ano_valido(ma_arq):
                        reg["mes_ano"] = ma_arq
                    completar_mes_ano_e_data(reg)
                    if not (reg.get("link") or "").strip():
                        reg["link"] = link
                    ok_docs += 1
                    print(
                        "  [OK] {0} | {1} | prev={2} real={3}".format(
                            caminho_final.name,
                            reg.get("mes_ano") or "?",
                            reg.get("valor_previsto") or "-",
                            reg.get("valor_realizado") or "-",
                        )
                    )
                else:
                    falha_docs += 1
                    print("  [AVISO] Nao baixou documento.")
            else:
                # só dados da planilha
                print("  [INFO] Sem link de documento — so dados da planilha.")
                if ANOS_FILTRO and ano and ano not in ANOS_FILTRO:
                    continue

            duplicar_valores(reg)
            completar_mes_ano_e_data(reg)
            sanitizar_valores_reg(reg)
            duplicar_valores(reg)
            sanitizar_valores_reg(reg)
            if not (reg.get("mes_ano") or "").strip():
                reg["mes_ano"] = parse_mes_ano("", reg.get("data_repasse") or "")
            if (reg.get("descricao") or "").strip():
                reg["descricao"] = normalizar_descricao(reg["descricao"])
            saida.append(reg)

    xlsx = salvar_planilha_repasses(saida, pasta)
    print()
    print("=" * 56)
    print(f" Concluidos: {len(saida)} registro(s)")
    print(f" Documentos salvos: {ok_docs}  |  falhas: {falha_docs}")
    if xlsx:
        print(f" Planilha (igual Sessões, para publicar): {xlsx}")
    print(f" Pasta docs: {pasta_docs}")
    print("=" * 56)


if __name__ == "__main__":
    # garante import do pacote irmão
    sys.path.insert(0, str(Path(__file__).resolve().parent))
    try:
        main()
    except Cancelado:
        print("[AVISO] Interrompido.")
        sys.exit(2)
