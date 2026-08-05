"""Geração da planilha Excel com openpyxl."""

from __future__ import annotations

import re
from datetime import datetime
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, numbers
from openpyxl.utils import get_column_letter

import config
from src.logger import get_logger

logger = get_logger()


def _parse_data(valor: str):
    if not valor:
        return None
    m = re.match(r"(\d{2})/(\d{2})/(\d{4})", valor.strip())
    if not m:
        return valor.strip()
    try:
        return datetime(int(m.group(3)), int(m.group(2)), int(m.group(1))).date()
    except ValueError:
        return valor.strip()


def _parse_valor(valor: str):
    if valor is None or valor == "":
        return None
    if isinstance(valor, (int, float, Decimal)):
        return float(valor)
    s = str(valor).strip()
    s = re.sub(r"[R$\s]", "", s)
    # 1.234.567,89 → 1234567.89
    if "," in s and "." in s:
        s = s.replace(".", "").replace(",", ".")
    elif "," in s:
        s = s.replace(",", ".")
    try:
        return float(Decimal(s))
    except (InvalidOperation, ValueError):
        return valor


def _estilo_cabecalho(ws) -> None:
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(wrap_text=True, vertical="center")
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions


def _ajustar_larguras(ws, minas: dict[str, int] | None = None) -> None:
    minas = minas or {}
    for idx, col in enumerate(ws.columns, start=1):
        letra = get_column_letter(idx)
        header = ws.cell(1, idx).value or ""
        max_len = len(str(header))
        for cell in col:
            if cell.value is not None:
                max_len = max(max_len, min(len(str(cell.value)), 60))
        ws.column_dimensions[letra].width = max(minas.get(str(header), 12), max_len + 2)


def gerar_planilha(
    registros: list[dict[str, Any]],
    arquivos_flat: list[dict[str, Any]],
    resumo_linhas: list[tuple[str, Any]],
    destino: Path | None = None,
) -> Path:
    destino = destino or config.PLANILHA_PATH
    destino.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()

    # --- Contratos ---
    ws_c = wb.active
    ws_c.title = "Contratos"
    headers_c = [
        "id_registro",
        "numero_contrato",
        "numero_contrato_slug",
        "tipo_registro",
        "data",
        "origem",
        "contratante",
        "contratada",
        "cnpj_cpf_contratada",
        "fiscal_responsavel",
        "valor",
        "inicio_vigencia",
        "fim_vigencia",
        "objeto",
        "qtd_itens",
        "qtd_arquivos_relacionados",
        "qtd_arquivos_baixados",
        "pasta_contrato",
        "url_detalhes",
        "contrato_original_url",
        "status",
        "mensagem_erro",
    ]
    ws_c.append(headers_c)

    for r in registros:
        itens = r.get("itens") or []
        arqs = [a for a in arquivos_flat if a.get("id_registro") == r.get("id_registro")]
        baixados = sum(1 for a in arqs if a.get("status_download") in ("baixado", "reaproveitado"))
        row = [
            r.get("id_registro"),
            r.get("numero_contrato"),
            r.get("numero_contrato_slug"),
            r.get("tipo_registro"),
            _parse_data(r.get("data_listagem") or ""),
            r.get("origem"),
            r.get("contratante"),
            r.get("contratada"),
            r.get("cnpj_cpf_contratada"),
            r.get("fiscal_responsavel"),
            _parse_valor(r.get("valor") or ""),
            _parse_data(r.get("inicio_vigencia") or ""),
            _parse_data(r.get("fim_vigencia") or ""),
            r.get("objeto"),
            len(itens),
            len(arqs),
            baixados,
            r.get("pasta_local"),
            r.get("url_detalhes"),
            r.get("contrato_original_url"),
            r.get("status"),
            r.get("mensagem_erro"),
        ]
        ws_c.append(row)

    for row in ws_c.iter_rows(min_row=2, min_col=11, max_col=11):
        for cell in row:
            if isinstance(cell.value, (int, float)):
                cell.number_format = '#,##0.00'
    for col in (5, 12, 13):
        for row in ws_c.iter_rows(min_row=2, min_col=col, max_col=col):
            for cell in row:
                if hasattr(cell.value, "year"):
                    cell.number_format = "DD/MM/YYYY"

    _estilo_cabecalho(ws_c)
    _ajustar_larguras(ws_c)

    # --- Itens ---
    ws_i = wb.create_sheet("Itens")
    headers_i = [
        "id_registro",
        "numero_contrato",
        "descricao",
        "quantidade",
        "unidade",
        "valor_unitario",
        "valor_total",
    ]
    ws_i.append(headers_i)
    for r in registros:
        for item in r.get("itens") or []:
            ws_i.append(
                [
                    r.get("id_registro"),
                    r.get("numero_contrato"),
                    item.get("descricao"),
                    item.get("quantidade"),
                    item.get("unidade"),
                    _parse_valor(item.get("valor_unitario") or ""),
                    _parse_valor(item.get("valor_total") or ""),
                ]
            )
    for col in (6, 7):
        for row in ws_i.iter_rows(min_row=2, min_col=col, max_col=col):
            for cell in row:
                if isinstance(cell.value, (int, float)):
                    cell.number_format = '#,##0.00'
    _estilo_cabecalho(ws_i)
    _ajustar_larguras(ws_i)

    # --- Arquivos ---
    ws_a = wb.create_sheet("Arquivos")
    headers_a = [
        "id_registro",
        "numero_contrato",
        "titulo_exibido",
        "nome_arquivo_original",
        "tipo_mime",
        "url_arquivo",
        "nome_salvo",
        "caminho_local",
        "status_download",
        "tamanho_bytes",
    ]
    ws_a.append(headers_a)
    for a in arquivos_flat:
        ws_a.append(
            [
                a.get("id_registro"),
                a.get("numero_contrato"),
                a.get("titulo_exibido"),
                a.get("nome_arquivo_original"),
                a.get("tipo_mime"),
                a.get("url_arquivo"),
                a.get("nome_salvo"),
                a.get("caminho_local"),
                a.get("status_download"),
                a.get("tamanho_bytes"),
            ]
        )
    _estilo_cabecalho(ws_a)
    _ajustar_larguras(ws_a)

    # --- Resumo ---
    ws_r = wb.create_sheet("Resumo")
    ws_r.append(["campo", "valor"])
    for campo, valor in resumo_linhas:
        ws_r.append([campo, valor])
    _estilo_cabecalho(ws_r)
    _ajustar_larguras(ws_r)

    wb.save(destino)
    logger.info("Planilha gerada: %s", destino)
    return destino
